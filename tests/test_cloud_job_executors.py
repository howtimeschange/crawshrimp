import tempfile
import sys
import unittest
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

import openpyxl

from core import data_sink
from core.cloud_approval_client import CloudApprovalError
from core.cloud_machine_agent import CloudMachineAgent


class FakeCloudClient:
    def __init__(self, responses=None, *, allowed_asset_uids=None):
        self.responses = list(responses or [])
        self.calls = []
        self.uploads = []
        self.downloads = []
        self.machine_token = "machine-secret"
        self.allowed_asset_uids = set(allowed_asset_uids or [])

    def request_json(self, method, path, body=None, *, token_type="machine"):
        self.calls.append({"method": method, "path": path, "body": dict(body or {})})
        if self.responses:
            response = self.responses.pop(0)
            if isinstance(response, Exception):
                raise response
            return response
        if path == "/api/assets/presign":
            if self.allowed_asset_uids and body.get("asset_uid") not in self.allowed_asset_uids:
                raise CloudApprovalError(f"asset {body.get('asset_uid')} is not leased for upload")
            return {"upload_url": f"/upload/{body.get('asset_uid')}", "object_key": f"objects/{body.get('asset_uid')}"}
        if path == "/api/material-test/import":
            return {
                "overview_rows": len(body.get("overview_rows") or []),
                "detail_rows": len(body.get("detail_rows") or []),
                "inserted_or_updated": len(body.get("overview_rows") or []) + len(body.get("detail_rows") or []),
                "source_uid": (body.get("source") or {}).get("source_uid", ""),
            }
        return {"ok": True}

    def upload_asset(self, upload_url, path, content_type, *, token_type="machine"):
        self.uploads.append({"upload_url": upload_url, "path": Path(path), "content_type": content_type})
        return {"ok": True}

    def download_asset(self, asset_uid, target_path, *, job_uid="", lease_id=""):
        target = Path(target_path)
        target.write_bytes(f"asset:{asset_uid}".encode("utf-8"))
        self.downloads.append({"asset_uid": asset_uid, "path": target, "job_uid": job_uid, "lease_id": lease_id})
        return target


class CloudJobExecutorTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self.tmp.cleanup)
        patcher = patch("core.runtime_paths.data_root", return_value=Path(self.tmp.name))
        patcher.start()
        self.addCleanup(patcher.stop)
        data_sink.init_db()

    def test_load_tmall_module_registers_module_for_dataclass_annotations(self):
        from core.cloud_job_executors import _load_tmall_module

        sys.modules.pop("cloud_tmall_ai_image_test_chain", None)

        module = _load_tmall_module()

        self.assertIs(sys.modules.get("cloud_tmall_ai_image_test_chain"), module)
        workflow = module.WorkflowItem(
            row_no=2,
            style_code="208326100202",
            item_id="1002178235142",
            category="长袖T恤",
            gender="中性",
        )
        self.assertEqual(workflow.style_code, "208326100202")

    def test_regenerate_ai_image_downloads_references_uploads_new_asset_and_completes_job(self):
        from core.cloud_job_executors import CloudJobExecutor

        captured = {}

        def regenerate(batch, asset_id, *, prompt="", reference_paths=None):
            captured["batch"] = batch
            captured["asset_id"] = asset_id
            captured["prompt"] = prompt
            captured["reference_paths"] = list(reference_paths or [])
            output = Path(batch["artifact_dir"]) / "regenerated.jpg"
            output.write_bytes(b"new image")
            return {
                "id": "cloud-ai-1",
                "kind": "ai",
                "path": str(output),
                "filename": "regenerated.jpg",
                "prompt": prompt,
                "generation_row": {"任务ID": "1xm-1"},
            }

        client = FakeCloudClient()
        executor = CloudJobExecutor(client, Path(self.tmp.name) / "cloud-jobs", tmall_module=SimpleNamespace(
            regenerate_approval_asset=regenerate,
        ))
        job = {
            "job_uid": "job-1",
            "batch_uid": "batch-1",
            "job_type": "regenerate_ai_image",
            "lease_id": "lease-1",
            "payload": {
                "batch_uid": "batch-1",
                "style_id": 101,
                "asset_uid": "regen-job-1",
                "rejected_asset_uid": "old-ai-1",
                "prompt_text": "new prompt",
                "reference_asset_uids": ["source-1", "ref-1"],
                "parent_asset_uid": "old-ai-1",
                "generation_row": {"完整Prompt": "old prompt"},
            },
        }

        result = executor.execute(job)

        self.assertEqual(result["status"], "succeeded")
        self.assertEqual(client.downloads[0]["asset_uid"], "source-1")
        self.assertEqual(client.downloads[1]["asset_uid"], "ref-1")
        self.assertEqual(captured["asset_id"], "old-ai-1")
        self.assertEqual(captured["prompt"], "new prompt")
        self.assertTrue(all(Path(path).is_relative_to(Path(self.tmp.name)) for path in captured["reference_paths"]))
        self.assertTrue(Path(captured["batch"]["artifact_dir"]).is_relative_to(Path(self.tmp.name) / "cloud-jobs" / "job-1"))
        self.assertEqual(client.uploads[0]["path"].name, "regenerated.jpg")
        presign_call = next(call for call in client.calls if call["path"] == "/api/assets/presign")
        self.assertEqual(presign_call["body"]["asset_uid"], "regen-job-1")
        self.assertEqual(presign_call["body"]["parent_asset_uid"], "old-ai-1")
        progress_bodies = [call["body"] for call in client.calls if call["path"].endswith("/progress")]
        renew_bodies = [call["body"] for call in client.calls if call["path"].endswith("/renew")]
        self.assertTrue(progress_bodies)
        self.assertTrue(renew_bodies)
        self.assertTrue(all(body["lease_id"] == "lease-1" for body in progress_bodies + renew_bodies))
        self.assertEqual(result["result"]["asset_uid"], "regen-job-1")

    def test_generate_ai_image_downloads_source_and_references_uploads_allowed_result_assets(self):
        from core.cloud_job_executors import CloudJobExecutor

        captured = {}

        def generate(batch, *, item_id="", style_code="", prompt="", main_image_path="", reference_paths=None):
            captured["batch"] = batch
            captured["item_id"] = item_id
            captured["style_code"] = style_code
            captured["prompt"] = prompt
            captured["main_image_path"] = main_image_path
            captured["reference_paths"] = list(reference_paths or [])
            output = Path(batch["artifact_dir"]) / "generated.jpg"
            output.write_bytes(b"new generated image")
            second = Path(batch["artifact_dir"]) / "generated-2.jpg"
            second.write_bytes(b"second generated image")
            return {
                "id": "generated-local-id",
                "kind": "ai",
                "path": str(output),
                "filename": "generated.jpg",
                "prompt": prompt,
                "generation_row": {"1XM任务ID": "1xm-new", "本地生成图文件": f"{output}\n{second}"},
            }

        client = FakeCloudClient(allowed_asset_uids={"cloud-result-1", "cloud-result-2"})
        executor = CloudJobExecutor(client, Path(self.tmp.name) / "cloud-jobs", tmall_module=SimpleNamespace(
            generate_approval_asset_for_item=generate,
        ))
        job = {
            "job_uid": "job-generate",
            "batch_uid": "batch-1",
            "job_type": "generate_ai_image",
            "lease_id": "lease-generate",
            "payload": {
                "request_uid": "gen-1",
                "batch_uid": "batch-1",
                "style_id": 101,
                "style_code": "208326100202",
                "item_id": "1001",
                "source_asset_uid": "source-1",
                "reference_asset_uids": ["ref-1", "ref-2"],
                "result_asset_uids": ["cloud-result-1", "cloud-result-2"],
                "prompt_template_version_id": 31,
                "prompt_text": "fresh prompt",
                "model": "gemini-3-pro-image-preview",
                "size": "2048x2048",
                "quality": "high",
                "output_format": "webp",
                "count": 2,
            },
        }

        result = executor.execute(job)

        self.assertEqual(result["status"], "succeeded")
        self.assertEqual([download["asset_uid"] for download in client.downloads], ["source-1", "ref-1", "ref-2"])
        self.assertEqual(captured["style_code"], "208326100202")
        self.assertEqual(captured["item_id"], "1001")
        self.assertEqual(captured["prompt"], "fresh prompt")
        self.assertEqual(Path(captured["main_image_path"]).name, "source-1.jpg")
        self.assertEqual([Path(path).name for path in captured["reference_paths"]], ["ref-1.jpg", "ref-2.jpg"])
        self.assertEqual(captured["batch"]["run_params"]["model"], "gemini-3-pro-image-preview")
        self.assertEqual(captured["batch"]["run_params"]["image_size"], "2048x2048")
        self.assertEqual(captured["batch"]["run_params"]["quality"], "high")
        self.assertEqual(captured["batch"]["run_params"]["output_format"], "webp")
        self.assertEqual(captured["batch"]["run_params"]["ai_image_count"], 2)
        presign_calls = [call for call in client.calls if call["path"] == "/api/assets/presign"]
        self.assertEqual([call["body"]["asset_uid"] for call in presign_calls], ["cloud-result-1", "cloud-result-2"])
        presign_call = presign_calls[0]
        self.assertEqual(presign_call["body"]["batch_uid"], "batch-1")
        self.assertEqual(presign_call["body"]["style_id"], 101)
        self.assertEqual(presign_call["body"]["prompt_template_version_id"], 31)
        self.assertEqual(presign_call["body"]["prompt_text"], "fresh prompt")
        self.assertEqual(presign_call["body"]["parent_asset_uid"], "source-1")
        self.assertEqual([asset["asset_uid"] for asset in result["result"]["assets"]], ["cloud-result-1", "cloud-result-2"])
        self.assertEqual(result["result"]["filename"], "generated.jpg")
        self.assertNotIn(str(Path(self.tmp.name)), str(result["result"]))

    def test_generate_ai_image_fails_when_requested_outputs_are_missing(self):
        from core.cloud_job_executors import CloudJobExecutor, CloudJobTerminalFailure

        def generate(batch, **_kwargs):
            output = Path(batch["artifact_dir"]) / "generated-only-one.jpg"
            output.write_bytes(b"one generated image")
            return {
                "id": "generated-local-id",
                "kind": "ai",
                "path": str(output),
                "filename": "generated-only-one.jpg",
                "prompt": "fresh prompt",
                "generation_row": {"1XM任务ID": "1xm-new"},
            }

        client = FakeCloudClient(allowed_asset_uids={"cloud-result-1", "cloud-result-2"})
        executor = CloudJobExecutor(client, Path(self.tmp.name) / "cloud-jobs", tmall_module=SimpleNamespace(
            generate_approval_asset_for_item=generate,
        ))
        job = {
            "job_uid": "job-generate-short",
            "batch_uid": "batch-1",
            "job_type": "generate_ai_image",
            "lease_id": "lease-generate",
            "payload": {
                "request_uid": "gen-1",
                "batch_uid": "batch-1",
                "style_id": 101,
                "style_code": "208326100202",
                "item_id": "1001",
                "source_asset_uid": "source-1",
                "result_asset_uids": ["cloud-result-1", "cloud-result-2"],
                "prompt_text": "fresh prompt",
                "count": 2,
            },
        }

        with self.assertRaises(CloudJobTerminalFailure) as raised:
            executor.execute(job)

        self.assertIn("expected 2 generated output", str(raised.exception))
        self.assertEqual([call["path"] for call in client.calls if call["path"] == "/api/assets/presign"], [])
        self.assertEqual(client.uploads, [])

    def test_submit_tmall_material_test_downloads_source_and_approved_images_and_completes_with_result_path(self):
        from core.cloud_job_executors import CloudJobExecutor

        captured = {}

        async def upload_batch(batch):
            captured["batch"] = batch
            result_path = Path(batch["artifact_dir"]) / "submit-result.json"
            result_path.write_text("{}", encoding="utf-8")
            return {"ok": True, "status": "created", "result_path": str(result_path), "submitted": 1}

        client = FakeCloudClient()
        executor = CloudJobExecutor(client, Path(self.tmp.name) / "cloud-jobs", tmall_module=SimpleNamespace(
            upload_approved_tmall_batch=upload_batch,
        ))
        job = {
            "job_uid": "job-submit",
            "batch_uid": "batch-submit",
            "job_type": "submit_tmall_material_test",
            "lease_id": "lease-submit",
            "payload": {
                "submit_plan": {
                    "batch_uid": "batch-submit",
                    "styles": [{"id": 7, "style_uid": "style-7", "style_code": "208326100202", "item_id": "1001"}],
                    "assets": [
                        {
                            "asset_uid": "source-main-1",
                            "style_id": 7,
                            "kind": "source",
                            "status": "uploaded",
                            "filename": "source.jpg",
                        },
                        {
                            "asset_uid": "ai-approved-1",
                            "style_id": 7,
                            "kind": "ai",
                            "status": "approved",
                            "filename": "ai.jpg",
                            "prompt_text": "prompt",
                            "meta": {"prompt_group": "front"},
                        },
                    ],
                },
            },
        }

        result = executor.execute(job)

        self.assertEqual(result["status"], "succeeded")
        self.assertEqual(result["result"]["result_path"], str(Path(self.tmp.name) / "cloud-jobs" / "job-submit" / "submit-result.json"))
        self.assertEqual([download["asset_uid"] for download in client.downloads], ["ai-approved-1", "source-main-1"])
        item = captured["batch"]["items"][0]
        self.assertEqual(item["style_code"], "208326100202")
        self.assertTrue(Path(item["origin_path"]).is_relative_to(Path(self.tmp.name) / "cloud-jobs" / "job-submit"))
        self.assertEqual(Path(item["origin_path"]).name, "source-main-1.jpg")
        self.assertEqual(item["assets"][0]["status"], "approved")
        self.assertTrue(Path(item["assets"][0]["path"]).is_relative_to(Path(self.tmp.name) / "cloud-jobs" / "job-submit"))
        self.assertTrue(any(call["path"] == "/api/jobs/job-submit/progress" for call in client.calls))

    def test_submit_tmall_material_test_cloud_batch_is_accepted_by_legacy_upload_planner(self):
        from core.cloud_job_executors import CloudJobExecutor, _load_tmall_module

        client = FakeCloudClient()
        executor = CloudJobExecutor(client, Path(self.tmp.name) / "cloud-jobs")
        job = {
            "job_uid": "job-submit-cloud",
            "batch_uid": "batch-submit",
            "job_type": "submit_tmall_material_test",
            "lease_id": "lease-submit",
            "payload": {},
        }
        submit_plan = {
            "batch_uid": "batch-submit",
            "styles": [{"id": 7, "style_uid": "style-7", "style_code": "208326100202", "item_id": "1001"}],
            "assets": [
                {"asset_uid": "source-main-1", "style_id": 7, "kind": "source", "status": "uploaded"},
                {
                    "asset_uid": "ai-approved-1",
                    "style_id": 7,
                    "kind": "ai",
                    "status": "approved",
                    "meta": {"generation": {"__task_run_uid": "original-local-run"}},
                },
            ],
        }

        batch = executor._submit_batch(job, submit_plan, Path(self.tmp.name) / "cloud-jobs" / "job-submit-cloud")
        upload_plan = _load_tmall_module().selected_upload_plan_from_approval_batch(batch)

        self.assertNotIn("task_run_uid", batch)
        self.assertEqual(len(upload_plan), 1)
        self.assertEqual([Path(path).name for path in upload_plan[0]["generated_paths"]], ["ai-approved-1.jpg"])

    def test_submit_tmall_material_test_prefers_source_when_reference_precedes_source(self):
        from core.cloud_job_executors import CloudJobExecutor

        captured = {}

        async def upload_batch(batch):
            captured["batch"] = batch
            result_path = Path(batch["artifact_dir"]) / "submit-result.json"
            result_path.write_text("{}", encoding="utf-8")
            return {"ok": True, "status": "created", "result_path": str(result_path), "submitted": 1}

        client = FakeCloudClient()
        executor = CloudJobExecutor(client, Path(self.tmp.name) / "cloud-jobs", tmall_module=SimpleNamespace(
            upload_approved_tmall_batch=upload_batch,
        ))
        job = {
            "job_uid": "job-submit-source-order",
            "batch_uid": "batch-submit",
            "job_type": "submit_tmall_material_test",
            "lease_id": "lease-submit",
            "payload": {
                "submit_plan": {
                    "batch_uid": "batch-submit",
                    "styles": [{"id": 7, "style_uid": "style-7"}],
                    "assets": [
                        {"asset_uid": "reference-1", "style_id": 7, "kind": "reference", "status": "uploaded"},
                        {"asset_uid": "source-main-1", "style_id": 7, "kind": "source", "status": "uploaded"},
                        {"asset_uid": "ai-approved-1", "style_id": 7, "kind": "ai", "status": "approved"},
                    ],
                },
            },
        }

        executor.execute(job)

        item = captured["batch"]["items"][0]
        self.assertEqual(Path(item["origin_path"]).name, "source-main-1.jpg")
        self.assertNotIn("reference-1", [download["asset_uid"] for download in client.downloads])

    def test_submit_tmall_material_test_raises_when_approved_style_has_no_source(self):
        from core.cloud_job_executors import CloudJobExecutor

        async def upload_batch(_batch):
            self.fail("submit should not run without a source origin")

        client = FakeCloudClient()
        executor = CloudJobExecutor(client, Path(self.tmp.name) / "cloud-jobs", tmall_module=SimpleNamespace(
            upload_approved_tmall_batch=upload_batch,
        ))
        job = {
            "job_uid": "job-submit-missing-source",
            "batch_uid": "batch-submit",
            "job_type": "submit_tmall_material_test",
            "lease_id": "lease-submit",
            "payload": {
                "submit_plan": {
                    "batch_uid": "batch-submit",
                    "styles": [{"id": 7, "style_uid": "style-7"}],
                    "assets": [
                        {"asset_uid": "reference-1", "style_id": 7, "kind": "reference", "status": "uploaded"},
                        {"asset_uid": "ai-approved-1", "style_id": 7, "kind": "ai", "status": "approved"},
                    ],
                },
            },
        }

        with self.assertRaises(ValueError) as raised:
            executor.execute(job)

        self.assertIn("missing a source asset", str(raised.exception))

    def test_submit_tmall_material_test_treats_real_uploader_create_failed_result_as_terminal_failure(self):
        from core.cloud_job_executors import CloudJobExecutor, CloudJobTerminalFailure

        async def upload_batch(_batch):
            return {"ok": False, "status": "create_failed", "attempted": 1, "failed": 1}

        client = FakeCloudClient()
        executor = CloudJobExecutor(client, Path(self.tmp.name) / "cloud-jobs", tmall_module=SimpleNamespace(
            upload_approved_tmall_batch=upload_batch,
        ))
        job = {
            "job_uid": "job-submit-failed",
            "batch_uid": "batch-submit",
            "job_type": "submit_tmall_material_test",
            "lease_id": "lease-submit",
            "payload": {
                "submit_plan": {
                    "batch_uid": "batch-submit",
                    "styles": [{"id": 7, "style_uid": "style-7"}],
                    "assets": [
                        {"asset_uid": "source-main-1", "style_id": 7, "kind": "source", "status": "uploaded"},
                        {"asset_uid": "ai-approved-1", "style_id": 7, "kind": "ai", "status": "approved"},
                    ],
                },
            },
        }

        with self.assertRaises(CloudJobTerminalFailure) as raised:
            executor.execute(job)

        self.assertIn("create_failed", str(raised.exception))
        self.assertIn("failed=1", str(raised.exception))

    def test_crawl_tmall_material_test_data_uploads_workbook_and_imports_rows(self):
        from core.cloud_job_executors import CloudJobExecutor

        captured = {}

        def runner(run_params, task_dir, job):
            captured["run_params"] = dict(run_params)
            captured["task_dir"] = task_dir
            captured["job_uid"] = job["job_uid"]
            path = Path(run_params["output_dir"]) / "material-export.xlsx"
            workbook = openpyxl.Workbook()
            overview = workbook.active
            overview.title = "概览"
            overview.append(["记录类型", "表格行号", "款号", "商品ID", "商品标题", "任务ID", "测试状态", "测试渠道", "测试素材数", "最优素材", "执行结果", "备注"])
            overview.append(["概览", 2, "208326", "1001", "标题", "T1", "完成", "搜索", 2, "M1", "成功", ""])
            detail = workbook.create_sheet("明细")
            detail.append(["记录类型", "表格行号", "款号", "商品ID", "商品标题", "任务ID", "测试状态", "测试渠道", "测试素材数", "统计口径", "统计日期", "图片类型", "素材ID", "素材比例", "素材占比", "素材URL", "搜索曝光", "搜索点击", "搜索点击率", "详情曝光", "详情点击", "详情点击率", "详情加购", "详情支付转化", "详情支付转化率", "数据下载链接", "执行结果", "备注"])
            for index in range(1001):
                detail.append(["明细", index + 2, "208326", f"100{index}", "标题", "T1", "完成", "搜索", 2, "ACCUMULATE_30_DAYS", "2026-07-01", "主图", f"M{index}", "1:1", "7.79%", f"https://img.test/{index}.jpg", 1000, 77, "7.79%", 500, 40, "8%", 12, 3, "2.50%", "", "成功", ""])
            workbook.save(path)
            return {"workbook_path": str(path)}

        client = FakeCloudClient()
        executor = CloudJobExecutor(client, Path(self.tmp.name) / "cloud-jobs", material_test_runner=runner)
        job = {
            "job_uid": "job-crawl",
            "batch_uid": "material-test",
            "job_type": "crawl_tmall_material_test_data",
            "lease_id": "lease-crawl",
            "payload": {"run_params": {"statistic_type": "ACCUMULATE_30_DAYS"}},
        }

        result = executor.execute(job)

        self.assertEqual(result["status"], "succeeded")
        self.assertEqual(result["result"]["overview_rows"], 1)
        self.assertEqual(result["result"]["detail_rows"], 1001)
        self.assertEqual(captured["run_params"]["output_dir"], str(Path(self.tmp.name) / "cloud-jobs" / "job-crawl" / "exports"))
        self.assertEqual(client.uploads[0]["path"].name, "material-export.xlsx")
        presign_call = next(call for call in client.calls if call["path"] == "/api/assets/presign")
        self.assertEqual(presign_call["body"]["batch_uid"], "material-test")
        self.assertEqual(presign_call["body"]["kind"], "result")
        import_calls = [call for call in client.calls if call["path"] == "/api/material-test/import"]
        self.assertEqual(len(import_calls), 2)
        self.assertEqual(len(import_calls[0]["body"]["overview_rows"]), 1)
        self.assertEqual(len(import_calls[0]["body"]["detail_rows"]), 1000)
        self.assertEqual(len(import_calls[1]["body"]["overview_rows"]), 0)
        self.assertEqual(len(import_calls[1]["body"]["detail_rows"]), 1)
        self.assertTrue(all(call["body"]["job_uid"] == "job-crawl" for call in import_calls))
        self.assertTrue(all(call["body"]["lease_id"] == "lease-crawl" for call in import_calls))
        self.assertEqual({call["body"]["source"]["source_uid"] for call in import_calls}, {"gen-material-test-job-crawl"})
        self.assertNotIn("workbook_path", result["result"])

    def test_crawl_tmall_material_test_data_finds_nested_export_workbook(self):
        from core.cloud_job_executors import CloudJobExecutor

        def runner(run_params, _task_dir, _job):
            table_dir = Path(run_params["output_dir"]) / "数据表格"
            table_dir.mkdir(parents=True, exist_ok=True)
            path = table_dir / "巴拉-AI测图数据抓取导出.xlsx"
            workbook = openpyxl.Workbook()
            overview = workbook.active
            overview.title = "概览"
            overview.append(["记录类型", "表格行号", "款号", "商品ID", "商品标题", "任务ID", "测试状态", "测试渠道", "测试素材数", "最优素材", "执行结果", "备注"])
            overview.append(["概览", 2, "208326", "1001", "标题", "T1", "完成", "搜索", 2, "M1", "成功", ""])
            detail = workbook.create_sheet("明细")
            detail.append(["记录类型", "表格行号", "款号", "商品ID", "商品标题", "任务ID", "测试状态", "测试渠道", "测试素材数", "统计口径", "统计日期", "图片类型", "素材ID", "素材比例", "素材占比", "素材URL", "搜索曝光", "搜索点击", "搜索点击率", "详情曝光", "详情点击", "详情点击率", "详情加购", "详情支付转化", "详情支付转化率", "数据下载链接", "执行结果", "备注"])
            detail.append(["明细", 2, "208326", "1001", "标题", "T1", "完成", "搜索", 2, "ACCUMULATE_30_DAYS", "2026-07-01", "主图", "M1", "1:1", "7.79%", "https://img.test/1.jpg", 1000, 77, "7.79%", 500, 40, "8%", 12, 3, "2.50%", "", "成功", ""])
            workbook.save(path)
            return {"status": "ok"}

        client = FakeCloudClient()
        executor = CloudJobExecutor(client, Path(self.tmp.name) / "cloud-jobs", material_test_runner=runner)
        result = executor.execute({
            "job_uid": "job-crawl-nested",
            "batch_uid": "material-test",
            "job_type": "crawl_tmall_material_test_data",
            "lease_id": "lease-crawl",
            "payload": {"run_params": {}},
        })

        self.assertEqual(result["status"], "succeeded")
        self.assertEqual(result["result"]["overview_rows"], 1)
        self.assertEqual(result["result"]["detail_rows"], 1)
        self.assertEqual(client.uploads[0]["path"].parent.name, "数据表格")

    def test_stale_lease_completion_is_not_retried_as_success(self):
        data_sink.save_cloud_machine_credentials("machine-1", "machine-secret", "任务机", ["regenerate_ai_image"])
        client = FakeCloudClient([
            {
                "job": {
                    "job_uid": "job-stale",
                    "batch_uid": "batch-1",
                    "job_type": "regenerate_ai_image",
                    "lease_id": "lease-old",
                    "payload": {"asset_uid": "ai-1", "generation_row": {"完整Prompt": "prompt"}},
                },
                "next_poll_after_seconds": 0,
            },
            CloudApprovalError("cloud request failed: HTTP 403; Stale lease"),
        ])
        executor = SimpleNamespace(execute=lambda _job: {"status": "succeeded", "result": {"ok": True}})
        agent = CloudMachineAgent(client, job_executor_factory=lambda _client: executor)

        result = agent.claim_once()

        self.assertEqual(result["job_result"]["status"], "stale_lease")
        complete_calls = [call for call in client.calls if call["path"] == "/api/jobs/job-stale/complete"]
        fail_calls = [call for call in client.calls if call["path"] == "/api/jobs/job-stale/fail"]
        self.assertEqual(len(complete_calls), 1)
        self.assertEqual(fail_calls, [])
        self.assertEqual(complete_calls[0]["body"]["lease_id"], "lease-old")

    def test_non_stale_completion_error_is_saved_for_retry_after_executor_success(self):
        data_sink.save_cloud_machine_credentials("machine-1", "machine-secret", "任务机", ["regenerate_ai_image"])
        client = FakeCloudClient([
            {
                "job": {
                    "job_uid": "job-forbidden",
                    "batch_uid": "batch-1",
                    "job_type": "regenerate_ai_image",
                    "lease_id": "lease-current",
                    "payload": {"asset_uid": "ai-1", "generation_row": {"完整Prompt": "prompt"}},
                },
                "next_poll_after_seconds": 0,
            },
            CloudApprovalError("cloud request failed: HTTP 403; Machine is disabled"),
        ])
        executor = SimpleNamespace(execute=lambda _job: {"status": "succeeded", "result": {"ok": True}})
        agent = CloudMachineAgent(client, job_executor_factory=lambda _client: executor)

        result = agent.claim_once()

        self.assertEqual(result["job_result"]["status"], "completion_pending")
        self.assertIn("Machine is disabled", result["job_result"]["message"])
        self.assertEqual(data_sink.get_pending_cloud_job_completion("job-forbidden"), {"ok": True})
        complete_calls = [call for call in client.calls if call["path"] == "/api/jobs/job-forbidden/complete"]
        fail_calls = [call for call in client.calls if call["path"] == "/api/jobs/job-forbidden/fail"]
        self.assertEqual(len(complete_calls), 1)
        self.assertEqual(fail_calls, [])

    def test_blocked_needs_login_is_returned_when_local_browser_readiness_fails(self):
        from core.cloud_job_executors import CloudJobBlocked

        data_sink.save_cloud_machine_credentials("machine-1", "machine-secret", "任务机", ["submit_tmall_material_test"])
        client = FakeCloudClient([
            {
                "job": {
                    "job_uid": "job-login",
                    "batch_uid": "batch-1",
                    "job_type": "submit_tmall_material_test",
                    "lease_id": "lease-login",
                    "payload": {"submit_plan": {"batch_uid": "batch-1", "styles": [], "assets": []}},
                },
                "next_poll_after_seconds": 0,
            },
            {"ok": True, "status": "retryable_failed"},
        ])
        executor = SimpleNamespace(execute=lambda _job: (_ for _ in ()).throw(CloudJobBlocked("needs_login", "Chrome CDP unavailable")))
        heartbeats = []
        agent = CloudMachineAgent(client, job_executor_factory=lambda _client: executor, heartbeat_callback=heartbeats.append)

        result = agent.claim_once()

        self.assertEqual(result["job_result"]["status"], "blocked_needs_login")
        self.assertEqual(heartbeats, ["online_busy", "needs_login"])
        fail_call = next(call for call in client.calls if call["path"] == "/api/jobs/job-login/fail")
        self.assertEqual(fail_call["body"]["lease_id"], "lease-login")
        self.assertEqual(fail_call["body"]["status"], "blocked_needs_login")
        self.assertEqual(fail_call["body"]["result"]["status"], "blocked_needs_login")
        self.assertFalse(fail_call["body"].get("terminal", False))

    def test_submit_tmall_material_test_does_not_auto_retry_terminal_failure(self):
        from core.cloud_job_executors import CloudJobTerminalFailure

        data_sink.save_cloud_machine_credentials("machine-1", "machine-secret", "任务机", ["submit_tmall_material_test"])
        client = FakeCloudClient([
            {
                "job": {
                    "job_uid": "job-terminal",
                    "batch_uid": "batch-1",
                    "job_type": "submit_tmall_material_test",
                    "lease_id": "lease-terminal",
                    "payload": {"submit_plan": {"batch_uid": "batch-1", "styles": [], "assets": []}},
                },
                "next_poll_after_seconds": 0,
            },
            {"ok": True, "status": "terminal_failed"},
        ])
        executor = SimpleNamespace(execute=lambda _job: (_ for _ in ()).throw(CloudJobTerminalFailure("tmall_create_failed")))
        agent = CloudMachineAgent(client, job_executor_factory=lambda _client: executor)

        result = agent.claim_once()

        self.assertEqual(result["job_result"]["status"], "terminal_failed")
        fail_call = next(call for call in client.calls if call["path"] == "/api/jobs/job-terminal/fail")
        self.assertEqual(fail_call["body"]["lease_id"], "lease-terminal")
        self.assertTrue(fail_call["body"]["terminal"])
        self.assertEqual(len([call for call in client.calls if call["path"] == "/api/machines/jobs/claim"]), 1)


if __name__ == "__main__":
    unittest.main()
