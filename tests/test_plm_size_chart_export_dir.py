import tempfile
import unittest
from importlib.util import find_spec
from pathlib import Path
from unittest.mock import patch

import yaml

from core import data_sink
from core.api_server import _finalize_plm_size_chart_downloader_outputs


ROOT = Path(__file__).resolve().parents[1]
MANIFEST_PATH = ROOT / "adapters" / "plm-ops-assistant" / "manifest.yaml"


class PlmSizeChartExportDirTests(unittest.TestCase):
    def test_manifest_declares_optional_output_directory(self):
        manifest = yaml.safe_load(MANIFEST_PATH.read_text(encoding="utf-8"))
        task = next(item for item in manifest["tasks"] if item["id"] == "size_chart_downloader")
        params = {item["id"]: item for item in task["params"]}

        self.assertEqual(params["output_dir"]["type"], "directory")
        self.assertIn("导出目录", params["output_dir"]["label"])
        self.assertIn("最终 Excel", params["output_dir"]["hint"])

    def test_finalize_copies_exported_excel_to_selected_output_dir(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            runtime_dir = base / "runtime"
            export_dir = base / "exports"
            runtime_dir.mkdir()
            exported = runtime_dir / "PLM尺码表下载_大货.xlsx"
            exported.write_bytes(b"excel")

            result = _finalize_plm_size_chart_downloader_outputs(
                runtime_files=[],
                exported_files=[str(exported)],
                run_params={"output_dir": str(export_dir)},
                log=lambda _: None,
            )

            copied = export_dir / exported.name
            self.assertEqual(result, [str(copied)])
            self.assertTrue(copied.is_file())
            self.assertTrue(exported.is_file())

    def test_finalize_copies_runtime_subdir_excel_to_selected_parent_dir(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            selected_dir = base / "Downloads"
            runtime_dir = selected_dir / "data" / "plm-ops-assistant" / "size_chart_downloader"
            runtime_dir.mkdir(parents=True)
            exported = runtime_dir / "PLM尺码表下载_大货.xlsx"
            exported.write_bytes(b"excel")

            result = _finalize_plm_size_chart_downloader_outputs(
                runtime_files=[],
                exported_files=[str(exported)],
                run_params={"output_dir": str(selected_dir)},
                log=lambda _: None,
            )

            copied = selected_dir / exported.name
            self.assertEqual(result, [str(copied)])
            self.assertTrue(copied.is_file())
            self.assertTrue(exported.is_file())

    def test_finalize_can_copy_runtime_files_when_exported_files_are_empty(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            runtime_dir = base / "runtime"
            export_dir = base / "exports"
            runtime_dir.mkdir()
            exported = runtime_dir / "PLM尺码表下载_大货.xlsx"
            exported.write_bytes(b"excel")

            result = _finalize_plm_size_chart_downloader_outputs(
                runtime_files=[str(exported)],
                exported_files=[],
                run_params={"output_dir": str(export_dir)},
                log=lambda _: None,
            )

            copied = export_dir / exported.name
            self.assertEqual(result, [str(copied)])
            self.assertTrue(copied.is_file())
            self.assertTrue(exported.is_file())

    def test_finalize_keeps_runtime_export_when_output_dir_is_blank(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            exported = Path(tmpdir) / "PLM尺码表下载_大货.xlsx"
            exported.write_bytes(b"excel")

            result = _finalize_plm_size_chart_downloader_outputs(
                runtime_files=[],
                exported_files=[str(exported)],
                run_params={},
                log=lambda _: None,
            )

            self.assertEqual(result, [str(exported)])

    def test_finalize_uses_unique_name_when_selected_dir_has_existing_file(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            runtime_dir = base / "runtime"
            export_dir = base / "exports"
            runtime_dir.mkdir()
            export_dir.mkdir()
            exported = runtime_dir / "PLM尺码表下载_大货.xlsx"
            exported.write_bytes(b"excel")
            existing = export_dir / exported.name
            existing.write_bytes(b"old")

            result = _finalize_plm_size_chart_downloader_outputs(
                runtime_files=[],
                exported_files=[str(exported)],
                run_params={"output_dir": str(export_dir)},
                log=lambda _: None,
            )

            copied = export_dir / "PLM尺码表下载_大货_2.xlsx"
            self.assertEqual(result, [str(copied)])
            self.assertEqual(existing.read_bytes(), b"old")
            self.assertEqual(copied.read_bytes(), b"excel")

    def test_export_excel_splits_wide_and_long_rows_into_configured_sheets(self):
        if find_spec("openpyxl") is None:
            self.skipTest("openpyxl is required for export_excel in this environment")

        from openpyxl import load_workbook

        manifest = yaml.safe_load(MANIFEST_PATH.read_text(encoding="utf-8"))
        task = next(item for item in manifest["tasks"] if item["id"] == "size_chart_downloader")
        output = next(item for item in task["output"] if item["type"] == "excel")
        rows = [
            {
                "__sheet_name": "宽表",
                "款号": "201326105103",
                "尺码表阶段": "大货",
                "尺码表名称": "326FY-BS-507大货_尺寸表",
                "修订版": "7",
                "测量点": "衣长",
                "测量点编码": "101",
                "80/": "35.5",
                "90/": "38.0",
            },
            {
                "__sheet_name": "长表",
                "款号": "201326105103",
                "尺码表阶段": "大货",
                "尺码表名称": "326FY-BS-507大货_尺寸表",
                "修订版": "7",
                "测量点": "衣长",
                "测量点编码": "101",
                "尺码": "80/",
                "尺码值": "35.5",
            },
            {
                "__sheet_name": "长表",
                "款号": "201326105103",
                "尺码表阶段": "大货",
                "尺码表名称": "326FY-BS-507大货_尺寸表",
                "修订版": "7",
                "测量点": "衣长",
                "测量点编码": "101",
                "尺码": "90/",
                "尺码值": "38.0",
            },
        ]

        with tempfile.TemporaryDirectory() as tmpdir:
            with patch.dict("os.environ", {"CRAWSHRIMP_DATA": tmpdir}, clear=False):
                out_path = data_sink.export_excel(
                    rows,
                    adapter_id="plm-ops-assistant",
                    task_id="size_chart_downloader",
                    filename_template="test.xlsx",
                    sheet_key=output["sheet_key"],
                    sheet_configs=output["sheets"],
                )

            wb = load_workbook(out_path, read_only=True, data_only=True)
            self.assertEqual(wb.sheetnames, ["宽表", "长表"])

            wide_headers = [cell.value for cell in next(wb["宽表"].iter_rows(min_row=1, max_row=1))]
            long_headers = [cell.value for cell in next(wb["长表"].iter_rows(min_row=1, max_row=1))]
            self.assertEqual(wb["宽表"].max_row, 2)
            self.assertEqual(wb["长表"].max_row, 3)
            self.assertIn("80/", wide_headers)
            self.assertIn("90/", wide_headers)
            self.assertNotIn("80/", long_headers)
            self.assertNotIn("90/", long_headers)
            self.assertIn("尺码", long_headers)
            self.assertIn("尺码值", long_headers)


if __name__ == "__main__":
    unittest.main()
