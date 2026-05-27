import unittest
from importlib.util import find_spec
from pathlib import Path


class DasenShowcaseTemplateTest(unittest.TestCase):
    def test_batch_create_template_has_current_dropdowns(self):
        if find_spec("openpyxl") is None:
            self.skipTest("openpyxl is required to inspect Excel data validation")

        from openpyxl import load_workbook

        path = Path("adapters/dasen-ops-assistant/templates/showcase-batch-template.xlsx")
        workbook = load_workbook(path)

        self.assertIn("案例批量导入", workbook.sheetnames)
        self.assertIn("下拉选项", workbook.sheetnames)

        sheet = workbook["案例批量导入"]
        headers = [cell.value for cell in sheet[1]]
        self.assertEqual(headers[2], "案例类型")
        self.assertEqual(headers[5], "所属AI纵队")

        validations = list(sheet.data_validations.dataValidation)
        ranges = {str(validation.sqref): validation.formula1 for validation in validations}
        self.assertIn("C2:C500", ranges)
        self.assertIn("F2:F500", ranges)

        options = workbook["下拉选项"]
        categories = [options[f"A{row}"].value for row in range(2, 13)]
        ai_teams = [options[f"B{row}"].value for row in range(2, 18)]

        self.assertIn("PL6 销售", categories)
        self.assertIn("数字中心", ai_teams)
        self.assertIn("供应链", ai_teams)
        self.assertIn("人资中心", ai_teams)
        self.assertNotIn("其他", ai_teams)


if __name__ == "__main__":
    unittest.main()
