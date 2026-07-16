import tempfile
import unittest
import asyncio
import json
from types import SimpleNamespace
from pathlib import Path
from unittest.mock import patch

import yaml
from PIL import Image

from core import api_server, data_sink, runtime_paths
from core.api_server import (
    _apply_bala_ai_face_background_generate,
    _cleanup_orphaned_runtime_artifacts,
    _finalize_bala_ai_video_assistant_outputs,
    _load_bala_video_template_catalog,
    _parse_seedance_cli_json_objects,
    _load_bala_model_library,
    _bala_models_for_generation,
)

ROOT = Path(__file__).resolve().parents[1]
MANIFEST_PATH = ROOT / "adapters" / "bala-ai-video-assistant" / "manifest.yaml"


class BalaAiVideoAssistantPackagingTests(unittest.TestCase):
    def test_happyhorse_cli_is_packaged_as_a_shared_integration(self):
        integration_root = ROOT / "integrations" / "bailianCLI"
        required_files = {
            "package.json",
            ".env.example",
            ".gitignore",
            "README.md",
            "bin/bailian.js",
            "src/bailian-client.js",
            "src/config.js",
            "examples/happyhorse-t2v.json",
            "examples/happyhorse-i2v.json",
            "examples/happyhorse-r2v.json",
            "test/bailian-client.test.js",
        }

        missing = sorted(
            relative_path
            for relative_path in required_files
            if not (integration_root / relative_path).is_file()
        )

        self.assertEqual(missing, [])

    def test_desktop_bundle_copies_shared_video_integrations_and_injects_node_runtime(self):
        build_config = (ROOT / "app" / "build.yml").read_text(encoding="utf-8")
        main_source = (ROOT / "app" / "src" / "main.js").read_text(encoding="utf-8")

        self.assertIn('from: "../integrations"', build_config)
        self.assertIn("to: python-scripts/integrations", build_config)
        self.assertIn("CRAWSHRIMP_NODE_EXECUTABLE: process.execPath", main_source)

    def test_video_provider_node_runtime_prefers_bundled_executable(self):
        resolver = getattr(api_server, "_bala_video_node_runtime", None)
        self.assertTrue(callable(resolver), "Video providers need a packaged Node runtime resolver")

        with patch.dict("os.environ", {"CRAWSHRIMP_NODE_EXECUTABLE": "/Applications/Crawshrimp"}, clear=False):
            executable, extra_env = resolver()

        self.assertEqual(executable, "/Applications/Crawshrimp")
        self.assertEqual(extra_env["ELECTRON_RUN_AS_NODE"], "1")

    def test_manifest_declares_material_prepare_task(self):
        manifest = yaml.safe_load(MANIFEST_PATH.read_text(encoding="utf-8"))

        self.assertEqual(manifest["id"], "bala-ai-video-assistant")
        task = next(item for item in manifest["tasks"] if item["id"] == "semir_video_material_prepare")

        self.assertEqual(task["script"], "semir-video-material-prepare.js")
        self.assertEqual(task["skip_auth"], False)
        params = {item["id"]: item for item in task["params"]}
        self.assertEqual(params["item_codes"]["type"], "textarea")
        self.assertEqual(params["cloud_path"]["default"], "巴拉营运BU-商品//巴拉货控/02 产品上新模块/2-2 巴拉产品上新/")
        self.assertEqual(params["max_image_mb"]["default"], 20)
        self.assertNotIn("auto_zip_package", params)

    def test_manifest_declares_face_background_generation_task(self):
        manifest = yaml.safe_load(MANIFEST_PATH.read_text(encoding="utf-8"))

        task = next(item for item in manifest["tasks"] if item["id"] == "bala_ai_face_background_generate")

        self.assertEqual(task["script"], "bala-ai-face-background-generate.js")
        self.assertEqual(task["entry_url"], "about:blank")
        self.assertEqual(task["skip_auth"], True)
        params = {item["id"]: item for item in task["params"]}
        self.assertEqual(params["source_images"]["type"], "file_images")
        self.assertEqual(params["material_root"]["include_file_listing"], True)
        self.assertEqual(params["background_prompt"]["default"], "换成马尔代夫的海边")
        self.assertEqual(params["model_key_tier"]["default"], "4k")
        self.assertIn("100女", params["model_groups"]["default"])
        output_columns = task["output"][0]["columns"]
        self.assertIn("AI任务UID", output_columns)
        self.assertIn("轮询URL", output_columns)

    def test_manifest_declares_qn_img2video_batch_task(self):
        manifest = yaml.safe_load(MANIFEST_PATH.read_text(encoding="utf-8"))

        task = next(item for item in manifest["tasks"] if item["id"] == "qn_img2video_batch")

        self.assertEqual(task["script"], "qn-img2video-batch.js")
        self.assertEqual(task["entry_url"], "https://quick.taobao.com/videostudio/img2video")
        self.assertEqual(task["skip_auth"], True)
        params = {item["id"]: item for item in task["params"]}
        self.assertEqual(params["execute_mode"]["default"], "plan")
        self.assertEqual(params["material_images"]["type"], "file_images")
        self.assertEqual(params["material_root"]["include_file_listing"], True)
        self.assertEqual(params["download_template_previews"]["default"], True)
        self.assertEqual(params["download_videos"]["default"], True)
        self.assertEqual(params["output_dir"]["type"], "directory")
        output_columns = task["output"][0]["columns"]
        self.assertIn("模板预览本地文件", output_columns)
        self.assertIn("提交任务ID", output_columns)
        self.assertIn("视频URL", output_columns)
        self.assertIn("本地视频文件", output_columns)

    def test_model_library_manifest_is_packaged_and_group_selects_standard_first(self):
        manifest_path = ROOT / "adapters" / "bala-ai-video-assistant" / "assets" / "model-library" / "manifest.json"

        with patch("core.adapter_loader.resolve_adapter_file", return_value=manifest_path):
            library, errors = _load_bala_model_library()
            models, warnings = _bala_models_for_generation({"model_groups": ["100女"], "models_per_group": 1})

        self.assertFalse(errors)
        self.assertGreaterEqual(len(library), 70)
        self.assertFalse(warnings)
        self.assertEqual(len(models), 1)
        self.assertEqual(models[0]["id"], "100女/标准.jpg")
        self.assertTrue(Path(models[0]["path"]).is_file())

    def test_template_catalog_loader_reads_local_software_manager_catalog(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            catalog = Path(tmpdir) / "template-catalog.json"
            catalog.write_text(json.dumps({
                "mainCategory": "童装/婴儿装/亲子装",
                "templates": [{
                    "templateId": "641241_62536236_21",
                    "title": "领口",
                    "localPreviewVideo": "/tmp/preview.mp4",
                    "localCoverImage": "/tmp/cover.png",
                }],
            }, ensure_ascii=False), encoding="utf-8")
            with patch("core.api_server.BALA_VIDEO_TEMPLATE_CATALOG_JSON", catalog), \
                    patch("core.api_server.BALA_VIDEO_TEMPLATE_CATALOG_CSV", Path(tmpdir) / "missing.csv"):
                payload = _load_bala_video_template_catalog()

        self.assertTrue(payload["ok"])
        self.assertEqual(payload["count"], 1)
        first = payload["templates"][0]
        self.assertIn("templateId", first)
        self.assertIn("localPreviewVideo", first)
        self.assertIn("localCoverImage", first)

    def test_seedance_cli_json_parser_handles_pretty_stdout(self):
        objects = _parse_seedance_cli_json_objects('noise\n{"id":"task-1"}\n{\n  "id": "task-1",\n  "status": "succeeded"\n}\n')

        self.assertEqual(objects[0]["id"], "task-1")
        self.assertEqual(objects[-1]["status"], "succeeded")

    def test_happyhorse_payload_builder_supports_text_image_and_reference_modes(self):
        builder = getattr(api_server, "_build_happyhorse_payload", None)
        self.assertTrue(callable(builder), "HappyHorse payload builder must be implemented")

        common = {
            "prompt": "童装模特自然转身展示服装",
            "image_paths": [],
            "resolution": "720P",
            "ratio": "3:4",
            "duration": 5,
        }
        text_payload = builder(SimpleNamespace(**common, mode="t2v", image_urls=[]))
        image_payload = builder(SimpleNamespace(
            **common,
            mode="i2v",
            image_urls=["https://example.com/approved.png"],
        ))
        reference_payload = builder(SimpleNamespace(
            **common,
            mode="r2v",
            image_urls=[
                "https://example.com/model.png",
                "https://example.com/detail.png",
            ],
        ))

        self.assertEqual(text_payload["model"], "happyhorse-1.1-t2v")
        self.assertEqual(text_payload["parameters"]["ratio"], "3:4")
        self.assertEqual(image_payload["model"], "happyhorse-1.1-i2v")
        self.assertNotIn("ratio", image_payload["parameters"])
        self.assertEqual(image_payload["input"]["media"][0]["type"], "first_frame")
        self.assertEqual(reference_payload["model"], "happyhorse-1.1-r2v")
        self.assertEqual(len(reference_payload["input"]["media"]), 2)

    def test_video_provider_payloads_reject_non_image_local_files(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            invalid_files = [
                Path(tmpdir) / "not-an-image.txt",
                Path(tmpdir) / "fake-image.png",
            ]
            invalid_files[0].write_text("not an image", encoding="utf-8")
            invalid_files[1].write_bytes(b"not a real png")

            for invalid_file in invalid_files:
                with self.subTest(path=invalid_file.name):
                    seedance = api_server.BalaSeedanceVideoRequest(
                        prompt="test",
                        image_paths=[str(invalid_file)],
                    )
                    happyhorse = SimpleNamespace(
                        mode="i2v",
                        prompt="test",
                        image_paths=[str(invalid_file)],
                        image_urls=[],
                        resolution="720P",
                        ratio="3:4",
                        duration=5,
                        watermark=False,
                    )

                    with self.assertRaisesRegex(Exception, "图片"):
                        api_server._build_seedance_payload(seedance)
                    with self.assertRaisesRegex(Exception, "图片"):
                        api_server._build_happyhorse_payload(happyhorse)

    def test_video_provider_process_communication_has_an_outer_timeout(self):
        communicator = getattr(api_server, "_communicate_video_provider_process", None)
        self.assertTrue(callable(communicator), "Video provider CLI processes need an outer timeout")

        class HangingProcess:
            def __init__(self):
                self.killed = False

            async def communicate(self):
                await asyncio.sleep(60)

            def kill(self):
                self.killed = True

            async def wait(self):
                return -9

        process = HangingProcess()
        with self.assertRaisesRegex(Exception, "超时"):
            asyncio.run(communicator(process, timeout_seconds=0.01, provider_label="测试供应商"))
        self.assertTrue(process.killed)

    def test_video_provider_status_ignores_runtime_environment_credentials(self):
        resolver = getattr(api_server, "_bala_video_provider_status", None)
        self.assertTrue(callable(resolver), "Video provider status resolver must be implemented")

        with patch.dict("os.environ", {
            "ARK_API_KEY": "runtime-placeholder-seedance",
            "DASHSCOPE_API_KEY": "runtime-placeholder-happyhorse",
        }, clear=False), patch("core.api_server.load_config", return_value={"ai": {"video": {}}}):
            status = resolver()

        self.assertFalse(status["seedance"]["configured"])
        self.assertFalse(status["happyhorse"]["configured"])
        self.assertEqual(status["seedance"]["source"], "未配置")
        self.assertEqual(status["happyhorse"]["source"], "未配置")
        self.assertNotIn("runtime-placeholder", json.dumps(status))

    def test_ai_capability_credentials_override_environment_fallbacks(self):
        config = {
            "ai": {
                "1xm": {"gpt_image_4k_key": "config-placeholder-image"},
                "video": {
                    "seedance_api_key": "config-placeholder-seedance",
                    "bailian_api_key": "config-placeholder-happyhorse",
                },
            },
        }
        with patch.dict("os.environ", {
            "ONE_XM_GPT_IMAGE_4K_KEY": "runtime-placeholder-image",
            "ARK_API_KEY": "runtime-placeholder-seedance",
            "DASHSCOPE_API_KEY": "runtime-placeholder-happyhorse",
        }, clear=False), patch("core.api_server.load_config", return_value=config):
            secrets = api_server._bala_video_provider_secrets()
            status = api_server._bala_video_provider_status()
            one_xm = api_server._resolve_one_xm_settings()

        self.assertEqual(secrets["seedance"], "config-placeholder-seedance")
        self.assertEqual(secrets["happyhorse"], "config-placeholder-happyhorse")
        self.assertEqual(status["seedance"]["source"], "AI 能力")
        self.assertEqual(status["happyhorse"]["source"], "AI 能力")
        self.assertEqual(one_xm["4k"], "config-placeholder-image")

    def test_video_provider_subprocess_environment_drops_runtime_credentials(self):
        config = {
            "ai": {
                "video": {
                    "seedance_api_key": "config-placeholder-seedance",
                    "bailian_api_key": "config-placeholder-happyhorse",
                },
            },
        }
        with patch.dict("os.environ", {
            "ARK_API_KEY": "runtime-placeholder-seedance",
            "VOLCENGINE_ARK_API_KEY": "runtime-placeholder-seedance-alias",
            "DASHSCOPE_API_KEY": "runtime-placeholder-happyhorse",
            "BAILIAN_API_KEY": "runtime-placeholder-happyhorse-alias",
        }, clear=False), patch("core.api_server.load_config", return_value=config):
            seedance_env, _ = api_server._bala_video_provider_env("seedance")
            happyhorse_env, _ = api_server._bala_video_provider_env("happyhorse")

        self.assertEqual(seedance_env["ARK_API_KEY"], "config-placeholder-seedance")
        self.assertNotIn("VOLCENGINE_ARK_API_KEY", seedance_env)
        self.assertEqual(happyhorse_env["DASHSCOPE_API_KEY"], "config-placeholder-happyhorse")
        self.assertNotIn("BAILIAN_API_KEY", happyhorse_env)

    def test_video_provider_preflight_validates_without_launching_a_cli(self):
        request_type = getattr(api_server, "BalaVideoProviderPreflightRequest", None)
        preflight = getattr(api_server, "_preflight_bala_video_provider", None)
        self.assertTrue(request_type, "Provider preflight request model must exist")
        self.assertTrue(callable(preflight), "Provider preflight validator must exist")

        with tempfile.TemporaryDirectory() as tmpdir:
            image = Path(tmpdir) / "source.png"
            Image.new("RGB", (8, 8), color=(255, 255, 255)).save(image)
            config = {
                "ai": {
                    "video": {
                        "seedance_api_key": "config-placeholder-seedance",
                        "bailian_api_key": "config-placeholder-happyhorse",
                    },
                },
            }
            request = request_type(
                provider="seedance",
                style_code="208326102205",
                prompt="童装模特自然转身展示服装",
                image_paths=[str(image)],
                output_dir=tmpdir,
            )
            with patch("core.api_server.load_config", return_value=config), patch(
                "core.api_server.asyncio.create_subprocess_exec",
                side_effect=AssertionError("preflight must not launch a provider CLI"),
            ):
                result = preflight(request)

        self.assertTrue(result["ok"])
        self.assertEqual(result["provider"], "seedance")
        self.assertEqual(result["status"], "ready")
        self.assertEqual(result["source"], "AI 能力")

    def test_missing_video_provider_credentials_direct_users_to_ai_capabilities_only(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            output_dir = Path(tmpdir) / "output"
            runtime_dir = Path(tmpdir) / "runtime"
            seedance_request = api_server.BalaSeedanceVideoRequest(
                prompt="童装模特自然转身展示服装",
                output_dir=str(output_dir),
            )
            happyhorse_request = api_server.BalaHappyHorseVideoRequest(
                mode="t2v",
                prompt="童装模特自然转身展示服装",
                output_dir=str(output_dir),
            )
            refresh_request = api_server.BalaVideoProviderTaskRequest(
                provider="seedance",
                task_id="seedance-task",
                output_dir=str(output_dir),
                download=False,
            )
            with patch("core.api_server.runtime_paths.child_dir", return_value=runtime_dir), \
                    patch("core.api_server._bala_video_provider_env", return_value=({}, [])), \
                    patch("core.api_server.asyncio.create_subprocess_exec", side_effect=AssertionError("missing credentials must not launch a provider CLI")):
                seedance_result = asyncio.run(api_server._run_seedance_cli(seedance_request))
                happyhorse_result = asyncio.run(api_server._run_happyhorse_cli(happyhorse_request))
                refresh_result = asyncio.run(api_server._run_video_provider_task_cli(refresh_request))

        self.assertEqual(seedance_result["error"], "请先在 AI 能力配置 Seedance 凭据")
        self.assertEqual(happyhorse_result["error"], "请先在 AI 能力配置 HappyHorse 凭据")
        self.assertEqual(refresh_result["error"], "请先在 AI 能力配置 Seedance 凭据")

    def test_provider_task_refresh_and_download_use_shared_cli_commands(self):
        runner = getattr(api_server, "_run_video_provider_task_cli", None)
        request_type = getattr(api_server, "BalaVideoProviderTaskRequest", None)
        self.assertTrue(callable(runner), "Provider task refresh must use the shared CLI")
        self.assertTrue(request_type, "Provider task refresh request model must exist")

        calls = []

        class FinishedProcess:
            returncode = 0

            async def communicate(self):
                return (b'{"id":"seedance-task","status":"succeeded","content":{"video_url":"https://example.test/video.mp4"}}', b'')

        async def fake_create_subprocess_exec(*command, **kwargs):
            calls.append((command, kwargs))
            return FinishedProcess()

        request = request_type(
            provider="seedance",
            task_id="seedance-task",
            output_dir="/tmp/provider-results",
            style_code="208326102205",
            download=False,
        )
        with patch("core.api_server._bala_video_node_runtime", return_value=("node", {})), \
                patch("core.api_server._bala_video_provider_env", return_value=({"ARK_API_KEY": "placeholder"}, ["placeholder"])), \
                patch("core.api_server.asyncio.create_subprocess_exec", side_effect=fake_create_subprocess_exec):
            result = asyncio.run(runner(request))

        self.assertEqual(calls[0][0][:4], ("node", "bin/seedance.js", "get", "seedance-task"))
        self.assertEqual(result["task_id"], "seedance-task")
        self.assertEqual(result["status"], "succeeded")

    def test_apply_face_background_creates_ai_image_job_with_source_and_model_assets(self):
        manifest_path = ROOT / "adapters" / "bala-ai-video-assistant" / "assets" / "model-library" / "manifest.json"

        async def wait_for_control(_status=None):
            return None

        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            data_root = base / "data"
            source = base / "208326102205" / "01_模拍原图" / "model.jpg"
            source.parent.mkdir(parents=True)
            source.write_bytes(b"fake jpg")

            with patch.dict("os.environ", {"CRAWSHRIMP_DATA": str(data_root)}, clear=False):
                runtime_paths.reset_runtime_data_root_cache()
                data_sink.init_db()
                with patch("core.adapter_loader.resolve_adapter_file", return_value=manifest_path):
                    rows = asyncio.run(_apply_bala_ai_face_background_generate(
                        {
                            "source_images": {"paths": [str(source)]},
                            "model_ref_ids": "100女/标准.jpg",
                            "background_prompt": "换成马尔代夫的海边",
                            "generation_mode": "create_only",
                            "model_key_tier": "4k",
                            "image_size": "1536x2048",
                            "quality": "high",
                        },
                        wait_for_control,
                        lambda _message: None,
                    ))

                self.assertEqual(len(rows), 1)
                row = rows[0]
                self.assertEqual(row["执行结果"], "已创建")
                self.assertEqual(row["提交状态"], "已创建，未提交")
                self.assertEqual(row["款号"], "208326102205")
                self.assertEqual(row["模特ID"], "100女/标准.jpg")
                job = data_sink.get_ai_image_job(row["AI任务UID"])
                self.assertIsNotNone(job)
                self.assertIn("马尔代夫", job["prompt"])
                self.assertEqual(job["params"]["main_image_path"], str(source.resolve(strict=False)))
                self.assertEqual(job["params"]["model_key_tier"], "4k")
                self.assertEqual(job["params"]["reference_image_paths"][0], str((manifest_path.parent / "100女" / "标准.jpg").resolve(strict=False)))
                assets = data_sink.list_ai_image_assets(row["AI任务UID"])
                self.assertEqual([asset["kind"] for asset in assets], ["main", "reference"])

            runtime_paths.reset_runtime_data_root_cache()

    def test_apply_face_background_records_async_submission_handles(self):
        manifest_path = ROOT / "adapters" / "bala-ai-video-assistant" / "assets" / "model-library" / "manifest.json"

        async def wait_for_control(_status=None):
            return None

        def fake_submit(job_uid, prompts, request_uid=""):
            return {
                "accepted": True,
                "batch_uid": "batch-1",
                "runs": [{
                    "run_uid": "run-1",
                    "task_id": "task-1",
                    "poll_url": "https://poll.example/task-1",
                    "error": "",
                }],
            }

        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            data_root = base / "data"
            source = base / "208326102205" / "01_模拍原图" / "model.jpg"
            source.parent.mkdir(parents=True)
            source.write_bytes(b"fake jpg")

            with patch.dict("os.environ", {"CRAWSHRIMP_DATA": str(data_root)}, clear=False):
                runtime_paths.reset_runtime_data_root_cache()
                data_sink.init_db()
                with patch("core.adapter_loader.resolve_adapter_file", return_value=manifest_path), \
                        patch("core.ai_image_service.submit_workbench_batch", side_effect=fake_submit):
                    rows = asyncio.run(_apply_bala_ai_face_background_generate(
                        {
                            "source_images": {"paths": [str(source)]},
                            "model_ref_ids": "100女/标准.jpg",
                            "background_prompt": "换成马尔代夫的海边",
                            "generation_mode": "submit_async",
                        },
                        wait_for_control,
                        lambda _message: None,
                    ))

                self.assertEqual(rows[0]["执行结果"], "已提交")
                self.assertEqual(rows[0]["提交状态"], "已异步提交")
                self.assertEqual(rows[0]["批次UID"], "batch-1")
                self.assertEqual(rows[0]["运行UID"], "run-1")
                self.assertEqual(rows[0]["1XM任务ID"], "task-1")
                self.assertEqual(rows[0]["轮询URL"], "https://poll.example/task-1")

            runtime_paths.reset_runtime_data_root_cache()

    def test_apply_background_swap_creates_review_batch_without_model_asset(self):
        async def wait_for_control(_status=None):
            return None

        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            data_root = base / "data"
            source = base / "208326100202" / "01_模拍原图" / "source.png"
            source.parent.mkdir(parents=True)
            source.write_bytes(b"\x89PNG\r\n\x1a\n")

            with patch.dict("os.environ", {"CRAWSHRIMP_DATA": str(data_root)}, clear=False):
                runtime_paths.reset_runtime_data_root_cache()
                data_sink.init_db()
                rows = asyncio.run(_apply_bala_ai_face_background_generate(
                    {
                        "operation_type": "background_swap",
                        "source_images": {"paths": [str(source)]},
                        "background_prompt": "换成马尔代夫的海边",
                        "generation_mode": "create_only",
                        "review_mode": "create_review_batch",
                        "approval_base_url": "http://127.0.0.1:18765",
                    },
                    wait_for_control,
                    lambda _message: None,
                ))

                self.assertEqual(rows[0]["操作类型"], "AI换背景")
                self.assertEqual(rows[0]["模特ID"], "")
                self.assertTrue(rows[0]["审批批次UID"])
                self.assertTrue(rows[0]["审批看板"].startswith("http://127.0.0.1:18765/bala-ai-video-review/"))
                self.assertEqual(rows[0]["下一步Provider"], "qn_img2video")
                assets = data_sink.list_ai_image_assets(rows[0]["AI任务UID"])
                self.assertEqual([asset["kind"] for asset in assets], ["main"])

            runtime_paths.reset_runtime_data_root_cache()

    def test_apply_outfit_swap_creates_review_batch_with_outfit_references(self):
        async def wait_for_control(_status=None):
            return None

        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            data_root = base / "data"
            source = base / "208326100202" / "01_模拍原图" / "source.png"
            garment = base / "garment.png"
            outfit = base / "outfit.png"
            variant = base / "variant.png"
            source.parent.mkdir(parents=True)
            for path in [source, garment, outfit, variant]:
                path.write_bytes(b"\x89PNG\r\n\x1a\n")

            with patch.dict("os.environ", {"CRAWSHRIMP_DATA": str(data_root)}, clear=False):
                runtime_paths.reset_runtime_data_root_cache()
                data_sink.init_db()
                rows = asyncio.run(_apply_bala_ai_face_background_generate(
                    {
                        "operation_type": "outfit_swap",
                        "source_images": {"paths": [str(source)]},
                        "garment_images": {"paths": [str(garment)]},
                        "outfit_reference_images": {"paths": [str(outfit)]},
                        "variant_reference_images": {"paths": [str(variant)]},
                        "generation_mode": "create_only",
                        "review_mode": "create_review_batch",
                        "approval_base_url": "http://127.0.0.1:18765",
                    },
                    wait_for_control,
                    lambda _message: None,
                ))

                self.assertEqual(rows[0]["操作类型"], "AI换装")
                self.assertEqual(rows[0]["服装图文件"], str(garment))
                self.assertEqual(rows[0]["搭配参考图文件"], str(outfit))
                self.assertEqual(rows[0]["同款不同色参考图文件"], str(variant))
                self.assertTrue(rows[0]["审批批次UID"])
                assets = data_sink.list_ai_image_assets(rows[0]["AI任务UID"])
                self.assertEqual([asset["kind"] for asset in assets], ["main", "reference", "reference", "reference"])

            runtime_paths.reset_runtime_data_root_cache()

    def test_finalize_material_prepare_groups_images_without_zip_by_default(self):
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            runtime_dir = base / "runtime"
            runtime_dir.mkdir()

            model_file = runtime_dir / "runtime-model.jpg"
            detail_file = runtime_dir / "runtime-detail.jpg"
            model_file.write_bytes(b"model")
            detail_file.write_bytes(b"detail")

            exported = base / "summary.xlsx"
            workbook = Workbook()
            workbook.active.append(["输入款号", "本地文件"])
            workbook.active.append(["208326102205", str(model_file)])
            workbook.active.append(["208326102205", str(detail_file)])
            workbook.save(exported)

            result = _finalize_bala_ai_video_assistant_outputs(
                task_id="semir_video_material_prepare",
                data_rows=[
                    {
                        "输入款号": "208326102205",
                        "输入编码": "208326102205",
                        "素材来源": "模拍图",
                        "文件名": "model.jpg",
                        "下载结果": "已下载",
                        "本地文件": str(model_file),
                        "__bala_group_code": "208326102205",
                        "__bala_source_type": "model",
                        "__package_filename": "model.jpg",
                        "__compress_threshold_bytes": 20 * 1024 * 1024,
                    },
                    {
                        "输入款号": "208326102205",
                        "输入编码": "208326102205",
                        "素材来源": "商品细节图",
                        "文件名": "detail.jpg",
                        "下载结果": "已下载",
                        "本地文件": str(detail_file),
                        "__bala_group_code": "208326102205",
                        "__bala_source_type": "detail",
                        "__package_filename": "detail.jpg",
                        "__compress_threshold_bytes": 20 * 1024 * 1024,
                    },
                ],
                runtime_files=[str(model_file), str(detail_file)],
                exported_files=[str(exported)],
                run_params={"package_name": "巴拉视频素材测试"},
                runtime_artifact_dir=str(runtime_dir),
                log=lambda _: None,
            )

            self.assertEqual(len(result), 2)
            package_dir = Path(result[0])
            self.assertTrue(package_dir.is_dir())
            self.assertEqual(package_dir.name, "巴拉视频素材测试")
            self.assertEqual(Path(result[1]), exported)
            self.assertTrue((package_dir / "208326102205" / "01_模拍原图" / "model.jpg").is_file())
            self.assertTrue((package_dir / "208326102205" / "02_商品细节图" / "detail.jpg").is_file())
            self.assertFalse(any(path.suffix == ".zip" for path in package_dir.parent.iterdir() if path.is_file()))
            self.assertFalse(model_file.exists())
            self.assertFalse(detail_file.exists())

    def test_finalize_material_prepare_invokes_compression_for_downloaded_images(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            runtime_dir = base / "runtime"
            runtime_dir.mkdir()

            image_file = runtime_dir / "runtime-model.jpg"
            image_file.write_bytes(b"large-image")

            def fake_compress(path, threshold_bytes, log):
                compressed = path.with_name("model-compressed.jpg")
                compressed.write_bytes(b"small")
                path.unlink()
                return compressed, "已压缩：测试"

            rows = [
                {
                    "输入款号": "208326102205",
                    "输入编码": "208326102205",
                    "素材来源": "模拍图",
                    "文件名": "model.png",
                    "下载结果": "已下载",
                    "本地文件": str(image_file),
                    "__bala_group_code": "208326102205",
                    "__bala_source_type": "model",
                    "__package_filename": "model.png",
                    "__compress_threshold_bytes": 1,
                }
            ]

            with patch("core.api_server._compress_bala_video_image_if_needed", side_effect=fake_compress) as mocked:
                result = _finalize_bala_ai_video_assistant_outputs(
                    task_id="semir_video_material_prepare",
                    data_rows=rows,
                    runtime_files=[str(image_file)],
                    exported_files=[],
                    run_params={"package_name": "巴拉视频素材压缩测试"},
                    runtime_artifact_dir=str(runtime_dir),
                    log=lambda _: None,
                )

            mocked.assert_called_once()
            package_dir = Path(result[0])
            self.assertTrue((package_dir / "208326102205" / "01_模拍原图" / "model-compressed.jpg").is_file())
            self.assertEqual(rows[0]["本地文件"], str(package_dir / "208326102205" / "01_模拍原图" / "model-compressed.jpg"))
            self.assertEqual(rows[0]["压缩结果"], "已压缩：测试")
            self.assertEqual(rows[0]["文件名"], "model-compressed.jpg")

    def test_finalize_material_prepare_rewrites_exported_excel_to_final_image_paths(self):
        from openpyxl import Workbook, load_workbook

        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            runtime_dir = base / "runtime"
            export_dir = base / "workspace"
            runtime_dir.mkdir()
            image_file = runtime_dir / "runtime-model.jpg"
            image_file.write_bytes(b"image")
            summary = base / "summary.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["输入款号", "本地文件"])
            sheet.append(["208326102205", str(image_file)])
            workbook.save(summary)
            rows = [{
                "输入款号": "208326102205",
                "输入编码": "208326102205",
                "素材来源": "模拍图",
                "文件名": "model.jpg",
                "下载结果": "已下载",
                "本地文件": str(image_file),
                "__bala_group_code": "208326102205",
                "__bala_source_type": "model",
                "__package_filename": "model.jpg",
                "__compress_threshold_bytes": 20 * 1024 * 1024,
            }]

            result = _finalize_bala_ai_video_assistant_outputs(
                task_id="semir_video_material_prepare",
                data_rows=rows,
                runtime_files=[str(image_file)],
                exported_files=[str(summary)],
                run_params={"package_name": "final-paths", "export_folder": str(export_dir)},
                runtime_artifact_dir=str(runtime_dir),
                log=lambda _: None,
            )

            copied_summary = next(Path(item) for item in result if str(item).endswith(".xlsx"))
            exported = load_workbook(copied_summary, read_only=True, data_only=True)
            exported_path = exported.active.cell(row=2, column=2).value
            exported.close()
            self.assertEqual(exported_path, rows[0]["本地文件"])
            self.assertTrue(Path(exported_path).is_file())

    def test_finalize_material_prepare_fails_closed_when_excel_paths_cannot_be_rewritten(self):
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            runtime_dir = base / "runtime"
            export_dir = base / "workspace"
            runtime_dir.mkdir()
            image_file = runtime_dir / "runtime-model.jpg"
            image_file.write_bytes(b"image")
            summary = base / "summary.xlsx"
            workbook = Workbook()
            workbook.active.append(["输入款号", "错误路径列"])
            workbook.active.append(["208326102205", str(image_file)])
            workbook.save(summary)
            rows = [{
                "输入款号": "208326102205",
                "输入编码": "208326102205",
                "素材来源": "模拍图",
                "文件名": "model.jpg",
                "下载结果": "已下载",
                "本地文件": str(image_file),
                "__bala_group_code": "208326102205",
                "__bala_source_type": "model",
                "__package_filename": "model.jpg",
                "__compress_threshold_bytes": 20 * 1024 * 1024,
            }]

            with self.assertRaisesRegex(RuntimeError, "本地文件"):
                _finalize_bala_ai_video_assistant_outputs(
                    task_id="semir_video_material_prepare",
                    data_rows=rows,
                    runtime_files=[str(image_file)],
                    exported_files=[str(summary)],
                    run_params={"package_name": "invalid-summary", "export_folder": str(export_dir)},
                    runtime_artifact_dir=str(runtime_dir),
                    log=lambda _: None,
                )

    def test_review_board_default_url_uses_the_active_backend_port(self):
        async def wait_for_control(_status=None):
            return None

        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            source = base / "208326102205" / "source.png"
            source.parent.mkdir(parents=True)
            source.write_bytes(b"\x89PNG\r\n\x1a\n")
            with patch.dict("os.environ", {
                "CRAWSHRIMP_DATA": str(base / "data"),
                "CRAWSHRIMP_PORT": "18766",
            }, clear=False):
                runtime_paths.reset_runtime_data_root_cache()
                data_sink.init_db()
                rows = asyncio.run(_apply_bala_ai_face_background_generate(
                    {
                        "operation_type": "background_swap",
                        "source_images": {"paths": [str(source)]},
                        "background_prompt": "纯色背景",
                        "generation_mode": "create_only",
                        "review_mode": "create_review_batch",
                    },
                    wait_for_control,
                    lambda _message: None,
                ))

            self.assertTrue(rows[0]["审批看板"].startswith("http://127.0.0.1:18766/bala-ai-video-review/"))
            runtime_paths.reset_runtime_data_root_cache()

    def test_backend_disables_query_string_access_logs(self):
        source = (ROOT / "core" / "api_server.py").read_text(encoding="utf-8")
        self.assertIn("access_log=False", source)

    def test_orphan_cleanup_includes_bala_video_runtime(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            runtime_dir = base / "data" / "bala-ai-video-assistant" / "semir_video_material_prepare" / "runtime" / "100"
            runtime_dir.mkdir(parents=True)
            stale = runtime_dir / "stale.jpg"
            stale.write_bytes(b"stale")

            def fake_artifact_dir_path(adapter_id, task_id, run_id, kind):
                self.assertEqual(adapter_id, "bala-ai-video-assistant")
                self.assertEqual(task_id, "semir_video_material_prepare")
                self.assertEqual(run_id, 100)
                self.assertEqual(kind, "runtime")
                return runtime_dir

            with patch("core.data_sink.artifact_dir_path", side_effect=fake_artifact_dir_path):
                _cleanup_orphaned_runtime_artifacts([
                    {
                        "id": 100,
                        "adapter_id": "bala-ai-video-assistant",
                        "task_id": "semir_video_material_prepare",
                    }
                ])

            self.assertFalse(stale.exists())


if __name__ == "__main__":
    unittest.main()
