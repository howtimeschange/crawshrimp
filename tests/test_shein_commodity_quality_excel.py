import os
import tempfile
import unittest
from importlib.util import find_spec
from pathlib import Path
from unittest.mock import patch

import yaml

from core import data_sink


ROOT = Path(__file__).resolve().parents[1]
MANIFEST_PATH = ROOT / "adapters" / "shein-helper" / "manifest.yaml"


class SheinCommodityQualityExcelTest(unittest.TestCase):
    def test_export_uses_quality_list_and_return_detail_sheets(self):
        if find_spec("openpyxl") is None:
            self.skipTest("openpyxl is required for export_excel in this environment")

        from openpyxl import load_workbook

        manifest = yaml.safe_load(MANIFEST_PATH.read_text(encoding="utf-8"))
        task = next(item for item in manifest["tasks"] if item["id"] == "commodity_quality")
        output = next(item for item in task["output"] if item["type"] == "excel")

        rows = [
            {
                "__sheet_name": "质量列表",
                "商品名称": "Tween Girl半身裙",
                "SKC": "sk25050817072795161",
                "SPU": "k250508170727",
                "品退数/品退率": "14 / 6.97%",
                "筛选摘要": "SKC=1项",
            },
            {
                "__sheet_name": "客退详情",
                "SKC": "sk25050817072795161",
                "SPU": "k250508170727",
                "商品名称": "Tween Girl半身裙",
                "退货时间": "2026-05-01 08:35:34",
                "是否品退": "否",
                "SKU": "I74tax65162p-粉红-12Y",
                "客退原因": "COD未妥投退回",
            },
        ]

        with tempfile.TemporaryDirectory() as tmpdir:
            with patch.dict(os.environ, {"CRAWSHRIMP_DATA": tmpdir}, clear=False):
                out_path = data_sink.export_excel(
                    rows,
                    adapter_id="shein-helper",
                    task_id="commodity_quality",
                    filename_template="test.xlsx",
                    sheet_key=output["sheet_key"],
                    sheet_configs=output["sheets"],
                )

            wb = load_workbook(out_path, read_only=True, data_only=True)
            self.assertEqual(wb.sheetnames, ["质量列表", "客退详情"])

            quality_headers = [cell.value for cell in next(wb["质量列表"].iter_rows(min_row=1, max_row=1))]
            detail_headers = [cell.value for cell in next(wb["客退详情"].iter_rows(min_row=1, max_row=1))]
            self.assertIn("商品名称", quality_headers)
            self.assertIn("品退数/品退率", quality_headers)
            self.assertIn("客退原因", detail_headers)
            self.assertNotIn("操作", quality_headers)
            self.assertNotIn("操作", detail_headers)


if __name__ == "__main__":
    unittest.main()
