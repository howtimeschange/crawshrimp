import threading
import time
import tempfile
import unittest
import zipfile
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from core import buyer_show_service, data_sink


PNG_1X1 = bytes.fromhex(
    "89504e470d0a1a0a0000000d4948445200000001000000010806000000"
    "1f15c4890000000a49444154789c6360000002000154a24f5d00000000"
    "49454e44ae426082"
)
JPEG_HEADER_BYTES = b"\xff\xd8\xff\xe0\x00\x10JFIF\x00\x01\x01\x00\x00\x01\x00\x01\x00\x00\xff\xd9"


class BuyerShowServiceTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self.tmp.cleanup)
        self.root = Path(self.tmp.name)
        patcher = patch("core.runtime_paths.data_root", return_value=self.root / "data")
        patcher.start()
        self.addCleanup(patcher.stop)
        home_patcher = patch("pathlib.Path.home", return_value=self.root / "home")
        home_patcher.start()
        self.addCleanup(home_patcher.stop)
        data_sink.init_db()

    def _source_rows(self, style_color_code="208326102205-00316", unique_value="ORD-1208326102205-00316"):
        model = self.root / "runtime" / "model.jpg"
        flat = self.root / "runtime" / "flat.jpg"
        model.parent.mkdir(parents=True, exist_ok=True)
        model.write_bytes(PNG_1X1)
        flat.write_bytes(PNG_1X1)
        return [{
            "表格行号": 2,
            "订单号": "ORD-1",
            "款色号": style_color_code,
            "货号": "",
            "尺码": "120",
            "唯一值": unique_value,
            "AI素材库路径": "巴拉搜索渠道-淘系//买家秀图库/冬季/上装/女/",
            "AI图包文件夹命名": unique_value,
            "存放地址": "巴拉搜索渠道-淘系//AI后买家秀图包/",
            "模拍文件": "model.jpg",
            "模拍云盘路径": "买家秀图库/冬季/上装/女/model.jpg",
            "模拍下载结果": "已下载",
            "模拍本地文件": str(model),
            "平铺参考图": "208326102205-00316.jpg",
            "平铺云盘路径": "平铺图库/208326102205-00316.jpg",
            "平铺下载结果": "已下载",
            "平铺本地文件": str(flat),
            "生图结果": "待生成",
        }]

    def test_finalize_generates_package_zip_summary_and_usage_record(self):
        generated = self.root / "generated.png"
        generated.write_bytes(PNG_1X1)
        created_jobs = []

        def fake_run(job_uid, settings=None):
            created_jobs.append(job_uid)
            return {
                "ok": True,
                "job_uid": job_uid,
                "summary": {
                    "task_id": "1xm-task-1",
                    "image_urls": ["https://cdn.example/generated.png"],
                    "runs": [{"task_id": "1xm-task-1", "image_urls": ["https://cdn.example/generated.png"]}],
                },
            }

        with (
            patch("core.buyer_show_service.ai_image_service.run_job_with_one_xm", side_effect=fake_run),
            patch("core.buyer_show_service.ai_image_service.materialize_remote_image", return_value={
                "ok": True,
                "path": str(generated),
                "url": "https://cdn.example/generated.png",
            }),
        ):
            refs = buyer_show_service.finalize_buyer_show_outputs(
                data_rows=self._source_rows(),
                runtime_files=[],
                exported_files=[],
                run_params={"export_folder": str(self.root / "exports"), "package_name": "AI买家秀测试"},
                runtime_artifact_dir=str(self.root / "runtime"),
                settings={"base_url": "https://api.example", "4k": "secret"},
                log=lambda _msg: None,
            )

        package_root = Path(refs[0])
        zip_path = Path(refs[1])
        summary_path = Path(refs[2])
        usage_path = Path(refs[3])
        self.assertTrue(package_root.is_dir())
        self.assertTrue(zip_path.is_file())
        self.assertTrue(summary_path.is_file())
        self.assertTrue(usage_path.is_file())
        output_files = list(package_root.rglob("*AI买家秀.png"))
        self.assertEqual(len(output_files), 1)
        self.assertIn("ORD-1208326102205-00316_001_208326102205-00316_AI买家秀", output_files[0].name)
        with zipfile.ZipFile(zip_path) as archive:
            self.assertIn(output_files[0].relative_to(package_root.parent).as_posix(), archive.namelist())

        wb = load_workbook(summary_path)
        values = list(wb.active.iter_rows(values_only=True))
        headers = list(values[0])
        row = dict(zip(headers, values[1]))
        self.assertEqual(row["生图结果"], "已生成")
        self.assertEqual(row["AI任务ID"], created_jobs[0])
        self.assertEqual(row["1XM任务ID"], "1xm-task-1")

        usage_rows = data_sink.list_buyer_show_material_usage()
        self.assertEqual(len(usage_rows), 1)
        self.assertEqual(usage_rows[0]["style_code"], "208326102205")
        self.assertEqual(usage_rows[0]["model_cloud_path"], "买家秀图库/冬季/上装/女/model.jpg")
        self.assertIn(str(output_files[0]), usage_rows[0]["output_file"])

    def test_finalize_skips_model_image_already_used_for_same_style_color(self):
        data_sink.create_buyer_show_material_usage({
            "style_code": "208326102205",
            "style_color_code": "208326102205-00316",
            "model_cloud_path": "买家秀图库/冬季/上装/女/model.jpg",
            "model_filename": "model.jpg",
        })
        with patch("core.buyer_show_service.ai_image_service.run_job_with_one_xm") as run_job:
            refs = buyer_show_service.finalize_buyer_show_outputs(
                data_rows=self._source_rows(),
                runtime_files=[],
                exported_files=[],
                run_params={"export_folder": str(self.root / "exports"), "package_name": "AI买家秀测试"},
                runtime_artifact_dir=str(self.root / "runtime"),
                settings={"base_url": "https://api.example", "4k": "secret"},
                log=lambda _msg: None,
            )

        run_job.assert_not_called()
        summary_path = Path(refs[2])
        wb = load_workbook(summary_path)
        values = list(wb.active.iter_rows(values_only=True))
        headers = list(values[0])
        row = dict(zip(headers, values[1]))
        self.assertEqual(row["生图结果"], "已跳过")
        self.assertIn("使用记录已存在", row["备注"])

    def test_finalize_allows_same_model_image_for_different_style_color(self):
        data_sink.create_buyer_show_material_usage({
            "style_code": "208326102205",
            "style_color_code": "208326102205-00316",
            "model_cloud_path": "买家秀图库/冬季/上装/女/model.jpg",
            "model_filename": "model.jpg",
        })
        generated = self.root / "generated-different-color.png"
        generated.write_bytes(PNG_1X1)

        with (
            patch("core.buyer_show_service.ai_image_service.run_job_with_one_xm", return_value={
                "ok": True,
                "summary": {"task_id": "1xm-task-different-color", "image_urls": ["https://cdn.example/generated.png"]},
            }) as run_job,
            patch("core.buyer_show_service.ai_image_service.materialize_remote_image", return_value={
                "ok": True,
                "path": str(generated),
                "url": "https://cdn.example/generated.png",
            }),
        ):
            refs = buyer_show_service.finalize_buyer_show_outputs(
                data_rows=self._source_rows(
                    style_color_code="208326102205-90001",
                    unique_value="ORD-1208326102205-90001",
                ),
                runtime_files=[],
                exported_files=[],
                run_params={"export_folder": str(self.root / "exports"), "package_name": "AI买家秀测试"},
                runtime_artifact_dir=str(self.root / "runtime"),
                settings={"base_url": "https://api.example", "4k": "secret"},
                log=lambda _msg: None,
            )

        run_job.assert_called_once()
        summary_path = Path(refs[2])
        wb = load_workbook(summary_path)
        values = list(wb.active.iter_rows(values_only=True))
        headers = list(values[0])
        row = dict(zip(headers, values[1]))
        self.assertEqual(row["生图结果"], "已生成")

    def test_finalize_retries_transient_generation_and_result_materialization(self):
        generated = self.root / "generated-after-retry.png"
        generated.write_bytes(PNG_1X1)
        transient = "[SSL: UNEXPECTED_EOF_WHILE_READING] EOF occurred in violation of protocol"

        with (
            patch("core.buyer_show_service.ai_image_service.run_job_with_one_xm", side_effect=[
                {"ok": False, "summary": {"error": transient, "image_urls": []}},
                {"ok": True, "summary": {"task_id": "1xm-after-retry", "image_urls": ["https://proxy.example/image.png"]}},
            ]) as run_job,
            patch("core.buyer_show_service.ai_image_service.materialize_remote_image", side_effect=[
                RuntimeError(transient),
                {"ok": True, "path": str(generated), "url": "https://proxy.example/image.png"},
            ]) as materialize,
            patch("core.buyer_show_service.time.sleep"),
        ):
            refs = buyer_show_service.finalize_buyer_show_outputs(
                data_rows=self._source_rows(),
                runtime_files=[],
                exported_files=[],
                run_params={"export_folder": str(self.root / "exports"), "package_name": "AI买家秀重试测试"},
                runtime_artifact_dir=str(self.root / "runtime"),
                settings={"base_url": "https://api.example", "4k": "secret"},
                log=lambda _msg: None,
            )

        self.assertEqual(run_job.call_count, 2)
        self.assertEqual(materialize.call_count, 2)
        package_root = Path(refs[0])
        output_files = list(package_root.rglob("*AI买家秀.png"))
        self.assertEqual(len(output_files), 1)

    def test_finalize_retries_non_transient_generation_failure_up_to_three_attempts(self):
        generated = self.root / "generated-after-non-transient-retry.png"
        generated.write_bytes(PNG_1X1)

        with (
            patch("core.buyer_show_service.ai_image_service.run_job_with_one_xm", side_effect=[
                {"ok": False, "summary": {"error": "provider rejected first attempt", "image_urls": []}},
                {"ok": False, "summary": {"error": "provider rejected second attempt", "image_urls": []}},
                {"ok": True, "summary": {"task_id": "1xm-third-attempt", "image_urls": ["https://proxy.example/image.png"]}},
            ]) as run_job,
            patch("core.buyer_show_service.ai_image_service.materialize_remote_image", return_value={
                "ok": True,
                "path": str(generated),
                "url": "https://proxy.example/image.png",
            }),
            patch("core.buyer_show_service.time.sleep"),
        ):
            refs = buyer_show_service.finalize_buyer_show_outputs(
                data_rows=self._source_rows(),
                runtime_files=[],
                exported_files=[],
                run_params={"export_folder": str(self.root / "exports"), "package_name": "AI买家秀非临时失败重试"},
                runtime_artifact_dir=str(self.root / "runtime"),
                settings={"base_url": "https://api.example", "4k": "secret"},
                log=lambda _msg: None,
            )

        self.assertEqual(run_job.call_count, 3)
        package_root = Path(refs[0])
        output_files = list(package_root.rglob("*AI买家秀.png"))
        self.assertEqual(len(output_files), 1)

    def test_finalize_runs_ai_generation_with_configured_concurrency_window(self):
        source_rows = []
        for index in range(5):
            row = self._source_rows(
                style_color_code=f"208326102205-0031{index}",
                unique_value=f"ORD-{index}",
            )[0]
            row["表格行号"] = index + 2
            row["模拍文件"] = f"model-{index}.jpg"
            row["模拍云盘路径"] = f"买家秀图库/冬季/上装/女/model-{index}.jpg"
            source_rows.append(row)

        active = 0
        max_active = 0
        lock = threading.Lock()

        def fake_prepare(row, **_kwargs):
            nonlocal active, max_active
            with lock:
                active += 1
                max_active = max(max_active, active)
            time.sleep(0.05)
            with lock:
                active -= 1
            row["生图结果"] = "待落图"
            row["AI任务ID"] = f"job-{row['表格行号']}"
            row["__generation_urls"] = [f"https://proxy.example/{row['表格行号']}.png"]
            row["__generation_prompt"] = "prompt"
            return row

        def fake_materialize(row, **_kwargs):
            row["生图结果"] = "已生成"
            row["生图文件"] = ""
            row["__usage_record_time"] = "2026-08-21T15:00:00"
            return row

        logs = []
        with (
            patch("core.buyer_show_service._prepare_buyer_show_generation_row", side_effect=fake_prepare) as prepare,
            patch("core.buyer_show_service._materialize_buyer_show_generation_row", side_effect=fake_materialize) as materialize,
        ):
            buyer_show_service.finalize_buyer_show_outputs(
                data_rows=source_rows,
                runtime_files=[],
                exported_files=[],
                run_params={
                    "export_folder": str(self.root / "exports"),
                    "package_name": "AI买家秀并发测试",
                    "ai_generation_concurrency": 3,
                    "ai_result_download_concurrency": 2,
                    "usage_record_mode": "ignore",
                },
                runtime_artifact_dir=str(self.root / "runtime"),
                settings={"base_url": "https://api.example", "4k": "secret"},
                log=logs.append,
            )

        self.assertEqual(prepare.call_count, 5)
        self.assertEqual(materialize.call_count, 5)
        self.assertGreaterEqual(max_active, 2)
        self.assertTrue(any("AI 生图并发窗口：3" in line for line in logs))
        self.assertTrue(any("AI 结果落图并发窗口：2" in line for line in logs))

    def test_finalize_downloads_ready_results_before_all_generation_links_are_collected(self):
        source_rows = []
        for index in range(6):
            row = self._source_rows(
                style_color_code=f"208326102205-0032{index}",
                unique_value=f"ORD-link-{index}",
            )[0]
            row["表格行号"] = index + 2
            row["模拍文件"] = f"model-link-{index}.jpg"
            row["模拍云盘路径"] = f"买家秀图库/冬季/上装/女/model-link-{index}.jpg"
            source_rows.append(row)

        events = []
        lock = threading.Lock()
        materialize_started = threading.Event()

        def fake_prepare(row, **_kwargs):
            with lock:
                events.append(f"prepare-{row['表格行号']}")
            if row["表格行号"] != 2:
                materialize_started.wait(2)
            row["生图结果"] = "待落图"
            row["AI任务ID"] = f"job-{row['表格行号']}"
            row["__generation_urls"] = [f"https://proxy.example/{row['表格行号']}.png"]
            row["__generation_prompt"] = "prompt"
            return row

        def fake_materialize(row, **_kwargs):
            with lock:
                events.append(f"materialize-{row['表格行号']}")
            materialize_started.set()
            row["生图结果"] = "已生成"
            row["生图文件"] = ""
            row["__usage_record_time"] = "2026-08-21T15:00:00"
            return row

        with (
            patch("core.buyer_show_service._prepare_buyer_show_generation_row", side_effect=fake_prepare),
            patch("core.buyer_show_service._materialize_buyer_show_generation_row", side_effect=fake_materialize),
        ):
            buyer_show_service.finalize_buyer_show_outputs(
                data_rows=source_rows,
                runtime_files=[],
                exported_files=[],
                run_params={
                    "export_folder": str(self.root / "exports"),
                    "package_name": "AI买家秀流水线落图测试",
                    "ai_generation_concurrency": 2,
                    "ai_result_download_concurrency": 2,
                    "usage_record_mode": "ignore",
                },
                runtime_artifact_dir=str(self.root / "runtime"),
                settings={"base_url": "https://api.example", "4k": "secret"},
                log=lambda _msg: None,
            )

        first_materialize = next(index for index, event in enumerate(events) if event.startswith("materialize-"))
        last_prepare = max(index for index, event in enumerate(events) if event.startswith("prepare-"))
        self.assertLess(first_materialize, last_prepare)

    def test_finalize_marks_materialization_failure_without_usage_record(self):
        with (
            patch("core.buyer_show_service.ai_image_service.run_job_with_one_xm", return_value={
                "ok": True,
                "summary": {"task_id": "1xm-task-download-fail", "image_urls": ["https://cdn.example/generated.png"]},
            }),
            patch("core.buyer_show_service.ai_image_service.materialize_remote_image", side_effect=RuntimeError("download failed")),
        ):
            refs = buyer_show_service.finalize_buyer_show_outputs(
                data_rows=self._source_rows(),
                runtime_files=[],
                exported_files=[],
                run_params={"export_folder": str(self.root / "exports"), "package_name": "AI买家秀落图失败测试"},
                runtime_artifact_dir=str(self.root / "runtime"),
                settings={"base_url": "https://api.example", "4k": "secret"},
                log=lambda _msg: None,
            )

        package_root = Path(refs[0])
        self.assertFalse(list(package_root.rglob("*AI买家秀.png")))
        self.assertEqual(data_sink.list_buyer_show_material_usage(), [])
        wb = load_workbook(Path(refs[2]))
        values = list(wb.active.iter_rows(values_only=True))
        row = dict(zip(values[0], values[1]))
        self.assertEqual(row["生图结果"], "生成失败")
        self.assertIn("download failed", row["备注"])

    def test_result_url_candidates_prefer_proxy_before_direct_image(self):
        url = (
            "https://one-xm-proxy.crawshrimp.com/v1/proxy-image"
            "?url=https%3A%2F%2Fimg.1xm.ai%2Fgenerated%2Ftask_1_0.png"
        )

        self.assertEqual(buyer_show_service._candidate_result_urls(url), [
            url,
            "https://img.1xm.ai/generated/task_1_0.png",
        ])

    def test_finalize_records_parent_style_code_for_joined_style_color(self):
        generated = self.root / "generated-joined.png"
        generated.write_bytes(PNG_1X1)

        with (
            patch("core.buyer_show_service.ai_image_service.run_job_with_one_xm", return_value={
                "ok": True,
                "summary": {"task_id": "1xm-task-joined", "image_urls": ["https://cdn.example/generated-joined.png"]},
            }),
            patch("core.buyer_show_service.ai_image_service.materialize_remote_image", return_value={
                "ok": True,
                "path": str(generated),
                "url": "https://cdn.example/generated-joined.png",
            }),
        ):
            buyer_show_service.finalize_buyer_show_outputs(
                data_rows=self._source_rows(
                    style_color_code="20842610420180915",
                    unique_value="120842610420180915",
                ),
                runtime_files=[],
                exported_files=[],
                run_params={"export_folder": str(self.root / "exports"), "package_name": "AI买家秀测试"},
                runtime_artifact_dir=str(self.root / "runtime"),
                settings={"base_url": "https://api.example", "4k": "secret"},
                log=lambda _msg: None,
            )

        usage_rows = data_sink.list_buyer_show_material_usage()
        self.assertEqual(len(usage_rows), 1)
        self.assertEqual(usage_rows[0]["style_code"], "208426104201")
        self.assertEqual(usage_rows[0]["style_color_code"], "20842610420180915")

    def test_finalize_uses_actual_image_header_for_package_extension(self):
        generated = self.root / "generated-mislabeled.png"
        generated.write_bytes(JPEG_HEADER_BYTES)

        with (
            patch("core.buyer_show_service.ai_image_service.run_job_with_one_xm", return_value={
                "ok": True,
                "summary": {"task_id": "1xm-task-jpeg", "image_urls": ["https://cdn.example/generated.png"]},
            }),
            patch("core.buyer_show_service.ai_image_service.materialize_remote_image", return_value={
                "ok": True,
                "path": str(generated),
                "url": "https://cdn.example/generated.png",
            }),
        ):
            refs = buyer_show_service.finalize_buyer_show_outputs(
                data_rows=self._source_rows(),
                runtime_files=[],
                exported_files=[],
                run_params={"export_folder": str(self.root / "exports"), "package_name": "AI买家秀测试"},
                runtime_artifact_dir=str(self.root / "runtime"),
                settings={"base_url": "https://api.example", "4k": "secret"},
                log=lambda _msg: None,
            )

        package_root = Path(refs[0])
        output_files = list(package_root.rglob("*AI买家秀.jpg"))
        self.assertEqual(len(output_files), 1)
        self.assertFalse(list(package_root.rglob("*AI买家秀.png")))

    def test_default_prompt_treats_sets_as_full_outfit_replacement(self):
        prompt = buyer_show_service.build_buyer_show_prompt({
            "AI素材库路径": "巴拉搜索渠道-淘系//买家秀图库/冬季/套装/婴幼童-长袖长裤两件套/男/",
            "款色号": "20842610420180915",
        }, {})

        self.assertIn("完整套装商品", prompt)
        self.assertIn("上衣、下装", prompt)
        self.assertIn("不要只替换上衣或只替换裤子", prompt)
        self.assertIn("不要保留原模拍图里原有服装的颜色和款式", prompt)

    def test_download_only_packages_downloaded_materials_for_review(self):
        refs = buyer_show_service.finalize_buyer_show_outputs(
            data_rows=self._source_rows(),
            runtime_files=[],
            exported_files=[],
            run_params={
                "export_folder": str(self.root / "exports"),
                "package_name": "AI买家秀测试",
                "execute_mode": "download_only",
            },
            runtime_artifact_dir=str(self.root / "runtime"),
            settings={"base_url": "https://api.example", "4k": "secret"},
            log=lambda _msg: None,
        )

        package_root = Path(refs[0])
        copied = sorted(path.name for path in package_root.rglob("*") if path.is_file() and "AI买家秀" not in path.name)
        self.assertEqual(len(copied), 2)
        self.assertTrue(all(name.endswith(".png") for name in copied))
        self.assertTrue(any("模拍原图" in name for name in copied))
        self.assertTrue(any("平铺参考图" in name for name in copied))
        wb = load_workbook(Path(refs[2]))
        values = list(wb.active.iter_rows(values_only=True))
        row = dict(zip(values[0], values[1]))
        self.assertEqual(row["生图结果"], "待生成")
        self.assertIn("已打包下载素材 2 个", row["备注"])

    def test_default_output_root_is_downloads_full_batch_folder(self):
        refs = buyer_show_service.finalize_buyer_show_outputs(
            data_rows=self._source_rows(),
            runtime_files=[],
            exported_files=[],
            run_params={
                "package_name": "AI买家秀默认目录",
                "execute_mode": "download_only",
            },
            runtime_artifact_dir=str(self.root / "runtime"),
            settings={"base_url": "https://api.example", "4k": "secret"},
            log=lambda _msg: None,
        )

        package_root = Path(refs[0])
        self.assertEqual(package_root.parent, self.root / "home" / "Downloads" / "AI 买家秀全量测试")
        self.assertTrue(package_root.is_dir())


if __name__ == "__main__":
    unittest.main()
