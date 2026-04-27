import tempfile
import unittest
import zipfile
from pathlib import Path
from unittest.mock import patch

from core.shenhui_pdf_screenshot import (
    build_pdf_yq_filename,
    convert_pdf_to_yq_images,
    crop_boxes_for_pdf_type,
    extract_style_color_code,
    finalize_pdf_batch_screenshot_outputs,
    parse_style_color_overrides,
)


class ShenhuiPdfScreenshotTests(unittest.TestCase):
    def test_extract_style_color_code_prefers_pdf_text_code(self):
        source_name = "559 夏季 K228043055 合格证-主线 通用（新版） 201226109105 兴仁市纬特西.pdf"
        text = "产品货号:201226109105\n201226109105-00322 80\nRMB199.00"

        code = extract_style_color_code(source_name=source_name, text=text)

        self.assertEqual(code, ("201226109105", "00322"))

    def test_extract_style_color_code_uses_wash_label_filename_color(self):
        code = extract_style_color_code(
            source_name="201226109105-2（88201牛仔中蓝）洗唛.pdf",
            text="",
        )

        self.assertEqual(code, ("201226109105", "88201"))

    def test_build_pdf_yq_filename_uses_default_role_name_for_single_color_style(self):
        self.assertEqual(
            build_pdf_yq_filename("201226109105", "00322", "hang_tag", 1, False),
            "yq(1).png",
        )
        self.assertEqual(
            build_pdf_yq_filename("201226109105", "00322", "wash_label", 1, False),
            "yq(2).png",
        )

    def test_build_pdf_yq_filename_uses_style_color_prefix_for_multi_color_style(self):
        self.assertEqual(
            build_pdf_yq_filename("201226109105", "00322", "hang_tag", 2, True),
            "201226109105-00322yq(1)-2.png",
        )
        self.assertEqual(
            build_pdf_yq_filename("201226109105", "", "wash_label", 1, True),
            "201226109105yq(2)-1.png",
        )

    def test_parse_style_color_overrides_accepts_filename_mapping(self):
        overrides = parse_style_color_overrides(
            "hang-tag.pdf=201226109105-00322\n201226109106-00456"
        )

        self.assertEqual(overrides["hang-tag.pdf"], "201226109105-00322")
        self.assertEqual(overrides["__default__"], "201226109106-00456")

    def test_crop_boxes_for_pdf_type_uses_default_templates(self):
        self.assertEqual(
            crop_boxes_for_pdf_type("wash_label", {}),
            [(0.0892, 0.2084, 0.4189, 0.7546)],
        )
        self.assertEqual(
            crop_boxes_for_pdf_type("hang_tag", {}),
            [(0.0113, 0.2352, 0.1535, 0.5058)],
        )

    def test_convert_pdf_to_yq_images_crops_every_rendered_page(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            pdf_path = base / "multi-page.pdf"
            pdf_path.write_bytes(b"%PDF-fake")
            page_one = base / "page-one.png"
            page_two = base / "page-two.png"
            crop_one = base / "crop-one.png"
            crop_two = base / "crop-two.png"
            for path in (page_one, page_two, crop_one, crop_two):
                path.write_bytes(b"png")

            with patch(
                "core.shenhui_pdf_screenshot.render_pdf_pages_with_pymupdf",
                return_value=[page_one, page_two],
            ), patch(
                "core.shenhui_pdf_screenshot.crop_image_to_yq_pages",
                side_effect=[[crop_one], [crop_two]],
            ) as crop_mock:
                outputs = convert_pdf_to_yq_images(
                    pdf_path,
                    base / "work",
                    log=lambda _: None,
                    crop_boxes=[(0.1, 0.2, 0.3, 0.4)],
                )

            self.assertEqual(outputs, [crop_one, crop_two])
            self.assertEqual(crop_mock.call_count, 2)
            self.assertEqual(crop_mock.call_args_list[0].args[0], page_one)
            self.assertEqual(crop_mock.call_args_list[1].args[0], page_two)

    def test_finalize_pdf_batch_uses_default_tag_and_wash_names_for_single_color_style(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            runtime_dir = base / "runtime"
            runtime_dir.mkdir()

            tag_pdf = base / "hang-tag.pdf"
            wash_pdf = base / "wash-label.pdf"
            tag_image = base / "tag.png"
            wash_image = base / "wash.png"
            exported = base / "summary.xlsx"
            tag_pdf.write_bytes(b"%PDF-fake")
            wash_pdf.write_bytes(b"%PDF-fake")
            tag_image.write_bytes(b"tag")
            wash_image.write_bytes(b"wash")
            from openpyxl import Workbook, load_workbook

            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["PDF文件", "PDF类型", "原始路径", "处理动作", "备注"])
            sheet.append(["hang-tag.pdf", "吊牌/合格证", str(tag_pdf), "等待后端截图", ""])
            sheet.append(["wash-label.pdf", "洗唛", str(wash_pdf), "等待后端截图", ""])
            workbook.save(exported)

            with patch(
                "core.shenhui_pdf_screenshot.extract_pdf_text",
                return_value="",
            ), patch(
                "core.shenhui_pdf_screenshot.convert_pdf_to_yq_images",
                side_effect=[[tag_image], [wash_image]],
            ):
                result = finalize_pdf_batch_screenshot_outputs(
                    data_rows=[
                        {
                            "PDF文件": "hang-tag.pdf",
                            "原始路径": str(tag_pdf),
                            "__pdf_path": str(tag_pdf),
                            "__pdf_type": "hang_tag",
                        },
                        {
                            "PDF文件": "wash-label.pdf",
                            "原始路径": str(wash_pdf),
                            "__pdf_path": str(wash_pdf),
                            "__pdf_type": "wash_label",
                        },
                    ],
                    exported_files=[str(exported)],
                    run_params={
                        "package_name": "PDF截图测试",
                        "style_color_overrides": "hang-tag.pdf=201226109105-00322\nwash-label.pdf=201226109105-00322",
                    },
                    runtime_artifact_dir=str(runtime_dir),
                    log=lambda _: None,
                )

            self.assertEqual(len(result), 2)
            zip_path = Path(result[0])
            self.assertTrue(zip_path.is_file())
            self.assertEqual(Path(result[1]), exported)

            with zipfile.ZipFile(zip_path) as archive:
                names = archive.namelist()
                self.assertTrue(any(name.endswith("PDF截图测试/201226109105/yq(1).png") for name in names))
                self.assertTrue(any(name.endswith("PDF截图测试/201226109105/yq(2).png") for name in names))

            exported_workbook = load_workbook(exported)
            exported_sheet = exported_workbook.active
            self.assertEqual(exported_sheet["D2"].value, "截图完成")
            self.assertIn("生成 1 张截图", exported_sheet["E2"].value)
            self.assertEqual(exported_sheet["D3"].value, "截图完成")
            self.assertIn("生成 1 张截图", exported_sheet["E3"].value)

    def test_finalize_pdf_batch_uses_style_color_prefix_when_style_has_multiple_colors(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            runtime_dir = base / "runtime"
            runtime_dir.mkdir()

            blue_tag_pdf = base / "blue-tag.pdf"
            blue_wash_pdf = base / "blue-wash.pdf"
            red_tag_pdf = base / "red-tag.pdf"
            red_wash_pdf = base / "red-wash.pdf"
            images = [base / f"image-{index}.png" for index in range(4)]
            for pdf_file in (blue_tag_pdf, blue_wash_pdf, red_tag_pdf, red_wash_pdf):
                pdf_file.write_bytes(b"%PDF-fake")
            for image in images:
                image.write_bytes(b"png")

            with patch(
                "core.shenhui_pdf_screenshot.extract_pdf_text",
                return_value="",
            ), patch(
                "core.shenhui_pdf_screenshot.convert_pdf_to_yq_images",
                side_effect=[[image] for image in images],
            ):
                result = finalize_pdf_batch_screenshot_outputs(
                    data_rows=[
                        {"原始路径": str(blue_tag_pdf), "__pdf_path": str(blue_tag_pdf), "__pdf_type": "hang_tag"},
                        {"原始路径": str(blue_wash_pdf), "__pdf_path": str(blue_wash_pdf), "__pdf_type": "wash_label"},
                        {"原始路径": str(red_tag_pdf), "__pdf_path": str(red_tag_pdf), "__pdf_type": "hang_tag"},
                        {"原始路径": str(red_wash_pdf), "__pdf_path": str(red_wash_pdf), "__pdf_type": "wash_label", "__style_code": "201226109105"},
                    ],
                    exported_files=[],
                    run_params={
                        "package_name": "PDF多款色测试",
                        "style_color_overrides": "\n".join([
                            "blue-tag.pdf=201226109105-00322",
                            "blue-wash.pdf=201226109105-00322",
                            "red-tag.pdf=201226109105-88201",
                        ]),
                    },
                    runtime_artifact_dir=str(runtime_dir),
                    log=lambda _: None,
                )

            zip_path = Path(result[0])
            with zipfile.ZipFile(zip_path) as archive:
                names = archive.namelist()
                self.assertTrue(any(name.endswith("PDF多款色测试/201226109105/201226109105-00322yq(1)-1.png") for name in names))
                self.assertTrue(any(name.endswith("PDF多款色测试/201226109105/201226109105-00322yq(2)-1.png") for name in names))
                self.assertTrue(any(name.endswith("PDF多款色测试/201226109105/201226109105-88201yq(1)-1.png") for name in names))
                self.assertTrue(any(name.endswith("PDF多款色测试/201226109105/201226109105yq(2)-1.png") for name in names))

    def test_finalize_pdf_batch_can_merge_outputs_into_existing_style_zip(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            runtime_dir = base / "runtime"
            export_dir = base / "export"
            runtime_dir.mkdir()
            export_dir.mkdir()

            style_zip = export_dir / "201226109105.zip"
            with zipfile.ZipFile(style_zip, "w", compression=zipfile.ZIP_DEFLATED) as archive:
                archive.writestr("model.jpg", b"model")
                archive.writestr("yq(1).png", b"old")

            tag_pdf = base / "hang-tag.pdf"
            wash_pdf = base / "wash-label.pdf"
            tag_image = base / "tag.png"
            wash_image = base / "wash.png"
            exported = base / "summary.xlsx"
            tag_pdf.write_bytes(b"%PDF-fake")
            wash_pdf.write_bytes(b"%PDF-fake")
            tag_image.write_bytes(b"tag")
            wash_image.write_bytes(b"wash")
            exported.write_bytes(b"excel")

            with patch(
                "core.shenhui_pdf_screenshot.extract_pdf_text",
                return_value="",
            ), patch(
                "core.shenhui_pdf_screenshot.convert_pdf_to_yq_images",
                side_effect=[[tag_image], [wash_image]],
            ):
                result = finalize_pdf_batch_screenshot_outputs(
                    data_rows=[
                        {"原始路径": str(tag_pdf), "__pdf_path": str(tag_pdf), "__pdf_type": "hang_tag"},
                        {"原始路径": str(wash_pdf), "__pdf_path": str(wash_pdf), "__pdf_type": "wash_label"},
                    ],
                    exported_files=[str(exported)],
                    run_params={
                        "package_name": "PDF合并测试",
                        "export_folder": str(export_dir),
                        "output_mode": "merge_existing_zips",
                        "style_color_overrides": "hang-tag.pdf=201226109105-00322\nwash-label.pdf=201226109105-00322",
                    },
                    runtime_artifact_dir=str(runtime_dir),
                    log=lambda _: None,
                )

            self.assertIn(str(style_zip), result)
            with zipfile.ZipFile(style_zip) as archive:
                names = archive.namelist()
                self.assertIn("model.jpg", names)
                self.assertIn("yq(1).png", names)
                self.assertIn("yq(2).png", names)
                self.assertEqual(names.count("yq(1).png"), 1)
                self.assertEqual(archive.read("yq(1).png"), b"tag")
                self.assertEqual(archive.read("yq(2).png"), b"wash")


if __name__ == "__main__":
    unittest.main()
