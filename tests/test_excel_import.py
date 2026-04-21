import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from core.api_server import _read_local_excel, _rows_raw_to_table


class ExcelImportTests(unittest.TestCase):
    def test_rows_raw_to_table_keeps_single_header_layout(self):
        table = _rows_raw_to_table([
            ["SPU", "外层站点", "列表时间范围"],
            ["123", "全球", "近7日"],
        ])

        self.assertEqual(table["headers"], ["SPU", "外层站点", "列表时间范围"])
        self.assertEqual(table["rows"][0]["SPU"], "123")
        self.assertEqual(table["rows"][0]["外层站点"], "全球")

    def test_rows_raw_to_table_collapses_grouped_headers(self):
        table = _rows_raw_to_table([
            ["外层站点", "SPU", "流量情况", None, "搜索数据", None],
            [None, None, "曝光量", "点击量", "支付订单数", "支付件数"],
            ["全球", "123", "100", "10", "2", "3"],
        ])

        self.assertEqual(
            table["headers"],
            [
                "外层站点",
                "SPU",
                "流量情况/曝光量",
                "流量情况/点击量",
                "搜索数据/支付订单数",
                "搜索数据/支付件数",
            ],
        )
        self.assertEqual(table["rows"][0]["SPU"], "123")
        self.assertEqual(table["rows"][0]["流量情况/点击量"], "10")

    def test_read_local_excel_round_trips_grouped_header_workbook(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "列表数据"
        ws.merge_cells("A1:A2")
        ws.merge_cells("B1:B2")
        ws.merge_cells("C1:F1")
        ws["A1"] = "外层站点"
        ws["B1"] = "SPU"
        ws["C1"] = "流量情况"
        ws["C2"] = "曝光量"
        ws["D2"] = "点击量"
        ws["E2"] = "支付订单数"
        ws["F2"] = "支付件数"
        ws.append(["全球", "123", "100", "10", "2", "3"])

        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "grouped.xlsx"
            wb.save(path)
            result = _read_local_excel(str(path))

        self.assertEqual(result["sheet_name"], "列表数据")
        self.assertEqual(result["headers"][:4], ["外层站点", "SPU", "流量情况/曝光量", "流量情况/点击量"])
        self.assertEqual(result["rows"][0]["SPU"], "123")
        self.assertEqual(result["rows"][0]["流量情况/支付件数"], "3")

    def test_rows_raw_to_table_ignores_trailing_blank_header_cells(self):
        table = _rows_raw_to_table([
            ["站点", "店铺", "优惠券品类", None, None],
            ["越南", "balabala2023.vn", "商店优惠券", None, None],
            ["泰国", "semir2022.th", "关注礼优惠券", None, None],
        ])

        self.assertEqual(table["headers"], ["站点", "店铺", "优惠券品类", "列4", "列5"])
        self.assertEqual(table["rows"][0]["站点"], "越南")
        self.assertEqual(table["rows"][0]["店铺"], "balabala2023.vn")
        self.assertEqual(table["rows"][1]["优惠券品类"], "关注礼优惠券")

    def test_shopee_template_exposes_custom_name_and_code_after_store(self):
        path = Path("adapters/shopee-plus-v2/Shopee优惠券批量设置模板（26.4.1）.xlsx")
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb[wb.sheetnames[0]]
            headers = [cell for cell in next(ws.iter_rows(values_only=True))]
        finally:
            wb.close()

        self.assertEqual(headers[:6], [
            "站点",
            "店铺",
            "优惠券名称（可选）",
            "优惠码（可选，最多5位 A-Z/0-9）",
            "优惠券领取期限（开始）精确到分",
            "优惠券领取期限（结束）精确到分",
        ])
        self.assertEqual(
            headers,
            [
                "站点",
                "店铺",
                "优惠券名称（可选）",
                "优惠码（可选，最多5位 A-Z/0-9）",
                "优惠券领取期限（开始）精确到分",
                "优惠券领取期限（结束）精确到分",
                "是否提前显示\n优惠券（是：1/否：0）",
                "优惠券品类",
                "奖励类型",
                "折扣类型",
                "优惠限额",
                "最高优惠金额(无限制填：nocap)",
                "最低消费金额",
                "可使用总数",
                "每个买家可用的优惠券数量上限",
                "是否覆盖已有券（如果有冲突日期券，则取消旧券以这次新建的为准；是：1/否：0）",
            ],
        )


if __name__ == "__main__":
    unittest.main()
