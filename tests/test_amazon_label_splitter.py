import os
import tempfile
import unittest
import zipfile
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from core.amazon_label_splitter import (
    extract_fnskus_from_text,
    normalize_mapping_rows,
    split_amazon_label_rows,
)
from core import data_sink


SAMPLE_DIR = Path("/Users/xingyicheng/Downloads/亚马逊标签测试-for易成")
SAMPLE_XLSX = SAMPLE_DIR / "FNSKU-SKU-标签映射表.xlsx"
SAMPLE_PDF = SAMPLE_DIR / "products.pdf"


def read_sample_mapping_rows():
    workbook = load_workbook(SAMPLE_XLSX, data_only=True, read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    headers = [str(value or "").strip() for value in rows[0]]
    result = []
    for raw_row in rows[1:]:
        row = {
            headers[index]: "" if value is None else str(value)
            for index, value in enumerate(raw_row)
            if index < len(headers)
        }
        if any(str(value or "").strip() for value in row.values()):
            result.append({
                "fnsku": row.get("FNSKU", ""),
                "sku": row.get("SKU", ""),
                "labelName": row.get("标签名称", ""),
            })
    return result


class AmazonLabelSplitterTests(unittest.TestCase):
    def test_extract_fnskus_from_text_dedupes_codes(self):
        self.assertEqual(
            extract_fnskus_from_text("x004qzeqzv\nX004QZEQZV\nX004QZFUF1"),
            ["X004QZEQZV", "X004QZFUF1"],
        )

    def test_normalize_mapping_rows_accepts_adapter_shape(self):
        mappings = normalize_mapping_rows([
            {"fnsku": "x004qzeqzv", "sku": "9950020568060", "labelName": "9950020568060+X004QZEQZV"},
        ])

        self.assertEqual(mappings["X004QZEQZV"]["sku"], "9950020568060")
        self.assertEqual(mappings["X004QZEQZV"]["label_name"], "9950020568060+X004QZEQZV")

    @unittest.skipUnless(SAMPLE_XLSX.is_file() and SAMPLE_PDF.is_file(), "sample Amazon label files are not available")
    def test_split_sample_pdf_by_fnsku_mapping(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            result_rows, refs = split_amazon_label_rows(
                data_rows=[{
                    "PDF文件": SAMPLE_PDF.name,
                    "__pdf_path": str(SAMPLE_PDF),
                    "__mapping_rows": read_sample_mapping_rows(),
                }],
                run_params={"package_name": "亚马逊标签测试"},
                runtime_artifact_dir=tmpdir,
                log=lambda _: None,
            )

            self.assertEqual(len(result_rows), 6)
            self.assertTrue(all(row["匹配结果"] == "已拆分" for row in result_rows))
            self.assertEqual(
                [row["识别FNSKU"] for row in result_rows],
                ["X004QZEQZV", "X004QZFUF1", "X004QZFUEH", "X004QZK1EL", "X004QZFS5D", "X004QZ88TL"],
            )
            self.assertEqual(
                [Path(row["输出PDF"]).name for row in result_rows],
                [
                    "9950020568060+X004QZEQZV.pdf",
                    "9950020567940+X004QZFUF1.pdf",
                    "9950020568237+X004QZFUEH.pdf",
                    "9950020568299+X004QZK1EL.pdf",
                    "9950020568145+X004QZFS5D.pdf",
                    "9950020568497+X004QZ88TL.pdf",
                ],
            )

            for row in result_rows:
                output_path = Path(row["输出PDF"])
                self.assertTrue(output_path.is_file())
                self.assertGreater(output_path.stat().st_size, 0)

            zip_path = Path(refs[0])
            self.assertTrue(zip_path.is_file())
            with zipfile.ZipFile(zip_path) as archive:
                names = archive.namelist()
                self.assertTrue(any(name.endswith("亚马逊标签测试/9950020568060+X004QZEQZV.pdf") for name in names))

            with patch.dict(os.environ, {"CRAWSHRIMP_DATA": tmpdir}):
                exported = data_sink.export_excel(
                    result_rows,
                    "amazon-ops-assistant",
                    "amazon_label_batch_process",
                    "amazon-test-{timestamp}.xlsx",
                    column_order=["PDF文件", "页码", "识别FNSKU", "SKU", "标签名称", "匹配结果", "输出PDF", "备注"],
                )
            exported_workbook = load_workbook(exported, read_only=True)
            sheet = exported_workbook.active
            self.assertEqual(sheet.max_row, 7)
            self.assertEqual(sheet["C2"].value, "X004QZEQZV")
            self.assertEqual(sheet["F2"].value, "已拆分")


if __name__ == "__main__":
    unittest.main()
