import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook

from core import odps_sync


class OdpsSyncTest(unittest.TestCase):
    def test_build_payload_maps_temu_mall_flux_excel_to_task_table(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "mall_flux.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["平台名称", "店铺名称", "店铺ID", "日期", "总数据/总浏览量"])
            ws.append(["Temu", "balabala Official Shop", "D80421", "2026-05-18", "4414"])
            wb.save(path)

            payload = odps_sync.build_sync_payload(
                adapter_id="temu",
                task_id="mall_flux",
                file_path=str(path),
            )

        self.assertEqual(payload["table_name"], "imp_ods_temu_mall_flux")
        self.assertEqual(payload["write_mode"], "append")
        self.assertEqual(payload["partition_spec"], {"dt": "2026-05-18"})
        self.assertEqual(payload["data"][0]["shop_id"], "D80421")
        self.assertEqual(payload["data"][0]["total_views"], 4414)
        field_names = [field["name"] for field in payload["fields"]]
        self.assertEqual(field_names[:5], ["platform_name", "shop_name", "shop_id", "stat_date", "total_views"])
        field_types = {field["name"]: field["type"] for field in payload["fields"]}
        self.assertEqual(field_types["shop_id"], "string")
        self.assertEqual(field_types["total_views"], "bigint")
        self.assertEqual(payload["fields"][4]["comment"], "总数据/总浏览量")

    def test_sync_file_posts_write_odps_payload_to_gateway_endpoint_with_app_code(self):
        captured = {}

        def fake_post_json(url, payload, timeout=30, headers=None):
            captured["url"] = url
            captured["payload"] = payload
            captured["timeout"] = timeout
            captured["headers"] = headers
            return {"success": True, "count": 1}

        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "mall_flux.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["平台名称", "店铺名称", "店铺ID", "日期"])
            ws.append(["Temu", "balabala Official Shop", "D80421", "2026-05-18"])
            wb.save(path)

            with patch.object(odps_sync, "_post_json", side_effect=fake_post_json):
                result = odps_sync.sync_file(
                    adapter_id="temu",
                    task_id="mall_flux",
                    file_path=str(path),
                    endpoint="http://dataworksapi.semirapp.com",
                    app_code="secret-code",
                )

        self.assertTrue(result["ok"])
        self.assertEqual(captured["url"], "http://dataworksapi.semirapp.com/api/v1/dataworks/write_odps")
        self.assertEqual(captured["headers"]["Authorization"], "APPCODE secret-code")
        self.assertEqual(captured["headers"]["Content-Type"], "application/json")
        self.assertEqual(captured["payload"]["table_name"], "imp_ods_temu_mall_flux")
        self.assertEqual(captured["payload"]["data"][0]["platform_name"], "Temu")

    def test_full_write_odps_url_is_preserved(self):
        url = odps_sync.resolve_write_odps_url("http://dataworksapi.semirapp.com/api/v1/dataworks/write_odps")
        self.assertEqual(url, "http://dataworksapi.semirapp.com/api/v1/dataworks/write_odps")

    def test_default_dataworks_endpoint_is_the_semir_gateway(self):
        self.assertEqual(
            odps_sync.DEFAULT_DATAWORKS_ENDPOINT,
            "http://dataworksapi.semirapp.com/api/v1/dataworks/write_odps",
        )

    def test_sync_file_requires_app_code(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "mall_flux.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["平台名称", "店铺名称", "店铺ID", "日期"])
            ws.append(["Temu", "balabala Official Shop", "D80421", "2026-05-18"])
            wb.save(path)

            with self.assertRaisesRegex(odps_sync.OdpsSyncError, "ODPS AppCode"):
                odps_sync.sync_file(
                    adapter_id="temu",
                    task_id="mall_flux",
                    file_path=str(path),
                    endpoint=odps_sync.DEFAULT_DATAWORKS_ENDPOINT,
                    app_code="",
                )

    def test_stat_date_range_is_synced_as_string_not_datetime(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "mall_flux.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["统计日期范围", "日期", "抓取时间"])
            ws.append(["2026-05-13 ~ 2026-05-19", "2026-05-18", "2026-05-19 12:34:56"])
            wb.save(path)

            payload = odps_sync.build_sync_payload(
                adapter_id="temu",
                task_id="mall_flux",
                file_path=str(path),
            )

        field_types = {field["name"]: field["type"] for field in payload["fields"]}
        self.assertEqual(field_types["stat_date_range"], "string")
        self.assertEqual(field_types["stat_date"], "string")
        self.assertEqual(field_types["captured_at"], "datetime")

    def test_build_payload_maps_tiktok_product_analytics_excel_to_task_table(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "tiktok_product_analytics.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["平台名称", "区域", "店铺ID", "统计日期范围", "抓取时间", "订单数", "商品曝光次数", "商品点击量"])
            ws.append(["TikTok", "US", "7496042382582647544", "2026-05-27 ~ 2026-06-02", "2026-06-03 12:00:00", "295", "751050", "23884"])
            wb.save(path)

            payload = odps_sync.build_sync_payload(
                adapter_id="tiktok-ops-assistant",
                task_id="product_analytics",
                file_path=str(path),
            )

        self.assertEqual(payload["table_name"], "imp_ods_tiktok_product_analytics")
        self.assertEqual(payload["partition_spec"], {"dt": "2026-05-27"})
        self.assertEqual(payload["data"][0]["platform_name"], "TikTok")
        self.assertEqual(payload["data"][0]["region"], "US")
        self.assertEqual(payload["data"][0]["shop_id"], "7496042382582647544")
        self.assertEqual(payload["data"][0]["orders"], 295)
        self.assertEqual(payload["data"][0]["product_impressions"], 751050)
        self.assertEqual(payload["data"][0]["product_clicks"], 23884)
        field_types = {field["name"]: field["type"] for field in payload["fields"]}
        self.assertEqual(field_types["stat_date_range"], "string")
        self.assertEqual(field_types["captured_at"], "datetime")
        self.assertEqual(field_types["orders"], "bigint")

    def test_build_payload_maps_shopify_traffic_data_excel_to_task_table(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "shopify_traffic_data.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append([
                "平台名称",
                "店铺标识",
                "报告名称",
                "数据类型",
                "统计日期范围",
                "统计日期",
                "在线商店访客",
                "访问",
                "访问变化",
                "抓取时间",
            ])
            ws.append([
                "Shopify",
                "balabala-global",
                "访问随时间变化",
                "汇总",
                "2026-05-27 ~ 2026-06-03",
                "",
                "72,471",
                "84,983",
                "2%",
                "2026-06-03 19:20:00",
            ])
            wb.save(path)

            payload = odps_sync.build_sync_payload(
                adapter_id="shopify-ops-assistant",
                task_id="traffic_data",
                file_path=str(path),
            )

        self.assertEqual(payload["table_name"], "imp_ods_shopify_traffic_data")
        self.assertEqual(payload["partition_spec"], {"dt": "2026-05-27"})
        self.assertEqual(payload["data"][0]["platform_name"], "Shopify")
        self.assertEqual(payload["data"][0]["store_key"], "balabala-global")
        self.assertEqual(payload["data"][0]["report_name"], "访问随时间变化")
        self.assertEqual(payload["data"][0]["data_type"], "汇总")
        self.assertEqual(payload["data"][0]["online_store_visitors"], 72471)
        self.assertEqual(payload["data"][0]["sessions"], 84983)
        self.assertEqual(payload["data"][0]["sessions_change"], "2%")
        field_types = {field["name"]: field["type"] for field in payload["fields"]}
        self.assertEqual(field_types["stat_date_range"], "string")
        self.assertEqual(field_types["stat_date"], "string")
        self.assertEqual(field_types["captured_at"], "datetime")

    def test_build_payload_maps_aliexpress_deal_analysis_excel_to_task_table(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "aliexpress_deal_analysis.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["平台名称", "店铺名称", "channelId", "数据类型", "统计日期", "统计日期范围", "指标编码", "指标名称", "指标值", "环比值", "同比值", "抓取时间"])
            ws.append(["AliExpress", "Semir Official Store", "125417", "核心指标", "2026-06-01", "2026-06-01 ~ 2026-06-01", "payAmt", "支付金额", "18811.33", "7.5362", "4.2461", "2026-06-03 18:20:30"])
            wb.save(path)

            payload = odps_sync.build_sync_payload(
                adapter_id="aliexpress-ops-assistant",
                task_id="deal_analysis",
                file_path=str(path),
            )

        self.assertEqual(payload["table_name"], "imp_ods_aliexpress_deal_analysis")
        self.assertEqual(payload["partition_spec"], {"dt": "2026-06-01"})
        self.assertEqual(payload["data"][0]["platform_name"], "AliExpress")
        self.assertEqual(payload["data"][0]["channel_id"], "125417")
        self.assertEqual(payload["data"][0]["metric_code"], "payAmt")
        self.assertEqual(payload["data"][0]["metric_value"], 18811.33)
        field_types = {field["name"]: field["type"] for field in payload["fields"]}
        self.assertEqual(field_types["stat_date"], "string")
        self.assertEqual(field_types["stat_date_range"], "string")
        self.assertEqual(field_types["metric_value"], "double")
        self.assertEqual(field_types["captured_at"], "datetime")

    def test_build_payload_maps_aliexpress_product_ranking_excel_to_task_table(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "aliexpress_product_ranking.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["平台名称", "店铺名称", "channelId", "榜单类型", "统计日期", "商品ID", "排行", "商品名称", "支付金额", "支付金额环比", "商品访客数", "抓取时间"])
            ws.append(["AliExpress", "Semir Official Store", "125417", "支付榜", "2026-06-01", "1005000000000000", "1", "Semir item", "100.5", "12.00%", "50", "2026-06-03 18:20:30"])
            wb.save(path)

            payload = odps_sync.build_sync_payload(
                adapter_id="aliexpress-ops-assistant",
                task_id="product_ranking",
                file_path=str(path),
            )

        self.assertEqual(payload["table_name"], "imp_ods_aliexpress_product_ranking")
        self.assertEqual(payload["partition_spec"], {"dt": "2026-06-01"})
        self.assertEqual(payload["data"][0]["platform_name"], "AliExpress")
        self.assertEqual(payload["data"][0]["item_id"], "1005000000000000")
        self.assertEqual(payload["data"][0]["rank_no"], 1)
        self.assertEqual(payload["data"][0]["pay_amt"], 100.5)
        field_types = {field["name"]: field["type"] for field in payload["fields"]}
        self.assertEqual(field_types["stat_date"], "string")
        self.assertEqual(field_types["item_id"], "string")
        self.assertEqual(field_types["rank_no"], "bigint")
        self.assertEqual(field_types["pay_amt"], "double")

    def test_build_payload_flattens_shopee_business_analysis_workbook(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "shopee_business_analysis.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "关键指标"
            ws.append(["统计时间", "总销售额（已下订单） (CNY)", "总订单数（已确认订单）", "商品点击量"])
            ws.append(["2026-06-03-2026-06-03", "45,190.88", "251", "20588"])
            trend = wb.create_sheet("趋势指标")
            trend.append(["统计时间", "总销售额（已下订单） (CNY)", "总订单数（已确认订单）"])
            trend.append(["2026-06-03 00:00", "2,391.86", "23"])
            shop_share = wb.create_sheet("店铺维度指标分析")
            shop_share.append(["店铺名称", "总销售额（已下订单） (CNY)", "总订单数（已确认订单）"])
            shop_share.append(["Balabala Official Store", "77.08%", "75.86%"])
            detail = wb.create_sheet("详情")
            detail.append(["店铺名称", "市场", "总销售额（已确认订单） (CNY)", "总订单数（已确认订单）"])
            detail.append(["Balabala Official Store", "VN", "29,990.60", "190"])
            wb.save(path)

            payload = odps_sync.build_sync_payload(
                adapter_id="shopee-plus-v2",
                task_id="business_analysis",
                file_path=str(path),
            )

        self.assertEqual(payload["table_name"], "imp_ods_shopee_business_analysis")
        self.assertEqual(payload["partition_spec"], {"dt": "2026-06-03"})
        self.assertEqual(payload["write_mode"], "append")
        self.assertEqual(payload["data"][0]["platform_name"], "Shopee")
        self.assertEqual(payload["data"][0]["sheet_name"], "关键指标")
        self.assertEqual(payload["data"][0]["stat_time"], "2026-06-03-2026-06-03")
        self.assertEqual(payload["data"][0]["metric_name"], "总销售额")
        self.assertEqual(payload["data"][0]["order_status"], "已下订单")
        self.assertEqual(payload["data"][0]["currency"], "CNY")
        self.assertEqual(payload["data"][0]["metric_value"], 45190.88)
        share_rows = [row for row in payload["data"] if row["sheet_name"] == "店铺维度指标分析"]
        self.assertEqual(share_rows[0]["dimension_type"], "shop_share")
        self.assertEqual(share_rows[0]["metric_value"], 77.08)
        detail_rows = [row for row in payload["data"] if row["sheet_name"] == "详情"]
        self.assertEqual(detail_rows[0]["shop_name"], "Balabala Official Store")
        self.assertEqual(detail_rows[0]["market"], "VN")
        self.assertEqual(detail_rows[0]["order_status"], "已确认订单")
        field_types = {field["name"]: field["type"] for field in payload["fields"]}
        self.assertEqual(field_types["metric_value"], "double")
        self.assertEqual(field_types["captured_at"], "datetime")

    def test_build_payload_flattens_lazada_business_advisor_workbook(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "Lazada生意参谋_PH_2026-06-02~2026-06-02.xls"
            wb = Workbook()
            ws = wb.active
            ws.title = "Key Metrics"
            ws.append(["Data Source: Lazada - Business Advisor - Dashboard  Date Range :02-06-2026~02-06-2026"])
            ws.append([])
            ws.append(["Date", "Revenue", "Visitors"])
            ws.append(["2026-06-02~2026-06-02", "1,234.56", "789"])
            ws.append(["2026-06-02", "100.00", "10"])
            metric = wb.create_sheet("Metric Definition")
            metric.append(["Metric Name", "Definition"])
            metric.append(["Revenue", "Total gross merchandise value"])
            wb.save(path)

            payload = odps_sync.build_sync_payload(
                adapter_id="lazada-plus-v1",
                task_id="business_advisor",
                file_path=str(path),
            )

        self.assertEqual(payload["table_name"], "imp_ods_lazada_business_advisor")
        self.assertEqual(payload["partition_spec"], {"dt": "2026-06-02"})
        self.assertEqual(payload["write_mode"], "append")
        self.assertEqual(payload["data"][0]["platform_name"], "Lazada")
        self.assertEqual(payload["data"][0]["sheet_name"], "Key Metrics")
        self.assertEqual(payload["data"][0]["country_code"], "PH")
        self.assertEqual(payload["data"][0]["stat_date_range"], "2026-06-02 ~ 2026-06-02")
        self.assertEqual(payload["data"][0]["metric_name"], "Revenue")
        self.assertEqual(payload["data"][0]["metric_value"], 1234.56)
        visitor_rows = [row for row in payload["data"] if row["metric_name"] == "Visitors"]
        self.assertEqual(visitor_rows[0]["metric_value"], 789)
        definition_rows = [row for row in payload["data"] if row["sheet_name"] == "Metric Definition"]
        self.assertEqual(definition_rows[0]["metric_name"], "Revenue")
        self.assertEqual(definition_rows[0]["definition"], "Total gross merchandise value")
        field_types = {field["name"]: field["type"] for field in payload["fields"]}
        self.assertEqual(field_types["metric_value"], "double")
        self.assertEqual(field_types["captured_at"], "datetime")


if __name__ == "__main__":
    unittest.main()
