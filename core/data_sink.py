"""
Data persistence layer
- Task run state stored in SQLite
- Data records exported to Excel / JSON
- Filename templates support {date}, {datetime}, {timestamp}, {adapter_id}, {task_id}
"""
import json
import logging
import os
import re
import sqlite3
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, List, Mapping, Optional

from core import runtime_paths
from core.models import TaskRun, TaskStatus

logger = logging.getLogger(__name__)

MAX_EXPORT_FILENAME_LENGTH = 140


class _SafeTemplateVars(dict):
    def __missing__(self, key):
        return ""


def _data_root() -> Path:
    return runtime_paths.child_dir("data")


def _db_path() -> Path:
    return runtime_paths.data_root() / "crawshrimp.db"


def _get_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(str(_db_path()))
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    """Create tables if not exists"""
    with _get_conn() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS task_runs (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                adapter_id   TEXT NOT NULL,
                task_id      TEXT NOT NULL,
                status       TEXT NOT NULL DEFAULT 'idle',
                started_at   TEXT,
                finished_at  TEXT,
                records_count INTEGER DEFAULT 0,
                error        TEXT,
                output_files TEXT DEFAULT '[]'
            )
        """)
        _ensure_column(conn, "task_runs", "last_seen_at", "TEXT")
        _ensure_column(conn, "task_runs", "phase", "TEXT DEFAULT ''")
        _ensure_column(conn, "task_runs", "current_row", "INTEGER DEFAULT 0")
        conn.execute("""
            CREATE INDEX IF NOT EXISTS idx_task_runs_adapter_task
            ON task_runs (adapter_id, task_id)
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS task_instances (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                instance_uid  TEXT NOT NULL UNIQUE,
                adapter_id    TEXT NOT NULL,
                task_id       TEXT NOT NULL,
                title         TEXT NOT NULL,
                status        TEXT NOT NULL DEFAULT 'draft',
                current_step  TEXT NOT NULL DEFAULT 'config',
                params_json   TEXT NOT NULL DEFAULT '{}',
                summary_json  TEXT NOT NULL DEFAULT '{}',
                last_run_id   INTEGER,
                archived      INTEGER NOT NULL DEFAULT 0,
                created_at    TEXT NOT NULL,
                updated_at    TEXT NOT NULL,
                completed_at  TEXT
            )
        """)
        conn.execute("""
            CREATE INDEX IF NOT EXISTS idx_task_instances_adapter_task_status
            ON task_instances (adapter_id, task_id, status, updated_at)
        """)
        conn.execute("""
            CREATE INDEX IF NOT EXISTS idx_task_instances_updated
            ON task_instances (updated_at)
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS task_instance_runs (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                instance_uid  TEXT NOT NULL,
                run_id        INTEGER NOT NULL,
                purpose       TEXT NOT NULL DEFAULT 'main',
                created_at    TEXT NOT NULL,
                UNIQUE(instance_uid, run_id, purpose)
            )
        """)
        conn.execute("""
            CREATE INDEX IF NOT EXISTS idx_task_instance_runs_instance
            ON task_instance_runs (instance_uid, created_at)
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS task_instance_artifacts (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                instance_uid  TEXT NOT NULL,
                kind          TEXT NOT NULL,
                label         TEXT NOT NULL,
                path          TEXT NOT NULL,
                meta_json     TEXT NOT NULL DEFAULT '{}',
                created_at    TEXT NOT NULL
            )
        """)
        conn.execute("""
            CREATE INDEX IF NOT EXISTS idx_task_instance_artifacts_instance
            ON task_instance_artifacts (instance_uid, created_at)
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS task_instance_events (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                instance_uid  TEXT NOT NULL,
                event_type    TEXT NOT NULL,
                message       TEXT NOT NULL,
                meta_json     TEXT NOT NULL DEFAULT '{}',
                created_at    TEXT NOT NULL
            )
        """)
        conn.execute("""
            CREATE INDEX IF NOT EXISTS idx_task_instance_events_instance
            ON task_instance_events (instance_uid, created_at)
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS task_schedules (
                id                 INTEGER PRIMARY KEY AUTOINCREMENT,
                schedule_uid       TEXT NOT NULL UNIQUE,
                adapter_id         TEXT NOT NULL,
                task_id            TEXT NOT NULL,
                title              TEXT NOT NULL,
                enabled            INTEGER NOT NULL DEFAULT 1,
                frequency          TEXT NOT NULL,
                time_of_day        TEXT NOT NULL,
                weekday            INTEGER,
                params_json        TEXT NOT NULL DEFAULT '{}',
                notify_channel     TEXT NOT NULL DEFAULT 'dingtalk',
                notify_template    TEXT NOT NULL DEFAULT '',
                last_run_id        INTEGER,
                last_instance_uid  TEXT,
                last_status        TEXT,
                last_error         TEXT,
                last_triggered_at  TEXT,
                archived           INTEGER NOT NULL DEFAULT 0,
                created_at         TEXT NOT NULL,
                updated_at         TEXT NOT NULL
            )
        """)
        conn.execute("""
            CREATE INDEX IF NOT EXISTS idx_task_schedules_adapter_task_enabled
            ON task_schedules (adapter_id, task_id, enabled, archived, updated_at)
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS cloud_machine_credentials (
                id INTEGER PRIMARY KEY CHECK (id = 1),
                machine_id TEXT NOT NULL DEFAULT '',
                machine_token TEXT NOT NULL DEFAULT '',
                machine_name TEXT NOT NULL DEFAULT '',
                capabilities_json TEXT NOT NULL DEFAULT '[]',
                updated_at TEXT NOT NULL
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS cloud_job_events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                job_uid TEXT NOT NULL,
                event_type TEXT NOT NULL,
                message TEXT NOT NULL DEFAULT '',
                payload_json TEXT NOT NULL DEFAULT '{}',
                created_at TEXT NOT NULL
            )
        """)
        conn.commit()


def _ensure_column(conn: sqlite3.Connection, table_name: str, column_name: str, definition: str) -> None:
    columns = {
        str(row["name"])
        for row in conn.execute(f"PRAGMA table_info({table_name})").fetchall()
    }
    if column_name not in columns:
        conn.execute(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {definition}")


def _now_iso() -> str:
    return datetime.now().isoformat()


def _json_dumps(value: Any) -> str:
    return json.dumps(value or {}, ensure_ascii=False)


def _row_to_dict(row: Optional[sqlite3.Row]) -> Optional[dict]:
    return dict(row) if row else None


def _json_loads_object(value: Any) -> dict:
    try:
        parsed = json.loads(value) if isinstance(value, str) else value
    except Exception:
        return {}
    return parsed if isinstance(parsed, dict) else {}


def save_cloud_machine_credentials(machine_id: str, machine_token: str, machine_name: str, capabilities: list[str]) -> dict:
    now = _now_iso()
    with _get_conn() as conn:
        conn.execute(
            """
            INSERT INTO cloud_machine_credentials (
                id, machine_id, machine_token, machine_name, capabilities_json, updated_at
            )
            VALUES (1, ?, ?, ?, ?, ?)
            ON CONFLICT(id) DO UPDATE SET
                machine_id = excluded.machine_id,
                machine_token = excluded.machine_token,
                machine_name = excluded.machine_name,
                capabilities_json = excluded.capabilities_json,
                updated_at = excluded.updated_at
            """,
            (machine_id, machine_token, machine_name, json.dumps(capabilities or [], ensure_ascii=False), now),
        )
        conn.commit()
    return get_cloud_machine_credentials() or {}


def get_cloud_machine_credentials() -> dict | None:
    with _get_conn() as conn:
        row = conn.execute("SELECT * FROM cloud_machine_credentials WHERE id = 1").fetchone()
    if not row:
        return None
    data = dict(row)
    try:
        data["capabilities"] = json.loads(data.pop("capabilities_json") or "[]")
    except Exception:
        data["capabilities"] = []
    return data


def clear_cloud_machine_credentials() -> None:
    with _get_conn() as conn:
        conn.execute("DELETE FROM cloud_machine_credentials WHERE id = 1")
        conn.commit()


def record_cloud_job_event(job_uid: str, event_type: str, message: str = "", payload: Optional[Mapping[str, Any]] = None) -> dict:
    now = _now_iso()
    with _get_conn() as conn:
        cursor = conn.execute(
            """
            INSERT INTO cloud_job_events (job_uid, event_type, message, payload_json, created_at)
            VALUES (?, ?, ?, ?, ?)
            """,
            (job_uid, event_type, message, _json_dumps(dict(payload or {})), now),
        )
        conn.commit()
        event_id = cursor.lastrowid
        row = conn.execute("SELECT * FROM cloud_job_events WHERE id = ?", (event_id,)).fetchone()
    data = dict(row)
    data["payload"] = _json_loads_object(data.pop("payload_json"))
    return data


def list_cloud_job_events(job_uid: str, limit: int = 100) -> list[dict]:
    with _get_conn() as conn:
        rows = conn.execute(
            """
            SELECT * FROM cloud_job_events
            WHERE job_uid = ?
            ORDER BY id DESC
            LIMIT ?
            """,
            (job_uid, max(1, int(limit or 100))),
        ).fetchall()
    result = []
    for row in reversed(rows):
        data = dict(row)
        data["payload"] = _json_loads_object(data.pop("payload_json"))
        result.append(data)
    return result


def create_task_instance(adapter_id: str, task_id: str, title: str, params: Optional[Mapping[str, Any]] = None) -> dict:
    """Insert a draft task instance and return the inserted row."""
    now = _now_iso()
    instance_uid = uuid.uuid4().hex
    with _get_conn() as conn:
        conn.execute("""
            INSERT INTO task_instances (
                instance_uid, adapter_id, task_id, title, status, current_step,
                params_json, summary_json, archived, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, 'draft', 'config', ?, '{}', 0, ?, ?)
        """, (
            instance_uid,
            str(adapter_id or "").strip(),
            str(task_id or "").strip(),
            str(title or "").strip() or "未命名任务",
            _json_dumps(dict(params or {})),
            now,
            now,
        ))
        conn.commit()
    return get_task_instance(instance_uid) or {}


def get_task_instance(instance_uid: str) -> Optional[dict]:
    """Return one task instance row by uid."""
    with _get_conn() as conn:
        row = conn.execute("""
            SELECT *
            FROM task_instances
            WHERE instance_uid=?
            LIMIT 1
        """, (str(instance_uid or "").strip(),)).fetchone()
        return _row_to_dict(row)


def get_task_instance_detail(instance_uid: str) -> dict:
    """Return one task instance plus parsed params, parsed summary, runs, artifacts, and events."""
    uid = str(instance_uid or "").strip()
    instance = get_task_instance(uid)
    if not instance:
        return {}

    with _get_conn() as conn:
        runs = conn.execute("""
            SELECT tir.*, tr.adapter_id, tr.task_id, tr.status, tr.started_at, tr.finished_at,
                   tr.records_count, tr.error, tr.output_files
            FROM task_instance_runs tir
            LEFT JOIN task_runs tr ON tr.id = tir.run_id
            WHERE tir.instance_uid=?
            ORDER BY tir.id DESC
        """, (uid,)).fetchall()
        artifacts = conn.execute("""
            SELECT *
            FROM task_instance_artifacts
            WHERE instance_uid=?
            ORDER BY id ASC
        """, (uid,)).fetchall()
        events = conn.execute("""
            SELECT *
            FROM task_instance_events
            WHERE instance_uid=?
            ORDER BY id ASC
        """, (uid,)).fetchall()

    detail = dict(instance)
    detail["params"] = _json_loads_object(detail.get("params_json"))
    detail["summary"] = _json_loads_object(detail.get("summary_json"))
    detail["runs"] = [dict(row) for row in runs]
    detail["artifacts"] = [dict(row) for row in artifacts]
    detail["events"] = [dict(row) for row in events]
    for artifact in detail["artifacts"]:
        artifact["meta"] = _json_loads_object(artifact.get("meta_json"))
    for event in detail["events"]:
        event["meta"] = _json_loads_object(event.get("meta_json"))
    return detail


def list_task_instances(
    status_group: str = "",
    adapter_id: str = "",
    task_id: str = "",
    keyword: str = "",
    limit: int = 100,
) -> list[dict]:
    """Return task instances filtered by status group and metadata."""
    status_groups = {
        "current": ("draft", "queued", "running", "generating", "creating", "waiting_approval"),
        "pending": ("waiting_approval", "failed", "create_failed", "partial_failed"),
        "history": ("completed", "stopped", "archived"),
    }
    clauses = []
    params: list[Any] = []

    group = str(status_group or "").strip()
    statuses = status_groups.get(group)
    if statuses:
        placeholders = ",".join("?" for _ in statuses)
        if group == "history":
            clauses.append(f"(status IN ({placeholders}) OR archived=1)")
        else:
            clauses.append(f"status IN ({placeholders})")
            clauses.append("archived=0")
        params.extend(statuses)
    elif group:
        clauses.append("status=?")
        clauses.append("archived=0")
        params.append(group)
    else:
        clauses.append("archived=0")

    adapter = str(adapter_id or "").strip()
    if adapter:
        clauses.append("adapter_id=?")
        params.append(adapter)

    task = str(task_id or "").strip()
    if task:
        clauses.append("task_id=?")
        params.append(task)

    term = str(keyword or "").strip()
    if term:
        clauses.append("(title LIKE ? OR instance_uid LIKE ? OR task_id LIKE ?)")
        like = f"%{term}%"
        params.extend([like, like, like])

    try:
        safe_limit = max(1, min(int(limit), 500))
    except Exception:
        safe_limit = 100

    where = " AND ".join(clauses) if clauses else "1=1"
    with _get_conn() as conn:
        rows = conn.execute(f"""
            SELECT *
            FROM task_instances
            WHERE {where}
            ORDER BY updated_at DESC, id DESC
            LIMIT ?
        """, [*params, safe_limit]).fetchall()
        return [dict(row) for row in rows]


def find_task_instance_by_approval_batch_id(batch_id: str) -> Optional[dict]:
    """Return the first task instance whose summary stores the approval batch id."""
    target = str(batch_id or "").strip()
    if not target:
        return None
    with _get_conn() as conn:
        rows = conn.execute("""
            SELECT *
            FROM task_instances
            WHERE summary_json LIKE ?
            ORDER BY updated_at DESC, id DESC
            LIMIT 20
        """, (f"%{target}%",)).fetchall()
    for row in rows:
        item = dict(row)
        summary = _json_loads_object(item.get("summary_json"))
        if str(summary.get("approval_batch_id") or "").strip() == target:
            return item
    return None


def update_task_instance(instance_uid: str, **fields) -> dict:
    """Update allowed task instance fields and return the updated row."""
    allowed = {
        "title",
        "status",
        "current_step",
        "params_json",
        "summary_json",
        "last_run_id",
        "archived",
        "completed_at",
    }
    updates = {}
    for key, value in fields.items():
        if key == "params":
            updates["params_json"] = _json_dumps(value)
        elif key == "summary":
            updates["summary_json"] = _json_dumps(value)
        elif key == "archived":
            archived = 1 if value else 0
            updates["archived"] = archived
            if archived:
                updates.setdefault("status", "archived")
                updates.setdefault("completed_at", _now_iso())
        elif key in allowed:
            updates[key] = value

    status = str(updates.get("status") or "").strip()
    if status in {"completed", "stopped", "archived"} and not updates.get("completed_at"):
        updates["completed_at"] = _now_iso()

    updates["updated_at"] = _now_iso()
    assignments = ", ".join(f"{key}=?" for key in updates.keys())
    values = list(updates.values())
    uid = str(instance_uid or "").strip()
    with _get_conn() as conn:
        conn.execute(
            f"UPDATE task_instances SET {assignments} WHERE instance_uid=?",
            [*values, uid],
        )
        conn.commit()
    return get_task_instance(uid) or {}


def link_task_instance_run(instance_uid: str, run_id: int, purpose: str = "main") -> None:
    """Associate a task_runs row with a task instance."""
    uid = str(instance_uid or "").strip()
    rid = int(run_id)
    now = _now_iso()
    with _get_conn() as conn:
        conn.execute("""
            INSERT OR IGNORE INTO task_instance_runs (instance_uid, run_id, purpose, created_at)
            VALUES (?, ?, ?, ?)
        """, (uid, rid, str(purpose or "main").strip() or "main", now))
        conn.execute("""
            UPDATE task_instances
            SET last_run_id=?, updated_at=?
            WHERE instance_uid=?
        """, (rid, now, uid))
        conn.commit()


def add_task_instance_artifact(
    instance_uid: str,
    kind: str,
    label: str,
    path: str,
    meta: Optional[Mapping[str, Any]] = None,
) -> dict:
    """Insert an artifact row and return it."""
    uid = str(instance_uid or "").strip()
    with _get_conn() as conn:
        cur = conn.execute("""
            INSERT INTO task_instance_artifacts (instance_uid, kind, label, path, meta_json, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (
            uid,
            str(kind or "").strip() or "file",
            str(label or "").strip() or "输出文件",
            str(path or "").strip(),
            _json_dumps(dict(meta or {})),
            _now_iso(),
        ))
        conn.execute("UPDATE task_instances SET updated_at=? WHERE instance_uid=?", (_now_iso(), uid))
        conn.commit()
        row = conn.execute("SELECT * FROM task_instance_artifacts WHERE id=?", (cur.lastrowid,)).fetchone()
        return dict(row) if row else {}


def add_task_instance_event(
    instance_uid: str,
    event_type: str,
    message: str,
    meta: Optional[Mapping[str, Any]] = None,
) -> dict:
    """Insert an event row and return it."""
    uid = str(instance_uid or "").strip()
    with _get_conn() as conn:
        cur = conn.execute("""
            INSERT INTO task_instance_events (instance_uid, event_type, message, meta_json, created_at)
            VALUES (?, ?, ?, ?, ?)
        """, (
            uid,
            str(event_type or "").strip() or "event",
            str(message or "").strip(),
            _json_dumps(dict(meta or {})),
            _now_iso(),
        ))
        conn.execute("UPDATE task_instances SET updated_at=? WHERE instance_uid=?", (_now_iso(), uid))
        conn.commit()
        row = conn.execute("SELECT * FROM task_instance_events WHERE id=?", (cur.lastrowid,)).fetchone()
        return dict(row) if row else {}


def create_task_schedule(
    adapter_id: str,
    task_id: str,
    title: str,
    frequency: str,
    time_of_day: str,
    weekday: Optional[int] = None,
    params: Optional[Mapping[str, Any]] = None,
    notify_channel: str = "dingtalk",
    notify_template: str = "",
    enabled: bool = True,
) -> dict:
    """Insert a persisted schedule definition and return it."""
    now = _now_iso()
    schedule_uid = uuid.uuid4().hex
    with _get_conn() as conn:
        conn.execute("""
            INSERT INTO task_schedules (
                schedule_uid, adapter_id, task_id, title, enabled, frequency,
                time_of_day, weekday, params_json, notify_channel, notify_template,
                archived, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0, ?, ?)
        """, (
            schedule_uid,
            str(adapter_id or "").strip(),
            str(task_id or "").strip(),
            str(title or "").strip() or "未命名定时任务",
            1 if enabled else 0,
            str(frequency or "").strip(),
            str(time_of_day or "").strip(),
            weekday,
            _json_dumps(dict(params or {})),
            str(notify_channel or "").strip() or "dingtalk",
            str(notify_template or "").strip(),
            now,
            now,
        ))
        conn.commit()
    return get_task_schedule(schedule_uid) or {}


def get_task_schedule(schedule_uid: str) -> Optional[dict]:
    """Return one schedule definition by uid."""
    with _get_conn() as conn:
        row = conn.execute("""
            SELECT *
            FROM task_schedules
            WHERE schedule_uid=?
            LIMIT 1
        """, (str(schedule_uid or "").strip(),)).fetchone()
        return _row_to_dict(row)


def get_task_schedule_detail(schedule_uid: str) -> dict:
    """Return one schedule plus parsed params."""
    schedule = get_task_schedule(schedule_uid)
    if not schedule:
        return {}
    detail = dict(schedule)
    detail["params"] = _json_loads_object(detail.get("params_json"))
    return detail


def list_task_schedules(
    adapter_id: str = "",
    task_id: str = "",
    enabled: Optional[bool] = None,
    keyword: str = "",
    include_archived: bool = False,
    limit: int = 100,
) -> list[dict]:
    """Return persisted schedule definitions filtered by metadata."""
    clauses = []
    params: list[Any] = []

    if not include_archived:
        clauses.append("archived=0")

    adapter = str(adapter_id or "").strip()
    if adapter:
        clauses.append("adapter_id=?")
        params.append(adapter)

    task = str(task_id or "").strip()
    if task:
        clauses.append("task_id=?")
        params.append(task)

    if enabled is not None:
        clauses.append("enabled=?")
        params.append(1 if enabled else 0)

    term = str(keyword or "").strip()
    if term:
        clauses.append("(title LIKE ? OR schedule_uid LIKE ? OR task_id LIKE ?)")
        like = f"%{term}%"
        params.extend([like, like, like])

    try:
        safe_limit = max(1, min(int(limit), 500))
    except Exception:
        safe_limit = 100

    where = " AND ".join(clauses) if clauses else "1=1"
    with _get_conn() as conn:
        try:
            rows = conn.execute(f"""
                SELECT *
                FROM task_schedules
                WHERE {where}
                ORDER BY updated_at DESC, id DESC
                LIMIT ?
            """, [*params, safe_limit]).fetchall()
        except sqlite3.OperationalError as exc:
            if "no such table: task_schedules" in str(exc):
                return []
            raise
        return [dict(row) for row in rows]


def update_task_schedule(schedule_uid: str, **fields) -> dict:
    """Update allowed schedule fields and return the updated row."""
    allowed = {
        "title",
        "enabled",
        "frequency",
        "time_of_day",
        "weekday",
        "params_json",
        "notify_channel",
        "notify_template",
        "last_run_id",
        "last_instance_uid",
        "last_status",
        "last_error",
        "last_triggered_at",
        "archived",
    }
    updates: dict[str, Any] = {}
    for key, value in fields.items():
        if key == "params":
            updates["params_json"] = _json_dumps(value)
        elif key in {"enabled", "archived"}:
            updates[key] = 1 if value else 0
        elif key in allowed:
            updates[key] = value

    updates["updated_at"] = _now_iso()
    assignments = ", ".join(f"{key}=?" for key in updates.keys())
    values = list(updates.values())
    uid = str(schedule_uid or "").strip()
    with _get_conn() as conn:
        conn.execute(
            f"UPDATE task_schedules SET {assignments} WHERE schedule_uid=?",
            [*values, uid],
        )
        conn.commit()
    return get_task_schedule(uid) or {}


def archive_task_schedule(schedule_uid: str) -> dict:
    """Archive a persisted schedule and disable it."""
    return update_task_schedule(schedule_uid, archived=True, enabled=False)


def record_task_schedule_run(
    schedule_uid: str,
    run_id: Optional[int] = None,
    instance_uid: str = "",
    status: str = "",
    error: str = "",
) -> dict:
    """Store the latest execution result on a schedule definition."""
    fields: dict[str, Any] = {
        "last_triggered_at": _now_iso(),
        "last_instance_uid": str(instance_uid or "").strip(),
        "last_status": str(status or "").strip(),
        "last_error": str(error or "").strip(),
    }
    if run_id is not None:
        fields["last_run_id"] = int(run_id)
    return update_task_schedule(schedule_uid, **fields)


def begin_run(adapter_id: str, task_id: str) -> int:
    """Record a task run start, return run_id"""
    now = datetime.now().isoformat()
    with _get_conn() as conn:
        cur = conn.execute("""
            INSERT INTO task_runs (adapter_id, task_id, status, started_at, last_seen_at, phase, current_row)
            VALUES (?, ?, 'running', ?, ?, '', 0)
        """, (adapter_id, task_id, now, now))
        conn.commit()
        return cur.lastrowid


def heartbeat_run(run_id: int, phase: str = "", current_row: int = 0, records_count: Optional[int] = None):
    fields = [
        "last_seen_at=?",
        "phase=?",
        "current_row=?",
    ]
    now = datetime.now().isoformat()
    values: list[Any] = [
        now,
        str(phase or "").strip(),
        max(0, int(current_row or 0)),
    ]
    if records_count is not None:
        fields.append("records_count=?")
        values.append(max(0, int(records_count or 0)))
    values.append(int(run_id))
    with _get_conn() as conn:
        conn.execute(f"""
            UPDATE task_runs
            SET {", ".join(fields)}
            WHERE id=?
        """, values)
        conn.commit()


def finish_run(run_id: int, records_count: int, output_files: List[str]):
    now = datetime.now().isoformat()
    with _get_conn() as conn:
        conn.execute("""
            UPDATE task_runs
            SET status='done', finished_at=?, last_seen_at=?, records_count=?, output_files=?, error=NULL
            WHERE id=?
        """, (now, now, records_count, json.dumps(output_files), run_id))
        conn.commit()


def fail_run(run_id: int, error: str):
    now = datetime.now().isoformat()
    with _get_conn() as conn:
        conn.execute("""
            UPDATE task_runs
            SET status='error', finished_at=?, last_seen_at=?, error=?
            WHERE id=?
        """, (now, now, error, run_id))
        conn.commit()


def stop_run(run_id: int, records_count: int, output_files: List[str], error: str = ""):
    now = datetime.now().isoformat()
    with _get_conn() as conn:
        conn.execute("""
            UPDATE task_runs
            SET status='stopped', finished_at=?, last_seen_at=?, records_count=?, output_files=?, error=?
            WHERE id=?
        """, (now, now, records_count, json.dumps(output_files), error, run_id))
        conn.commit()


def stop_orphaned_active_runs(error: str = "任务运行时后端已重启，已自动标记为停止") -> int:
    """Mark runs that cannot have an in-memory worker after backend startup as stopped."""
    with _get_conn() as conn:
        cur = conn.execute("""
            UPDATE task_runs
            SET status='stopped', finished_at=?, last_seen_at=?, error=COALESCE(NULLIF(error, ''), ?)
            WHERE status IN ('running', 'pausing', 'paused', 'stopping')
        """, (datetime.now().isoformat(), datetime.now().isoformat(), error))
        conn.commit()
        return int(cur.rowcount or 0)


def list_active_runs(statuses: Optional[Iterable[str]] = None) -> List[dict]:
    """Return runs that are still considered active by the desktop runtime."""
    active_statuses = [
        str(item or "").strip()
        for item in (statuses or ("running", "pausing", "paused", "stopping"))
        if str(item or "").strip()
    ]
    if not active_statuses:
        return []
    placeholders = ",".join("?" for _ in active_statuses)
    with _get_conn() as conn:
        rows = conn.execute(f"""
            SELECT *
            FROM task_runs
            WHERE status IN ({placeholders})
        """, active_statuses).fetchall()
        return [dict(row) for row in rows]


def get_latest_run(adapter_id: str, task_id: str) -> Optional[dict]:
    with _get_conn() as conn:
        row = conn.execute("""
            SELECT * FROM task_runs
            WHERE adapter_id=? AND task_id=?
            ORDER BY id DESC LIMIT 1
        """, (adapter_id, task_id)).fetchone()
        return dict(row) if row else None


def get_latest_task_instance_run(instance_uid: str) -> Optional[dict]:
    """Return the latest task_runs row linked to one task instance."""
    uid = str(instance_uid or "").strip()
    if not uid:
        return None
    with _get_conn() as conn:
        row = conn.execute("""
            SELECT tr.*, tir.purpose, tir.created_at AS linked_at
            FROM task_instance_runs tir
            LEFT JOIN task_runs tr ON tr.id = tir.run_id
            WHERE tir.instance_uid=?
            ORDER BY tir.id DESC
            LIMIT 1
        """, (uid,)).fetchone()
        return dict(row) if row else None


def list_runs(adapter_id: str, task_id: str, limit: int = 20) -> List[dict]:
    with _get_conn() as conn:
        rows = conn.execute("""
            SELECT * FROM task_runs
            WHERE adapter_id=? AND task_id=?
            ORDER BY id DESC LIMIT ?
        """, (adapter_id, task_id, limit)).fetchall()
        return [dict(r) for r in rows]


def _normalize_output_file_path(path: str) -> str:
    try:
        return str(Path(str(path)).expanduser().resolve(strict=False))
    except Exception:
        return str(path or '').strip()


def remove_output_files(paths: List[str]) -> dict:
    target_paths = {
        _normalize_output_file_path(path)
        for path in (paths or [])
        if str(path or '').strip()
    }
    target_paths.discard('')
    if not target_paths:
        return {"updated_runs": 0, "removed_refs": 0}

    updated_runs = 0
    removed_refs = 0

    with _get_conn() as conn:
        rows = conn.execute("""
            SELECT id, output_files
            FROM task_runs
            WHERE output_files IS NOT NULL AND output_files != '[]'
        """).fetchall()

        for row in rows:
            raw_files = row["output_files"]
            try:
                files = json.loads(raw_files) if isinstance(raw_files, str) else raw_files
            except Exception:
                continue
            if not isinstance(files, list):
                continue

            kept_files = []
            changed = False
            for item in files:
                file_path = str(item or '').strip()
                if not file_path:
                    changed = True
                    continue
                if _normalize_output_file_path(file_path) in target_paths:
                    removed_refs += 1
                    changed = True
                    continue
                kept_files.append(file_path)

            if changed:
                conn.execute(
                    "UPDATE task_runs SET output_files=? WHERE id=?",
                    (json.dumps(kept_files), row["id"]),
                )
                updated_runs += 1

        conn.commit()

    return {"updated_runs": updated_runs, "removed_refs": removed_refs}


def is_output_file_path(path: str) -> bool:
    target_path = _normalize_output_file_path(path)
    if not target_path:
        return False

    with _get_conn() as conn:
        rows = conn.execute("""
            SELECT output_files
            FROM task_runs
            WHERE output_files IS NOT NULL AND output_files != '[]'
        """).fetchall()

        for row in rows:
            raw_files = row["output_files"]
            try:
                files = json.loads(raw_files) if isinstance(raw_files, str) else raw_files
            except Exception:
                continue
            if not isinstance(files, list):
                continue
            for item in files:
                file_path = str(item or '').strip()
                if file_path and _normalize_output_file_path(file_path) == target_path:
                    return True

    return False


# ─── Export functions ───

def _sanitize_filename(text: Any, fallback: str = "output") -> str:
    value = str(text or "").strip()
    value = re.sub(r"\s+", " ", value)
    value = re.sub(r"[\x00-\x1f]+", "", value)
    value = re.sub(r'[\\/:*?"<>|]+', "_", value)
    value = value.strip(" .")
    return value or fallback


def _shorten_filename(filename: str, max_length: int = MAX_EXPORT_FILENAME_LENGTH) -> str:
    value = _sanitize_filename(filename)
    if len(value) <= max_length:
        return value

    suffix = Path(value).suffix
    stem = value[:-len(suffix)] if suffix else value
    budget = max_length - len(suffix)
    if budget <= 0:
        return value[:max_length].strip(" ._") or "output"

    tail_budget = min(64, max(24, budget // 2))
    head_budget = max(12, budget - tail_budget - 1)
    shortened = f"{stem[:head_budget].rstrip(' ._')}_{stem[-tail_budget:].lstrip(' ._')}{suffix}"
    return shortened[:max_length].strip(" ._") or f"output{suffix}"


def _render_filename(template: str, adapter_id: str, task_id: str,
                     filename_vars: Optional[Mapping[str, Any]] = None) -> str:
    now = datetime.now()
    vars_map = {
        "date": now.strftime("%Y%m%d"),
        "datetime": now.strftime("%Y%m%d_%H%M%S"),
        "timestamp": now.strftime("%Y%m%d-%H%M%S"),
        "adapter_id": adapter_id,
        "task_id": task_id,
    }
    if filename_vars:
        for key, value in filename_vars.items():
            vars_map[str(key)] = _sanitize_filename(value, "")
    return _shorten_filename(template.format_map(_SafeTemplateVars(vars_map)))


def _ensure_unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    idx = 2
    while True:
        candidate = path.with_name(f"{stem}_{idx}{suffix}")
        if not candidate.exists():
            return candidate
        idx += 1


def _normalize_sheet_title(name: Any, fallback: str) -> str:
    value = _sanitize_filename(name, fallback).strip()
    if not value:
        value = fallback
    return value[:31] or fallback[:31] or "Sheet1"


def _clean_excel_row(row: Mapping[str, Any]) -> dict[str, Any]:
    cleaned: dict[str, Any] = {}
    for key, value in (row or {}).items():
        key_text = str(key or "")
        if key_text.startswith("__"):
            continue
        cleaned[key_text] = value
    return cleaned


def _normalize_column_groups(column_groups: Optional[Iterable[Mapping[str, Any]]]) -> list[dict[str, Any]]:
    normalized_groups = []
    for group in column_groups or []:
        if hasattr(group, "model_dump"):
            group = group.model_dump()
        label = str((group or {}).get("label") or "").strip()
        columns = [
            str(col or "").strip()
            for col in ((group or {}).get("columns") or [])
            if str(col or "").strip()
        ]
        if not label or not columns:
            continue
        normalized_groups.append({"label": label, "columns": columns})
    return normalized_groups


def _resolve_headers(data: List[dict], column_order: Optional[List[str]] = None) -> list[str]:
    headers: list[str] = []
    seen_headers = set()
    for key in column_order or []:
        key_text = str(key or "").strip()
        if not key_text or key_text.startswith("__") or key_text in seen_headers:
            continue
        seen_headers.add(key_text)
        headers.append(key_text)
    for row in data:
        for key in row.keys():
            key_text = str(key or "").strip()
            if not key_text or key_text.startswith("__") or key_text in seen_headers:
                continue
            seen_headers.add(key_text)
            headers.append(key_text)
    return headers


def _write_excel_sheet(ws, data: List[dict], column_order: Optional[List[str]] = None,
                       column_groups: Optional[Iterable[Mapping[str, Any]]] = None) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    headers = _resolve_headers(data, column_order)
    if not headers:
        headers = ["提示"]
        data = [{"提示": "无数据"}]

    header_fill = PatternFill(fill_type="solid", fgColor="F5F7FA")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    def _leaf_header_label(column_key: str) -> str:
        value = str(column_key or "").strip()
        if "/" in value:
            return value.split("/")[-1].strip() or value
        return value

    normalized_groups = _normalize_column_groups(column_groups)
    group_lookup = {}
    for group in normalized_groups:
        for column in group["columns"]:
            group_lookup[column] = group["label"]

    has_grouped_header = bool(normalized_groups)
    if has_grouped_header:
        top_row = []
        leaf_row = []
        for header in headers:
            group_label = group_lookup.get(header, "")
            if group_label:
                top_row.append(group_label)
                leaf_row.append(_leaf_header_label(header))
            else:
                top_row.append(_leaf_header_label(header))
                leaf_row.append("")

        ws.append(top_row)
        ws.append(leaf_row)

        col_index = 1
        while col_index <= len(headers):
            header = headers[col_index - 1]
            group_label = group_lookup.get(header, "")
            if not group_label:
                ws.merge_cells(start_row=1, start_column=col_index, end_row=2, end_column=col_index)
                col_index += 1
                continue

            end_col = col_index
            while end_col < len(headers) and group_lookup.get(headers[end_col], "") == group_label:
                end_col += 1
            if end_col > col_index:
                ws.merge_cells(start_row=1, start_column=col_index, end_row=1, end_column=end_col)
            col_index = end_col + 1

        for row_no in (1, 2):
            for cell in ws[row_no]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
        ws.freeze_panes = "A3"
    else:
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        ws.freeze_panes = "A2"

    for row in data:
        ws.append([str(row.get(h, "")) for h in headers])

    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 10
        for row_idx in range(1, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            max_len = max(max_len, len(str(value or "")))
        ws.column_dimensions[letter].width = min(max_len + 4, 60)


def prepare_artifact_dir(adapter_id: str, task_id: str, run_id: int, kind: str = "artifacts") -> str:
    """Create and return a per-run artifact directory."""
    out_dir = artifact_dir_path(adapter_id, task_id, run_id, kind)
    out_dir.mkdir(parents=True, exist_ok=True)
    return str(out_dir)


def artifact_dir_path(adapter_id: str, task_id: str, run_id: int, kind: str = "artifacts") -> Path:
    """Return the per-run artifact directory path without creating it."""
    safe_kind = _sanitize_filename(kind, "artifacts")
    return _data_root() / adapter_id / task_id / safe_kind / str(run_id)


def export_excel(
    data: List[dict],
    adapter_id: str,
    task_id: str,
    filename_template: str = "{task_id}_{date}.xlsx",
    filename_vars: Optional[Mapping[str, Any]] = None,
    column_order: Optional[List[str]] = None,
    column_groups: Optional[List[Mapping[str, Any]]] = None,
    sheet_key: Optional[str] = None,
    sheet_configs: Optional[List[Mapping[str, Any]]] = None,
) -> str:
    """
    Export data to Excel file.
    Returns absolute path of written file.
    """
    from openpyxl import Workbook

    if not data:
        raise ValueError("No data to export")

    filename = _render_filename(filename_template, adapter_id, task_id, filename_vars)
    out_dir = _data_root() / adapter_id / task_id
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = _ensure_unique_path(out_dir / filename)

    wb = Workbook()
    cleaned_data = [_clean_excel_row(row) for row in data]
    normalized_sheet_key = str(sheet_key or "").strip()

    if not normalized_sheet_key:
        ws = wb.active
        ws.title = _normalize_sheet_title(task_id, task_id[:31] or "Sheet1")
        _write_excel_sheet(ws, cleaned_data, column_order, column_groups)
        total_rows = len(cleaned_data)
    else:
        sheet_rows: dict[str, list[dict[str, Any]]] = {}
        sheet_order: list[str] = []
        for raw_row in data:
            sheet_name = str((raw_row or {}).get(normalized_sheet_key) or "").strip() or "Sheet1"
            if sheet_name not in sheet_rows:
                sheet_rows[sheet_name] = []
                sheet_order.append(sheet_name)
            sheet_rows[sheet_name].append(_clean_excel_row(raw_row))

        config_by_name: dict[str, Mapping[str, Any]] = {}
        ordered_names: list[str] = []
        for config in sheet_configs or []:
            if hasattr(config, "model_dump"):
                config = config.model_dump()
            name = str((config or {}).get("name") or "").strip()
            if not name:
                continue
            config_by_name[name] = config
            ordered_names.append(name)

        final_sheet_names = [name for name in ordered_names if name in sheet_rows]
        final_sheet_names.extend(name for name in sheet_order if name not in final_sheet_names)
        if not final_sheet_names:
            final_sheet_names = ["Sheet1"]
            sheet_rows["Sheet1"] = cleaned_data

        ws = wb.active
        first_name = final_sheet_names[0]
        ws.title = _normalize_sheet_title(first_name, "Sheet1")
        first_config = config_by_name.get(first_name) or {}
        _write_excel_sheet(
            ws,
            sheet_rows.get(first_name) or [],
            list((first_config or {}).get("columns") or column_order or []),
            (first_config or {}).get("column_groups") or column_groups,
        )

        for name in final_sheet_names[1:]:
            config = config_by_name.get(name) or {}
            sheet = wb.create_sheet(title=_normalize_sheet_title(name, "Sheet"))
            _write_excel_sheet(
                sheet,
                sheet_rows.get(name) or [],
                list((config or {}).get("columns") or column_order or []),
                (config or {}).get("column_groups") or column_groups,
            )

        total_rows = sum(len(rows) for rows in sheet_rows.values())

    wb.save(str(out_path))
    logger.info(f"Excel exported: {out_path} ({total_rows} rows)")
    return str(out_path)


def export_json(data: List[dict], adapter_id: str, task_id: str,
                filename_template: str = "{task_id}_{date}.json",
                filename_vars: Optional[Mapping[str, Any]] = None) -> str:
    """
    Export data to JSON file.
    Returns absolute path of written file.
    """
    filename = _render_filename(filename_template, adapter_id, task_id, filename_vars)
    out_dir = _data_root() / adapter_id / task_id
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = _ensure_unique_path(out_dir / filename)

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    logger.info(f"JSON exported: {out_path} ({len(data)} records)")
    return str(out_path)


def export_desktop_excel(data: List[dict], adapter_id: str, task_id: str,
                         filename_template: str = "{task_id}_{date}.xlsx",
                         filename_vars: Optional[Mapping[str, Any]] = None) -> str:
    """Export to user Desktop (mirrors temu-assistant behavior)"""
    from openpyxl import Workbook
    desktop = Path.home() / "Desktop"
    filename = _render_filename(filename_template, adapter_id, task_id, filename_vars)
    out_path = _ensure_unique_path(desktop / filename)
    wb = Workbook()
    ws = wb.active
    ws.title = task_id[:31]
    if data:
        headers = list(data[0].keys())
        ws.append(headers)
        for row in data:
            ws.append([str(row.get(h, "")) for h in headers])
    wb.save(str(out_path))
    logger.info(f"Excel exported to Desktop: {out_path}")
    return str(out_path)
