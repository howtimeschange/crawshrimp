"""
FastAPI service - complete task execution pipeline
Electron spawns this process; all GUI calls go through HTTP on localhost:18765

Bundling notes (for electron-builder / python-build-standalone):
  - All deps (fastapi, uvicorn, apscheduler, openpyxl, websockets, pyyaml) must be
    pre-installed in the bundled Python env (see app/scripts/bundle-python.sh)
  - Built-in adapters (adapters/) are shipped as extraResources alongside Python scripts
  - User-installed adapters live under the selected runtime data dir
"""
import asyncio
import base64
import csv
import hashlib
import html
import hmac
import json
import logging
import mimetypes
import os
import re
import secrets
import shutil
import tempfile
import threading
import time
import zipfile
from contextlib import asynccontextmanager
from datetime import datetime, timedelta
from pathlib import Path
from typing import Callable, List, Optional
from urllib.parse import urlencode, parse_qs, urlparse, unquote
from uuid import uuid4

from fastapi import FastAPI, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse
from pydantic import BaseModel, ConfigDict, Field

from core import runtime_paths
from core import script_favorites
from core import bala_ai_model_library
from core import bala_ai_video_materials
from core import bala_ai_video_review
from core.config import load_config, patch_config, save_config
from core.cloud_approval_client import CloudApprovalClient, CloudApprovalError
from core.cloud_approval_url import (
    invalidate_cloud_approval_url_cache,
    resolve_cloud_approval_url,
)
from core.cloud_batch_sync import sync_local_approval_batch
from core.cloud_machine_agent import CloudMachineAgent
from core import adapter_loader
from core import ai_image_service
from core import ai_video_generation_service
from core import data_sink
from core import notifier
from core import odps_sync
from core import scheduler as sched_module
from core.cdp_bridge import CDPBridge, get_bridge, reset_bridge
from core.browser_session import open_browser_session
from core.dev_harness import run_harness_capture, run_harness_eval, run_harness_snapshot
from core.dev_harness_models import DevHarnessCaptureRequest, DevHarnessEvalRequest, DevHarnessSnapshotRequest
from core.knowledge_service import ensure_knowledge_index, rebuild_knowledge_index, search_knowledge
from core.one_xm_image import DEFAULT_BASE_URL as ONE_XM_DEFAULT_BASE_URL
from core.one_xm_image import OneXMImageClient, file_to_data_url, run_image_task_until_done
from core.probe_models import ProbeRequest
from core.probe_service import read_probe_bundle, read_probe_bundle_full, run_probe_request
from core.runtime_install_guard import InstallRuntimeBusy, RuntimeInstallGuard, UpdateDrainActive
from core.shenhui_pdf_screenshot import finalize_pdf_batch_screenshot_outputs, convert_pdf_rows_to_yq_output_root
from core.amazon_label_splitter import (
    copy_amazon_label_outputs_to_export_folder,
    split_amazon_label_rows,
)

logger = logging.getLogger(__name__)

API_VERSION = "1.4.16"
API_TOKEN_HEADER = "x-crawshrimp-token"
PUBLIC_API_PATHS = {"/health", "/docs", "/openapi.json", "/redoc", "/docs/oauth2-redirect"}
PRIVATE_API_PATHS = {"/bala-ai-video-materials/api/from-rows"}
PUBLIC_API_PREFIXES = (
    "/adapter-assets/",
    "/tmall-ai-image-approval/",
    "/bala-ai-video-materials/",
    "/bala-ai-video-model-library/",
    "/bala-ai-video-review/",
)

# In-memory task run state for live status / logs
_run_logs: dict = {}   # job_id -> list[str]
_run_status: dict = {} # job_id -> {'status', 'run_id', 'records'}
_run_controls: dict = {}  # job_id -> {'task', 'pause_requested', 'stop_requested', 'resume_event', ...}
_task_locks: dict[str, asyncio.Lock] = {}

ACTIVE_LIVE_STATUSES = {"running", "pausing", "paused", "stopping"}
TMALL_OPS_ADAPTER_ID = "tmall-ops-assistant"
TMALL_MATERIAL_EXPORT_TASK_ID = "tmall_material_test_data_export"
DEFAULT_TASK_SCHEDULE_NOTIFY_TEMPLATE = """巴拉-AI测图数据抓取导出执行通知
定时任务：{{schedule_title}}
执行状态：{{status}}
导出记录：{{records}}
输出文件：{{output_files}}
导出目录：{{export_dir}}
完成时间：{{finished_at}}
错误信息：{{error}}"""

RUNTIME_CLEANUP_TASKS = {
    ("amazon-ops-assistant", "amazon_label_batch_process"),
    ("mop-ops-assistant", "cloud_folder_download"),
    ("semir-cloud-drive", "batch_ai_generate"),
    ("semir-cloud-drive", "batch_image_download"),
    ("semir-cloud-drive", "tmall_material_match_buy"),
    ("semir-cloud-drive", "tmall_material_new_624"),
    ("bala-ai-video-assistant", "semir_video_material_prepare"),
    ("bala-ai-video-assistant", "bala_ai_face_background_generate"),
    ("shenhui-new-arrival", "prepare_upload_package"),
    ("tiktok-ops-assistant", "creator_video_download"),
    ("tmall-ops-assistant", "tmall_packaging_upload"),
}
runtime_install_guard = RuntimeInstallGuard()

def _backend_lock_dir() -> Path:
    explicit = str(os.environ.get("CRAWSHRIMP_BACKEND_LOCK_DIR") or "").strip()
    if explicit:
        path = Path(explicit).expanduser()
        path.mkdir(parents=True, exist_ok=True)
        return path
    return runtime_paths.data_root()


def _backend_lock_path() -> Path:
    return _backend_lock_dir() / "backend.lock"


def _get_api_token() -> str:
    env_token = str(os.environ.get("CRAWSHRIMP_API_TOKEN") or "").strip()
    if env_token:
        return env_token

    token_path = _backend_lock_dir() / "api-token"
    try:
        if token_path.exists():
            return token_path.read_text(encoding="utf-8").strip()
        token_path.parent.mkdir(parents=True, exist_ok=True)
        token = secrets.token_hex(32)
        token_path.write_text(token, encoding="utf-8")
        try:
            token_path.chmod(0o600)
        except OSError:
            pass
        return token
    except Exception as exc:
        logger.exception("Failed to read or create crawshrimp API token")
        raise RuntimeError("Failed to prepare crawshrimp API token") from exc


def _api_auth_required() -> bool:
    return os.environ.get("CRAWSHRIMP_ALLOW_UNAUTHENTICATED") != "1"


def _allowed_cors_origins() -> list[str]:
    raw = str(os.environ.get("CRAWSHRIMP_ALLOWED_ORIGINS") or "").strip()
    if raw:
        origins = [item.strip() for item in raw.split(",") if item.strip()]
        if origins:
            return origins
    return [
        "http://127.0.0.1:5173",
        "http://localhost:5173",
    ]


def _add_local_cors_headers(request: Request, response: JSONResponse) -> JSONResponse:
    origin = str(request.headers.get("origin") or "").strip()
    allowed = _allowed_cors_origins()
    if origin and ("*" in allowed or origin in allowed):
        response.headers["Access-Control-Allow-Origin"] = "*" if "*" in allowed else origin
        response.headers["Vary"] = "Origin"
        response.headers["Access-Control-Allow-Headers"] = "Content-Type, X-Crawshrimp-Token"
        response.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, PATCH, DELETE, OPTIONS"
    return response


def _external_call_timeout_seconds(default: float = 20.0) -> float:
    raw = str(os.environ.get("CRAWSHRIMP_EXTERNAL_CALL_TIMEOUT_SECONDS") or "").strip()
    if not raw:
        return default
    try:
        value = float(raw)
    except ValueError:
        return default
    return max(0.0, value)


async def _bridge_call_async(async_name: str, sync_name: str, *args, **kwargs):
    bridge = get_bridge()
    async_method = getattr(bridge, async_name, None)
    if async_method:
        awaitable = async_method(*args, **kwargs)
    else:
        sync_method = getattr(bridge, sync_name)
        awaitable = asyncio.to_thread(sync_method, *args, **kwargs)

    timeout_seconds = _external_call_timeout_seconds()
    if timeout_seconds <= 0:
        return await awaitable
    try:
        return await asyncio.wait_for(awaitable, timeout=timeout_seconds)
    except asyncio.TimeoutError as exc:
        raise TimeoutError(f"External dependency timed out: {sync_name}") from exc


async def _bridge_get_tabs() -> list:
    return await _bridge_call_async("get_tabs_async", "get_tabs")


async def _bridge_new_tab(url: str) -> dict:
    return await _bridge_call_async("new_tab_async", "new_tab", url)


async def _bridge_find_tab(url: str) -> Optional[dict]:
    return await _bridge_call_async("find_tab_async", "find_tab", url)


async def _bridge_get_tab(tab_id: str) -> Optional[dict]:
    return await _bridge_call_async("get_tab_async", "get_tab", tab_id)


async def _await_cleanup_after_cancel(awaitable):
    """Keep partial export/finalize cleanup running even if the outer task is cancelled again."""
    cleanup_task = asyncio.ensure_future(awaitable)
    while True:
        try:
            return await asyncio.shield(cleanup_task)
        except asyncio.CancelledError:
            if cleanup_task.done():
                return cleanup_task.result()


def _resolve_allowed_file_delete_path(raw_path: str) -> Path:
    try:
        data_root = runtime_paths.data_root().resolve(strict=False)
        path = Path(raw_path).expanduser().resolve(strict=False)
        if path == data_root:
            raise HTTPException(400, f"Refusing to delete crawshrimp data directory: {raw_path}")
        path.relative_to(data_root)
        return path
    except ValueError:
        raise HTTPException(400, f"File delete is limited to crawshrimp data directory: {raw_path}")
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(400, f"Invalid file path: {raw_path}") from exc


def _ensure_registered_output_file_path(path: Path) -> None:
    try:
        is_registered = data_sink.is_output_file_path(str(path))
    except Exception as exc:
        raise HTTPException(500, f"Unable to validate file delete request: {path}") from exc
    if not is_registered:
        raise HTTPException(400, f"File delete is limited to registered task output files: {path}")


def _task_lock(jid: str) -> asyncio.Lock:
    lock = _task_locks.get(jid)
    if lock is None:
        lock = asyncio.Lock()
        _task_locks[jid] = lock
    return lock


def _task_jid(adapter_id: str, task_id: str) -> str:
    return f"{adapter_id}::{task_id}"


def _instance_jid(instance_uid: str) -> str:
    uid = str(instance_uid or "").strip()
    return f"instance::{uid}" if uid else ""


def _run_jid(adapter_id: str, task_id: str, instance_uid: str = "") -> str:
    return _instance_jid(instance_uid) or _task_jid(adapter_id, task_id)


def _run_jids(adapter_id: str, task_id: str, instance_uid: str = "") -> list[str]:
    task_jid = _task_jid(adapter_id, task_id)
    instance_jid = _instance_jid(instance_uid)
    if instance_jid and instance_jid != task_jid:
        return [task_jid, instance_jid]
    return [task_jid]


def _set_live_status(jids: list[str], status: dict) -> None:
    for job_id in jids:
        if job_id:
            _run_status[job_id] = dict(status or {})


def _merge_live_status(jids: list[str], patch: dict) -> None:
    for job_id in jids:
        if not job_id:
            continue
        live = dict(_run_status.get(job_id) or {})
        _run_status[job_id] = {**live, **dict(patch or {})}


def _append_run_log(jids: list[str], message: str) -> None:
    for job_id in jids:
        if job_id:
            _run_logs.setdefault(job_id, []).append(message)


def _latest_instance_run(instance_uid: str) -> Optional[dict]:
    run = data_sink.get_latest_task_instance_run(instance_uid)
    if not run:
        return None
    return {**run, "id": run.get("id") or run.get("run_id")}


def _task_is_active(jid: str) -> bool:
    live = _run_status.get(jid) or {}
    control = _run_controls.get(jid)
    task = control.get("task") if control else None
    if task and not task.done():
        return True
    if _task_lock(jid).locked():
        return True
    return live.get("status") in ACTIVE_LIVE_STATUSES and ((task and not task.done()) or not control)


def _task_has_live_control(jid: str) -> bool:
    live = _run_status.get(jid) or {}
    control = _run_controls.get(jid)
    task = control.get("task") if control else None
    if task and not task.done():
        return True
    return live.get("status") in ACTIVE_LIVE_STATUSES and not control


def _default_tmall_material_export_dir() -> str:
    return str(Path.home() / "Downloads" / "抓虾导出" / "天猫运营助手" / "巴拉-AI测图数据抓取导出")


def _default_tmall_material_export_params(params: Optional[dict] = None) -> dict:
    merged = {
        "mode": "new",
        "output_dir": _default_tmall_material_export_dir(),
        "test_status": "1",
        "statistic_type": "ACCUMULATE_30_DAYS",
        "page_size": 20,
    }
    for key, value in (params or {}).items():
        merged[key] = value
    if not str(merged.get("output_dir") or "").strip():
        merged["output_dir"] = _default_tmall_material_export_dir()
    if not str(merged.get("mode") or "").strip():
        merged["mode"] = "new"
    return merged


def _validate_schedule_time(value: str) -> str:
    text = str(value or "").strip()
    if not re.match(r"^\d{1,2}:\d{2}$", text):
        raise HTTPException(400, "定时任务时间必须使用 HH:mm")
    hour_raw, minute_raw = text.split(":", 1)
    hour = int(hour_raw)
    minute = int(minute_raw)
    if hour < 0 or hour > 23 or minute < 0 or minute > 59:
        raise HTTPException(400, "定时任务时间超出范围")
    return f"{hour:02d}:{minute:02d}"


def _validate_schedule_frequency(frequency: str, weekday: Optional[int]) -> tuple[str, Optional[int]]:
    value = str(frequency or "").strip().lower()
    if value not in {"daily", "weekly"}:
        raise HTTPException(400, "定时任务频次必须是 daily 或 weekly")
    if value == "weekly":
        try:
            day = int(weekday or 0)
        except Exception:
            day = 0
        if day < 1 or day > 7:
            raise HTTPException(400, "每周定时任务必须设置 weekday=1..7")
        return value, day
    return value, None


def _task_display_name(adapter_id: str, task_id: str) -> tuple[str, str]:
    adapter_name = adapter_id
    task_name = task_id
    try:
        adapter_loader.scan_all()
        manifest = adapter_loader.get_adapter(adapter_id)
        if manifest:
            adapter_name = manifest.name or adapter_id
            task = next((item for item in manifest.tasks if item.id == task_id), None)
            if task:
                task_name = task.name or task_id
    except Exception:
        logger.debug("failed to resolve task display name", exc_info=True)
    return adapter_name, task_name


def _render_task_schedule_template(template: str, values: dict) -> str:
    text = str(template or DEFAULT_TASK_SCHEDULE_NOTIFY_TEMPLATE)
    for key, value in (values or {}).items():
        text = text.replace("{{" + str(key) + "}}", str(value if value is not None else ""))
    return text


def _schedule_next_run_map() -> dict[str, str]:
    try:
        return {
            str(job.get("schedule_uid") or ""): str(job.get("next_run") or "")
            for job in sched_module.list_jobs()
            if str(job.get("kind") or "") == "task_schedule" and str(job.get("schedule_uid") or "")
        }
    except Exception:
        return {}


class BackendInstanceLock:
    def __init__(self, file_path: Path):
        self.file_path = file_path
        self.handle = None
        self.acquired = False

    def acquire(self) -> bool:
        self.file_path.parent.mkdir(parents=True, exist_ok=True)
        self.handle = self.file_path.open("a+")
        if os.name == "nt":
            try:
                import msvcrt

                self.handle.seek(0)
                msvcrt.locking(self.handle.fileno(), msvcrt.LK_NBLCK, 1)
            except OSError:
                self.acquired = False
                try:
                    self.handle.close()
                finally:
                    self.handle = None
                return False
        else:
            try:
                import fcntl

                fcntl.flock(self.handle.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
            except BlockingIOError:
                self.acquired = False
                try:
                    self.handle.close()
                finally:
                    self.handle = None
                return False
            except OSError:
                self.acquired = False
                try:
                    self.handle.close()
                finally:
                    self.handle = None
                return False
        try:
            self.handle.seek(0)
            self.handle.truncate()
            self.handle.write(str(os.getpid()))
            self.handle.flush()
        except Exception:
            logger.debug("Failed to write backend instance lock metadata", exc_info=True)
            self.acquired = False
            try:
                self.handle.close()
            finally:
                self.handle = None
            return False
        self.acquired = True
        return True

    def close(self) -> None:
        if not self.handle:
            return
        try:
            if self.acquired:
                if os.name == "nt":
                    import msvcrt

                    self.handle.seek(0)
                    msvcrt.locking(self.handle.fileno(), msvcrt.LK_UNLCK, 1)
                else:
                    import fcntl

                    fcntl.flock(self.handle.fileno(), fcntl.LOCK_UN)
        except Exception:
            logger.debug("Failed to release backend instance lock", exc_info=True)
        try:
            self.handle.close()
        finally:
            self.handle = None
            self.acquired = False


def _acquire_backend_instance_lock() -> BackendInstanceLock:
    lock = BackendInstanceLock(_backend_lock_path())
    lock.acquire()
    return lock


def _backend_runtime_kind() -> str:
    explicit = str(os.environ.get("CRAWSHRIMP_BACKEND_KIND") or "").strip()
    if explicit:
        return explicit
    scripts_dir = Path(__file__).resolve().parent.parent
    return "packaged" if scripts_dir.name == "python-scripts" else "source"


def _backend_runtime_info() -> dict:
    data_dir = runtime_paths.data_root()
    lock_dir = _backend_lock_dir()
    lock_pid = ""
    try:
        lock_pid = _backend_lock_path().read_text(encoding="utf-8").strip()
    except Exception:
        lock_pid = ""
    return {
        "kind": _backend_runtime_kind(),
        "pid": os.getpid(),
        "backend_instance_id": str(os.environ.get("CRAWSHRIMP_BACKEND_INSTANCE_ID") or ""),
        "backend_lock_pid": lock_pid,
        "scripts_dir": str(Path(__file__).resolve().parent.parent),
        "data_dir": str(data_dir.resolve() if data_dir.exists() else data_dir),
        "lock_dir": str(lock_dir.expanduser().resolve() if lock_dir.exists() else lock_dir),
        "owns_backend_instance": bool(getattr(app.state, "owns_backend_instance", False)) if "app" in globals() else False,
    }


def _current_local_api_base_url() -> str:
    port = str(os.environ.get("CRAWSHRIMP_PORT") or "18765").strip() or "18765"
    return f"http://127.0.0.1:{port}"


def _inject_runtime_task_params(run_params: dict, runtime_artifact_dir: Optional[str] = None) -> dict:
    if "__crawshrimp_api_base_url" not in run_params or not str(run_params.get("__crawshrimp_api_base_url") or "").strip():
        run_params["__crawshrimp_api_base_url"] = _current_local_api_base_url()
    if runtime_artifact_dir and not str(run_params.get("__crawshrimp_runtime_artifact_dir") or "").strip():
        run_params["__crawshrimp_runtime_artifact_dir"] = str(runtime_artifact_dir)
    return run_params


def _build_run_control() -> dict:
    resume_event = asyncio.Event()
    resume_event.set()
    return {
        "task": None,
        "pause_requested": False,
        "stop_requested": False,
        "stop_reason": "任务已停止",
        "resume_event": resume_event,
        "pause_logged": False,
        "last_progress_exec_no": 0,
        "total_rows": 0,
        "shared_progress": {},
    }


def _int_from_mapping(source: Optional[dict], key: str, default: int = 0) -> int:
    if not source:
        return default
    try:
        return int(source.get(key) or default)
    except (TypeError, ValueError):
        return default


def _str_from_mapping(source: Optional[dict], key: str) -> str:
    if not source:
        return ""
    return str(source.get(key) or "").strip()


def _build_live_progress(payload: Optional[dict] = None, run_control: Optional[dict] = None) -> dict:
    payload = dict(payload or {})
    incoming_shared = payload.get('shared') if isinstance(payload.get('shared'), dict) else {}
    if run_control is not None:
        previous_shared = dict(run_control.get('shared_progress') or {})
        shared_state = {**previous_shared, **dict(incoming_shared or {})}
        if incoming_shared:
            run_control['shared_progress'] = dict(shared_state)
    else:
        shared_state = dict(incoming_shared or {})
    total_rows = _int_from_mapping(shared_state, 'total_rows', _int_from_mapping(run_control, 'total_rows') if run_control else 0)
    current_exec_no = _int_from_mapping(shared_state, 'current_exec_no')
    current_row_no = _int_from_mapping(shared_state, 'current_row_no')
    batch_no = _int_from_mapping(shared_state, 'batch_no')
    total_batches = _int_from_mapping(shared_state, 'total_batches')
    search_total_codes = _int_from_mapping(shared_state, 'search_total_codes')
    search_completed_codes = _int_from_mapping(shared_state, 'search_completed_codes')
    generation_total_jobs = _int_from_mapping(shared_state, 'generation_total_jobs')
    generation_submitted_jobs = _int_from_mapping(shared_state, 'generation_submitted_jobs')
    generation_completed_jobs = _int_from_mapping(shared_state, 'generation_completed_jobs')
    buyer_id = _str_from_mapping(shared_state, 'current_buyer_id')
    store = _str_from_mapping(shared_state, 'current_store')
    current_source_filename = _str_from_mapping(shared_state, 'current_source_filename')
    phase_name = str(payload.get('phase') or '').strip()
    records = _int_from_mapping(payload, 'records')

    download_total = _int_from_mapping(run_control, 'download_total') if run_control else _int_from_mapping(payload, 'download_total', _int_from_mapping(payload, 'download_item_total'))
    download_completed = _int_from_mapping(run_control, 'download_completed') if run_control else _int_from_mapping(payload, 'download_completed')
    download_success = _int_from_mapping(run_control, 'download_success') if run_control else _int_from_mapping(payload, 'download_success')
    download_failed = _int_from_mapping(run_control, 'download_failed') if run_control else _int_from_mapping(payload, 'download_failed')
    download_concurrency = _int_from_mapping(run_control, 'download_concurrency') if run_control else _int_from_mapping(payload, 'download_concurrency')
    download_retry_attempts = _int_from_mapping(run_control, 'download_retry_attempts') if run_control else _int_from_mapping(payload, 'download_retry_attempts')
    download_started = bool(run_control.get('download_started')) if run_control else bool(download_total > 0)
    download_active = bool(run_control.get('download_active')) if run_control else bool(payload.get('download_active'))
    download_last_label = str((run_control.get('download_last_label') if run_control else payload.get('download_last_label')) or '').strip()
    download_current_label = str((run_control.get('download_current_label') if run_control else payload.get('download_current_label')) or '').strip()
    download_active_labels = run_control.get('download_active_labels') if run_control else payload.get('download_active_labels')
    if not isinstance(download_active_labels, list):
        download_active_labels = []
    download_active_labels = [str(item or '').strip() for item in download_active_labels if str(item or '').strip()]
    download_active_count = _int_from_mapping(run_control, 'download_active_count') if run_control else _int_from_mapping(payload, 'download_active_count')
    download_speed_bps = _int_from_mapping(run_control, 'download_speed_bps') if run_control else _int_from_mapping(payload, 'download_speed_bps')
    download_bytes_completed = _int_from_mapping(run_control, 'download_bytes_completed') if run_control else _int_from_mapping(payload, 'download_bytes_completed')
    download_total_bytes = _int_from_mapping(run_control, 'download_total_bytes') if run_control else _int_from_mapping(payload, 'download_total_bytes')

    if run_control and total_rows:
        run_control['total_rows'] = total_rows

    percent = 0.0
    if total_rows > 0 and current_exec_no > 0:
        percent = min(100.0, round((current_exec_no / total_rows) * 100, 1))

    progress_text = ""
    if current_exec_no > 0 and total_rows > 0:
        progress_text = f"{current_exec_no}/{total_rows}"

    return {
        "current": current_exec_no,
        "total": total_rows,
        "row_no": current_row_no,
        "batch_no": batch_no,
        "total_batches": total_batches,
        "search_total_codes": search_total_codes,
        "search_completed_codes": search_completed_codes,
        "generation_total_jobs": generation_total_jobs,
        "generation_submitted_jobs": generation_submitted_jobs,
        "generation_completed_jobs": generation_completed_jobs,
        "buyer_id": buyer_id,
        "store": store,
        "current_source_filename": current_source_filename,
        "phase": phase_name,
        "completed": records,
        "percent": percent,
        "progress_text": progress_text,
        "download_total": download_total,
        "download_completed": download_completed,
        "download_success": download_success,
        "download_failed": download_failed,
        "download_concurrency": download_concurrency,
        "download_retry_attempts": download_retry_attempts,
        "download_started": download_started,
        "download_active": download_active,
        "download_last_label": download_last_label,
        "download_current_label": download_current_label,
        "download_active_labels": download_active_labels,
        "download_active_count": download_active_count,
        "download_speed_bps": download_speed_bps,
        "download_bytes_completed": download_bytes_completed,
        "download_total_bytes": download_total_bytes,
        "list_total_rows": _int_from_mapping(shared_state, 'list_total_rows'),
        "list_completed_rows": _int_from_mapping(shared_state, 'list_completed_rows'),
        "list_total_batches": _int_from_mapping(shared_state, 'list_total_batches'),
        "list_completed_batches": _int_from_mapping(shared_state, 'list_completed_batches'),
        "detail_total_targets": _int_from_mapping(shared_state, 'detail_total_targets'),
        "detail_completed_targets": _int_from_mapping(shared_state, 'detail_completed_targets'),
        "detail_current_target_index": _int_from_mapping(shared_state, 'detail_current_target_index'),
        "detail_current_target": _str_from_mapping(shared_state, 'detail_current_target'),
        "detail_dimension_total": _int_from_mapping(shared_state, 'detail_dimension_total'),
        "detail_dimension_index": _int_from_mapping(shared_state, 'detail_dimension_index'),
        "detail_dimension_label": _str_from_mapping(shared_state, 'detail_dimension_label'),
        "detail_current_page": _int_from_mapping(shared_state, 'detail_current_page'),
        "detail_total_pages": _int_from_mapping(shared_state, 'detail_total_pages'),
        "detail_total_rows": _int_from_mapping(shared_state, 'detail_total_rows'),
        "detail_request_count": _int_from_mapping(shared_state, 'detail_request_count'),
        "detail_records_collected": _int_from_mapping(shared_state, 'detail_records_collected'),
        "doudian_stage": _str_from_mapping(shared_state, 'doudian_stage'),
        "doudian_activity_total": _int_from_mapping(shared_state, 'doudian_activity_total'),
        "doudian_activity_completed": _int_from_mapping(shared_state, 'doudian_activity_completed'),
        "doudian_current_activity": _str_from_mapping(shared_state, 'doudian_current_activity'),
        "doudian_current_product_total": _int_from_mapping(shared_state, 'doudian_current_product_total'),
        "doudian_current_product_completed": _int_from_mapping(shared_state, 'doudian_current_product_completed'),
        "doudian_detail_rows": _int_from_mapping(shared_state, 'doudian_detail_rows'),
        "doudian_signup_total": _int_from_mapping(shared_state, 'doudian_signup_total'),
        "doudian_signup_completed": _int_from_mapping(shared_state, 'doudian_signup_completed'),
        "doudian_order_window_total": _int_from_mapping(shared_state, 'doudian_order_window_total'),
        "doudian_order_window_completed": _int_from_mapping(shared_state, 'doudian_order_window_completed'),
        "doudian_mixed_rows": _int_from_mapping(shared_state, 'doudian_mixed_rows'),
    }


def _extract_candidate_urls_from_text(value: object) -> list[str]:
    text = str(value or "").strip()
    if not text:
        return []
    matches = re.findall(r"https?://[\s\S]*?(?=https?://|$)", text, flags=re.IGNORECASE)
    if matches:
        return [item.rstrip("、，,;；").strip() for item in matches if item.strip()]
    return [
        item.strip()
        for item in re.split(r"[\n\r\t 、，,;；]+", text)
        if item.strip()
    ]


def _normalize_tmall_item_link(raw_url: object) -> str:
    text = str(raw_url or "").strip().rstrip("、，,;；")
    if not text:
        return ""
    try:
        parsed = urlparse(text)
    except Exception:
        return ""
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        return ""
    query = parse_qs(parsed.query or "", keep_blank_values=False)
    item_id = ""
    for key in ("id", "itemId", "item_id", "itemNumId", "item_num_id"):
        candidate = (query.get(key) or [""])[0].strip()
        if candidate:
            item_id = candidate
            break
    if not re.fullmatch(r"\d{6,}", item_id):
        return ""
    sku_id = ""
    for key in ("skuId", "sku_id"):
        candidate = (query.get(key) or [""])[0].strip()
        if candidate:
            sku_id = candidate
            break
    params = {"id": item_id}
    if re.fullmatch(r"\d{6,}", sku_id):
        params["skuId"] = sku_id
    return f"https://detail.tmall.com/item.htm?{urlencode(params)}"


def _resolve_tmall_buyer_review_item_urls(run_params: dict) -> list[str]:
    urls: list[str] = []
    seen: set[str] = set()
    for key in ("item_links", "item_url", "links"):
        for candidate in _extract_candidate_urls_from_text(run_params.get(key)):
            normalized = _normalize_tmall_item_link(candidate)
            if not normalized or normalized in seen:
                continue
            seen.add(normalized)
            urls.append(normalized)
    return urls


def _normalize_amazon_reviews_url(raw_url: object) -> str:
    text = str(raw_url or "").strip().rstrip("、，,;；")
    if not text:
        return ""
    if text.startswith("//"):
        text = f"https:{text}"
    if not re.match(r"^https?://", text, flags=re.IGNORECASE):
        text = f"https://{text.lstrip('/')}"
    try:
        parsed = urlparse(text)
    except Exception:
        return ""
    host = (parsed.netloc or "").lower()
    if "amazon." not in host:
        return ""
    match = re.search(r"/(?:dp|gp/product|product-reviews)/([A-Z0-9]{10})(?:[/?#]|$)", parsed.path or "", flags=re.IGNORECASE)
    if not match:
        query = parse_qs(parsed.query or "", keep_blank_values=False)
        asin = (query.get("asin") or query.get("ASIN") or [""])[0].strip().upper()
    else:
        asin = match.group(1).upper()
    if not re.fullmatch(r"[A-Z0-9]{10}", asin or ""):
        return ""
    origin = "https://www.amazon.com" if host == "amazon.com" else f"{parsed.scheme or 'https'}://{parsed.netloc}"
    return f"{origin}/product-reviews/{asin}?sortBy=recent"


def _resolve_amazon_review_urls(run_params: dict) -> list[str]:
    urls: list[str] = []
    seen_asins: set[str] = set()
    for key in ("review_urls", "product_urls", "item_links", "links"):
        for candidate in _extract_candidate_urls_from_text(run_params.get(key)):
            normalized = _normalize_amazon_reviews_url(candidate)
            if not normalized:
                continue
            asin_match = re.search(r"/product-reviews/([A-Z0-9]{10})(?:[/?#]|$)", normalized, flags=re.IGNORECASE)
            asin = asin_match.group(1).upper() if asin_match else normalized
            if asin in seen_asins:
                continue
            seen_asins.add(asin)
            urls.append(normalized)
    return urls


def _first_url_from_value(value) -> str:
    values = value if isinstance(value, list) else [value]
    for item in values:
        if isinstance(item, dict):
            for key in ("url", "link", "product_url", "productUrl", "goods_url", "goodsUrl", "item_url", "itemUrl"):
                found = _first_url_from_value(item.get(key))
                if found:
                    return found
            for key in ("rows", "urls", "links", "product_urls", "productUrls"):
                found = _first_url_from_value(item.get(key))
                if found:
                    return found
            continue
        text = str(item or "").strip()
        if not text:
            continue
        match = re.search(r"https?://[^\s,，;；、]+", text)
        if match:
            return match.group(0).rstrip(",，;；、")
        return text
    return ""


def _resolve_task_target_entry_url(adapter_id: str, task_id: str, run_params: dict, fallback_url: str) -> str:
    fallback = str(fallback_url or "").strip()
    if (adapter_id, task_id) == ("temu", "single_product_reviews"):
        product_url = _first_url_from_value(
            run_params.get("product_url")
            or run_params.get("product_urls")
            or run_params.get("productUrls")
            or run_params.get("goods_url")
            or run_params.get("item_url")
            or run_params.get("url")
            or ""
        )
        if product_url:
            return product_url
        return fallback
    if (adapter_id, task_id) == ("shopify-ops-assistant", "traffic_data"):
        parsed = urlparse(fallback or "https://admin.shopify.com/store/balabala-global/analytics/reports/sessions_over_time")
        query = parse_qs(parsed.query or "", keep_blank_values=False)
        store_key = str(run_params.get("store_key") or run_params.get("store") or "").strip()
        if not store_key:
            match = re.search(r"/store/([^/]+)", parsed.path or "")
            store_key = match.group(1) if match else "balabala-global"
        ql = (query.get("ql") or [""])[0].strip()
        if not ql:
            ql = "\n".join([
                "FROM sessions",
                "SHOW online_store_visitors,",
                "  sessions",
                "WHERE human_or_bot_session IN ('human', 'bot')",
                "TIMESERIES day",
                "WITH TOTALS, PERCENT_CHANGE",
                "SINCE startOfDay(-7d)",
                "UNTIL today",
                "COMPARE TO previous_period",
                "ORDER BY day ASC",
                "LIMIT 1000",
                "VISUALIZE sessions TYPE line",
            ])
        params = urlencode({"ql": ql})
        return f"https://admin.shopify.com/store/{store_key}/analytics/reports/sessions_over_time?{params}"
    if adapter_id == "tiktok-ops-assistant" and task_id in {"product_analytics", "product_rating", "product_management_export", "creator_video_download"}:
        raw_regions = run_params.get("shop_regions") or run_params.get("regions")
        if isinstance(raw_regions, list):
            selected_regions = [str(item) for item in raw_regions]
        else:
            selected_regions = _extract_candidate_urls_from_text(raw_regions)
        region = ""
        for candidate in selected_regions:
            normalized = str(candidate or "").strip().upper()
            if normalized == "ALL":
                continue
            if re.fullmatch(r"[A-Z]{2}", normalized):
                region = normalized
                break
        if not region:
            region = "US"
        parsed = urlparse(fallback)
        query = parse_qs(parsed.query or "", keep_blank_values=False)
        shop_id = (query.get("shop_id") or [""])[0].strip()
        if task_id == "creator_video_download":
            suffix = f"?shop_region={region}"
            if shop_id:
                suffix += f"&shop_id={shop_id}"
            return f"https://affiliate.tiktokshopglobalselling.com/insights/transaction-analysis{suffix}"
        host = "seller.us.tiktokshopglobalselling.com" if region == "US" else "seller.eu.tiktokshopglobalselling.com"
        suffix = f"?shop_region={region}"
        if shop_id:
            suffix += f"&shop_id={shop_id}"
        path_map = {
            "product_analytics": "/compass/product-traffic-analysis",
            "product_management_export": "/product/manage",
            "product_rating": "/product/rating",
        }
        path = path_map.get(task_id, "/product/rating")
        return f"https://{host}{path}{suffix}"
    if (adapter_id, task_id) == ("amazon-ops-assistant", "amazon_reviews_full_export"):
        urls = _resolve_amazon_review_urls(run_params)
        if urls:
            return urls[0]
        raise ValueError("请先填写有效的 Amazon 商品链接或评论页链接")
    if (adapter_id, task_id) != ("tmall-ops-assistant", "buyer_reviews"):
        return fallback
    urls = _resolve_tmall_buyer_review_item_urls(run_params)
    if urls:
        return urls[0]
    raise ValueError("请先填写有效的天猫商品链接")


def _resolve_task_open_mode(adapter_id: str, task_id: str, run_params: dict, task_param_ids: set[str]) -> str:
    if (adapter_id, task_id) == ("tmall-ops-assistant", "buyer_reviews"):
        return "new"
    return str(run_params.get('mode') or ('new' if 'mode' not in task_param_ids else 'current')).strip().lower()


async def _evaluate_filename_context(runner, expression: str) -> dict:
    try:
        result = await runner.evaluate(expression)
    except Exception:
        return {}
    if not result.success or not result.data or not isinstance(result.data, list):
        return {}
    row = result.data[0] or {}
    return row if isinstance(row, dict) else {}


async def _build_export_filename_context(adapter_id: str, task_id: str, run_params: dict, runner) -> dict:
    def normalize_range_text(value: str) -> str:
        return str(value or '').strip().replace(' ~ ', '~')

    def normalize_list_scope(value, empty_label: str) -> str:
        if not isinstance(value, list):
            return empty_label
        clean_values = [str(item).strip() for item in value if str(item).strip()]
        return '+'.join(clean_values) if clean_values else empty_label

    def extract_temu_goods_id(value) -> str:
        text = str(value or '').strip()
        if not text:
            return ''
        try:
            parsed = urlparse(text)
            query = parse_qs(parsed.query)
            for key in ('goods_id', 'goodsId', 'goodsid', 'product_id'):
                candidate = str((query.get(key) or [''])[0]).strip()
                if re.fullmatch(r'\d{8,}', candidate):
                    return candidate
        except Exception:
            pass
        match = (
            re.search(r'(?:^|[?&#])goods[_-]?id=(\d{8,})', text, re.I)
            or re.search(r'-g-(\d{8,})(?:[./?#]|$)', text, re.I)
        )
        return match.group(1) if match else ''

    if adapter_id == 'shein-helper':
        shein_context_js = r"""
(() => {
  const textOf = el => (el?.textContent || '').replace(/\s+/g, ' ').trim()
  const clean = value => String(value || '').replace(/\s+/g, ' ').trim()
  const normalizeHyphenDate = value => {
    const match = clean(value).match(/\d{4}[-/]\d{2}[-/]\d{2}/)
    return match ? match[0].replace(/\//g, '-') : ''
  }
  const normalizeCompactDate = value => {
    const digits = clean(value).replace(/[^\d]/g, '')
    return digits.length >= 8
      ? `${digits.slice(0, 4)}-${digits.slice(4, 6)}-${digits.slice(6, 8)}`
      : ''
  }
  const formatRange = (start, end) => {
    if (start && end) return `${start}~${end}`
    return start || end || ''
  }
  const shared = window.__CRAWSHRIMP_SHARED__ || {}
  const feedbackTemplate = shared.api_template && typeof shared.api_template === 'object'
    ? shared.api_template
    : {}
  const feedbackFilter = feedbackTemplate.filter_payload && typeof feedbackTemplate.filter_payload === 'object'
    ? feedbackTemplate.filter_payload
    : (feedbackTemplate.payload && typeof feedbackTemplate.payload === 'object' ? feedbackTemplate.payload : {})
  const feedbackStart = normalizeHyphenDate(feedbackFilter.startCommentTime || feedbackFilter.commentStartTime || '')
  const feedbackEnd = normalizeHyphenDate(feedbackFilter.commentEndTime || feedbackFilter.endCommentTime || '')

  const templateQueue = Array.isArray(shared.api_templates)
    ? shared.api_templates.filter(item => item && typeof item === 'object')
    : []
  const merchandiseTemplate = templateQueue[0]
    || (shared.api_template && typeof shared.api_template === 'object' ? shared.api_template : {})
  const merchandiseFilter = merchandiseTemplate.filter_payload && typeof merchandiseTemplate.filter_payload === 'object'
    ? merchandiseTemplate.filter_payload
    : (merchandiseTemplate.payload && typeof merchandiseTemplate.payload === 'object' ? merchandiseTemplate.payload : {})
  const merchandiseStart = normalizeCompactDate(merchandiseFilter.startDate || merchandiseFilter.dt || '')
  const merchandiseEnd = normalizeCompactDate(merchandiseFilter.endDate || merchandiseFilter.dt || merchandiseFilter.startDate || '')

  const isCheckedButton = button => {
    if (!button) return false
    const className = String(button.className || '')
    if (/checked|Checked|active|Active/.test(className)) return true
    const ariaPressed = String(button.getAttribute?.('aria-pressed') || '').toLowerCase()
    if (ariaPressed === 'true') return true
    const radio = button.querySelector?.('input[type="radio"], input')
    return !!radio?.checked
  }
  const merchTimeScope = [...document.querySelectorAll('button')]
    .map(button => ({ button, label: clean(textOf(button)) }))
    .find(item => ['昨天', '近7天', '近30天', '自定义'].includes(item.label) && isCheckedButton(item.button))
    ?.label || ''

  return {
    success: true,
    data: [{
      review_date_range: formatRange(feedbackStart, feedbackEnd),
      merch_time_scope: merchTimeScope,
      merch_date_range: formatRange(merchandiseStart, merchandiseEnd),
    }],
    meta: { has_more: false }
  }
})()
"""

        page_ctx = await _evaluate_filename_context(runner, shein_context_js)
        ctx = {'shop_name': 'Shein'}

        if task_id == 'product_feedback':
            custom_range = run_params.get('review_date_range') or {}
            start = str(custom_range.get('start') or '').strip()
            end = str(custom_range.get('end') or '').strip()
            if start and end:
                ctx.update({
                    'time_scope': '自定义',
                    'date_range': f"{start}~{end}",
                })
            else:
                ctx.update({
                    'time_scope': '当前筛选',
                    'date_range': normalize_range_text(page_ctx.get('review_date_range') or '') or '当前筛选',
                })
            return ctx

        if task_id == 'merchandise_details':
            time_mode = str(run_params.get('time_mode') or '').strip().lower()
            scope_map = {
                'yesterday': '昨天',
                'last7': '近7天',
                'last30': '近30天',
                'custom': '自定义',
            }
            custom_range = run_params.get('custom_date_range') or {}
            start = str(custom_range.get('start') or '').strip()
            end = str(custom_range.get('end') or '').strip()
            if time_mode == 'custom' and start and end:
                ctx.update({
                    'time_scope': '自定义',
                    'date_range': f"{start}~{end}",
                })
            else:
                ctx.update({
                    'time_scope': scope_map.get(time_mode) or str(page_ctx.get('merch_time_scope') or '').strip() or '当前筛选',
                    'date_range': normalize_range_text(page_ctx.get('merch_date_range') or '') or '当前筛选',
                })
            return ctx

        ctx.setdefault('time_scope', '当前筛选')
        ctx.setdefault('date_range', '当前筛选')
        return ctx

    if adapter_id == 'tiktok-ops-assistant':
        ctx = {
            'time_scope': '当前筛选',
            'date_range': '当前筛选',
        }
        tiktok_regions = run_params.get('shop_regions') or run_params.get('regions') or []
        ctx['region_scope'] = normalize_list_scope(tiktok_regions, '全部地区')
        if task_id == 'product_management_export':
            status_label_map = {
                'all': '全部',
                'active': '在售',
                'reviewing': '审核中',
                'violation': '需要关注',
                'deactivate': '已下架',
                'draft': '草稿',
                'deleted': '已删除',
                'sold-out': '售罄',
                'listing-failed': '审核失败',
                'suspended': '平台下架',
                'frozen': '已冻结',
                'traffic-deboost': '流量管控',
            }
            statuses = run_params.get('product_statuses') or run_params.get('statuses') or []
            if isinstance(statuses, list):
                clean_statuses = [
                    status_label_map.get(str(item).strip(), str(item).strip())
                    for item in statuses
                    if str(item).strip()
                ]
                ctx['status_scope'] = '+'.join(clean_statuses) if clean_statuses else '默认状态'
            else:
                raw_status = str(statuses or '').strip()
                ctx['status_scope'] = status_label_map.get(raw_status, raw_status) or '默认状态'
        return ctx

    if adapter_id == 'aliexpress-ops-assistant':
        ctx = {
            'time_scope': '当前筛选',
            'date_range': '当前筛选',
        }
        date_context_js = r"""
(() => {
  const clean = value => String(value || '').replace(/\s+/g, ' ').trim()
  const normalizeDate = value => {
    const match = clean(value).match(/(\d{4})[-/](\d{1,2})[-/](\d{1,2})/)
    if (!match) return ''
    return `${match[1]}-${String(match[2]).padStart(2, '0')}-${String(match[3]).padStart(2, '0')}`
  }
  const dates = [...document.querySelectorAll('input')]
    .map(input => normalizeDate(input.value || input.getAttribute('value') || ''))
    .filter(Boolean)
  return {
    success: true,
    data: [{
      stat_date: dates[0] || '',
      end_date: dates[1] || dates[0] || '',
    }],
    meta: { has_more: false }
  }
})()
"""
        page_ctx = await _evaluate_filename_context(runner, date_context_js)
        time_label_map = {
            'page': '当前筛选',
            'recent1': '最近1天',
            'recent7': '最近7天',
            'recent30': '最近30天',
            'day': '自然日',
            'week': '自然周',
            'month': '自然月',
        }
        time_range = str(run_params.get('time_range') or '').strip()
        if time_range:
            ctx['time_scope'] = time_label_map.get(time_range, time_range)
        stat_date = str(run_params.get('stat_date') or run_params.get('date') or '').strip()
        start = stat_date or str(page_ctx.get('stat_date') or '').strip()
        end = stat_date or str(page_ctx.get('end_date') or start).strip()
        if start and end:
            ctx['date_range'] = f"{start}~{end}"
        elif start:
            ctx['date_range'] = start
        if task_id == 'product_ranking':
            rank_label_map = {
                'pay_amt': '支付榜',
                'uv': '访客榜',
                'wishlist': '收藏量',
                'cart': '加购榜',
            }
            rank_type = str(run_params.get('rank_type') or '').strip()
            ctx['rank_scope'] = rank_label_map.get(rank_type, rank_type) or '支付榜'
        return ctx

    if adapter_id != 'temu':
        return {}

    backend_context_js = r"""
(() => {
  const textOf = el => (el?.textContent || '').replace(/\s+/g, ' ').trim()
  const normalizeShopName = value => String(value || '')
    .replace(/\s+/g, ' ')
    .replace(/\s+\d+\s*人关注.*$/, '')
    .trim()
  const shared = window.__CRAWSHRIMP_SHARED__ || {}
  const root = document.querySelector('[class*="userInfo"], [class*="seller-name"], [class*="account-info_userInfo"]')
  const blacklist = new Set(['服务市场', '履约中心', '学习', '运营对接', '规则中心', '消息', '客服', '反馈', '99+', '首页', '数据中心', '商品数据', 'TEMU Agent Center', 'Beta', 'Seller Central'])
  const seen = new Set()
  const candidates = []
  const push = text => {
    const clean = normalizeShopName(text)
    if (!clean || seen.has(clean)) return
    seen.add(clean)
    candidates.push(clean)
  }
  const sharedShopName = normalizeShopName(shared.shopName || shared.shop_name || '')
  const strictShopName = [...document.querySelectorAll('[class*="account-info_mallInfo__"], [class*="account-info_accountInfo__"]')]
    .map(el => normalizeShopName(textOf(el)))
    .find(text => text && text.length <= 80 && !blacklist.has(text)) || ''
  if (root) {
    push(textOf(root.querySelector('[class*="elli_outerWrapper"], [class*="seller-name"], [class*="shopName"], [class*="userName"]')))
    for (const el of root.querySelectorAll('*')) {
      const text = textOf(el)
      if (!text || text.length > 80) continue
      if ([...el.children].some(c => textOf(c) === text)) continue
      push(text)
    }
  }
  for (const el of document.querySelectorAll(
    '[class*="elli_outerWrapper"], [class*="mallInfo"], [class*="accountInfo"], [class*="account-info_mallInfo"], [class*="account-info_accountInfo"]'
  )) {
    const text = textOf(el)
    if (!text || text.length > 80) continue
    push(text)
  }
  const shopName =
    sharedShopName ||
    strictShopName ||
    candidates.find(text => /shop/i.test(text)) ||
    candidates.find(text => /[A-Za-z]/.test(text) && !blacklist.has(text)) ||
    candidates.find(text => !blacklist.has(text)) ||
    ''
  const rowContainsLabel = (row, label) => {
    if (!row || !label) return false
    if (textOf(row).includes(label)) return true
    return [...row.querySelectorAll('div, label, span, td, th, button, a')]
      .some(node => textOf(node).includes(label))
  }
  const timeRow = [...document.querySelectorAll('[class*="index-module__row___"]')].find(row => {
    const label = textOf(row.querySelector('[class*="index-module__row_label___"]'))
    if (label === '时间区间') return true
    if (label === '时间' && row.querySelector('input[data-testid="beast-core-rangePicker-htmlInput"]')) return true
    return false
  }) || null
  const reviewTimeRow = [...document.querySelectorAll('[class*="flat-field_item__"]')]
    .find(row => rowContainsLabel(row, '评价时间')) || null
  const timeScope = textOf(timeRow?.querySelector('input[data-testid="beast-core-select-htmlInput"]'))
  const dateRange = textOf(timeRow?.querySelector('input[data-testid="beast-core-rangePicker-htmlInput"], input[class*="RPR_input_"]'))
  const reviewActive = [...(reviewTimeRow?.querySelectorAll('[class*="flat-field_capsule__"]') || [])]
    .find(el => String(el.className || '').includes('flat-field_active__')) || null
  const reviewTimeScope = textOf(reviewActive)
  const reviewInput = reviewTimeRow?.querySelector('input[data-testid="beast-core-rangePicker-htmlInput"], input[class*="RPR_input_"]') || null
  const reviewDateRange = String(reviewInput?.value || '').trim() || (/^\d{4}-\d{2}-\d{2}\s*~\s*\d{4}-\d{2}-\d{2}$/.test(reviewTimeScope) ? reviewTimeScope : '')
  const activeOuterSite = textOf(
    [...document.querySelectorAll('a[class*="index-module__drItem___"]')]
      .find(el => String(el.className || '').includes('index-module__active___')) || null
  )
  return {
    success: true,
    data: [{
      shared_shop_name: sharedShopName,
      strict_shop_name: strictShopName,
      shop_name: shopName,
      time_scope: timeScope,
      date_range: dateRange,
      review_time_scope: reviewTimeScope,
      review_date_range: reviewDateRange,
      active_outer_site: activeOuterSite,
    }],
    meta: { has_more: false }
  }
})()
"""

    storefront_context_js = r"""
(() => {
  const textOf = el => (el?.textContent || '').replace(/\s+/g, ' ').trim()
  const h1 = textOf(document.querySelector('h1'))
  const title = String(document.title || '').replace(/\s*-\s*Great Offers.*$/i, '').trim()
  const goodsMatch = String(location.href || '').match(/(?:[?&]goods[_-]?id=|-g-)(\d{8,})/i)
  return {
    success: true,
    data: [{
      shop_name: h1 || title || '',
      current_url: String(location.href || ''),
      goods_id: goodsMatch ? goodsMatch[1] : '',
    }],
    meta: { has_more: false }
  }
})()
"""

    page_ctx = {}
    if task_id not in {'reviews', 'store_items', 'single_product_reviews'}:
        page_ctx = await _evaluate_filename_context(runner, backend_context_js)
    elif task_id in {'reviews', 'store_items', 'single_product_reviews'}:
        page_ctx = await _evaluate_filename_context(runner, storefront_context_js)

    shared_shop_name = str(page_ctx.get('shared_shop_name') or '').strip()
    strict_shop_name = str(page_ctx.get('strict_shop_name') or '').strip()
    shop_name = shared_shop_name or strict_shop_name or str(page_ctx.get('shop_name') or '').strip()
    if not shop_name and task_id in {'reviews', 'store_items'}:
        shop_url = str(run_params.get('shop_url') or '').strip()
        if shop_url:
            try:
                parsed = urlparse(shop_url)
                shop_name = parse_qs(parsed.query).get('mall_id', [''])[0].strip()
            except Exception:
                shop_name = ''

    ctx = {'shop_name': shop_name or 'Temu店铺'}

    if task_id == 'single_product_reviews':
        product_url_for_filename = _first_url_from_value(
            run_params.get('product_url')
            or run_params.get('product_urls')
            or run_params.get('productUrls')
            or ''
        )
        goods_url_for_filename = _first_url_from_value(
            run_params.get('goods_url')
            or run_params.get('goodsUrl')
            or run_params.get('item_url')
            or run_params.get('itemUrl')
            or run_params.get('url')
            or ''
        )
        goods_id = (
            extract_temu_goods_id(run_params.get('goods_id'))
            or extract_temu_goods_id(product_url_for_filename)
            or extract_temu_goods_id(goods_url_for_filename)
            or extract_temu_goods_id(page_ctx.get('goods_id'))
            or extract_temu_goods_id(page_ctx.get('current_url'))
        )
        ctx['goods_id'] = goods_id or '未知商品'

    def normalize_temporal_value(value) -> str:
        if isinstance(value, str):
            return str(value or '').strip()
        if not value or not isinstance(value, dict):
            return ''
        direct = str(value.get('value') or '').strip()
        start = str(value.get('start') or '').strip()
        end = str(value.get('end') or '').strip()
        if direct:
            return direct
        if start and end:
            return f"{start}~{end}"
        return start or end

    def resolve_relative_range(label: str, mapping: dict[str, int]) -> str:
        label = str(label or '').strip()
        if not label:
            return ''
        today = datetime.now().date()
        if label == '今日':
            return f"{today.isoformat()}~{today.isoformat()}"
        if label == '昨日':
            day = today - timedelta(days=1)
            return f"{day.isoformat()}~{day.isoformat()}"
        if label == '本周':
            start_day = today - timedelta(days=today.weekday())
            return f"{start_day.isoformat()}~{today.isoformat()}"
        if label == '本月':
            start_day = today.replace(day=1)
            return f"{start_day.isoformat()}~{today.isoformat()}"
        days = mapping.get(label)
        if not days:
            return ''
        start_day = today - timedelta(days=days - 1)
        return f"{start_day.isoformat()}~{today.isoformat()}"

    def resolve_time_filename_context() -> dict:
        candidates = [
            {
                'time_param': 'review_time_range',
                'custom_param': 'custom_review_time_range',
                'page_scope_key': 'review_time_scope',
                'page_range_key': 'review_date_range',
                'relative_map': {'近30天': 30, '近60天': 60, '近90天': 90},
                'custom_scope_key': '',
                'custom_scope_label': '自定义',
            },
            {
                'time_param': 'time_range',
                'custom_param': 'custom_range',
                'page_scope_key': 'time_scope',
                'page_range_key': 'date_range',
                'relative_map': {'昨日': 1, '近7日': 7, '近30日': 30},
                'custom_scope_key': '',
                'custom_scope_label': '自定义',
            },
            {
                'time_param': 'return_time_range',
                'custom_param': 'custom_return_time_range',
                'page_scope_key': 'time_scope',
                'page_range_key': 'date_range',
                'relative_map': {'今日': 0, '昨日': 1, '近7日': 7, '近30日': 30},
                'custom_scope_key': 'return_time_field',
                'custom_scope_label': '自定义',
            },
            {
                'time_param': 'qc_time_range',
                'custom_param': 'custom_qc_time_range',
                'page_scope_key': 'time_scope',
                'page_range_key': 'date_range',
                'relative_map': {'今日': 0, '昨日': 1, '近7日': 7, '近30日': 30},
                'custom_scope_key': '',
                'custom_scope_label': '最新抽检时间',
            },
            {
                'time_param': '',
                'custom_param': 'custom_qc_time_range',
                'page_scope_key': '',
                'page_range_key': '',
                'relative_map': {},
                'custom_scope_key': '',
                'custom_scope_label': '最新抽检时间',
            },
            {
                'time_param': '',
                'custom_param': 'custom_return_time_range',
                'page_scope_key': '',
                'page_range_key': '',
                'relative_map': {},
                'custom_scope_key': 'return_time_field',
                'custom_scope_label': '自定义',
            },
            {
                'time_param': 'list_time_range',
                'custom_param': None,
                'page_scope_key': '',
                'page_range_key': '',
                'relative_map': {'近7日': 7, '近30日': 30},
                'custom_scope_key': '',
                'custom_scope_label': '自定义',
            },
            {
                'time_param': 'detail_time_range',
                'custom_param': None,
                'page_scope_key': '',
                'page_range_key': '',
                'relative_map': {'近7日': 7, '近30日': 30},
                'custom_scope_key': '',
                'custom_scope_label': '自定义',
            },
        ]

        for spec in candidates:
            time_param = spec['time_param']
            custom_param = spec['custom_param']
            if time_param not in run_params and custom_param not in run_params:
                continue

            time_range = str(run_params.get(time_param) or '').strip() if time_param else ''
            custom_range = run_params.get(custom_param) or {}
            start = str(custom_range.get('start') or '').strip()
            end = str(custom_range.get('end') or '').strip()
            page_time_scope = str(page_ctx.get(spec['page_scope_key']) or '').strip()
            page_date_range = normalize_range_text(page_ctx.get(spec['page_range_key']) or '')
            custom_scope = str(run_params.get(spec.get('custom_scope_key') or '') or '').strip()

            if start and end:
                return {
                    'time_scope': custom_scope or page_time_scope or spec.get('custom_scope_label') or '自定义',
                    'date_range': f"{start}~{end}",
                }

            if time_range == '自定义':
                return {
                    'time_scope': custom_scope or page_time_scope or spec.get('custom_scope_label') or '自定义',
                    'date_range': f"{start}~{end}" if start and end else (page_date_range or '自定义'),
                }
            if time_range:
                return {
                    'time_scope': time_range,
                    'date_range': page_date_range or resolve_relative_range(time_range, spec['relative_map']) or time_range,
                }
            return {
                'time_scope': page_time_scope or '当前筛选',
                'date_range': page_date_range or resolve_relative_range(page_time_scope, spec['relative_map']) or page_time_scope or '当前筛选',
            }

        date_range_specs = [
            ('stat_date_range', '统计日期'),
            ('stat_week', '按周'),
            ('stat_month', '按月'),
            ('bill_date_range', '对账日期'),
        ]
        for param_key, scope_label in date_range_specs:
            if param_key not in run_params:
                continue
            value = normalize_temporal_value(run_params.get(param_key))
            if not value:
                continue
            return {'time_scope': scope_label, 'date_range': normalize_range_text(value)}

        return {}

    ctx.update(resolve_time_filename_context())
    ctx.setdefault('time_scope', '当前筛选')
    ctx.setdefault('date_range', '当前筛选')

    regions = run_params.get('regions') or []
    ctx['region_scope'] = normalize_list_scope(regions, '全部地区')
    ctx['site_scope'] = normalize_list_scope(run_params.get('outer_sites') or [], '全部站点')
    ctx['detail_site_scope'] = normalize_list_scope(run_params.get('detail_sites') or [], '全部详情站点')
    ctx['detail_grain_scope'] = normalize_list_scope(run_params.get('detail_grains') or [], '全部粒度')

    stat_grain = str(run_params.get('stat_grain') or '').strip()
    if stat_grain:
        ctx['grain_scope'] = stat_grain
    elif 'stat_grain' in run_params:
        ctx['grain_scope'] = '当前粒度'
    else:
        ctx['grain_scope'] = '当前粒度'

    goods_statuses = run_params.get('goods_statuses') or []
    ctx['goods_status_scope'] = normalize_list_scope(goods_statuses, '全部状态')
    qc_statuses = run_params.get('qc_statuses') or []
    ctx['qc_status_scope'] = normalize_list_scope(qc_statuses, '全部结果')
    if task_id == 'recommended_retail_price':
        ctx['status_scope'] = '全部状态'
    if task_id == 'quality_dashboard':
        ctx['quality_tab_scope'] = '品质分析+品质优化'

    return ctx


def _url_matches_prefix(url: str, prefix: str) -> bool:
    if not url or not prefix:
        return False
    if url.startswith(prefix):
        return True

    try:
        url_p = urlparse(url)
        prefix_p = urlparse(prefix)
    except Exception:
        return False

    url_host = (url_p.hostname or '').lower()
    prefix_host = (prefix_p.hostname or '').lower()
    url_path = url_p.path or '/'
    prefix_path = prefix_p.path or '/'

    if not url_host or not prefix_host:
        return False

    if url_host == prefix_host:
        return url_path.startswith(prefix_path)

    # Temu seller 现在会按区域切到 agentseller-us.temu.com / agentseller-xx.temu.com
    if prefix_host == 'agentseller.temu.com' and url_host.endswith('.temu.com') and url_host.startswith('agentseller'):
        normalized_prefix = prefix_path if prefix_path.endswith('/') else prefix_path + '/'
        normalized_url = url_path if url_path.endswith('/') else url_path + '/'
        return normalized_url.startswith(normalized_prefix) or normalized_prefix == '//'

    return False


def _is_temu_agentseller_url(url: str) -> bool:
    try:
        parsed = urlparse(str(url or ""))
    except Exception:
        return False
    host = (parsed.hostname or "").lower()
    return host == "agentseller.temu.com" or (
        host.endswith(".temu.com") and host.startswith("agentseller-")
    )


def _is_temu_kuajingmaihuo_url(url: str) -> bool:
    try:
        parsed = urlparse(str(url or ""))
    except Exception:
        return False
    return (parsed.hostname or "").lower() == "seller.kuajingmaihuo.com"


def _is_temu_backend_url(url: str) -> bool:
    return _is_temu_agentseller_url(url) or _is_temu_kuajingmaihuo_url(url)


def _is_temu_no_auth_url(url: str) -> bool:
    try:
        parsed = urlparse(str(url or ""))
    except Exception:
        return False
    host = (parsed.hostname or "").lower()
    path = parsed.path or ""
    return host == "seller.temu.com" and path == "/no-auth.html"


def _temu_new_tab_entry_url(target_entry_url: str) -> str:
    """Open Temu backend from its shell first; direct protected deep links can hit no-auth."""
    if not _is_temu_backend_url(target_entry_url):
        return target_entry_url
    try:
        parsed = urlparse(str(target_entry_url or ""))
    except Exception:
        return "https://agentseller.temu.com"
    host = (parsed.hostname or "agentseller.temu.com").lower()
    if host == "agentseller.temu.com" or (host.endswith(".temu.com") and host.startswith("agentseller-")):
        return f"{parsed.scheme or 'https'}://{host}/"
    if host == "seller.kuajingmaihuo.com":
        return "https://seller.kuajingmaihuo.com/main/order-manager/shipping-list"
    return "https://agentseller.temu.com"


def _is_temu_opener_tab_url(url: str) -> bool:
    return _is_temu_backend_url(url) and not _is_temu_no_auth_url(url)


def _is_doudian_fxg_url(url: str) -> bool:
    try:
        parsed = urlparse(str(url or ""))
    except Exception:
        return False
    return (parsed.hostname or "").lower() == "fxg.jinritemai.com"


def _is_doudian_backend_url(url: str) -> bool:
    if not _is_doudian_fxg_url(url):
        return False
    try:
        parsed = urlparse(str(url or ""))
    except Exception:
        return False
    return (parsed.path or "").startswith("/ffa/")


def _is_doudian_opener_tab_url(url: str) -> bool:
    return _is_doudian_backend_url(url)


def _doudian_opener_tab_sort_key(target_entry_url: str, tab_url: str) -> tuple[int, int]:
    try:
        target_path = urlparse(str(target_entry_url or "")).path or ""
        tab_path = urlparse(str(tab_url or "")).path or ""
    except Exception:
        return (2, 0)
    target_section = "/".join(target_path.strip("/").split("/")[:2])
    tab_section = "/".join(tab_path.strip("/").split("/")[:2])
    return (
        0 if target_section and target_section == tab_section else 1,
        0 if "/campaign" in tab_path or "/merchant" in tab_path else 1,
    )


def _temu_opener_tab_sort_key(target_entry_url: str, tab_url: str) -> tuple[int, int]:
    if _is_temu_kuajingmaihuo_url(target_entry_url):
        return (0 if _is_temu_kuajingmaihuo_url(tab_url) else 1, 0)
    if _is_temu_agentseller_url(target_entry_url):
        return (0 if _is_temu_agentseller_url(tab_url) else 1, 0)
    return (0, 0)


def _is_compatible_current_tab_for_task(adapter_id: str, task_id: str, target_entry_url: str, tab_url: str) -> bool:
    if adapter_id != "temu":
        return False
    if _is_temu_agentseller_url(target_entry_url):
        return _is_temu_agentseller_url(tab_url)
    if _is_temu_kuajingmaihuo_url(target_entry_url):
        return _is_temu_kuajingmaihuo_url(tab_url)
    return False


def _normalize_export_guard_value(value) -> str:
    return str(value or '').strip()


def _dedupe_temu_goods_traffic_detail_rows(data_rows):
    if not isinstance(data_rows, list):
        return data_rows

    deduped = []
    seen = set()

    for row in data_rows:
        if not isinstance(row, dict):
            deduped.append(row)
            continue

        if _normalize_export_guard_value(row.get('记录类型')) != '明细':
            deduped.append(row)
            continue

        outer_site = _normalize_export_guard_value(row.get('外层站点'))
        spu = _normalize_export_guard_value(row.get('SPU'))
        detail_site = _normalize_export_guard_value(row.get('详情站点筛选'))
        detail_grain = _normalize_export_guard_value(row.get('详情时间粒度'))
        date = _normalize_export_guard_value(row.get('日期'))
        site = _normalize_export_guard_value(row.get('站点'))

        # 仅对明细行做最终去重保险，避免误伤异常行或上下文不完整的记录。
        if not all([spu, detail_site, detail_grain, date, site]):
            deduped.append(row)
            continue

        key = (outer_site, spu, detail_site, detail_grain, date, site)
        if key in seen:
            continue
        seen.add(key)
        deduped.append(row)

    return deduped


def _dedupe_amazon_reviews_full_export_rows(data_rows):
    if not isinstance(data_rows, list):
        return data_rows

    deduped = []
    seen = set()

    for row in data_rows:
        if not isinstance(row, dict):
            deduped.append(row)
            continue

        asin = _normalize_export_guard_value(row.get('ASIN'))
        review_id = _normalize_export_guard_value(row.get('评价ID'))
        if asin and review_id:
            key = (asin, review_id)
        else:
            key_parts = [
                asin,
                _normalize_export_guard_value(row.get('买家昵称')),
                _normalize_export_guard_value(row.get('评价时间')),
                _normalize_export_guard_value(row.get('评分')),
                _normalize_export_guard_value(row.get('评价标题')),
                _normalize_export_guard_value(row.get('评价内容')),
            ]
            if not asin or not any(key_parts[1:]):
                deduped.append(row)
                continue
            key = tuple(key_parts)

        if key in seen:
            continue
        seen.add(key)
        deduped.append(row)

    return deduped


def _apply_final_export_guards(adapter_id: str, task_id: str, data_rows):
    guarded = data_rows if isinstance(data_rows, list) else list(data_rows or [])
    if adapter_id == 'temu' and task_id == 'goods_traffic_detail':
        guarded = _dedupe_temu_goods_traffic_detail_rows(guarded)
    if adapter_id == 'amazon-ops-assistant' and task_id == 'amazon_reviews_full_export':
        guarded = _dedupe_amazon_reviews_full_export_rows(guarded)
    return guarded


def _safe_local_name(value: str, fallback: str = "item") -> str:
    text = str(value or "").strip()
    text = re.sub(r'[\x00-\x1f]+', '', text)
    text = re.sub(r'[\\/:*?"<>|]+', '_', text)
    text = re.sub(r'\s+', ' ', text).strip(' ._')
    return text or fallback


def _expand_user_configured_local_path(value: str) -> Path:
    text = str(value or "").strip()
    text = re.sub(
        r"%([^%]+)%",
        lambda match: os.environ.get(match.group(1), match.group(0)),
        text,
    )
    text = os.path.expandvars(text)
    if os.sep == "/" and "\\" in text and not re.match(r"^[A-Za-z]:\\", text):
        text = text.replace("\\", os.sep)
    return Path(text).expanduser()


def _ensure_unique_local_path(path: Path) -> Path:
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    index = 2
    while True:
        candidate = path.with_name(f"{stem}_{index}{suffix}")
        if not candidate.exists():
            return candidate
        index += 1


def _ensure_unique_local_dir(path: Path) -> Path:
    if not path.exists():
        return path
    index = 2
    while True:
        candidate = path.with_name(f"{path.name}_{index}")
        if not candidate.exists():
            return candidate
        index += 1


def _parse_semir_cloud_path(raw_value: str) -> dict:
    raw = str(raw_value or "").strip()
    divider = raw.find("//")
    if divider < 0:
      return {"mount_name": "", "relative_path": "", "relative_prefix": ""}
    mount_name = str(raw[:divider]).strip()
    relative_path = "/".join(
        part.strip()
        for part in raw[divider + 2:].replace("\\", "/").split("/")
        if str(part or "").strip()
    )
    return {
        "mount_name": mount_name,
        "relative_path": relative_path,
        "relative_prefix": f"{relative_path}/" if relative_path else "",
    }


def _normalize_semir_package_layout(raw_value: str) -> str:
    return "flat" if str(raw_value or "").strip().lower() == "flat" else "by_code"


def _semir_output_folder_and_filename(
    fullpath: str,
    relative_path: str,
    fallback_name: str,
    preferred_filename: str = "",
) -> tuple[str, str]:
    normalized_fullpath = str(fullpath or "").replace("\\", "/").strip("/")
    normalized_base = str(relative_path or "").replace("\\", "/").strip("/")
    preferred_clean_name = (
        _safe_local_name(preferred_filename, "")
        if str(preferred_filename or "").strip()
        else ""
    )

    if normalized_fullpath:
        tail = normalized_fullpath
    elif normalized_base:
        tail = f"{normalized_base}/{fallback_name}"
    else:
        tail = fallback_name

    safe_parts = [_safe_local_name(part, "item") for part in tail.split("/") if str(part or "").strip()]
    if not safe_parts:
        clean_name = preferred_clean_name or _safe_local_name(fallback_name, "file")
        return ("未分类路径", clean_name)

    if len(safe_parts) == 1:
        clean_name = preferred_clean_name or _safe_local_name(safe_parts[0], _safe_local_name(fallback_name, "file"))
        return ("未分类路径", clean_name)

    clean_name = preferred_clean_name or _safe_local_name(safe_parts[-1], _safe_local_name(fallback_name, "file"))
    folder_name = "__".join(part for part in safe_parts[:-1] if str(part or "").strip()) or "未分类路径"
    return (_safe_local_name(folder_name, "未分类路径"), clean_name)


def _build_semir_package_target(
    package_root: Path,
    input_code: str,
    folder_name: str,
    clean_filename: str,
    *,
    package_layout: str,
    preserve_path: bool,
) -> Path:
    if _normalize_semir_package_layout(package_layout) == "flat":
        return package_root / clean_filename
    if preserve_path:
        return package_root / input_code / folder_name / clean_filename
    return package_root / input_code / clean_filename


def _copy_file_to_unique_target(source: Path, target: Path) -> Path:
    target.parent.mkdir(parents=True, exist_ok=True)
    unique_target = _ensure_unique_local_path(target)
    shutil.copy2(source, unique_target)
    return unique_target


def _move_file_to_unique_target(source: Path, target: Path) -> Path:
    target.parent.mkdir(parents=True, exist_ok=True)
    unique_target = _ensure_unique_local_path(target)
    try:
        shutil.move(str(source), str(unique_target))
    except Exception:
        shutil.copy2(source, unique_target)
        try:
            source.unlink()
        except Exception:
            logger.debug("Failed to clean source after fallback copy %s", source, exc_info=True)
    return unique_target


def _move_dir_to_unique_target(source: Path, target: Path) -> Path:
    target.parent.mkdir(parents=True, exist_ok=True)
    unique_target = _ensure_unique_local_dir(target)
    try:
        shutil.move(str(source), str(unique_target))
    except Exception:
        shutil.copytree(source, unique_target)
        shutil.rmtree(source, ignore_errors=True)
    return unique_target


def _relocate_runtime_file_to_unique_target(source: Path, target: Path, runtime_dir: Path) -> Path:
    try:
        if _is_within_directory(source, runtime_dir):
            return _move_file_to_unique_target(source, target)
    except Exception:
        logger.debug("Failed to decide runtime file relocation for %s", source, exc_info=True)
    return _copy_file_to_unique_target(source, target)


def _cleanup_semir_runtime_artifacts(runtime_files: list, package_root: Optional[Path]) -> None:
    for file_path in runtime_files or []:
        source = Path(str(file_path or "")).expanduser()
        try:
            if source.exists() and source.is_file():
                source.unlink()
        except Exception:
            logger.debug("Failed to clean semir runtime file %s", source, exc_info=True)

    try:
        if package_root and package_root.exists() and package_root.is_dir():
            shutil.rmtree(package_root)
    except Exception:
        logger.debug("Failed to clean semir package root %s", package_root, exc_info=True)


def _rewrite_bala_material_summary_excels(exported_files: list, data_rows: list, log) -> None:
    rows = [row for row in (data_rows or []) if isinstance(row, dict)]
    if not rows:
        return
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("openpyxl 不可用，无法刷新巴拉素材结果表最终路径") from exc

    for raw_path in exported_files or []:
        path = Path(str(raw_path or "")).expanduser()
        if path.suffix.lower() != ".xlsx":
            continue
        if not path.is_file():
            raise RuntimeError(f"巴拉素材结果表不存在，无法刷新本地文件路径: {path}")
        workbook = None
        try:
            workbook = load_workbook(path)
            changed = False
            for sheet in workbook.worksheets:
                header_row = 0
                path_column = 0
                for row_index in range(1, min(sheet.max_row, 3) + 1):
                    for column_index in range(1, sheet.max_column + 1):
                        if str(sheet.cell(row=row_index, column=column_index).value or "").strip() == "本地文件":
                            header_row = row_index
                            path_column = column_index
                            break
                    if path_column:
                        break
                if not path_column:
                    continue
                for offset, row in enumerate(rows, start=1):
                    sheet.cell(row=header_row + offset, column=path_column).value = str(row.get("本地文件") or "")
                changed = True
            if not changed:
                raise RuntimeError(f"巴拉素材结果表缺少本地文件列: {path}")
            workbook.save(path)
            log(f"Bala video material Excel paths refreshed: {path}")
        except Exception as exc:
            raise RuntimeError(f"巴拉素材结果表最终路径刷新失败: {path}: {exc}") from exc
        finally:
            if workbook is not None:
                workbook.close()


def _cleanup_shenhui_runtime_artifacts(runtime_files: list, package_root: Optional[Path], work_dir: Optional[Path]) -> None:
    _cleanup_semir_runtime_artifacts(runtime_files, package_root)
    try:
        if work_dir and work_dir.exists() and work_dir.is_dir():
            shutil.rmtree(work_dir)
    except Exception:
        logger.debug("Failed to clean shenhui work dir %s", work_dir, exc_info=True)


def _is_within_directory(path: Path, directory: Path) -> bool:
    try:
        path.resolve(strict=False).relative_to(directory.resolve(strict=False))
        return True
    except Exception:
        return False


def _is_direct_child_file(path: Path, directory: Path) -> bool:
    try:
        return path.resolve(strict=False).parent == directory.resolve(strict=False)
    except Exception:
        return False


def _cleanup_runtime_artifact_dir(runtime_artifact_dir: str, preserve_paths: Optional[list] = None) -> None:
    runtime_dir = Path(str(runtime_artifact_dir or "")).expanduser()
    if not runtime_dir.exists() or not runtime_dir.is_dir():
        return

    preserved: set[Path] = set()
    for raw_path in preserve_paths or []:
        if not str(raw_path or "").strip():
            continue
        try:
            path = Path(str(raw_path)).expanduser().resolve(strict=False)
        except Exception:
            continue
        if _is_within_directory(path, runtime_dir):
            preserved.add(path)

    for path in sorted(runtime_dir.rglob("*"), key=lambda item: len(item.parts), reverse=True):
        path_resolved = path.resolve(strict=False)
        if path_resolved in preserved:
            continue
        if path.is_dir() and any(_is_within_directory(preserved_path, path_resolved) for preserved_path in preserved):
            try:
                next(path.iterdir())
            except StopIteration:
                try:
                    path.rmdir()
                except Exception:
                    logger.debug("Failed to remove empty runtime dir %s", path, exc_info=True)
            except Exception:
                logger.debug("Failed to inspect runtime dir %s", path, exc_info=True)
            continue

        try:
            if path.is_dir():
                path.rmdir()
            else:
                path.unlink(missing_ok=True)
        except Exception:
            logger.debug("Failed to clean runtime artifact %s", path, exc_info=True)

    try:
        next(runtime_dir.iterdir())
    except StopIteration:
        try:
            runtime_dir.rmdir()
        except Exception:
            logger.debug("Failed to remove empty runtime dir %s", runtime_dir, exc_info=True)
    except Exception:
        logger.debug("Failed to inspect runtime dir %s", runtime_dir, exc_info=True)


def _cleanup_orphaned_runtime_artifacts(runs: Optional[list] = None) -> None:
    for run in runs or []:
        adapter_id = str((run or {}).get("adapter_id") or "").strip()
        task_id = str((run or {}).get("task_id") or "").strip()
        run_id = (run or {}).get("id")
        if (adapter_id, task_id) not in RUNTIME_CLEANUP_TASKS or run_id is None:
            continue
        try:
            runtime_dir = data_sink.artifact_dir_path(adapter_id, task_id, int(run_id), "runtime")
        except Exception:
            continue
        _cleanup_runtime_artifact_dir(str(runtime_dir), preserve_paths=[])


def _append_note(existing: object, note: str) -> str:
    current = str(existing or "").strip()
    clean_note = str(note or "").strip()
    if not current:
        return clean_note
    if not clean_note or clean_note in current:
        return current
    return f"{current}; {clean_note}"


def _truthy_param(value: object) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (list, tuple, set)):
        return any(_truthy_param(item) for item in value)
    normalized = str(value or "").strip().lower()
    return normalized in {"1", "true", "yes", "y", "on", "enabled", "zip", "auto_zip", "auto_zip_package", "是", "压缩"}


def _shenhui_auto_zip_enabled(run_params: Optional[dict]) -> bool:
    return _truthy_param((run_params or {}).get("auto_zip_package"))


def _bala_video_source_folder(source_type: object) -> str:
    normalized = str(source_type or "").strip().lower()
    if normalized == "model":
        return "01_模拍原图"
    if normalized == "detail":
        return "02_商品细节图"
    return "03_其他素材"


def _bala_video_image_threshold_bytes(run_params: Optional[dict], row: Optional[dict] = None) -> int:
    raw = (row or {}).get("__compress_threshold_bytes")
    if raw in (None, ""):
        raw = (run_params or {}).get("max_image_mb")
        try:
            mb = float(raw)
        except Exception:
            mb = 20
        return max(1, min(80, int(mb))) * 1024 * 1024
    try:
        threshold = int(float(raw))
    except Exception:
        threshold = 20 * 1024 * 1024
    return max(1, threshold)


def _format_mb(size: int) -> str:
    try:
        return f"{size / 1024 / 1024:.2f}MB"
    except Exception:
        return "0.00MB"


def _bala_video_rgb_image(image):
    from PIL import Image

    if image.mode in {"RGBA", "LA"} or "transparency" in getattr(image, "info", {}):
        rgba = image.convert("RGBA")
        canvas = Image.new("RGB", rgba.size, "white")
        canvas.paste(rgba, mask=rgba.split()[-1])
        return canvas
    return image.convert("RGB")


def _compress_bala_video_image_if_needed(path: Path, threshold_bytes: int, log) -> tuple[Path, str]:
    if not path.is_file():
        return path, "压缩跳过：文件不存在"
    before = path.stat().st_size
    if before <= threshold_bytes:
        return path, "无需压缩"

    try:
        from PIL import Image, ImageOps
    except Exception:
        return path, "压缩失败：Pillow 不可用"

    suffix = path.suffix.lower()
    if suffix not in {".jpg", ".jpeg", ".png", ".webp", ".bmp", ".tif", ".tiff"}:
        return path, "压缩跳过：非支持图片格式"

    temp_dir = _ensure_unique_local_dir(path.parent / ".__bala_compress_tmp")
    temp_dir.mkdir(parents=True, exist_ok=True)
    best_temp = None
    best_size = before
    target_suffix = suffix if suffix in {".jpg", ".jpeg"} else ".jpg"
    try:
        with Image.open(path) as source:
            image = _bala_video_rgb_image(ImageOps.exif_transpose(source))
            original_width, original_height = image.size
            quality_steps = [92, 88, 84, 80, 76, 72, 68, 64, 60, 56, 52, 48]
            scale_steps = [1.0, 0.92, 0.84, 0.76, 0.68, 0.60, 0.52]

            for scale in scale_steps:
                if scale == 1.0:
                    candidate_image = image
                else:
                    width = max(1, int(original_width * scale))
                    height = max(1, int(original_height * scale))
                    candidate_image = image.resize((width, height))

                for quality in quality_steps:
                    temp_path = temp_dir / f"{path.stem}-{int(scale * 100)}-{quality}{target_suffix}"
                    candidate_image.save(
                        temp_path,
                        "JPEG",
                        quality=quality,
                        optimize=True,
                        progressive=True,
                    )
                    size = temp_path.stat().st_size
                    if size < best_size:
                        if best_temp and best_temp.exists():
                            best_temp.unlink(missing_ok=True)
                        best_temp = temp_path
                        best_size = size
                    else:
                        temp_path.unlink(missing_ok=True)
                    if size <= threshold_bytes:
                        raise StopIteration
    except StopIteration:
        pass
    except Exception as exc:
        shutil.rmtree(temp_dir, ignore_errors=True)
        return path, f"压缩失败：{exc}"

    if not best_temp or not best_temp.is_file() or best_size >= before:
        shutil.rmtree(temp_dir, ignore_errors=True)
        return path, f"压缩失败：未生成更小文件（原始 {_format_mb(before)}）"

    final_path = path if target_suffix == suffix else _ensure_unique_local_path(path.with_suffix(target_suffix))
    try:
        if final_path == path:
            best_temp.replace(path)
        else:
            shutil.move(str(best_temp), str(final_path))
            path.unlink(missing_ok=True)
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)

    after = final_path.stat().st_size if final_path.is_file() else best_size
    status = f"已压缩：{_format_mb(before)} -> {_format_mb(after)}"
    if after > threshold_bytes:
        status = f"{status}，仍超过阈值 {_format_mb(threshold_bytes)}"
    if log:
        log(f"Bala video image compressed: {final_path} ({status})")
    return final_path, status


def _verify_tiktok_creator_video_download_rows(data_rows: list, log=None) -> list:
    success_total = 0
    missing_rows = []
    for row in data_rows or []:
        if not isinstance(row, dict):
            continue
        if str(row.get("下载结果") or "").strip() != "已下载":
            continue
        success_total += 1
        local_path_text = str(row.get("本地文件") or "").strip()
        local_path = Path(local_path_text).expanduser() if local_path_text else None
        if local_path and local_path.is_file() and local_path.stat().st_size > 0:
            continue
        missing_rows.append(row)

    for row in missing_rows:
        row["下载结果"] = "下载失败"
        row["下载备注"] = _append_note(row.get("下载备注"), "下载完成后本地文件缺失，未打包")
        row["本地文件"] = ""

    if missing_rows and log:
        available = success_total - len(missing_rows)
        log(
            "TikTok creator video local file check: "
            f"{available}/{success_total} downloaded files still exist; "
            f"{len(missing_rows)} missing rows marked failed before export"
        )
    return data_rows


def _default_output_root_for_runtime(runtime_dir: Path, exported_files: list) -> Path:
    exported_refs = [str(path or "").strip() for path in exported_files or [] if str(path or "").strip()]
    if exported_refs:
        return Path(exported_refs[0]).expanduser().parent
    if runtime_dir.parent.name == "runtime":
        return runtime_dir.parent.parent
    return runtime_dir.parent


def _semir_output_roots(runtime_dir: Path, exported_files: list, run_params: dict) -> tuple[Path, Optional[Path], list[str]]:
    exported_refs = [str(path) for path in exported_files or [] if str(path or "").strip()]
    default_root = _default_output_root_for_runtime(runtime_dir, exported_refs)
    export_folder = str((run_params or {}).get("export_folder") or "").strip()
    export_root = Path(export_folder).expanduser() if export_folder else None
    if export_root:
        export_root.mkdir(parents=True, exist_ok=True)
    return default_root, export_root, exported_refs


BALA_AI_VIDEO_ADAPTER_ID = "bala-ai-video-assistant"
BALA_AI_FACE_BACKGROUND_TASK_ID = "bala_ai_face_background_generate"
BALA_MODEL_LIBRARY_MANIFEST = "assets/model-library/manifest.json"
BALA_VIDEO_TEMPLATE_DIR = Path.home() / "Downloads" / "巴拉AI视频模板库"
BALA_VIDEO_TEMPLATE_CATALOG_JSON = BALA_VIDEO_TEMPLATE_DIR / "template-catalog.json"
BALA_VIDEO_TEMPLATE_CATALOG_CSV = BALA_VIDEO_TEMPLATE_DIR / "template-catalog.csv"
BALA_SEEDANCE_CLI_DIR = Path(__file__).resolve().parents[1] / "integrations" / "seedanceCLI"
BALA_SEEDANCE_DEFAULT_MODEL = "doubao-seedance-2-0-260128"
BALA_HAPPYHORSE_CLI_DIR = Path(__file__).resolve().parents[1] / "integrations" / "bailianCLI"
BALA_HAPPYHORSE_MODELS = {
    "t2v": "happyhorse-1.1-t2v",
    "i2v": "happyhorse-1.1-i2v",
    "r2v": "happyhorse-1.1-r2v",
}
BALA_KLING_VIDEO_MODELS = {
    "kling-v3": "kling/kling-v3-video-generation",
    "kling-omni": "kling/kling-v3-omni-video-generation",
}
BALA_PIXVERSE_PROVIDER = "pixverse-motioncontrol"
BALA_PIXVERSE_MODEL = "pixverse/pixverse-motioncontrol"
BALA_BAILIAN_VIDEO_PROVIDERS = frozenset({
    "happyhorse",
    *BALA_KLING_VIDEO_MODELS.keys(),
    BALA_PIXVERSE_PROVIDER,
})
BALA_KLING_MODES = frozenset({"std", "pro"})
BALA_KLING_RATIOS = frozenset({"16:9", "9:16", "1:1"})
BALA_PIXVERSE_RESOLUTIONS = frozenset({"360P", "540P", "720P"})
BALA_AI_IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp"}


def _bala_operation_type(run_params: dict) -> str:
    return bala_ai_video_review.normalize_operation_type((run_params or {}).get("operation_type"))


def _bala_operation_label(operation_type: str) -> str:
    return {
        "face_swap": "AI换脸",
        "background_swap": "AI换背景",
        "outfit_swap": "AI换装",
        "pose_swap": "AI换姿势",
    }.get(operation_type, "AI换脸")


def _bala_string_list(value: object) -> list[str]:
    if isinstance(value, list):
        source = value
    else:
        source = re.split(r"[\n,，、;；]+", str(value or ""))
    result: list[str] = []
    seen: set[str] = set()
    for item in source:
        text = str(item or "").strip()
        if not text or text in seen:
            continue
        result.append(text)
        seen.add(text)
    return result


def _bala_image_path_allowed(path: Path) -> bool:
    return path.suffix.lower() in BALA_AI_IMAGE_EXTS and path.is_file()


def _bala_material_listing_paths(run_params: dict) -> list[Path]:
    listing = (run_params or {}).get("material_root_files")
    items = listing.get("paths") if isinstance(listing, dict) else listing
    paths: list[Path] = []
    for item in items or []:
        raw_path = item.get("path") if isinstance(item, dict) else item
        value = str(raw_path or "").strip()
        if value:
            paths.append(Path(value).expanduser())
    return paths


def _bala_selected_source_paths(run_params: dict) -> list[Path]:
    source_images = (run_params or {}).get("source_images")
    raw_paths = source_images.get("paths") if isinstance(source_images, dict) else source_images
    return [Path(path).expanduser() for path in _bala_string_list(raw_paths)]


def _bala_scan_material_root(run_params: dict) -> list[Path]:
    root_value = str((run_params or {}).get("material_root") or "").strip()
    if not root_value:
        return []
    root = Path(root_value).expanduser()
    if not root.is_dir():
        return []
    paths: list[Path] = []
    for path in root.rglob("*"):
        if path.is_file() and path.suffix.lower() in BALA_AI_IMAGE_EXTS:
            paths.append(path)
    return paths


def _bala_prefer_model_source(path: Path) -> int:
    text = str(path).replace("\\", "/")
    name = path.name
    if re.search(r"(?i)(?:^|[-_\s])AI(?:$|[-_\s.])|AI换", name):
        return -1
    if "01_模拍原图" in text:
        return 0
    if "模拍" in text:
        return 1
    if "02_商品细节图" in text or "商品细节" in text or "平拍" in text:
        return 9
    return 3


def _bala_extract_style_code(path: Path) -> str:
    for part in reversed(path.parts):
        match = re.search(r"\b(\d{12})\b", part)
        if match:
            return match.group(1)
    match = re.search(r"\b(\d{12})\b", str(path))
    return match.group(1) if match else ""


def _bala_source_images_for_generation(run_params: dict) -> tuple[list[Path], list[str]]:
    errors: list[str] = []
    selected = _bala_selected_source_paths(run_params)
    if selected:
        candidates = selected
    else:
        candidates = _bala_material_listing_paths(run_params) or _bala_scan_material_root(run_params)
    if not candidates:
        return [], ["请先选择待处理图片，或选择包含 01_模拍原图 的素材根目录"]

    valid: list[Path] = []
    seen: set[str] = set()
    for path in candidates:
        resolved = path.expanduser().resolve(strict=False)
        identity = str(resolved)
        if identity in seen:
            continue
        seen.add(identity)
        if not _bala_image_path_allowed(resolved):
            continue
        valid.append(resolved)
    if not valid:
        return [], ["没有找到可用的 JPG/PNG/WEBP 素材图"]

    if not selected:
        preferred = [path for path in valid if _bala_prefer_model_source(path) <= 1]
        if preferred:
            valid = preferred
    valid.sort(key=lambda path: (_bala_prefer_model_source(path), str(path).lower()))
    source_limit = _task_int_param(run_params, "source_limit", 3, min_value=1, max_value=100)
    return valid[:source_limit], errors


def _load_bala_model_library() -> tuple[list[dict], list[str]]:
    try:
        manifest_path = adapter_loader.resolve_adapter_file(
            BALA_AI_VIDEO_ADAPTER_ID,
            BALA_MODEL_LIBRARY_MANIFEST,
            allowed_suffixes=None,
        )
    except Exception as exc:
        return [], [f"无法读取巴拉模特库 manifest: {exc}"]

    try:
        payload = json.loads(manifest_path.read_text(encoding="utf-8"))
    except Exception as exc:
        return [], [f"巴拉模特库 manifest 解析失败: {exc}"]

    items = payload.get("items") if isinstance(payload, dict) else []
    root = manifest_path.parent
    result: list[dict] = []
    for item in items or []:
        if not isinstance(item, dict):
            continue
        relative_path = str(item.get("relative_path") or "").strip()
        if not relative_path:
            continue
        path = (root / relative_path).resolve(strict=False)
        if not _bala_image_path_allowed(path):
            continue
        result.append({**item, "path": str(path)})
    if not result:
        return [], ["巴拉模特库为空或图片文件缺失"]
    return result, []


def _bala_model_rank(item: dict) -> tuple[int, str]:
    expression = str(item.get("expression") or "")
    if re.search(r"标准|原始|正脸", expression):
        rank = 0
    elif re.search(r"微笑|大笑|可爱", expression):
        rank = 1
    elif re.search(r"侧脸|低头|抬头", expression):
        rank = 2
    else:
        rank = 3
    return rank, str(item.get("id") or "")


def _bala_models_for_generation(run_params: dict) -> tuple[list[dict], list[str]]:
    library, errors = _load_bala_model_library()
    if errors:
        return [], errors
    by_id = {str(item.get("id") or ""): item for item in library}
    explicit_ids = _bala_string_list((run_params or {}).get("model_ref_ids"))
    if explicit_ids:
        selected: list[dict] = []
        missing: list[str] = []
        for model_id in explicit_ids:
            item = by_id.get(model_id)
            if item:
                selected.append(item)
            else:
                missing.append(model_id)
        if selected:
            return selected, ([f"未找到指定模特图: {'、'.join(missing)}"] if missing else [])
        return [], [f"指定模特图均不存在: {'、'.join(missing)}"]

    groups = _bala_string_list((run_params or {}).get("model_groups")) or ["100女"]
    per_group = _task_int_param(run_params, "models_per_group", 1, min_value=1, max_value=12)
    selected = []
    missing_groups = []
    for group in groups:
        candidates = [item for item in library if str(item.get("group") or "") == group]
        candidates.sort(key=_bala_model_rank)
        if not candidates:
            missing_groups.append(group)
            continue
        selected.extend(candidates[:per_group])
    warnings = [f"未找到模特分组: {'、'.join(missing_groups)}"] if missing_groups else []
    if not selected:
        return [], warnings or ["请至少选择一个可用的模特分组或具体模特图 ID"]
    return selected, warnings


def _bala_param_image_paths(run_params: dict, key: str) -> list[Path]:
    value = (run_params or {}).get(key)
    raw = value.get("paths") if isinstance(value, dict) else value
    result: list[Path] = []
    for item in _bala_string_list(raw):
        path = Path(item).expanduser()
        if _bala_image_path_allowed(path):
            result.append(path)
    return result


def _bala_ai_prompt(source_path: Path, model_item: dict, background_prompt: str, extra: str = "") -> str:
    model_label = (
        f"{model_item.get('group_label') or model_item.get('group') or ''}"
        f" {model_item.get('expression') or ''}"
    ).strip()
    lines = [
        "请基于输入的第一张童装商品模拍图进行真实电商照片级图像编辑。",
        "必须保留第一张图中的童装商品、版型、颜色、图案、材质、穿搭关系、身体姿态和主体构图，不要改款、不要换衣服、不要增加无关商品。",
        f"参考第二张巴拉 AI 模特头像素材（{model_label}），只对人物脸部五官、年龄气质和表情进行自然替换；脸部需要贴合原图角度、光线、肤色和清晰度。",
        f"将背景替换为：{background_prompt}。背景要真实自然、干净高级、符合儿童服装商业图，不要出现文字、水印、Logo、吊牌、合格证或多余人物。",
        "输出完整儿童模特穿着原商品的照片，真实摄影质感，画面边缘完整，不要拼贴感。",
    ]
    extra_text = str(extra or "").strip()
    if extra_text:
        lines.append(f"补充要求：{extra_text}")
    return "\n".join(lines)


def _bala_ai_operation_prompt(source_path: Path, model_item: dict, run_params: dict, operation_type: str) -> str:
    extra = str((run_params or {}).get("prompt_extra") or "").strip()
    if operation_type == "face_swap":
        background_prompt = str((run_params or {}).get("background_prompt") or "保持原图自然背景").strip()
        return _bala_ai_prompt(source_path, model_item or {}, background_prompt, extra)
    if operation_type == "background_swap":
        lines = [
            "请基于输入的童装商品模拍图进行真实电商照片级图像编辑。",
            "必须保留童装商品、版型、颜色、图案、材质、穿搭关系、人物姿态和主体构图，不要改款、不要换衣服。",
            f"将背景替换为：{str((run_params or {}).get('background_prompt') or '').strip()}。背景真实自然，符合儿童服装商业图，不要出现文字、水印、Logo、吊牌、合格证或多余人物。",
        ]
    elif operation_type == "outfit_swap":
        lines = [
            "请基于输入的童装模特图进行真实电商照片级换装编辑。",
            "将服装图中的商品自然替换到模特身上，保留童装版型、颜色、图案、材质和穿搭比例。",
            "参考搭配图和同款不同色图理解衣服结构，输出自然、清晰、无拼贴感的完整儿童模特图。",
        ]
    elif operation_type == "pose_swap":
        lines = [
            "请基于输入的童装商品模拍图进行真实电商照片级姿势调整。",
            f"姿势要求：{str((run_params or {}).get('pose_prompt') or '').strip()}。",
            "必须保留童装商品、版型、颜色、图案、材质和人物年龄气质，不要出现文字、水印、Logo 或多余人物。",
        ]
    else:
        lines = ["请基于输入图片生成真实自然的儿童服装电商图。"]
    if extra:
        lines.append(f"补充要求：{extra}")
    return "\n".join(lines)


def _bala_ai_error_row(message: str, note: str = "") -> dict:
    return {
        "源图序号": "",
        "源图文件": "",
        "款号": "",
        "操作类型": "",
        "模特ID": "",
        "模特分组": "",
        "模特年龄段": "",
        "模特性别": "",
        "模特表情": "",
        "背景Prompt": "",
        "服装图文件": "",
        "搭配参考图文件": "",
        "同款不同色参考图文件": "",
        "姿势Prompt": "",
        "完整Prompt": "",
        "AI任务UID": "",
        "AI任务标题": "",
        "提交状态": "未创建",
        "批次UID": "",
        "运行UID": "",
        "1XM任务ID": "",
        "轮询URL": "",
        "审批批次UID": "",
        "审批看板": "",
        "下一步Provider": "",
        "AI生图工作台": "AI 生图",
        "执行结果": "失败",
        "备注": _append_note(message, note),
    }


def _bala_create_ai_image_job_row(
    *,
    source_index: int,
    source_path: Path,
    model_item: dict,
    run_params: dict,
    operation_type: str,
    prompt: str,
    submit_async: bool,
) -> dict:
    style_code = _bala_extract_style_code(source_path)
    model_id = str(model_item.get("id") or "")
    operation_label = _bala_operation_label(operation_type)
    background_prompt = str(run_params.get("background_prompt") or "").strip()
    garment_paths = _bala_param_image_paths(run_params, "garment_images")
    outfit_reference_paths = _bala_param_image_paths(run_params, "outfit_reference_images")
    variant_reference_paths = _bala_param_image_paths(run_params, "variant_reference_images")
    pose_prompt = str(run_params.get("pose_prompt") or "").strip()
    reference_paths: list[str] = []
    reference_roles: list[str] = []
    if operation_type == "face_swap" and str(model_item.get("path") or "").strip():
        reference_paths.append(str(model_item.get("path") or ""))
        reference_roles.append("bala_ai_model_face")
    elif operation_type == "outfit_swap":
        for role, paths in [
            ("garment", garment_paths),
            ("outfit_reference", outfit_reference_paths),
            ("variant_reference", variant_reference_paths),
        ]:
            for path in paths:
                reference_paths.append(str(path))
                reference_roles.append(role)
    title_parts = [
        f"巴拉{operation_label}",
        style_code or source_path.stem,
        str(model_item.get("group") or ""),
        str(model_item.get("expression") or ""),
    ]
    title = _safe_local_name(" ".join(part for part in title_parts if part), f"巴拉{operation_label}")
    output_dir = str(run_params.get("output_dir") or "").strip()
    workspace_dir = str(run_params.get("workspace_dir") or output_dir).strip()
    params = {
        "prompt": prompt,
        "size": str(run_params.get("image_size") or "1536x2048").strip() or "1536x2048",
        "quality": str(run_params.get("quality") or "high").strip() or "high",
        "output_format": str(run_params.get("output_format") or "png").strip() or "png",
        "n": 1,
        "model_key_tier": str(run_params.get("model_key_tier") or "4k").strip() or "4k",
        "main_image_path": str(source_path),
        "reference_image_paths": reference_paths,
        "workflow": BALA_AI_FACE_BACKGROUND_TASK_ID,
        "surface": "ai-video-workflow",
        "workspace_dir": workspace_dir,
        "operation_type": operation_type,
        "background_prompt": background_prompt,
        "garment_image_paths": [str(path) for path in garment_paths],
        "outfit_reference_image_paths": [str(path) for path in outfit_reference_paths],
        "variant_reference_image_paths": [str(path) for path in variant_reference_paths],
        "pose_prompt": pose_prompt,
        "style_code": style_code,
        "bala_model": {
            "id": model_id,
            "group": model_item.get("group"),
            "age_label": model_item.get("age_label"),
            "gender": model_item.get("gender"),
            "expression": model_item.get("expression"),
        },
    }
    job = data_sink.create_ai_image_job({
        "title": title,
        "prompt": prompt,
        "model_key": str(run_params.get("model") or "gpt-image-2").strip() or "gpt-image-2",
        "status": "draft",
        "output_dir": output_dir,
        "params": params,
        "summary": {
            "workflow": BALA_AI_FACE_BACKGROUND_TASK_ID,
            "surface": "ai-video-workflow",
            "workspace_dir": workspace_dir,
            "operation_type": operation_type,
            "source_path": str(source_path),
            "model_id": model_id,
            "background_prompt": background_prompt,
        },
    })
    job_uid = str(job.get("job_uid") or "")
    data_sink.create_ai_image_asset({
        "job_uid": job_uid,
        "kind": "main",
        "source_type": "local",
        "path": str(source_path),
        "sort_order": 0,
        "meta": {"role": "source_product_model", "style_code": style_code},
    })
    for index, path in enumerate(reference_paths, start=1):
        role = reference_roles[index - 1] if index - 1 < len(reference_roles) else "reference"
        data_sink.create_ai_image_asset({
            "job_uid": job_uid,
            "kind": "reference",
            "source_type": "local",
            "path": path,
            "sort_order": index,
            "meta": {"role": role, "model_id": model_id if role == "bala_ai_model_face" else ""},
        })

    submit_status = "已创建，未提交"
    batch_uid = ""
    run_uid = ""
    task_id = ""
    poll_url = ""
    result = "已创建"
    note = ""
    if submit_async:
        try:
            batch = ai_image_service.submit_workbench_batch(
                job_uid,
                [{"title": operation_label, "prompt": prompt, "count": 1}],
                request_uid=f"bala-ai-face-bg-{uuid4().hex}",
            )
            submit_status = "已异步提交" if batch.get("accepted") else "提交未接受"
            batch_uid = str(batch.get("batch_uid") or "")
            first_run = next(iter(batch.get("runs") or []), {})
            run_uid = str(first_run.get("run_uid") or "")
            task_id = str(first_run.get("task_id") or "")
            poll_url = str(first_run.get("poll_url") or "")
            if first_run.get("error"):
                note = str(first_run.get("error") or "")
                result = "提交失败"
            else:
                result = "已提交"
        except ai_image_service.MissingModelKeyError as exc:
            submit_status = "提交失败"
            result = "已创建，提交失败"
            note = str(exc)
        except Exception as exc:
            submit_status = "提交失败"
            result = "已创建，提交失败"
            note = str(exc)

    return {
        "源图序号": source_index,
        "源图文件": str(source_path),
        "款号": style_code,
        "操作类型": operation_label,
        "模特ID": model_id,
        "模特分组": str(model_item.get("group_label") or model_item.get("group") or ""),
        "模特年龄段": str(model_item.get("age_label") or ""),
        "模特性别": str(model_item.get("gender") or ""),
        "模特表情": str(model_item.get("expression") or ""),
        "背景Prompt": background_prompt,
        "服装图文件": "\n".join(str(path) for path in garment_paths),
        "搭配参考图文件": "\n".join(str(path) for path in outfit_reference_paths),
        "同款不同色参考图文件": "\n".join(str(path) for path in variant_reference_paths),
        "姿势Prompt": pose_prompt,
        "完整Prompt": prompt,
        "AI任务UID": job_uid,
        "AI任务标题": title,
        "提交状态": submit_status,
        "批次UID": batch_uid,
        "运行UID": run_uid,
        "1XM任务ID": task_id,
        "轮询URL": poll_url,
        "审批批次UID": "",
        "审批看板": "",
        "下一步Provider": "",
        "AI生图工作台": "AI 生图",
        "执行结果": result,
        "备注": note,
    }


async def _apply_bala_ai_face_background_generate(run_params: dict, wait_for_control, log) -> list[dict]:
    operation_type = _bala_operation_type(run_params)
    operation_label = _bala_operation_label(operation_type)
    background_prompt = str((run_params or {}).get("background_prompt") or "").strip()
    if operation_type == "background_swap" and not background_prompt:
        return [{**_bala_ai_error_row("请填写换背景 Prompt"), "操作类型": operation_label}]
    if operation_type == "outfit_swap" and not _bala_param_image_paths(run_params, "garment_images"):
        return [{**_bala_ai_error_row("请至少选择一张服装图"), "操作类型": operation_label}]
    if operation_type == "pose_swap" and not str((run_params or {}).get("pose_prompt") or "").strip():
        return [{**_bala_ai_error_row("请填写换姿势 Prompt"), "操作类型": operation_label}]

    sources, source_errors = _bala_source_images_for_generation(run_params)
    models: list[dict] = [{}]
    model_warnings: list[str] = []
    if operation_type == "face_swap":
        models, model_warnings = _bala_models_for_generation(run_params)
    if source_errors and not sources:
        return [{**_bala_ai_error_row("；".join(source_errors)), "操作类型": operation_label}]
    if operation_type == "face_swap" and not models:
        return [{**_bala_ai_error_row("；".join(model_warnings or ["没有可用的巴拉 AI 模特素材"])), "操作类型": operation_label}]

    max_combinations = _task_int_param(run_params, "max_combinations", 12, min_value=1, max_value=100)
    submit_async = str((run_params or {}).get("generation_mode") or "submit_async").strip() != "create_only"
    rows: list[dict] = []
    combinations: list[tuple[int, Path, dict]] = []
    for source_index, source_path in enumerate(sources, start=1):
        for model_item in models:
            combinations.append((source_index, source_path, model_item))
            if len(combinations) >= max_combinations:
                break
        if len(combinations) >= max_combinations:
            break

    if log:
        mode = "异步提交" if submit_async else "仅创建"
        log(f"Bala AI image generation: operation={operation_type}, {len(combinations)} job(s), mode={mode}")
    for index, (source_index, source_path, model_item) in enumerate(combinations, start=1):
        await wait_for_control({"records": len(rows), "current_row": index, "phase": "create-ai-image-jobs"})
        prompt = _bala_ai_operation_prompt(source_path, model_item, run_params, operation_type)
        try:
            row = await asyncio.to_thread(
                _bala_create_ai_image_job_row,
                source_index=source_index,
                source_path=source_path,
                model_item=model_item,
                run_params=run_params,
                operation_type=operation_type,
                prompt=prompt,
                submit_async=submit_async,
            )
        except Exception as exc:
            row = {
                **_bala_ai_error_row(str(exc)),
                "源图序号": source_index,
                "源图文件": str(source_path),
                "款号": _bala_extract_style_code(source_path),
                "操作类型": operation_label,
                "模特ID": str(model_item.get("id") or ""),
                "模特分组": str(model_item.get("group_label") or model_item.get("group") or ""),
                "模特年龄段": str(model_item.get("age_label") or ""),
                "模特性别": str(model_item.get("gender") or ""),
                "模特表情": str(model_item.get("expression") or ""),
                "背景Prompt": background_prompt,
                "完整Prompt": prompt,
            }
        if model_warnings:
            row["备注"] = _append_note(row.get("备注"), "；".join(model_warnings))
        rows.append(row)

    await wait_for_control({"records": len(rows), "current_row": len(rows), "phase": "create-ai-image-jobs"})
    if not rows:
        return [{**_bala_ai_error_row("没有生成任何任务组合"), "操作类型": operation_label}]

    review_mode = str((run_params or {}).get("review_mode") or "").strip()
    if review_mode == "create_review_batch":
        review_rows = [row for row in rows if str(row.get("AI任务UID") or "").strip()]
        if review_rows:
            api_base_url = str((run_params or {}).get("approval_base_url") or _bala_ai_video_base_url()).strip()
            batch = bala_ai_video_review.build_review_batch(
                review_rows,
                run_params or {},
                str(runtime_paths.child_dir("bala-ai-video-review")),
                api_base_url,
            )
            bala_ai_video_review.save_review_batch(batch)
            for row in rows:
                if str(row.get("AI任务UID") or "").strip():
                    row["审批批次UID"] = batch.get("batch_id", "")
                    row["审批看板"] = batch.get("board_url", "")
                    row["下一步Provider"] = batch.get("video_provider", "qn_img2video")
    return rows


def _finalize_bala_ai_video_assistant_outputs(
    task_id: str,
    data_rows: list,
    runtime_files: list,
    exported_files: list,
    run_params: dict,
    runtime_artifact_dir: str,
    log,
) -> list[str]:
    if task_id != "semir_video_material_prepare":
        return [str(path) for path in (runtime_files or []) + (exported_files or []) if str(path or "").strip()]

    def rewrite_row_paths(source_root: Path, target_root: Path) -> None:
        for row in data_rows or []:
            if not isinstance(row, dict):
                continue
            raw_path = str(row.get("本地文件") or "").strip()
            if not raw_path:
                continue
            path = Path(raw_path).expanduser()
            try:
                relative = path.resolve(strict=False).relative_to(source_root.resolve(strict=False))
            except Exception:
                continue
            row["本地文件"] = str(target_root / relative)

    runtime_dir = Path(runtime_artifact_dir)
    runtime_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    package_base = _safe_local_name(
        run_params.get("package_name") or f"巴拉AI视频素材_{timestamp}",
        f"巴拉AI视频素材_{timestamp}",
    )
    package_root = _ensure_unique_local_dir(runtime_dir / package_base)

    successful_rows = []
    for row in data_rows or []:
        if not isinstance(row, dict):
            continue
        local_path = Path(str(row.get("本地文件") or "")).expanduser()
        if str(row.get("下载结果") or "").strip() != "已下载" or not local_path.is_file():
            continue
        successful_rows.append((row, local_path))

    if successful_rows:
        started_at = time.monotonic()
        log(f"Bala video material finalize started: {len(successful_rows)} downloaded files")
        for row, local_path in successful_rows:
            group_code = _safe_local_name(
                row.get("__bala_group_code") or row.get("输入款号") or row.get("输入编码") or "未分类",
                "未分类",
            )
            source_folder = _bala_video_source_folder(row.get("__bala_source_type") or row.get("素材来源"))
            clean_filename = _safe_local_name(
                row.get("__package_filename") or row.get("文件名") or local_path.name,
                local_path.name,
            )
            target = package_root / group_code / source_folder / clean_filename
            relocated = _relocate_runtime_file_to_unique_target(local_path, target, runtime_dir)
            threshold_bytes = _bala_video_image_threshold_bytes(run_params, row)
            final_path, compress_status = _compress_bala_video_image_if_needed(relocated, threshold_bytes, log)
            row["本地文件"] = str(final_path)
            row["压缩结果"] = compress_status
            if final_path.name != clean_filename:
                row["文件名"] = final_path.name
                row["__package_filename"] = final_path.name
        log(
            "Bala video material folder prepared without ZIP compression: "
            f"{package_root} ({time.monotonic() - started_at:.1f}s)"
        )

    exported_refs = [str(path) for path in exported_files or [] if str(path or "").strip()]
    final_refs = [*exported_refs]
    export_folder = str(run_params.get("export_folder") or "").strip()

    if successful_rows and package_root.exists():
        if export_folder:
            target_root = Path(export_folder).expanduser()
            target_root.mkdir(parents=True, exist_ok=True)
            external_dir = _move_dir_to_unique_target(package_root, target_root / package_root.name)
            rewrite_row_paths(package_root, external_dir)
            _rewrite_bala_material_summary_excels(exported_files, data_rows, log)
            external_refs = [str(external_dir)]
            log(f"Bala video material folder moved to export folder: {external_dir}")
            for file_path in exported_files or []:
                source = Path(str(file_path or "")).expanduser()
                if not source.is_file():
                    continue
                copied = _copy_file_to_unique_target(source, target_root / source.name)
                external_refs.append(str(copied))
            final_refs = external_refs
        else:
            target_root = _default_output_root_for_runtime(runtime_dir, exported_files)
            target_root.mkdir(parents=True, exist_ok=True)
            external_dir = _move_dir_to_unique_target(package_root, target_root / package_root.name)
            rewrite_row_paths(package_root, external_dir)
            _rewrite_bala_material_summary_excels(exported_files, data_rows, log)
            log(f"Bala video material folder moved to default output folder: {external_dir}")
            final_refs = [str(external_dir), *exported_refs]
    elif package_root.exists():
        shutil.rmtree(package_root, ignore_errors=True)

    _cleanup_semir_runtime_artifacts(runtime_files, None)
    _cleanup_runtime_artifact_dir(str(runtime_dir), preserve_paths=final_refs)

    return final_refs


def _finalize_shenhui_new_arrival_outputs(
    task_id: str,
    data_rows: list,
    runtime_files: list,
    exported_files: list,
    run_params: dict,
    runtime_artifact_dir: str,
    log,
) -> list[str]:
    if task_id == "pdf_batch_screenshot":
        return finalize_pdf_batch_screenshot_outputs(
            data_rows=data_rows,
            exported_files=exported_files,
            run_params=run_params,
            runtime_artifact_dir=runtime_artifact_dir,
            log=log,
        )

    if task_id != "prepare_upload_package":
        return [str(path) for path in (runtime_files or []) + (exported_files or []) if str(path or "").strip()]

    runtime_dir = Path(runtime_artifact_dir)
    runtime_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    package_base = _safe_local_name(
        run_params.get("package_name") or f"深绘上新图包_{timestamp}",
        f"深绘上新图包_{timestamp}",
    )
    package_root = _ensure_unique_local_dir(runtime_dir / package_base)
    pdf_work_dir = _ensure_unique_local_dir(runtime_dir / "_pdf_work")
    auto_zip_package = _shenhui_auto_zip_enabled(run_params)

    successful_rows = []
    for row in data_rows or []:
        if not isinstance(row, dict):
            continue
        local_path = Path(str(row.get("本地文件") or "")).expanduser()
        if str(row.get("下载结果") or "").strip() != "已下载" or not local_path.is_file():
            continue
        successful_rows.append((row, local_path))

    style_zip_paths = []
    if successful_rows:
        finalize_started_at = time.monotonic()
        pdf_rows = []
        image_count = 0
        pdf_count = 0
        log(
            "Shenhui package finalize started: "
            f"{len(successful_rows)} downloaded files, auto_zip={auto_zip_package}"
        )
        for row, local_path in successful_rows:
            group_code = _safe_local_name(
                row.get("__shenhui_group_code") or row.get("输入款号") or row.get("输入编码") or "未分类",
                "未分类",
            )
            role = str(row.get("__shenhui_asset_role") or "").strip().lower()

            if role == "pdf_yq" or local_path.suffix.lower() == ".pdf":
                row["__pdf_path"] = str(local_path)
                row["原始路径"] = row.get("原始路径") or str(local_path)
                row["__shenhui_group_code"] = group_code
                pdf_rows.append((row, local_path, group_code))
                pdf_count += 1
                continue

            clean_filename = _safe_local_name(
                row.get("__package_filename") or row.get("文件名") or local_path.name,
                local_path.name,
            )
            target = package_root / group_code / clean_filename
            relocated = _relocate_runtime_file_to_unique_target(local_path, target, runtime_dir)
            row["本地文件"] = str(relocated)
            image_count += 1
        log(
            "Shenhui package files arranged: "
            f"{image_count} image/material files, {pdf_count} PDF files, "
            f"{time.monotonic() - finalize_started_at:.1f}s"
        )

        if pdf_rows:
            pdf_data_rows = [row for row, _local_path, _group_code in pdf_rows]
            pdf_started_at = time.monotonic()
            converted_count = convert_pdf_rows_to_yq_output_root(
                data_rows=pdf_data_rows,
                output_root=package_root,
                pdf_work_dir=pdf_work_dir,
                run_params=run_params or {},
                log=log,
            )
            if converted_count:
                log(f"Shenhui downloaded PDF screenshots added to style packages: {converted_count}")
            for row, local_path, group_code in pdf_rows:
                if str(row.get("处理动作") or "").strip() != "截图失败":
                    continue
                pending_target = package_root / group_code / "_PDF待裁图" / _safe_local_name(
                    row.get("__package_filename") or row.get("文件名") or local_path.name,
                    local_path.name,
                )
                if local_path.is_file():
                    preserved = _copy_file_to_unique_target(local_path, pending_target)
                    row["本地文件"] = str(preserved)
                    row["备注"] = _append_note(row.get("备注"), f"原 PDF 已保留：{preserved.name}")
            log(
                "Shenhui package PDF processing finished: "
                f"{converted_count} generated images from {len(pdf_rows)} PDF files, "
                f"{time.monotonic() - pdf_started_at:.1f}s"
            )

        if auto_zip_package:
            zip_started_at = time.monotonic()
            style_zip_dir = _ensure_unique_local_dir(runtime_dir / f"{package_root.name}_deepdraw_zips")
            style_zip_dir.mkdir(parents=True, exist_ok=True)
            for style_dir in sorted(path for path in package_root.iterdir() if path.is_dir()):
                if style_dir.name.startswith("_"):
                    continue
                style_zip_path = _ensure_unique_local_path(style_zip_dir / f"{style_dir.name}.zip")
                with zipfile.ZipFile(style_zip_path, "w", compression=zipfile.ZIP_STORED) as archive:
                    for file_path in sorted(style_dir.rglob("*")):
                        if not file_path.is_file():
                            continue
                        archive.write(file_path, arcname=str(file_path.relative_to(style_dir)))
                style_zip_paths.append(style_zip_path)
            if style_zip_paths:
                log(
                    "Shenhui DeepDraw style ZIPs created: "
                    f"{len(style_zip_paths)} in {time.monotonic() - zip_started_at:.1f}s"
                )
        else:
            log(
                "Shenhui package folder prepared without ZIP compression: "
                f"{package_root} ({time.monotonic() - finalize_started_at:.1f}s)"
            )

    export_folder = str(run_params.get("export_folder") or "").strip()
    exported_refs = [str(path) for path in exported_files or [] if str(path or "").strip()]
    final_refs = [*([str(path) for path in style_zip_paths if str(path or "").strip()]), *exported_refs]

    if export_folder:
        target_root = Path(export_folder).expanduser()
        target_root.mkdir(parents=True, exist_ok=True)
        external_refs = []

        if successful_rows and auto_zip_package:
            for style_zip_path in style_zip_paths:
                if style_zip_path.exists():
                    copied_style_zip = _copy_file_to_unique_target(style_zip_path, target_root / style_zip_path.name)
                    external_refs.append(str(copied_style_zip))
            if style_zip_paths:
                log(f"Shenhui style ZIPs copied to export folder: {target_root}")
        elif successful_rows and package_root.exists():
            external_dir = _move_dir_to_unique_target(package_root, target_root / package_root.name)
            external_refs.append(str(external_dir))
            log(f"Shenhui package folder copied to export folder: {external_dir}")

        for file_path in exported_files or []:
            source = Path(str(file_path or "")).expanduser()
            if not source.is_file():
                continue
            copied = _copy_file_to_unique_target(source, target_root / source.name)
            external_refs.append(str(copied))

        if external_refs:
            final_refs = external_refs
    elif style_zip_paths:
        target_root = _default_output_root_for_runtime(runtime_dir, exported_files)
        target_root.mkdir(parents=True, exist_ok=True)
        default_refs = []
        for style_zip_path in style_zip_paths:
            if style_zip_path.exists():
                copied_style_zip = _copy_file_to_unique_target(style_zip_path, target_root / style_zip_path.name)
                default_refs.append(str(copied_style_zip))
        if default_refs:
            log(f"Shenhui style ZIPs moved to default output folder: {target_root}")
            final_refs = [*default_refs, *exported_refs]
    elif successful_rows and package_root.exists():
        target_root = _default_output_root_for_runtime(runtime_dir, exported_files)
        target_root.mkdir(parents=True, exist_ok=True)
        external_dir = _move_dir_to_unique_target(package_root, target_root / package_root.name)
        log(f"Shenhui package folder moved to default output folder: {external_dir}")
        final_refs = [str(external_dir), *exported_refs]

    if successful_rows:
        _cleanup_shenhui_runtime_artifacts(runtime_files, package_root, pdf_work_dir)
    else:
        if package_root.exists():
            shutil.rmtree(package_root, ignore_errors=True)
        if pdf_work_dir.exists():
            shutil.rmtree(pdf_work_dir, ignore_errors=True)

    _cleanup_runtime_artifact_dir(str(runtime_dir), preserve_paths=final_refs)

    return final_refs


def _build_semir_ai_material_entries(data_rows: list) -> list[dict]:
    entries = []
    for row in data_rows or []:
        if not isinstance(row, dict):
            continue
        input_code = _safe_local_name(str(row.get("输入编码") or "").strip(), "未分类")
        materials = row.get("__素材明细") or []
        if not isinstance(materials, list):
            continue
        for material in materials:
            if not isinstance(material, dict):
                continue
            local_path = Path(str(material.get("local_path") or "")).expanduser()
            if not local_path.is_file():
                continue
            entries.append({
                "input_code": input_code,
                "filename": str(material.get("filename") or local_path.name).strip() or local_path.name,
                "cloud_path": str(material.get("cloud_path") or "").strip(),
                "local_path": local_path,
                "preserve_path": True,
            })
    return entries


def _build_semir_ai_generated_entries(data_rows: list) -> list[dict]:
    entries = []
    for row in data_rows or []:
        if not isinstance(row, dict):
            continue
        input_code = _safe_local_name(str(row.get("输入编码") or "").strip(), "未分类")
        generated = row.get("__生成图明细") or []
        if not isinstance(generated, list):
            continue
        for item in generated:
            if not isinstance(item, dict):
                continue
            local_path = Path(str(item.get("local_path") or "")).expanduser()
            if not local_path.is_file():
                continue
            entries.append({
                "input_code": input_code,
                "filename": str(item.get("filename") or local_path.name).strip() or local_path.name,
                "cloud_path": "",
                "local_path": local_path,
                "preserve_path": False,
            })
    return entries


def _create_semir_ai_zip(
    entries: list[dict],
    *,
    runtime_dir: Path,
    output_root: Path,
    package_base: str,
    relative_path: str,
    flatten_code_folder: bool,
    log,
) -> Optional[str]:
    if not entries:
        return None

    package_root = _ensure_unique_local_dir(runtime_dir / package_base)
    for entry in entries:
        folder_name, clean_filename = _semir_output_folder_and_filename(
            entry.get("cloud_path") or "",
            relative_path,
            entry.get("filename") or entry["local_path"].name,
        )
        target = _build_semir_package_target(
            package_root,
            entry["input_code"],
            folder_name,
            clean_filename,
            package_layout="by_code",
            preserve_path=bool(entry.get("preserve_path")) and not flatten_code_folder,
        )
        _copy_file_to_unique_target(entry["local_path"], target)

    output_root.mkdir(parents=True, exist_ok=True)
    zip_path = _ensure_unique_local_path(output_root / f"{package_root.name}.zip")
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for file_path in package_root.rglob("*"):
            if not file_path.is_file():
                continue
            archive.write(file_path, arcname=str(file_path.relative_to(package_root.parent)))
    shutil.rmtree(package_root, ignore_errors=True)
    log(f"Semir AI package created: {zip_path}")
    return str(zip_path)


def _finalize_semir_tmall_material_match_buy_outputs(
    data_rows: list,
    runtime_files: list,
    exported_files: list,
    run_params: dict,
    runtime_artifact_dir: str,
    log,
) -> list[str]:
    runtime_dir = Path(runtime_artifact_dir)
    runtime_dir.mkdir(parents=True, exist_ok=True)
    output_root, target_root, exported_refs = _semir_output_roots(runtime_dir, exported_files, run_params)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    package_base = _safe_local_name(
        run_params.get("package_name") or f"搭配购素材包_{timestamp}",
        f"搭配购素材包_{timestamp}",
    )
    package_root = _ensure_unique_local_dir(runtime_dir / package_base)

    successful_rows = []
    for row in data_rows or []:
        if not isinstance(row, dict):
            continue
        local_path = Path(str(row.get("本地文件") or "")).expanduser()
        if str(row.get("下载结果") or "").strip() != "已下载" or not local_path.is_file():
            continue
        successful_rows.append((row, local_path))

    zip_path = None
    if successful_rows:
        package_started_at = time.monotonic()
        for row, local_path in successful_rows:
            clean_filename = _safe_local_name(
                row.get("__package_filename") or row.get("文件名") or local_path.name,
                local_path.name,
            )
            target = package_root / clean_filename
            # A matched material can be reused by multiple target IDs, so keep the
            # runtime source in place until the final cleanup step.
            _copy_file_to_unique_target(local_path, target)

        zip_output_root = target_root or output_root
        zip_output_root.mkdir(parents=True, exist_ok=True)
        zip_path = _ensure_unique_local_path(zip_output_root / f"{package_root.name}.zip")
        with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_STORED) as archive:
            for file_path in package_root.rglob("*"):
                if not file_path.is_file():
                    continue
                archive.write(file_path, arcname=str(file_path.relative_to(package_root.parent)))
        log(
            "Semir Tmall match-buy package created: "
            f"{zip_path} ({len(successful_rows)} files, {time.monotonic() - package_started_at:.1f}s)"
        )

    runtime_refs = [*exported_refs]
    if zip_path:
        runtime_refs.insert(0, str(zip_path))
    final_refs = runtime_refs

    if target_root:
        external_refs = []

        if successful_rows and zip_path and zip_path.exists():
            if _is_within_directory(zip_path, target_root):
                external_refs.append(str(zip_path))
            else:
                copied_zip = _copy_file_to_unique_target(zip_path, target_root / zip_path.name)
                external_refs.append(str(copied_zip))
            log(f"Semir Tmall match-buy ZIP ready in export folder: {target_root}")

        for file_path in exported_files or []:
            source = Path(str(file_path or "")).expanduser()
            if not source.is_file():
                continue
            copied = _copy_file_to_unique_target(source, target_root / source.name)
            external_refs.append(str(copied))

        if external_refs:
            final_refs = external_refs

    _cleanup_semir_runtime_artifacts(runtime_files, package_root)
    _cleanup_runtime_artifact_dir(str(runtime_dir), preserve_paths=final_refs)

    return final_refs


def _finalize_semir_cloud_drive_outputs(
    task_id: str,
    data_rows: list,
    runtime_files: list,
    exported_files: list,
    run_params: dict,
    runtime_artifact_dir: str,
    log,
) -> list[str]:
    if task_id != "batch_image_download":
        if task_id in {"tmall_material_match_buy", "tmall_material_new_624"}:
            return _finalize_semir_tmall_material_match_buy_outputs(
                data_rows=data_rows,
                runtime_files=runtime_files,
                exported_files=exported_files,
                run_params=run_params,
                runtime_artifact_dir=runtime_artifact_dir,
                log=log,
            )

        if task_id == "batch_ai_generate":
            runtime_dir = Path(runtime_artifact_dir)
            runtime_dir.mkdir(parents=True, exist_ok=True)
            output_root, _, exported_refs = _semir_output_roots(runtime_dir, exported_files, run_params)
            cloud = _parse_semir_cloud_path(run_params.get("cloud_path"))
            relative_path = cloud.get("relative_path") or ""
            duplicate_mode = str(run_params.get("duplicate_mode") or "first_per_stem").strip().lower()
            flatten_code_folder = duplicate_mode != "all"
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            final_refs: list[str] = []

            generated_zip = _create_semir_ai_zip(
                _build_semir_ai_generated_entries(data_rows),
                runtime_dir=runtime_dir,
                output_root=output_root,
                package_base=_safe_local_name(
                    run_params.get("generated_package_name") or f"森马云盘AI生成图包_{timestamp}",
                    f"森马云盘AI生成图包_{timestamp}",
                ),
                relative_path="",
                flatten_code_folder=True,
                log=log,
            )
            if generated_zip:
                final_refs.append(generated_zip)

            package_mode = str(run_params.get("material_package_mode") or "discard").strip().lower()
            if package_mode == "zip":
                material_zip = _create_semir_ai_zip(
                    _build_semir_ai_material_entries(data_rows),
                    runtime_dir=runtime_dir,
                    output_root=output_root,
                    package_base=_safe_local_name(
                        run_params.get("material_package_name") or f"森马云盘AI素材包_{timestamp}",
                        f"森马云盘AI素材包_{timestamp}",
                    ),
                    relative_path=relative_path,
                    flatten_code_folder=flatten_code_folder,
                    log=log,
                )
                if material_zip:
                    final_refs.append(material_zip)

            final_refs = [*final_refs, *exported_refs]
            _cleanup_semir_runtime_artifacts(runtime_files, None)
            _cleanup_runtime_artifact_dir(str(runtime_dir), preserve_paths=final_refs)
            return final_refs

        return [str(path) for path in (runtime_files or []) + (exported_files or []) if str(path or "").strip()]

    runtime_dir = Path(runtime_artifact_dir)
    runtime_dir.mkdir(parents=True, exist_ok=True)

    output_root, target_root, exported_refs = _semir_output_roots(runtime_dir, exported_files, run_params)

    cloud = _parse_semir_cloud_path(run_params.get("cloud_path"))
    relative_path = cloud.get("relative_path") or ""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    package_base = _safe_local_name(
        run_params.get("package_name") or f"森马云盘图片包_{timestamp}",
        f"森马云盘图片包_{timestamp}",
    )
    package_root = _ensure_unique_local_dir(runtime_dir / package_base)
    duplicate_mode = str(run_params.get("duplicate_mode") or "first_per_stem").strip().lower()
    package_layout = _normalize_semir_package_layout(run_params.get("package_layout"))
    preserve_path = duplicate_mode == "all"

    successful_rows = []
    for row in data_rows or []:
        if not isinstance(row, dict):
            continue
        local_path = Path(str(row.get("本地文件") or "")).expanduser()
        if str(row.get("下载结果") or "").strip() != "已下载" or not local_path.is_file():
            continue
        successful_rows.append((row, local_path))

    zip_path = None
    if successful_rows:
        for row, local_path in successful_rows:
            input_code = _safe_local_name(row.get("输入编码") or "未分类", "未分类")
            folder_name, clean_filename = _semir_output_folder_and_filename(
                row.get("云盘路径") or "",
                relative_path,
                row.get("文件名") or local_path.name,
                row.get("__package_filename") or "",
            )
            target = _build_semir_package_target(
                package_root,
                input_code,
                folder_name,
                clean_filename,
                package_layout=package_layout,
                preserve_path=preserve_path,
            )
            _copy_file_to_unique_target(local_path, target)

        zip_output_root = target_root or output_root
        zip_output_root.mkdir(parents=True, exist_ok=True)
        zip_path = _ensure_unique_local_path(zip_output_root / f"{package_root.name}.zip")
        with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            for file_path in package_root.rglob("*"):
                if not file_path.is_file():
                    continue
                archive.write(file_path, arcname=str(file_path.relative_to(package_root.parent)))
        log(f"Semir package created: {zip_path}")

    runtime_refs = [*exported_refs]
    if zip_path:
        runtime_refs.insert(0, str(zip_path))
    final_refs = runtime_refs

    if target_root:
        external_refs = []

        if successful_rows and package_root.exists():
            external_dir = _ensure_unique_local_dir(target_root / package_root.name)
            shutil.copytree(package_root, external_dir)
            if zip_path and zip_path.exists() and _is_within_directory(zip_path, target_root):
                external_refs.append(str(zip_path))
            elif zip_path and zip_path.exists():
                copied_zip = _copy_file_to_unique_target(zip_path, target_root / zip_path.name)
                external_refs.append(str(copied_zip))
            log(f"Semir package copied to export folder: {external_dir}")

        for file_path in exported_files or []:
            source = Path(str(file_path or "")).expanduser()
            if not source.is_file():
                continue
            copied = _copy_file_to_unique_target(source, target_root / source.name)
            external_refs.append(str(copied))

        if external_refs:
            final_refs = external_refs

    _cleanup_semir_runtime_artifacts(runtime_files, package_root)
    _cleanup_runtime_artifact_dir(str(runtime_dir), preserve_paths=final_refs)

    return final_refs


def _mop_package_relative_parts(value: str, fallback_name: str = "file") -> list[str]:
    parts = [
        _safe_local_name(part, "")
        for part in str(value or "").replace("\\", "/").split("/")
        if str(part or "").strip() and str(part or "").strip() not in {".", ".."}
    ]
    if parts:
        return parts
    return [_safe_local_name(fallback_name, "file")]


def _infer_mop_direct_package_root(row: dict, local_path: Path, runtime_dir: Path) -> Optional[Path]:
    try:
        if _is_within_directory(local_path, runtime_dir):
            return None
    except Exception:
        return None

    rel_parts = _mop_package_relative_parts(
        row.get("__mop_package_path") or row.get("本地目录内路径") or row.get("文件名"),
        row.get("文件名") or local_path.name,
    )
    actual_parts = list(local_path.parts)
    if len(actual_parts) < len(rel_parts):
        return None
    if actual_parts[-len(rel_parts):] != rel_parts:
        return None
    root = local_path
    for _part in rel_parts:
        root = root.parent
    return root if root.exists() and root.is_dir() else None


def _finalize_mop_cloud_folder_download_outputs(
    data_rows: list,
    runtime_files: list,
    exported_files: list,
    run_params: dict,
    runtime_artifact_dir: str,
    log,
) -> list[str]:
    runtime_dir = Path(runtime_artifact_dir)
    runtime_dir.mkdir(parents=True, exist_ok=True)
    output_root, target_root, exported_refs = _semir_output_roots(runtime_dir, exported_files, run_params)
    final_root = target_root or output_root
    final_root.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    package_base = _safe_local_name(
        run_params.get("package_name") or f"MOP云盘模拍图包_{timestamp}",
        f"MOP云盘模拍图包_{timestamp}",
    )

    successful_rows = []
    for row in data_rows or []:
        if not isinstance(row, dict):
            continue
        if str(row.get("__sheet_name") or "").strip() != "下载明细":
            continue
        local_path = Path(str(row.get("__runtime_local_path") or row.get("本地文件") or "")).expanduser()
        if str(row.get("下载结果") or "").strip() != "已下载" or not local_path.is_file():
            continue
        successful_rows.append((row, local_path))

    final_refs: list[str] = []
    package_root: Optional[Path] = None
    if successful_rows:
        direct_root_by_key: dict[str, Path] = {}
        all_rows_are_direct = True
        for row, local_path in successful_rows:
            direct_root = _infer_mop_direct_package_root(row, local_path, runtime_dir)
            if direct_root is None:
                all_rows_are_direct = False
                break
            direct_root_by_key[str(direct_root.resolve(strict=False))] = direct_root

        if all_rows_are_direct and len(direct_root_by_key) == 1:
            package_root = next(iter(direct_root_by_key.values()))
            for row, local_path in successful_rows:
                row["本地文件"] = str(local_path)
            final_refs.append(str(package_root))
            log(f"MOP cloud folder direct output ready: {package_root} ({len(successful_rows)} files)")
        else:
            package_started_at = time.monotonic()
            package_root = _ensure_unique_local_dir(final_root / package_base)
            for row, local_path in successful_rows:
                rel_parts = _mop_package_relative_parts(
                    row.get("__mop_package_path") or row.get("本地目录内路径") or row.get("文件名"),
                    row.get("文件名") or local_path.name,
                )
                target = package_root.joinpath(*rel_parts)
                relocated = _relocate_runtime_file_to_unique_target(local_path, target, runtime_dir)
                row["本地文件"] = str(relocated)

            final_refs.append(str(package_root))
            log(
                "MOP cloud folder package prepared: "
                f"{package_root} ({len(successful_rows)} files, {time.monotonic() - package_started_at:.1f}s)"
            )

        if _truthy_param(run_params.get("create_zip")):
            zip_path = _ensure_unique_local_path(final_root / f"{package_root.name}.zip")
            with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
                for file_path in sorted(package_root.rglob("*")):
                    if not file_path.is_file():
                        continue
                    archive.write(file_path, arcname=str(file_path.relative_to(package_root.parent)))
            final_refs.insert(0, str(zip_path))
            log(f"MOP cloud folder ZIP created: {zip_path}")

    if target_root:
        copied_exports = []
        for file_path in exported_files or []:
            source = Path(str(file_path or "")).expanduser()
            if not source.is_file():
                continue
            copied = _copy_file_to_unique_target(source, target_root / source.name)
            copied_exports.append(str(copied))
        final_refs.extend(copied_exports)
    else:
        final_refs.extend(exported_refs)

    runtime_files_to_cleanup = []
    for file_path in runtime_files or []:
        source = Path(str(file_path or "")).expanduser()
        try:
            if _is_within_directory(source, runtime_dir):
                runtime_files_to_cleanup.append(str(source))
        except Exception:
            logger.debug("Failed to classify MOP runtime file %s", source, exc_info=True)
    _cleanup_semir_runtime_artifacts(runtime_files_to_cleanup, None)
    _cleanup_runtime_artifact_dir(str(runtime_dir), preserve_paths=final_refs)

    return final_refs


def _nested_config_value(source: dict, path: str, default: str = "") -> str:
    current = source or {}
    for part in path.split("."):
        if not isinstance(current, dict):
            return default
        current = current.get(part)
    return str(current or default).strip()


def _resolve_one_xm_settings() -> dict:
    cfg = load_config()
    base_url = (
        _nested_config_value(cfg, "ai.1xm.base_url")
        or os.environ.get("ONE_XM_BASE_URL", "")
        or ONE_XM_DEFAULT_BASE_URL
    )
    return {
        "base_url": base_url,
        "2k": (
            _nested_config_value(cfg, "ai.1xm.gpt_image_2k_key")
            or os.environ.get("ONE_XM_GPT_IMAGE_2K_KEY", "").strip()
            or os.environ.get("ONE_XM_API_KEY", "").strip()
        ),
        "4k": (
            _nested_config_value(cfg, "ai.1xm.gpt_image_4k_key")
            or os.environ.get("ONE_XM_GPT_IMAGE_4K_KEY", "").strip()
            or os.environ.get("ONE_XM_API_KEY", "").strip()
        ),
        "ai.1xm.gpt_image_2k_key": (
            _nested_config_value(cfg, "ai.1xm.gpt_image_2k_key")
            or os.environ.get("ONE_XM_GPT_IMAGE_2K_KEY", "").strip()
            or os.environ.get("ONE_XM_API_KEY", "").strip()
        ),
        "ai.1xm.gpt_image_4k_key": (
            _nested_config_value(cfg, "ai.1xm.gpt_image_4k_key")
            or os.environ.get("ONE_XM_GPT_IMAGE_4K_KEY", "").strip()
            or os.environ.get("ONE_XM_API_KEY", "").strip()
        ),
        "ai.1xm.gemini_3_1_flash_image_preview_key": (
            _nested_config_value(cfg, "ai.1xm.gemini_3_1_flash_image_preview_key")
            or os.environ.get("ONE_XM_GEMINI_3_1_FLASH_IMAGE_PREVIEW_KEY", "").strip()
        ),
        "ai.1xm.gemini_3_pro_image_preview_key": (
            _nested_config_value(cfg, "ai.1xm.gemini_3_pro_image_preview_key")
            or os.environ.get("ONE_XM_GEMINI_3_PRO_IMAGE_PREVIEW_KEY", "").strip()
        ),
    }


def _normalize_one_xm_key_tier(row: dict, run_params: dict) -> str:
    explicit = str(
        row.get("__1xm_key_tier")
        or run_params.get("one_xm_key_tier")
        or "auto"
    ).strip().lower()
    if explicit in {"2k", "4k"}:
        return explicit
    payload = row.get("__1xm_payload") if isinstance(row.get("__1xm_payload"), dict) else {}
    size = str(payload.get("size") or row.get("尺寸") or "").strip()
    match = re.match(r"^(\d+)x(\d+)$", size, flags=re.IGNORECASE)
    if match and max(int(match.group(1)), int(match.group(2))) > 2048:
        return "4k"
    return "2k"


def _parse_internal_list(value) -> list[str]:
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    if not value:
        return []
    return [item.strip() for item in re.split(r"[\n\r,，、；;]+", str(value)) if item.strip()]


def _prepare_one_xm_payload(row: dict) -> dict:
    raw_payload = row.get("__1xm_payload") or {}
    if isinstance(raw_payload, str):
        try:
            raw_payload = json.loads(raw_payload)
        except Exception:
            raw_payload = {}
    if not isinstance(raw_payload, dict):
        raw_payload = {}

    quality = str(raw_payload.get("quality") or row.get("质量") or "auto").strip().lower()
    if quality not in {"auto", "standard", "low", "medium", "high"}:
        quality = "auto"
    try:
        requested_count = int(raw_payload.get("n") or row.get("生成数量") or 1)
    except Exception:
        requested_count = 1
    payload = {
        "model": str(raw_payload.get("model") or row.get("模型") or "gpt-image-2").strip() or "gpt-image-2",
        "prompt": str(raw_payload.get("prompt") or row.get("最终提示词") or "").strip(),
        "size": str(raw_payload.get("size") or row.get("尺寸") or "1024x1024").strip(),
        "quality": quality,
        "output_format": str(raw_payload.get("output_format") or row.get("格式") or "png").strip(),
        "n": max(1, min(8, requested_count)),
    }
    ratio = str(raw_payload.get("ratio") or row.get("比例") or row.get("ratio") or "").strip()
    if ratio:
        payload["ratio"] = ratio
    for optional_key in ("webhook_url", "webhook_secret", "mask"):
        if raw_payload.get(optional_key):
            payload[optional_key] = raw_payload.get(optional_key)

    reference_paths = _parse_internal_list(row.get("__1xm_reference_paths")) or _parse_internal_list(row.get("参考图文件"))
    images = []
    image_errors = []
    for raw_path in reference_paths[:10]:
        try:
            images.append(file_to_data_url(raw_path))
        except Exception as exc:
            image_errors.append(f"{Path(raw_path).name}: {exc}")
    if images:
        payload["image"] = images
    if image_errors:
        row["备注"] = "；".join([str(row.get("备注") or "").strip(), f"参考图读取失败：{'；'.join(image_errors)}"]).strip("；")

    return payload


def _run_one_xm_generation_row(row: dict, run_params: dict, one_xm_settings: dict) -> dict:
    patched = dict(row)
    tier = _normalize_one_xm_key_tier(patched, run_params)
    payload = _prepare_one_xm_payload(patched)
    try:
        selected_key_id, api_key = ai_image_service.select_model_key({
            "model_key": payload.get("model"),
            "size": payload.get("size"),
            "model_key_tier": tier,
        }, one_xm_settings)
    except ai_image_service.MissingModelKeyError as exc:
        patched["执行结果"] = "配置缺失"
        patched["备注"] = str(exc)
        return patched

    if not payload.get("prompt"):
        patched["执行结果"] = "参数缺失"
        patched["备注"] = "缺少最终提示词"
        return patched

    client = OneXMImageClient(api_key, base_url=one_xm_settings.get("base_url") or ONE_XM_DEFAULT_BASE_URL)
    result = run_image_task_until_done(
        client,
        payload,
        idempotency_key=str(patched.get("__1xm_idempotency_key") or "").strip() or f"tmall_ai_{secrets.token_hex(8)}",
        task_attempts=max(1, int(run_params.get("compensate_attempts") or 2)),
        request_retries=max(1, int(run_params.get("retry_attempts") or 3)),
        request_timeout_seconds=max(30, int(float(run_params.get("request_timeout_seconds") or 120))),
        retry_delay_seconds=1.5,
        poll_timeout_seconds=max(60, int(float(run_params.get("poll_timeout_minutes") or 10) * 60)),
    )

    if result.get("ok"):
        urls = result.get("image_urls") or []
        patched["1XM任务ID"] = result.get("task_id") or ""
        patched["1XM轮询URL"] = result.get("poll_url") or ""
        patched["生成图URL"] = "\n".join(urls)
        patched["生成图数量"] = len(urls)
        patched["执行结果"] = "已生成"
        patched["__1xm_key_tier"] = selected_key_id
        retry_note = []
        if int(result.get("compensation_attempts") or 0) > 0:
            retry_note.append(f"补偿 {result.get('compensation_attempts')} 次")
        if int(result.get("poll_attempts") or 0) > 0:
            retry_note.append(f"轮询 {result.get('poll_attempts')} 次")
        patched["备注"] = "；".join(retry_note)
        return patched

    patched["执行结果"] = "生成失败"
    patched["备注"] = str(result.get("error") or "1XM 任务失败")
    return patched


async def _apply_tmall_ai_image_generation(data_rows: list, run_params: dict, wait_for_control, log) -> list:
    if str(run_params.get("execute_mode") or "plan").strip().lower() != "generate":
        return data_rows

    planned_indexes = [
        index
        for index, row in enumerate(data_rows or [])
        if isinstance(row, dict) and row.get("__1xm_generate")
    ]
    if not planned_indexes:
        return data_rows

    max_jobs = max(1, int(run_params.get("max_generate_jobs") or 1))
    generation_concurrency = max(1, min(100, int(run_params.get("generation_concurrency") or 100)))
    settings = _resolve_one_xm_settings()
    patched_rows = [dict(row) if isinstance(row, dict) else row for row in (data_rows or [])]
    generation_total = min(len(planned_indexes), max_jobs)
    log(f"[1xm] 准备并发执行 {generation_total}/{len(planned_indexes)} 条 GPT-Image-2 异步生图任务，并发上限 {generation_concurrency}")

    selected_indexes = planned_indexes[:generation_total]
    for row_index in planned_indexes[generation_total:]:
        row = patched_rows[row_index]
        row["执行结果"] = "已跳过"
        row["备注"] = f"超过本次最多真实生成 {max_jobs} 条的费用保护阈值"

    semaphore = asyncio.Semaphore(generation_concurrency)

    async def generate_one(row_index: int, ordinal: int) -> tuple[int, dict]:
        row = patched_rows[row_index]
        async with semaphore:
            log(f"[1xm] 提交生图 {ordinal}/{generation_total}: 款号 {row.get('款号') or ''} / {row.get('提示词字段名') or ''}")
            try:
                patched = await asyncio.to_thread(_run_one_xm_generation_row, row, run_params, settings)
            except Exception as exc:
                patched = dict(row)
                patched["执行结果"] = "生成失败"
                patched["备注"] = str(exc)
                log(f"[1xm][warn] 生图失败：{exc}")
            return row_index, patched

    tasks = [
        asyncio.create_task(generate_one(row_index, ordinal))
        for ordinal, row_index in enumerate(selected_indexes, start=1)
    ]
    for completed, task in enumerate(asyncio.as_completed(tasks), start=1):
        row_index, patched = await task
        patched_rows[row_index] = patched
        await wait_for_control({
            "records": completed,
            "shared": {
                "total_rows": generation_total,
                "current_exec_no": completed,
                "current_row_no": int(patched.get("表格行号") or 0),
                "current_buyer_id": str(patched.get("款号") or ""),
                "current_store": "1XM GPT-Image-2 生图",
                "generation_total_jobs": generation_total,
                "generation_completed_jobs": completed,
            },
            "phase": "one_xm_generate",
        })

    await wait_for_control({
        "records": generation_total,
        "shared": {
            "total_rows": generation_total,
            "current_exec_no": generation_total,
            "current_store": "1XM GPT-Image-2 生图完成",
            "generation_total_jobs": generation_total,
            "generation_completed_jobs": generation_total,
        },
        "phase": "one_xm_done",
    })
    return patched_rows


def _task_file_param_path(run_params: dict, key: str, default: str = "") -> str:
    value = (run_params or {}).get(key)
    if isinstance(value, dict):
        for field in ("path", "file_path", "local_path", "absolute_path"):
            candidate = str(value.get(field) or "").strip()
            if candidate:
                return candidate
        return default
    candidate = str(value or "").strip()
    return candidate or default


def _normalize_prompt_source(value: object) -> str:
    text = str(value or "").strip().lower()
    return "cloud_prompt_library" if text == "cloud_prompt_library" else "local_excel"


def _task_int_param(run_params: dict, key: str, default: int, *, min_value: int | None = None, max_value: int | None = None) -> int:
    try:
        value = int(float(str((run_params or {}).get(key, default)).strip()))
    except Exception:
        value = default
    if min_value is not None:
        value = max(min_value, value)
    if max_value is not None:
        value = min(max_value, value)
    return value


def _task_float_param(run_params: dict, key: str, default: float, *, min_value: float | None = None, max_value: float | None = None) -> float:
    try:
        value = float(str((run_params or {}).get(key, default)).strip())
    except Exception:
        value = default
    if min_value is not None:
        value = max(min_value, value)
    if max_value is not None:
        value = min(max_value, value)
    return value


def _task_bool_param(run_params: dict, key: str, default: bool = False) -> bool:
    value = (run_params or {}).get(key, default)
    if isinstance(value, bool):
        return value
    text = str(value or "").strip().lower()
    if not text:
        return default
    return text in {"1", "true", "yes", "y", "on", "是", "启用", "开启"}


def _tmall_ai_model_params(run_params: dict) -> tuple[str, str, str]:
    model_id = str((run_params or {}).get("model_id") or "").strip()
    model = str((run_params or {}).get("model") or "").strip()
    tier = str((run_params or {}).get("one_xm_key_tier") or "").strip().lower()

    model_map = {
        "gpt-image-2k": ("gpt-image-2", "2k"),
        "gpt-image-4k": ("gpt-image-2", "4k"),
        "gemini-3.1-flash-image-preview": ("gemini-3.1-flash-image-preview", "auto"),
        "gemini-3-pro-image-preview": ("gemini-3-pro-image-preview", "auto"),
    }
    if model_id in model_map:
        mapped_model, mapped_tier = model_map[model_id]
        return model_id, mapped_model, mapped_tier

    if model in {"gemini-3.1-flash-image-preview", "gemini-3-pro-image-preview"}:
        return model, model, "auto"
    if model == "gpt-image-2" and tier in {"2k", "4k"}:
        return f"gpt-image-{tier}", model, tier
    return "gpt-image-4k", "gpt-image-2", "4k"


def _extract_first_approval_board_url(data_rows: list) -> str:
    return _extract_tmall_approval_board_urls(data_rows).get("approval_board_url", "")


def _extract_tmall_approval_board_urls(data_rows: list) -> dict:
    urls = {"approval_board_url": "", "cloud_board_url": "", "local_board_url": ""}
    for row in data_rows or []:
        if not isinstance(row, dict):
            continue
        cloud_url = str(row.get("云端审批看板") or "").strip()
        local_url = str(row.get("本地审批看板") or "").strip()
        board_url = str(row.get("审批看板") or "").strip()
        if cloud_url.startswith(("http://", "https://")) and not urls["cloud_board_url"]:
            urls["cloud_board_url"] = cloud_url
        if local_url.startswith(("http://", "https://")) and not urls["local_board_url"]:
            urls["local_board_url"] = local_url
        if board_url.startswith(("http://", "https://")):
            if "/tmall-ai-image-approval/" in board_url and not urls["local_board_url"]:
                urls["local_board_url"] = board_url
            elif "batch_uid=" in board_url and not urls["cloud_board_url"]:
                urls["cloud_board_url"] = board_url
            if not urls["approval_board_url"]:
                urls["approval_board_url"] = board_url
    urls["approval_board_url"] = urls["cloud_board_url"] or urls["approval_board_url"] or urls["local_board_url"]
    return urls


def _extract_bala_review_board_url(data_rows: list) -> str:
    for row in data_rows or []:
        if not isinstance(row, dict):
            continue
        board_url = str(row.get("审批看板") or row.get("bala_review_board_url") or "").strip()
        if board_url.startswith(("http://", "https://")) and "/bala-ai-video-review/" in board_url:
            return board_url
    return ""


def _is_tmall_ai_generation_confirmation_result(adapter_id: str, task_id: str, data_rows: list) -> bool:
    if (adapter_id, task_id) != ("tmall-ops-assistant", "tmall_ai_image_test_chain"):
        return False
    rows = [row for row in (data_rows or []) if isinstance(row, dict)]
    if not rows:
        return False
    return any(
        str(row.get("阶段") or "").strip() == "确认提交生图"
        or str(row.get("审批状态") or "").strip() == "pending_generation_confirmation"
        for row in rows
    )


def _is_table_output_file_ref(value: object) -> bool:
    text = str(value or "").strip().split("?", 1)[0].lower()
    return text.endswith((".xlsx", ".xlsm", ".xls", ".csv"))


def _parse_tmall_approval_board_url(url: str) -> dict:
    try:
        parsed = urlparse(str(url or "").strip())
        parts = [part for part in parsed.path.split("/") if part]
        batch_id = parts[-1] if parts else ""
        token = (parse_qs(parsed.query).get("token") or [""])[0]
        if "tmall-ai-image-approval" not in parsed.path:
            return {}
        return {"approval_batch_id": batch_id, "approval_token": token}
    except Exception:
        return {}


def _task_instance_artifact_kind(path: str) -> str:
    suffix = Path(str(path or "").split("?", 1)[0]).suffix.lower()
    if suffix in {".xlsx", ".xls", ".xlsm", ".csv"}:
        return "excel"
    if suffix in {".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp", ".tif", ".tiff"}:
        return "image"
    if not suffix:
        return "directory"
    return "file"


def _sync_task_instance_artifacts(instance_uid: str, output_files: list, run_id: int) -> None:
    if not instance_uid:
        return
    for path in output_files or []:
        text = str(path or "").strip()
        if not text:
            continue
        try:
            label = Path(text.split("?", 1)[0]).name or text
            data_sink.add_task_instance_artifact(
                instance_uid,
                kind=_task_instance_artifact_kind(text),
                label=label,
                path=text,
                meta={"run_id": run_id},
            )
        except Exception:
            logger.debug("Failed to sync task instance artifact: %s", text, exc_info=True)


def _load_tmall_ai_image_chain_module():
    import importlib.util
    import sys

    script_path = Path(__file__).resolve().parents[1] / "adapters" / "tmall-ops-assistant" / "tools" / "run_tmall_ai_image_test_chain.py"
    spec = importlib.util.spec_from_file_location("crawshrimp_tmall_ai_image_test_chain", script_path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"无法加载天猫 AI 测图全链路脚本：{script_path}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def _tmall_ai_image_prompt_library_params(run_params: dict) -> tuple[str, str]:
    cloud_prompt_library_id = str((run_params or {}).get("cloud_prompt_library_id") or "").strip()
    if not cloud_prompt_library_id:
        raise RuntimeError("缺少云端 Prompt 库")
    embedded_templates_json = str((run_params or {}).get("cloud_prompt_templates_json") or "").strip()
    if embedded_templates_json:
        return cloud_prompt_library_id, embedded_templates_json
    source_type = str((run_params or {}).get("cloud_prompt_library_source") or "").strip().lower()
    if source_type == "local" or cloud_prompt_library_id.startswith("local:"):
        raise RuntimeError("本地 Prompt 库未返回可用模板，请重新选择")
    if cloud_prompt_library_id.startswith("cloud:"):
        cloud_prompt_library_id = cloud_prompt_library_id.removeprefix("cloud:")
    templates_payload = _cloud_prompt_library_export(cloud_prompt_library_id)
    return cloud_prompt_library_id, json.dumps(templates_payload, ensure_ascii=False)


async def _apply_tmall_ai_image_test_chain(run_params: dict, runtime_artifact_dir: str, wait_for_control, log) -> list:
    import argparse

    module = _load_tmall_ai_image_chain_module()
    execute_mode = str((run_params or {}).get("execute_mode") or "approval_then_create").strip().lower()
    execution = module.normalize_chain_execution_mode(execute_mode)
    prompt_source = _normalize_prompt_source((run_params or {}).get("prompt_source") or "local_excel")
    model_id, model, one_xm_key_tier = _tmall_ai_model_params(run_params or {})

    workflow_file = _task_file_param_path(run_params, "workflow_file", getattr(module, "DEFAULT_WORKFLOW_FILE", ""))
    prompt_file = _task_file_param_path(run_params, "prompt_file", getattr(module, "DEFAULT_PROMPT_FILE", ""))
    cloud_prompt_library_id = str((run_params or {}).get("cloud_prompt_library_id") or "").strip()
    cloud_prompt_templates_json = ""
    if not workflow_file:
        raise RuntimeError("缺少测图任务导入模板文件")
    if prompt_source == "local_excel" and not prompt_file:
        raise RuntimeError("缺少 AI 测图提示词库文件")
    if prompt_source == "cloud_prompt_library":
        cloud_prompt_library_id, cloud_prompt_templates_json = _tmall_ai_image_prompt_library_params(run_params or {})

    args = argparse.Namespace(
        execute_mode=execution["mode"],
        workflow_file=workflow_file,
        prompt_source=prompt_source,
        prompt_file=prompt_file,
        cloud_prompt_library_id=cloud_prompt_library_id,
        cloud_prompt_templates_json=cloud_prompt_templates_json,
        limit=0,
        cloud_path=str((run_params or {}).get("cloud_path") or getattr(module, "DEFAULT_CLOUD_PATH", "")).strip(),
        cdp_url=str((run_params or {}).get("cdp_url") or "http://127.0.0.1:9222").strip(),
        confirm_generation=bool(execution["confirm_generation"]),
        generate=bool(execution["generate"]),
        approval_required=bool(execution["approval_required"]),
        live_upload=bool(execution["live_upload"]),
        live_create=bool(execution["live_create"]),
        live_online=bool(execution["live_online"]),
        model_id=model_id,
        model=model,
        ratio=str((run_params or {}).get("ratio") or "").strip(),
        image_size=str((run_params or {}).get("image_size") or "1536x2048").strip(),
        ai_image_count=_task_int_param(run_params, "ai_image_count", 4, min_value=1, max_value=20),
        generation_concurrency=_task_int_param(run_params, "generation_concurrency", 100, min_value=1, max_value=100),
        reference_mode=str((run_params or {}).get("reference_mode") or "main_only").strip().lower(),
        allow_global_semir_fallback=_task_bool_param(run_params, "allow_global_semir_fallback", False),
        quality=str((run_params or {}).get("quality") or "auto").strip().lower(),
        output_format=str((run_params or {}).get("output_format") or "jpeg").strip().lower(),
        one_xm_key_tier=one_xm_key_tier,
        retry_attempts=_task_int_param(run_params, "retry_attempts", 3, min_value=1, max_value=8),
        compensate_attempts=_task_int_param(run_params, "compensate_attempts", 2, min_value=1, max_value=6),
        poll_timeout_minutes=_task_float_param(run_params, "poll_timeout_minutes", 12.0, min_value=1.0, max_value=60.0),
        tmall_upload_delay_seconds=_task_float_param(run_params, "tmall_upload_delay_seconds", 1.5, min_value=0.0, max_value=60.0),
        tmall_create_delay_seconds=_task_float_param(run_params, "tmall_create_delay_seconds", 5.0, min_value=0.0, max_value=120.0),
        tmall_batch_delay_seconds=_task_float_param(run_params, "tmall_batch_delay_seconds", 12.0, min_value=0.0, max_value=180.0),
        tmall_readback_delay_seconds=_task_float_param(run_params, "tmall_readback_delay_seconds", 5.0, min_value=0.0, max_value=120.0),
        approval_message_template=str((run_params or {}).get("approval_message_template") or "").strip(),
        approval_notify_channel=str((run_params or {}).get("approval_notify_channel") or "dingtalk").strip().lower(),
        approval_base_url=str((run_params or {}).get("approval_base_url") or f"http://127.0.0.1:{os.environ.get('CRAWSHRIMP_PORT', '18765')}").strip(),
        task_instance_uid=str((run_params or {}).get("__task_instance_uid") or (run_params or {}).get("task_instance_uid") or "").strip(),
        task_run_uid=str((run_params or {}).get("task_run_uid") or (run_params or {}).get("__task_run_id") or (run_params or {}).get("run_id") or (run_params or {}).get("__task_instance_uid") or "").strip(),
    )
    if args.quality not in {"auto", "low", "medium", "high"}:
        args.quality = "auto"
    if args.one_xm_key_tier not in {"auto", "2k", "4k"}:
        args.one_xm_key_tier = "auto"
    if args.reference_mode not in {"main_only", "main_and_detail"}:
        args.reference_mode = "main_only"

    log(
        "[tmall-ai-chain] 后端编排执行："
        f"mode={args.execute_mode}, confirm_generation={args.confirm_generation}, generate={args.generate}, "
        f"approval={args.approval_required}, upload={args.live_upload}, create={args.live_create}, online={args.live_online}, "
        f"ai_image_count={args.ai_image_count}, quality={args.quality}"
    )
    return await module.run_chain_rows(
        args,
        Path(runtime_artifact_dir),
        log=log,
        wait_for_control=wait_for_control,
    )


def _tmall_local_export_run_folder_name(run_params: dict) -> str:
    started_raw = str((run_params or {}).get("__task_started_at") or "").strip()
    try:
        started = datetime.fromisoformat(started_raw)
    except Exception:
        started = datetime.now()
    run_id = str((run_params or {}).get("__task_run_id") or (run_params or {}).get("task_run_uid") or "").strip()
    suffix = f"_run{run_id}" if run_id else ""
    return _safe_local_name(f"运行_{started.strftime('%Y%m%d-%H%M%S')}{suffix}", "运行结果")


def _tmall_local_export_root(run_params: dict, task_folder: str, *, run_scoped: bool = False) -> Path:
    output_dir = str((run_params or {}).get("output_dir") or "").strip()
    target_root = _expand_user_configured_local_path(output_dir) if output_dir else (
        Path.home() / "Downloads" / "抓虾导出" / "天猫运营助手" / _safe_local_name(task_folder, "巴拉-AI测图")
    )
    if run_scoped:
        target_root = _ensure_unique_local_dir(target_root / _tmall_local_export_run_folder_name(run_params))
    target_root.mkdir(parents=True, exist_ok=True)
    return target_root


def _split_tmall_local_path_values(value) -> list[str]:
    if isinstance(value, (list, tuple, set)):
        values = []
        for item in value:
            values.extend(_split_tmall_local_path_values(item))
        return values
    text = str(value or "").strip()
    if not text:
        return []
    return [item.strip() for item in re.split(r"[\n\r|；;]+", text) if item.strip()]


def _tmall_local_item_folder(row: dict) -> str:
    style_code = _safe_local_name(
        row.get("款号") or row.get("style_code") or row.get("款式编码") or "未填款号",
        "未填款号",
    )
    item_id = _safe_local_name(
        row.get("商品ID") or row.get("天猫商品ID") or row.get("ID（用于测图的ID）") or "未填商品ID",
        "未填商品ID",
    )
    return f"{style_code}_{item_id}"


def _tmall_row_has_semir_material_file(row: dict) -> bool:
    stage = str(row.get("阶段") or "").strip()
    result = str(row.get("执行结果") or "").strip()
    return any(token in stage for token in ("森马", "找图", "参考图")) or any(
        token in result for token in ("参考图", "已下载")
    )


def _tmall_row_has_ai_generated_file(row: dict) -> bool:
    stage = str(row.get("阶段") or "").strip()
    result = str(row.get("执行结果") or "").strip()
    return any(token in stage for token in ("1XM", "AI", "生图")) or "已生成" in result


def _copy_tmall_data_tables_to_local_export(exported_files: list, target_root: Path) -> list[str]:
    copied_refs = []
    table_root = target_root / "数据表格"
    for file_path in exported_files or []:
        source = Path(str(file_path or "")).expanduser()
        if not source.is_file():
            continue
        if _is_within_directory(source, table_root):
            copied_refs.append(str(source))
            continue
        copied = _copy_file_to_unique_target(source, table_root / source.name)
        copied_refs.append(str(copied))
    return copied_refs


def _finalize_plm_size_chart_downloader_outputs(
    runtime_files: list,
    exported_files: list,
    run_params: dict,
    log,
) -> list[str]:
    fallback_refs = [str(path) for path in [*(runtime_files or []), *(exported_files or [])] if str(path or "").strip()]
    output_dir = str((run_params or {}).get("output_dir") or "").strip()
    if not output_dir:
        return fallback_refs

    target_root = _expand_user_configured_local_path(output_dir)
    target_root.mkdir(parents=True, exist_ok=True)

    copied_refs = []
    source_refs = []
    seen_sources = set()
    for file_path in [*(exported_files or []), *(runtime_files or [])]:
        source_key = str(file_path or "").strip()
        if not source_key or source_key in seen_sources:
            continue
        seen_sources.add(source_key)
        source_refs.append(source_key)

    for file_path in source_refs:
        source = Path(str(file_path or "")).expanduser()
        if not source.is_file():
            continue
        if _is_direct_child_file(source, target_root):
            copied_refs.append(str(source))
            continue
        copied_refs.append(str(_copy_file_to_unique_target(source, target_root / source.name)))

    if copied_refs and log:
        log(f"PLM 尺码表已复制到导出目录：{target_root}")
    return copied_refs or fallback_refs


def _copy_tmall_row_local_files(
    data_rows: list,
    target_root: Path,
    source_folder: str,
    fields: tuple[str, ...],
    *,
    include_row,
    seen: set[tuple[str, str, str]],
) -> list[str]:
    copied_refs = []
    for row in data_rows or []:
        if not isinstance(row, dict) or not include_row(row):
            continue
        item_folder = _tmall_local_item_folder(row)
        target_dir = target_root / source_folder / item_folder
        for field in fields:
            for raw_path in _split_tmall_local_path_values(row.get(field)):
                lowered = raw_path.lower()
                if lowered.startswith(("http://", "https://", "data:", "blob:")):
                    continue
                source = Path(raw_path).expanduser()
                if not source.is_file():
                    continue
                key = (str(source.resolve(strict=False)), source_folder, item_folder)
                if key in seen:
                    continue
                seen.add(key)
                if _is_within_directory(source, target_root):
                    copied_refs.append(str(source))
                    continue
                copied = _copy_file_to_unique_target(source, target_dir / source.name)
                copied_refs.append(str(copied))
    return copied_refs


def _copy_tmall_ai_chain_images_to_local_export(data_rows: list, target_root: Path) -> list[str]:
    seen: set[tuple[str, str, str]] = set()
    copied_refs = []
    copied_refs.extend(_copy_tmall_row_local_files(
        data_rows,
        target_root,
        "网盘素材图",
        ("主图原图文件", "主参考图文件", "细节参考图文件", "参考图文件"),
        include_row=lambda _row: True,
        seen=seen,
    ))
    copied_refs.extend(_copy_tmall_row_local_files(
        data_rows,
        target_root,
        "网盘素材图",
        ("本地文件",),
        include_row=_tmall_row_has_semir_material_file,
        seen=seen,
    ))
    copied_refs.extend(_copy_tmall_row_local_files(
        data_rows,
        target_root,
        "AI生成图",
        ("本地生成图文件",),
        include_row=lambda _row: True,
        seen=seen,
    ))
    copied_refs.extend(_copy_tmall_row_local_files(
        data_rows,
        target_root,
        "AI生成图",
        ("本地文件",),
        include_row=_tmall_row_has_ai_generated_file,
        seen=seen,
    ))
    return copied_refs


def _finalize_tmall_ai_image_test_chain_outputs(
    data_rows: list,
    runtime_files: list,
    exported_files: list,
    run_params: dict,
    runtime_artifact_dir: str,
    log,
) -> list[str]:
    target_root = _tmall_local_export_root(run_params, "巴拉-AI测图全链路", run_scoped=True)
    fallback_refs = [str(path) for path in [*(runtime_files or []), *(exported_files or [])] if str(path or "").strip()]

    table_refs = _copy_tmall_data_tables_to_local_export(exported_files, target_root)
    image_refs = _copy_tmall_ai_chain_images_to_local_export(data_rows, target_root)
    if log:
        log(f"巴拉-AI测图全链路本地导出：表格 {len(table_refs)} 个，图片 {len(image_refs)} 张，目录 {target_root}")
    return [*table_refs, *image_refs] or fallback_refs


def _finalize_tmall_material_test_data_export_outputs(
    runtime_files: list,
    exported_files: list,
    run_params: dict,
    log,
) -> list[str]:
    target_root = _tmall_local_export_root(run_params, "巴拉-AI测图数据抓取导出")
    fallback_refs = [str(path) for path in [*(runtime_files or []), *(exported_files or [])] if str(path or "").strip()]

    table_refs = _copy_tmall_data_tables_to_local_export(exported_files, target_root)
    if log:
        log(f"巴拉-AI测图数据抓取导出本地导出：表格 {len(table_refs)} 个，目录 {target_root}")
    final_refs = table_refs or fallback_refs
    if _truthy_param((run_params or {}).get("sync_to_cloud")):
        workbook_path = _first_excel_output_path(final_refs)
        if workbook_path is None:
            if log:
                log("[warn] 云端测图数据看板同步跳过：未找到可同步的 Excel 文件")
        else:
            try:
                sync_result = sync_tmall_material_test_workbook_to_cloud(workbook_path)
                if log:
                    log(
                        "云端测图数据看板同步完成："
                        f"概览 {sync_result.get('overview_rows', 0)} 条，"
                        f"明细 {sync_result.get('detail_rows', 0)} 条，"
                        f"写入/更新 {sync_result.get('inserted_or_updated', 0)} 条"
                    )
            except Exception as exc:
                if log:
                    log(f"[warn] 云端测图数据看板同步失败：{exc}")
    return final_refs


def _finalize_tmall_compete_paid_monitor_outputs(
    runtime_files: list,
    exported_files: list,
    run_params: dict,
    log,
) -> list[str]:
    target_root = _tmall_local_export_root(run_params, "竞品付费投放数据监控")
    fallback_refs = [str(path) for path in [*(runtime_files or []), *(exported_files or [])] if str(path or "").strip()]

    table_refs = _copy_tmall_data_tables_to_local_export(exported_files, target_root)
    if log:
        log(f"竞品付费投放数据监控本地导出：表格 {len(table_refs)} 个，目录 {target_root}")
    return table_refs or fallback_refs


def _first_excel_output_path(paths: list) -> Path | None:
    for item in paths or []:
        path = Path(str(item or "")).expanduser()
        if path.suffix.lower() in {".xlsx", ".xlsm", ".xls"} and path.is_file():
            return path
    return None


def sync_tmall_material_test_workbook_to_cloud(workbook_path: Path, *, client: CloudApprovalClient | None = None) -> dict:
    from core.cloud_job_executors import MATERIAL_IMPORT_DETAIL_CHUNK_SIZE, parse_tmall_material_test_workbook

    path = Path(workbook_path).expanduser().resolve()
    if not path.is_file():
        raise ValueError(f"material test workbook not found: {path}")
    cloud_client = client or _build_cloud_client()
    parsed = parse_tmall_material_test_workbook(path)
    asset_uid = _local_material_test_workbook_asset_uid(path)
    content_type = mimetypes.guess_type(path.name)[0] or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    presign = cloud_client.request_json("POST", "/api/assets/presign", {
        "batch_uid": "material-test",
        "style_id": 1,
        "asset_uid": asset_uid,
        "kind": "result",
        "filename": path.name,
        "content_type": content_type,
        "meta": {
            "job_type": "tmall_material_test_data_export",
            "sync_mode": "local_manual_export",
        },
    })
    upload_url = str(presign.get("upload_url") or "")
    if not upload_url:
        raise CloudApprovalError("cloud presign response missing upload_url for material test workbook")
    cloud_client.upload_asset(upload_url, path, content_type)
    source = {
        "source_uid": asset_uid,
        "filename": path.name,
        "object_key": str(presign.get("object_key") or ""),
        "sync_mode": "local_manual_export",
    }
    totals = {"overview_rows": 0, "detail_rows": 0, "inserted_or_updated": 0}
    detail_rows = parsed.get("detail_rows") or []
    chunks = [detail_rows[index:index + MATERIAL_IMPORT_DETAIL_CHUNK_SIZE] for index in range(0, len(detail_rows), MATERIAL_IMPORT_DETAIL_CHUNK_SIZE)] or [[]]
    for index, detail_chunk in enumerate(chunks):
        response = cloud_client.request_json("POST", "/api/material-test/import", {
            "source": source,
            "overview_rows": parsed.get("overview_rows") or [] if index == 0 else [],
            "detail_rows": detail_chunk,
        })
        totals["overview_rows"] += int(response.get("overview_rows") or 0)
        totals["detail_rows"] += int(response.get("detail_rows") or 0)
        totals["inserted_or_updated"] += int(response.get("inserted_or_updated") or 0)
    return {
        **totals,
        "workbook_asset_uid": asset_uid,
        "workbook_filename": path.name,
        "workbook_object_key": str(presign.get("object_key") or ""),
    }


def _local_material_test_workbook_asset_uid(path: Path) -> str:
    try:
        stat = path.stat()
        seed = f"{path.name}:{stat.st_size}:{stat.st_mtime_ns}"
    except OSError:
        seed = str(path)
    digest = hashlib.sha256(seed.encode("utf-8")).hexdigest()[:24]
    return f"local-material-test-{digest}"


def _tmall_packaging_style_folder(row: dict) -> str:
    return _safe_local_name(
        row.get("款号") or row.get("输入编码") or row.get("style_code") or "未填款号",
        "未填款号",
    )


def _tmall_packaging_usage_folder(row: dict) -> str:
    category = str(row.get("__category") or "").strip()
    usage = str(row.get("图片用途") or "").strip()
    filename = str(row.get("__package_filename") or row.get("文件名") or "").strip()
    haystack = f"{category} {usage} {filename}"
    if category in {"main_1x1", "micro_1x1"} or "1:1" in usage or "1比1" in filename:
        return "01_1比1主图"
    if category in {"main_3x4", "micro_3x4"} or "3:4" in usage or "3比4" in filename:
        return "02_3比4主图"
    if category == "vertical" or "商品竖图" in haystack or "竖图" in haystack:
        return "03_商品竖图"
    if category == "pc_detail" or any(token in haystack for token in ("PC详情", "商详", "详情")):
        return "04_商详页"
    return "99_未分类"


def _finalize_tmall_packaging_upload_outputs(
    data_rows: list,
    runtime_files: list,
    exported_files: list,
    run_params: dict,
    runtime_artifact_dir: str,
    log,
) -> list[str]:
    runtime_dir = Path(runtime_artifact_dir)
    runtime_dir.mkdir(parents=True, exist_ok=True)
    output_root, target_root, exported_refs = _semir_output_roots(runtime_dir, exported_files, run_params)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    package_base = _safe_local_name(
        run_params.get("package_name") or f"天猫包装图片包_{timestamp}",
        f"天猫包装图片包_{timestamp}",
    )
    package_root = _ensure_unique_local_dir(runtime_dir / package_base)

    successful_rows: list[tuple[dict, Path]] = []
    for row in data_rows or []:
        if not isinstance(row, dict):
            continue
        local_path = Path(str(row.get("本地文件") or "")).expanduser()
        if str(row.get("下载结果") or "").strip() != "已下载" or not local_path.is_file():
            continue
        successful_rows.append((row, local_path))

    zip_path = None
    if successful_rows:
        seen_assets: set[tuple[str, str, str]] = set()
        for row, local_path in successful_rows:
            style_folder = _tmall_packaging_style_folder(row)
            usage_folder = _tmall_packaging_usage_folder(row)
            clean_filename = _safe_local_name(
                row.get("__package_filename") or row.get("文件名") or local_path.name,
                local_path.name,
            )
            source_key = str(row.get("云盘路径") or row.get("原文件名") or clean_filename)
            dedupe_key = (style_folder, usage_folder, source_key)
            if dedupe_key in seen_assets:
                continue
            seen_assets.add(dedupe_key)
            target = package_root / style_folder / usage_folder / clean_filename
            _copy_file_to_unique_target(local_path, target)

        zip_output_root = target_root or output_root
        zip_output_root.mkdir(parents=True, exist_ok=True)
        zip_path = _ensure_unique_local_path(zip_output_root / f"{package_root.name}.zip")
        with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            for file_path in package_root.rglob("*"):
                if not file_path.is_file():
                    continue
                archive.write(file_path, arcname=str(file_path.relative_to(package_root.parent)))
        if log:
            log(f"天猫包装图片包已整理：{zip_path} ({len(seen_assets)} files)")

    final_refs = [*exported_refs]
    if zip_path:
        final_refs.insert(0, str(zip_path))

    if target_root:
        external_refs: list[str] = []
        if successful_rows and package_root.exists():
            external_dir = _ensure_unique_local_dir(target_root / package_root.name)
            shutil.copytree(package_root, external_dir)
            if log:
                log(f"天猫包装图片包已复制到导出目录：{external_dir}")
            if zip_path and zip_path.exists() and _is_within_directory(zip_path, target_root):
                external_refs.append(str(zip_path))
            elif zip_path and zip_path.exists():
                external_refs.append(str(_copy_file_to_unique_target(zip_path, target_root / zip_path.name)))

        for file_path in exported_files or []:
            source = Path(str(file_path or "")).expanduser()
            if not source.is_file():
                continue
            if _is_within_directory(source, target_root):
                external_refs.append(str(source))
            else:
                external_refs.append(str(_copy_file_to_unique_target(source, target_root / source.name)))

        if external_refs:
            final_refs = external_refs

    runtime_files_to_cleanup = []
    for file_path in runtime_files or []:
        source = Path(str(file_path or "")).expanduser()
        if _is_within_directory(source, runtime_dir):
            runtime_files_to_cleanup.append(str(source))
    _cleanup_semir_runtime_artifacts(runtime_files_to_cleanup, package_root)
    _cleanup_runtime_artifact_dir(str(runtime_dir), preserve_paths=final_refs)
    return final_refs


def _finalize_tmall_ops_assistant_outputs(
    task_id: str,
    data_rows: list,
    runtime_files: list,
    exported_files: list,
    run_params: dict,
    runtime_artifact_dir: str,
    log,
) -> list[str]:
    if task_id == "tmall_packaging_upload":
        return _finalize_tmall_packaging_upload_outputs(
            data_rows=data_rows,
            runtime_files=runtime_files,
            exported_files=exported_files,
            run_params=run_params,
            runtime_artifact_dir=runtime_artifact_dir,
            log=log,
        )
    if task_id == "tmall_ai_image_test_chain":
        return _finalize_tmall_ai_image_test_chain_outputs(
            data_rows=data_rows,
            runtime_files=runtime_files,
            exported_files=exported_files,
            run_params=run_params,
            runtime_artifact_dir=runtime_artifact_dir,
            log=log,
        )
    if task_id == "tmall_material_test_data_export":
        return _finalize_tmall_material_test_data_export_outputs(
            runtime_files=runtime_files,
            exported_files=exported_files,
            run_params=run_params,
            log=log,
        )
    if task_id == "tmall_compete_paid_monitor":
        return _finalize_tmall_compete_paid_monitor_outputs(
            runtime_files=runtime_files,
            exported_files=exported_files,
            run_params=run_params,
            log=log,
        )
    return [*(runtime_files or []), *(exported_files or [])]


def _finalize_tiktok_creator_video_outputs(
    data_rows: list,
    runtime_files: list,
    exported_files: list,
    run_params: dict,
    runtime_artifact_dir: str,
    log,
) -> list[str]:
    runtime_dir = Path(runtime_artifact_dir)
    runtime_dir.mkdir(parents=True, exist_ok=True)

    exported_refs = [str(path) for path in exported_files or [] if str(path or "").strip()]
    default_root = _default_output_root_for_runtime(runtime_dir, exported_files)
    output_dir = str(run_params.get("output_dir") or "").strip()
    target_root = Path(output_dir).expanduser() if output_dir else None
    if target_root:
        target_root.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    package_base = _safe_local_name(
        run_params.get("package_name") or f"达人视频下载_{timestamp}",
        f"达人视频下载_{timestamp}",
    )
    package_root = _ensure_unique_local_dir(runtime_dir / package_base)

    successful_rows = []
    downloaded_total = 0
    missing_rows = []
    for row in data_rows or []:
        if not isinstance(row, dict):
            continue
        if str(row.get("下载结果") or "").strip() == "已下载":
            downloaded_total += 1
        local_path = Path(str(row.get("本地文件") or "")).expanduser()
        if str(row.get("下载结果") or "").strip() != "已下载":
            continue
        if not local_path.is_file():
            missing_rows.append(row)
            continue
        successful_rows.append((row, local_path))

    if downloaded_total:
        log(f"TikTok creator video package includes {len(successful_rows)}/{downloaded_total} downloaded files")
    if missing_rows:
        log(f"[warn] TikTok creator video package skipped {len(missing_rows)} rows because local files were missing")

    zip_path = None
    if successful_rows:
        for row, local_path in successful_rows:
            filename = _safe_local_name(
                row.get("计划文件名") or row.get("视频ID") or local_path.name,
                local_path.name,
            )
            _copy_file_to_unique_target(local_path, package_root / filename)

        zip_output_root = target_root or default_root
        zip_output_root.mkdir(parents=True, exist_ok=True)
        zip_path = _ensure_unique_local_path(zip_output_root / f"{package_root.name}.zip")
        with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            for file_path in package_root.rglob("*"):
                if not file_path.is_file():
                    continue
                archive.write(file_path, arcname=str(file_path.relative_to(package_root.parent)))
        log(f"TikTok creator video package created: {zip_path}")

    final_refs = [*([str(zip_path)] if zip_path else []), *exported_refs]

    if target_root:
        external_refs = []
        if zip_path and zip_path.exists() and _is_within_directory(zip_path, target_root):
            external_refs.append(str(zip_path))
        elif zip_path and zip_path.exists():
            external_refs.append(str(_copy_file_to_unique_target(zip_path, target_root / zip_path.name)))
        for file_path in exported_files or []:
            source = Path(str(file_path or "")).expanduser()
            if not source.is_file():
                continue
            copied = _copy_file_to_unique_target(source, target_root / source.name)
            external_refs.append(str(copied))
        if external_refs:
            final_refs = external_refs

    _cleanup_semir_runtime_artifacts(runtime_files, package_root)
    _cleanup_runtime_artifact_dir(str(runtime_dir), preserve_paths=final_refs)

    return final_refs


def _rows_raw_to_table(rows_raw, header_row: int = 1):
    if not rows_raw:
        return {"headers": [], "rows": [], "total": 0}

    hi = header_row - 1
    if hi < 0 or hi >= len(rows_raw):
        return {"error": f"Header row index {hi} out of range", "headers": [], "rows": [], "total": 0}

    def _stringify_row(raw_row):
        return [str(c) if c is not None else '' for c in raw_row]

    def _dedupe_headers(raw_headers):
        used = {}
        headers = []
        for index, raw_header in enumerate(raw_headers):
            header = str(raw_header or '').strip() or f"列{index + 1}"
            if header not in used:
                used[header] = 1
                headers.append(header)
                continue
            used[header] += 1
            headers.append(f"{header}_{used[header]}")
        return headers

    def _build_headers():
        primary = _stringify_row(rows_raw[hi])
        if hi + 1 >= len(rows_raw):
            return _dedupe_headers(primary), hi + 1

        secondary = _stringify_row(rows_raw[hi + 1])
        width = max(len(primary), len(secondary))
        # Only collapse when row 2 is clearly filling blanks from row 1.
        # Trailing blank header cells in wide templates should not treat row 2 as another header row.
        has_grouped_children = any(
            not str(primary[index] if index < len(primary) else '').strip()
            and str(secondary[index] if index < len(secondary) else '').strip()
            for index in range(width)
        )
        if not has_grouped_children:
            return _dedupe_headers(primary), hi + 1

        merged_headers = []
        last_parent = ''
        for index in range(width):
            parent = str(primary[index] if index < len(primary) else '').strip()
            child = str(secondary[index] if index < len(secondary) else '').strip()
            if parent:
                last_parent = parent
            effective_parent = parent or (last_parent if child else '')
            if child and effective_parent and child != effective_parent:
                merged_headers.append(f"{effective_parent}/{child}")
            else:
                merged_headers.append(effective_parent or child or f"列{index + 1}")

        return _dedupe_headers(merged_headers), hi + 2

    headers, data_start_index = _build_headers()
    rows = []
    for raw in rows_raw[data_start_index:]:
        if all(c is None or str(c).strip() == '' for c in raw):
            continue
        row = {}
        for i, h in enumerate(headers):
            v = raw[i] if i < len(raw) else None
            row[h] = str(v) if v is not None else ''
        rows.append(row)

    return {"headers": headers, "rows": rows, "total": len(rows)}


def _read_local_excel(path: str, sheet: Optional[str] = None, header_row: int = 1):
    """Internal helper to read Excel/CSV without raising HTTPExceptions"""
    p = Path(path)
    if not p.exists():
        return {"error": f"File not found: {path}", "headers": [], "rows": [], "total": 0}

    def _worksheet_rows(ws):
        try:
            ws.reset_dimensions()
        except Exception:
            pass
        return list(ws.iter_rows(values_only=True))

    suffix = p.suffix.lower()
    try:
        if suffix in ('.xlsx', '.xls', '.xlsm'):
            import openpyxl

            wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
            try:
                if sheet:
                    ws = wb[sheet]
                    table = _rows_raw_to_table(_worksheet_rows(ws), header_row)
                    table["sheet_name"] = ws.title
                    table["sheets"] = {
                        ws.title: {
                            "headers": table.get("headers", []),
                            "rows": table.get("rows", []),
                            "total": table.get("total", 0),
                        }
                    }
                    return table

                workbook_tables = {}
                for ws in wb.worksheets:
                    workbook_tables[ws.title] = _rows_raw_to_table(_worksheet_rows(ws), header_row)

                active_name = wb.active.title if wb.active else (wb.sheetnames[0] if wb.sheetnames else "Sheet1")
                active_table = workbook_tables.get(active_name) or {"headers": [], "rows": [], "total": 0}
                return {
                    "headers": active_table.get("headers", []),
                    "rows": active_table.get("rows", []),
                    "total": active_table.get("total", 0),
                    "sheet_name": active_name,
                    "sheets": workbook_tables,
                }
            finally:
                wb.close()

        if suffix == '.csv':
            import csv

            with open(p, newline='', encoding='utf-8-sig') as f:
                rows_raw = list(csv.reader(f))
            table = _rows_raw_to_table(rows_raw, header_row)
            table["sheet_name"] = "Sheet1"
            table["sheets"] = {
                "Sheet1": {
                    "headers": table.get("headers", []),
                    "rows": table.get("rows", []),
                    "total": table.get("total", 0),
                }
            }
            return table

        return {"error": f"Unsupported format: {suffix}", "headers": [], "rows": [], "total": 0}
    except Exception as e:
        return {"error": str(e), "headers": [], "rows": [], "total": 0}


def _template_search_roots(adapter_id: str) -> List[Path]:
    roots: List[Path] = []

    adapter_dir = adapter_loader.get_adapter_dir(adapter_id)
    if adapter_dir:
        roots.append(adapter_dir.resolve())

    built_in_dir = (Path(__file__).parent.parent / "adapters" / adapter_id).resolve()
    if built_in_dir.exists():
        roots.append(built_in_dir)

    seen = set()
    deduped = []
    for root in roots:
        marker = str(root)
        if marker in seen:
            continue
        seen.add(marker)
        deduped.append(root)
    return deduped


def _resolve_template_path(adapter_id: str, template_file: Optional[str]) -> Optional[str]:
    if not template_file:
        return None

    for base_dir in _template_search_roots(adapter_id):
        try:
            candidate = (base_dir / template_file).resolve()
            candidate.relative_to(base_dir)
        except Exception:
            continue

        if candidate.exists() and candidate.is_file():
            return str(candidate)

    return None


def _serialize_task_param(adapter_id: str, param) -> dict:
    data = param.model_dump()
    templates = []
    for template in data.get("templates") or []:
        t = dict(template)
        t["path"] = _resolve_template_path(adapter_id, t.get("file"))
        templates.append(t)

    if not templates and data.get("template_file"):
        templates.append({
            "file": data.get("template_file"),
            "label": data.get("template_label"),
            "description": None,
            "version": None,
            "path": _resolve_template_path(adapter_id, data.get("template_file")),
        })

    data["templates"] = templates
    data["template_path"] = next((t.get("path") for t in templates if t.get("path")), None)
    return data


async def _execute_task(adapter_id: str, task_id: str, params: Optional[dict] = None, runtime_options: Optional[dict] = None, run_control: Optional[dict] = None):
    """
    Core task execution pipeline:
    1. Find / create matching Chrome tab via CDP
    2. Optional auth check + login wait
    3. Inject + execute JS script
    4. Persist results (Excel/JSON)
    5. Send notifications if configured
    6. Record run state in SQLite
    """
    from core.js_runner import JSRunner, RunAbortedError
    from core.models import OutputType

    instance_uid = str((params or {}).get("__task_instance_uid") or "").strip()
    schedule_uid = str((params or {}).get("__task_schedule_uid") or "").strip()
    jid = _run_jid(adapter_id, task_id, instance_uid)
    live_jids = _run_jids(adapter_id, task_id, instance_uid)
    # 保留历史日志，新一轮运行用分隔线追加（不覆盖）
    from datetime import datetime as _dt
    _sep = f"─── 新运行 {_dt.now().strftime('%m-%d %H:%M:%S')} ───────────────────────"
    for job_id in live_jids:
        if job_id not in _run_logs:
            _run_logs[job_id] = []
        else:
            _run_logs[job_id].append('')
            _run_logs[job_id].append(_sep)
    _set_live_status(live_jids, {
        'status': 'running',
        'run_id': None,
        'adapter_id': adapter_id,
        'task_id': task_id,
        'instance_uid': instance_uid,
        'records': 0,
        'last_seen_at': datetime.now().isoformat(),
        'current_row': 0,
        'phase': '',
    })

    def log(msg: str):
        logger.info(msg)
        _append_run_log(live_jids, msg)

    def build_progress(payload: Optional[dict] = None) -> dict:
        return _build_live_progress(payload, run_control=run_control)

    async def wait_for_control(payload: Optional[dict] = None):
        if not run_control:
            return

        payload = dict(payload or {})
        kind = str(payload.get('kind') or '').strip()

        if kind == 'before_download_urls':
            download_total = int(payload.get('download_item_total') or 0)
            run_control['download_total'] = download_total
            run_control['download_completed'] = int(payload.get('download_completed') or 0)
            run_control['download_success'] = int(payload.get('download_success') or 0)
            run_control['download_failed'] = int(payload.get('download_failed') or 0)
            run_control['download_concurrency'] = int(payload.get('download_concurrency') or 0)
            run_control['download_retry_attempts'] = int(payload.get('download_retry_attempts') or 0)
            run_control['download_started'] = download_total > 0
            run_control['download_active'] = download_total > 0
            run_control['download_last_label'] = ''
            run_control['download_current_label'] = ''
            run_control['download_active_labels'] = []
            run_control['download_active_count'] = 0
            run_control['download_speed_bps'] = 0
            run_control['download_bytes_completed'] = int(payload.get('download_bytes_completed') or 0)
            run_control['download_total_bytes'] = int(payload.get('download_total_bytes') or 0)
        elif kind == 'download_urls_progress':
            if 'download_total' in payload:
                run_control['download_total'] = int(payload.get('download_total') or 0)
            if 'download_completed' in payload:
                run_control['download_completed'] = int(payload.get('download_completed') or 0)
            if 'download_success' in payload:
                run_control['download_success'] = int(payload.get('download_success') or 0)
            if 'download_failed' in payload:
                run_control['download_failed'] = int(payload.get('download_failed') or 0)
            if 'download_active' in payload:
                run_control['download_active'] = bool(payload.get('download_active'))
            if 'download_last_label' in payload:
                run_control['download_last_label'] = str(payload.get('download_last_label') or '').strip()
            if 'download_current_label' in payload:
                run_control['download_current_label'] = str(payload.get('download_current_label') or '').strip()
            if 'download_active_labels' in payload:
                labels = payload.get('download_active_labels') or []
                run_control['download_active_labels'] = [str(item or '').strip() for item in labels if str(item or '').strip()] if isinstance(labels, list) else []
            if 'download_active_count' in payload:
                run_control['download_active_count'] = int(payload.get('download_active_count') or 0)
            if 'download_speed_bps' in payload:
                run_control['download_speed_bps'] = int(payload.get('download_speed_bps') or 0)
            if 'download_bytes_completed' in payload:
                run_control['download_bytes_completed'] = int(payload.get('download_bytes_completed') or 0)
            if 'download_total_bytes' in payload:
                run_control['download_total_bytes'] = int(payload.get('download_total_bytes') or 0)
            run_control['download_started'] = bool(run_control.get('download_total') or run_control.get('download_completed'))
        elif kind == 'before_phase' and str(payload.get('phase') or '').strip() == 'finalize_all':
            download_total = int(run_control.get('download_total') or 0)
            if download_total > 0:
                run_control['download_completed'] = download_total
            run_control['download_active'] = False
            run_control['download_current_label'] = ''
            run_control['download_active_labels'] = []
            run_control['download_active_count'] = 0
            run_control['download_speed_bps'] = 0

        records = int(payload.get('records') or 0)
        progress = build_progress(payload)
        current_row = int(progress['row_no'] or progress['current'] or 0)
        last_seen_at = datetime.now().isoformat()
        base_status = {
            'run_id': run_id,
            'adapter_id': adapter_id,
            'task_id': task_id,
            'instance_uid': instance_uid,
            'records': records,
            'last_seen_at': last_seen_at,
            'current_row': current_row,
            'current': progress['current'],
            'total': progress['total'],
            'row_no': progress['row_no'],
            'batch_no': progress['batch_no'],
            'total_batches': progress['total_batches'],
            'search_total_codes': progress['search_total_codes'],
            'search_completed_codes': progress['search_completed_codes'],
            'generation_total_jobs': progress['generation_total_jobs'],
            'generation_submitted_jobs': progress['generation_submitted_jobs'],
            'generation_completed_jobs': progress['generation_completed_jobs'],
            'buyer_id': progress['buyer_id'],
            'store': progress['store'],
            'current_source_filename': progress['current_source_filename'],
            'phase': progress['phase'],
            'completed': progress['completed'],
            'percent': progress['percent'],
            'progress_text': progress['progress_text'],
            'download_total': progress['download_total'],
            'download_completed': progress['download_completed'],
            'download_success': progress['download_success'],
            'download_failed': progress['download_failed'],
            'download_concurrency': progress['download_concurrency'],
            'download_retry_attempts': progress['download_retry_attempts'],
            'download_started': progress['download_started'],
            'download_active': progress['download_active'],
            'download_last_label': progress['download_last_label'],
            'download_current_label': progress['download_current_label'],
            'download_active_labels': progress['download_active_labels'],
            'download_active_count': progress['download_active_count'],
            'download_speed_bps': progress['download_speed_bps'],
            'download_bytes_completed': progress['download_bytes_completed'],
            'download_total_bytes': progress['download_total_bytes'],
            'list_total_rows': progress['list_total_rows'],
            'list_completed_rows': progress['list_completed_rows'],
            'list_total_batches': progress['list_total_batches'],
            'list_completed_batches': progress['list_completed_batches'],
            'detail_total_targets': progress['detail_total_targets'],
            'detail_completed_targets': progress['detail_completed_targets'],
            'detail_current_target_index': progress['detail_current_target_index'],
            'detail_current_target': progress['detail_current_target'],
            'detail_dimension_total': progress['detail_dimension_total'],
            'detail_dimension_index': progress['detail_dimension_index'],
            'detail_dimension_label': progress['detail_dimension_label'],
            'detail_current_page': progress['detail_current_page'],
            'detail_total_pages': progress['detail_total_pages'],
            'detail_total_rows': progress['detail_total_rows'],
            'detail_request_count': progress['detail_request_count'],
            'detail_records_collected': progress['detail_records_collected'],
            'doudian_stage': progress['doudian_stage'],
            'doudian_activity_total': progress['doudian_activity_total'],
            'doudian_activity_completed': progress['doudian_activity_completed'],
            'doudian_current_activity': progress['doudian_current_activity'],
            'doudian_current_product_total': progress['doudian_current_product_total'],
            'doudian_current_product_completed': progress['doudian_current_product_completed'],
            'doudian_detail_rows': progress['doudian_detail_rows'],
            'doudian_signup_total': progress['doudian_signup_total'],
            'doudian_signup_completed': progress['doudian_signup_completed'],
            'doudian_order_window_total': progress['doudian_order_window_total'],
            'doudian_order_window_completed': progress['doudian_order_window_completed'],
            'doudian_mixed_rows': progress['doudian_mixed_rows'],
        }
        if run_id:
            try:
                data_sink.heartbeat_run(
                    run_id,
                    phase=progress['phase'],
                    current_row=current_row,
                    records_count=records,
                )
            except Exception as heartbeat_error:
                logger.debug("task heartbeat update failed for %s: %s", jid, heartbeat_error)

        # 通用批处理进度：只要行号前进，就打印一次，不绑定具体 adapter / phase。
        if (
            payload.get('kind') == 'before_phase' and
            progress['current'] > 0 and
            progress['total'] > 0 and
            progress['current'] != int(run_control.get('last_progress_exec_no') or 0)
        ):
            extra_parts = []
            if progress['batch_no'] > 0 and progress['total_batches'] > 0:
                extra_parts.append(f"批次 {progress['batch_no']}/{progress['total_batches']}")
            if progress['row_no'] > 0:
                extra_parts.append(f"源表行 {progress['row_no']}")
            if progress['buyer_id']:
                extra_parts.append(f"目标 {progress['buyer_id']}")
            suffix = f" · {' · '.join(extra_parts)}" if extra_parts else ""
            log(f"[progress] 第 {progress['current']}/{progress['total']} 条 · 已完成 {records} 条{suffix}")
            run_control['last_progress_exec_no'] = progress['current']

        if (
            payload.get('kind') == 'before_download_urls' and
            progress['current'] > 0 and
            progress['total'] > 0
        ):
            download_total = int(payload.get('download_item_total') or 0)
            download_batch_total = int(payload.get('download_batch_total') or 0)
            concurrency = int(payload.get('download_concurrency') or 0)
            retry_attempts = int(payload.get('download_retry_attempts') or 0)
            marker = f"{progress['current']}::{download_total}::{download_batch_total}::{concurrency}::{retry_attempts}"
            if marker != str(run_control.get('last_download_marker') or ''):
                target_text = f" · 目标 {progress['buyer_id']}" if progress['buyer_id'] else ""
                count_text = (
                    f" · 本批 {download_batch_total} 个文件 · 累计 {download_total} 个文件"
                    if download_batch_total > 0 and download_total > download_batch_total
                    else (f" · {download_total} 个文件" if download_total > 0 else "")
                )
                concurrency_text = f" · 并发 {concurrency}" if concurrency > 0 else ""
                retry_text = f" · 重试 {retry_attempts} 次" if retry_attempts > 1 else ""
                log(f"[download] 第 {progress['current']}/{progress['total']} 条{target_text}{count_text}{concurrency_text}{retry_text}，开始下载")
                run_control['last_download_marker'] = marker

        if run_control.get('stop_requested'):
            raise RunAbortedError(str(run_control.get('stop_reason') or '任务已停止'))

        if run_control.get('pause_requested'):
            _set_live_status(live_jids, {'status': 'paused', **base_status})
            if not run_control.get('pause_logged'):
                log("[control] 任务已暂停，等待继续…")
                run_control['pause_logged'] = True
            while run_control.get('pause_requested'):
                if run_control.get('stop_requested'):
                    raise RunAbortedError(str(run_control.get('stop_reason') or '任务已停止'))
                try:
                    await asyncio.wait_for(run_control['resume_event'].wait(), timeout=0.5)
                except asyncio.TimeoutError:
                    continue
            if run_control.get('pause_logged'):
                log("[control] 任务已继续")
                run_control['pause_logged'] = False

        _set_live_status(live_jids, {'status': 'running', **base_status})

    # 任务执行前重扫一次适配包，确保磁盘上的 manifest 改动立即生效
    adapter_loader.scan_all()
    m = adapter_loader.get_adapter(adapter_id)
    if not m:
        raise ValueError(f"Adapter not found: {adapter_id}")

    task = next((t for t in m.tasks if t.id == task_id), None)
    if not task:
        raise ValueError(f"Task not found: {task_id}")

    run_params = dict(params or {})
    run_started_at = datetime.now().isoformat()
    run_id = data_sink.begin_run(adapter_id, task_id)
    run_params["__task_run_id"] = run_id
    run_params["__task_started_at"] = run_started_at
    try:
        data_sink.heartbeat_run(run_id, phase="starting", current_row=0, records_count=0)
    except Exception as heartbeat_error:
        logger.debug("task heartbeat update failed for %s: %s", jid, heartbeat_error)
    if instance_uid:
        data_sink.link_task_instance_run(instance_uid, run_id, purpose="main")
    _merge_live_status(live_jids, {
        'run_id': run_id,
        'adapter_id': adapter_id,
        'task_id': task_id,
        'instance_uid': instance_uid,
        'last_seen_at': datetime.now().isoformat(),
        'current_row': 0,
        'phase': 'starting',
    })
    runner = None
    data = []
    output_files = []
    runtime_artifact_dir = data_sink.prepare_artifact_dir(adapter_id, task_id, run_id, "runtime")

    try:
        log(f"[{adapter_id}/{task_id}] Starting...")

        _inject_runtime_task_params(run_params, runtime_artifact_dir)
        per_item_review_urls = _resolve_tmall_buyer_review_item_urls(run_params) if (adapter_id, task_id) == ("tmall-ops-assistant", "buyer_reviews") else []
        runtime_options = runtime_options or {}
        task_param_ids = {p.id for p in task.params}
        for p in task.params:
            if p.id not in run_params and p.default is not None:
                run_params[p.id] = p.default

        mode = _resolve_task_open_mode(adapter_id, task_id, run_params, task_param_ids)
        current_tab_id = str(runtime_options.get('current_tab_id') or '').strip()
        shop_url = str(run_params.get('shop_url') or '').strip()
        configured_entry_url = shop_url or str(task.entry_url or m.entry_url or '').strip()
        target_entry_url = _resolve_task_target_entry_url(adapter_id, task_id, run_params, configured_entry_url)
        configured_match_prefixes = list(task.tab_match_prefixes or m.tab_match_prefixes or [])
        platform_name = m.name or adapter_id
        tmall_new_mode_entry_guard_tasks = {
            'tmall_material_test_data_export',
            'tmall_compete_paid_monitor',
        }

        def matches_task_entry_url(url: str) -> bool:
            prefixes = configured_match_prefixes or [target_entry_url]
            return any(_url_matches_prefix(url, prefix) for prefix in prefixes) or _is_compatible_current_tab_for_task(
                adapter_id,
                task_id,
                target_entry_url,
                url,
            )

        async def settle_new_mode_target_tab(initial_tab: Optional[dict], baseline_tab_ids: set[str]) -> Optional[dict]:
            if (
                mode == 'current'
                or not initial_tab
                or adapter_id != 'tmall-ops-assistant'
                or task_id not in tmall_new_mode_entry_guard_tasks
            ):
                return initial_tab
            initial_tab_id = str(initial_tab.get('id') or '')
            deadline = time.monotonic() + 8.0
            latest_tab = initial_tab
            while time.monotonic() < deadline:
                try:
                    tabs = [item for item in await _bridge_get_tabs() if item.get('type') == 'page']
                except Exception:
                    tabs = []

                if initial_tab_id:
                    refreshed = next((item for item in tabs if str(item.get('id') or '') == initial_tab_id), None)
                    if refreshed:
                        latest_tab = refreshed
                        if matches_task_entry_url(str(refreshed.get('url') or '')):
                            return refreshed

                new_target_tabs = [
                    item for item in tabs
                    if str(item.get('id') or '') not in baseline_tab_ids
                    and str(item.get('id') or '')
                    and matches_task_entry_url(str(item.get('url') or ''))
                ]
                if new_target_tabs:
                    selected = new_target_tabs[-1]
                    if str(selected.get('id') or '') != initial_tab_id:
                        log(f"新建页面实际落到目标标签页：{str(selected.get('url') or '')[:120]}")
                    return selected

                await asyncio.sleep(0.5)
            return latest_tab

        # 自动解析 file_excel 参数：如果传了 path 但没有 rows，就在后台读出来注入
        for pk, pv in run_params.items():
            if isinstance(pv, dict) and pv.get('path') and not pv.get('rows'):
                log(f"Resolving Excel rows for param '{pk}' from path: {pv['path']}")
                resolved = _read_local_excel(pv['path'])
                if "error" in resolved:
                    log(f"[warn] Failed to resolve Excel rows: {resolved['error']}")
                else:
                    pv['rows'] = resolved['rows']
                    pv['headers'] = resolved['headers']
                    if resolved.get('sheet_name') is not None:
                        pv['sheet_name'] = resolved['sheet_name']
                    if resolved.get('sheets') is not None:
                        pv['sheets'] = resolved['sheets']
                    log(f"Successfully resolved {len(pv['rows'])} rows.")

        is_shopee_marketing_adapter = target_entry_url.startswith('https://seller.shopee.cn/portal/marketing')

        tab = None
        current_mode_preferred_prefixes: list[str] = []
        new_mode_baseline_tab_ids: set[str] = set()
        if mode == 'current':
            all_tabs = [t for t in await _bridge_get_tabs() if t.get('type') == 'page']
            preferred_prefixes = configured_match_prefixes or [
                'https://seller.shopee.cn/portal/marketing',
                'https://seller.shopee.cn/portal/marketing/vouchers/list',
                'https://seller.shopee.cn/portal/marketing/vouchers/new',
                'https://seller.shopee.cn/portal/marketing/follow-prize/new',
            ] if is_shopee_marketing_adapter else configured_match_prefixes or [target_entry_url]
            current_mode_preferred_prefixes = list(preferred_prefixes)
            if adapter_id == 'temu' and shop_url:
                preferred_prefixes = list(dict.fromkeys([
                    *preferred_prefixes,
                    'https://www.temu.com/bgn_verification.html',
                ]))
                current_mode_preferred_prefixes = list(preferred_prefixes)

            if current_tab_id:
                tab = next((t for t in all_tabs if str(t.get('id')) == current_tab_id), None)
                if not tab:
                    raise RuntimeError("未找到当前页面对应的 Chrome 标签页。请重新聚焦目标业务页面后重试。")
                tab_url = str(tab.get('url', ''))
                if not any(_url_matches_prefix(tab_url, prefix) for prefix in preferred_prefixes) and not _is_compatible_current_tab_for_task(adapter_id, task_id, target_entry_url, tab_url):
                    raise RuntimeError(f"当前页面不是目标业务页：{tab_url[:120]}")
            else:
                candidate_tabs = [
                    t for t in all_tabs
                    if any(_url_matches_prefix(str(t.get('url', '')), prefix) for prefix in preferred_prefixes)
                    or _is_compatible_current_tab_for_task(adapter_id, task_id, target_entry_url, str(t.get('url', '')))
                ]
                if len(candidate_tabs) == 1:
                    tab = candidate_tabs[0]
                elif len(candidate_tabs) > 1:
                    raise RuntimeError(
                        "current 模式检测到多个目标页面，无法判断当前标签页。"
                        "请关闭多余相关标签页后重试，或改用“新页面开启（重新导航）”。"
                    )
                else:
                    raise RuntimeError(f"未找到已打开的目标页面：{target_entry_url}。请选择“新页面开启（重新导航）”或先手动打开并登录。")
        else:
            open_entry_url = _temu_new_tab_entry_url(target_entry_url) if adapter_id == 'temu' else target_entry_url
            new_mode_baseline_tab_ids = {
                str(item.get('id') or '')
                for item in await _bridge_get_tabs()
                if item.get('type') == 'page' and str(item.get('id') or '')
            }
            if open_entry_url != target_entry_url:
                log(f"mode=new，Temu 使用后台入口新建页面：{open_entry_url}，稍后导航业务页：{target_entry_url}")
            else:
                log(f"mode=new，尝试新建页面：{target_entry_url}")
            if adapter_id == 'doudian-ops-assistant' and _is_doudian_backend_url(target_entry_url):
                opener_candidates = [
                    item for item in await _bridge_get_tabs()
                    if item.get('type') == 'page' and _is_doudian_opener_tab_url(str(item.get('url') or ''))
                ]
                opener_candidates.sort(key=lambda item: _doudian_opener_tab_sort_key(target_entry_url, str(item.get('url') or '')))
                opener_tab = opener_candidates[0] if opener_candidates else None
                if opener_tab:
                    log(f"抖店复用已登录后台页打开新标签：{str(opener_tab.get('url') or '')[:120]}")
                    opener_runner = JSRunner(
                        get_bridge().get_tab_ws_url(opener_tab),
                        tab_id=str(opener_tab.get('id') or ''),
                        tab_url=str(opener_tab.get('url') or ''),
                        artifact_dir=runtime_artifact_dir,
                    )
                    before_ids = {
                        str(item.get('id') or '')
                        for item in await _bridge_get_tabs()
                        if item.get('type') == 'page'
                    }
                    open_expr = (
                        "(() => {\n"
                        f"  const url = {json.dumps(open_entry_url, ensure_ascii=False)};\n"
                        "  const win = window.open(url, '_blank');\n"
                        "  return { success: !!win, data: [], meta: { has_more: false } };\n"
                        "})()\n"
                    )
                    open_result = await opener_runner.evaluate_user_gesture(open_expr)
                    if open_result.success:
                        await asyncio.sleep(1.5)
                        new_tabs = [
                            item for item in await _bridge_get_tabs()
                            if item.get('type') == 'page' and str(item.get('id') or '') not in before_ids
                        ]
                        tab = next((item for item in new_tabs if _is_doudian_fxg_url(str(item.get('url') or ''))), None)
                        if not tab and new_tabs:
                            tab = new_tabs[-1]
                    else:
                        log(f"抖店页面上下文打开新标签失败，回退 CDP 新建：{open_result.error or 'unknown'}")
            if adapter_id == 'temu' and _is_temu_backend_url(target_entry_url):
                opener_candidates = [
                    item for item in await _bridge_get_tabs()
                    if item.get('type') == 'page' and _is_temu_opener_tab_url(str(item.get('url') or ''))
                ]
                opener_candidates.sort(key=lambda item: _temu_opener_tab_sort_key(target_entry_url, str(item.get('url') or '')))
                opener_tab = opener_candidates[0] if opener_candidates else None
                if opener_tab:
                    log(f"Temu 复用已登录后台页打开新标签：{str(opener_tab.get('url') or '')[:120]}")
                    opener_runner = JSRunner(
                        get_bridge().get_tab_ws_url(opener_tab),
                        tab_id=str(opener_tab.get('id') or ''),
                        tab_url=str(opener_tab.get('url') or ''),
                        artifact_dir=runtime_artifact_dir,
                    )
                    before_ids = {
                        str(item.get('id') or '')
                        for item in await _bridge_get_tabs()
                        if item.get('type') == 'page'
                    }
                    open_expr = (
                        "(() => {\n"
                        f"  const url = {json.dumps(open_entry_url, ensure_ascii=False)};\n"
                        "  const win = window.open(url, '_blank');\n"
                        "  return { success: !!win, data: [], meta: { has_more: false } };\n"
                        "})()\n"
                    )
                    open_result = await opener_runner.evaluate_user_gesture(open_expr)
                    if open_result.success:
                        await asyncio.sleep(1.5)
                        new_tabs = [
                            item for item in await _bridge_get_tabs()
                            if item.get('type') == 'page' and str(item.get('id') or '') not in before_ids
                        ]
                        tab = next((item for item in new_tabs if _is_temu_opener_tab_url(str(item.get('url') or ''))), None)
                        if not tab and new_tabs:
                            tab = new_tabs[-1]
                    else:
                        log(f"Temu 页面上下文打开新标签失败，回退 CDP 新建：{open_result.error or 'unknown'}")
            if not tab:
                try:
                    tab = await _bridge_new_tab(open_entry_url)
                except Exception as e:
                    log(f"新建 tab 失败，尝试复用现有 tab: {e}")
                    tab = await _bridge_find_tab(open_entry_url) or await _bridge_find_tab(target_entry_url)
            if not tab:
                raise RuntimeError(f"无法打开或找到目标页面：{target_entry_url}")

        tab = await settle_new_mode_target_tab(tab, new_mode_baseline_tab_ids)
        if not tab:
            raise RuntimeError(f"无法打开或找到目标页面：{target_entry_url}")

        log(f"Found tab: {tab.get('url', '')[:120]}")
        runner = JSRunner(
            get_bridge().get_tab_ws_url(tab),
            tab_id=str(tab.get('id') or ''),
            tab_url=str(tab.get('url') or ''),
            artifact_dir=runtime_artifact_dir,
        )

        async def navigate_runner_to(url: str, wait_seconds: float = 2.0, via_page: bool = False):
            if via_page:
                expression = (
                    "(() => {\n"
                    f"  location.href = {json.dumps(str(url or ''), ensure_ascii=False)};\n"
                    "  return { success: true, data: [], meta: { has_more: false } };\n"
                    "})()\n"
                )
                nav_result = await runner.evaluate(expression)
                if nav_result.success and wait_seconds > 0:
                    await asyncio.sleep(wait_seconds)
                    try:
                        await runner._refresh_ws_url()
                    except Exception:
                        logger.debug("refresh ws url failed after page-context navigate", exc_info=True)
            else:
                nav_result = await runner.navigate(url, wait_seconds=wait_seconds)
            if nav_result.success:
                runner.tab_url = str(url or runner.tab_url or '')
            return nav_result

        async def recover_temu_no_auth(stage: str = "入口"):
            if adapter_id != 'temu':
                return
            current_tab = await _bridge_get_tab(str(tab.get('id') or runner.tab_id or '')) if (tab or runner.tab_id) else None
            current_url = str((current_tab or {}).get('url') or '')
            if not _is_temu_no_auth_url(current_url):
                return
            shell_url = _temu_new_tab_entry_url(target_entry_url)
            log(f"检测到 Temu 反爬/无授权页（{stage}）：{current_url}，改从后台入口恢复：{shell_url}")
            nav_result = await navigate_runner_to(shell_url, wait_seconds=3)
            if not nav_result.success:
                raise RuntimeError(nav_result.error or f"无法从 Temu no-auth 恢复到后台入口：{shell_url}")

        async def get_runner_current_url() -> str:
            current_tab = None
            tab_id = str(tab.get('id') or runner.tab_id or '') if (tab or runner.tab_id) else ''
            if tab_id:
                try:
                    current_tab = await _bridge_get_tab(tab_id)
                except Exception:
                    current_tab = None
            return str((current_tab or {}).get('url') or runner.tab_url or (tab or {}).get('url') or '')

        async def get_runner_page_href() -> str:
            href_result = await runner.evaluate(
                "(() => ({ success: true, data: [{ href: String(location.href || '') }], meta: { has_more: false } }))()"
            )
            if not href_result.success or not href_result.data:
                return ""
            return str((href_result.data[0] or {}).get("href") or "")

        def is_task_entry_url(url: str) -> bool:
            prefixes = current_mode_preferred_prefixes or configured_match_prefixes or [target_entry_url]
            return any(_url_matches_prefix(url, prefix) for prefix in prefixes) or _is_compatible_current_tab_for_task(
                adapter_id,
                task_id,
                target_entry_url,
                url,
            )

        def merge_output_file_refs(*groups):
            merged = []
            seen = set()
            for group in groups:
                for item in group or []:
                    path = str(item or '').strip()
                    if not path or path in seen:
                        continue
                    seen.add(path)
                    merged.append(path)
            return merged

        if mode == 'new' and adapter_id == 'tmall-ops-assistant' and task_id in tmall_new_mode_entry_guard_tasks:
            page_href = await get_runner_page_href()
            if not is_task_entry_url(page_href):
                log(f"新建页运行上下文尚未进入业务入口，导航到：{target_entry_url}（当前：{page_href[:120] or 'unknown'}）")
                entry_nav = await navigate_runner_to(target_entry_url, wait_seconds=3)
                if not entry_nav.success:
                    raise RuntimeError(entry_nav.error or f"无法导航到业务入口：{target_entry_url}")
                await asyncio.sleep(1)

        async def export_outputs(data_rows):
            files = []
            filename_vars = await _build_export_filename_context(adapter_id, task_id, run_params, runner)
            for out in task.output:
                if out.type == OutputType.excel:
                    if not data_rows:
                        log("No data rows returned, skip Excel export")
                        continue
                    fname = out.filename or "{task_id}_{date}.xlsx"
                    path = await asyncio.to_thread(
                        data_sink.export_excel,
                        data_rows,
                        adapter_id,
                        task_id,
                        fname,
                        filename_vars=filename_vars,
                        column_order=getattr(out, 'columns', None),
                        column_groups=getattr(out, 'column_groups', None),
                        sheet_key=getattr(out, 'sheet_key', None),
                        sheet_configs=getattr(out, 'sheets', None),
                    )
                    files.append(path)
                    log(f"Excel exported: {path}")

                elif out.type == OutputType.json:
                    fname = out.filename or "{task_id}_{date}.json"
                    path = await asyncio.to_thread(
                        data_sink.export_json,
                        data_rows,
                        adapter_id,
                        task_id,
                        fname,
                        filename_vars=filename_vars,
                    )
                    files.append(path)
                    log(f"JSON exported: {path}")

                elif out.type == OutputType.notify:
                    if notifier.should_notify(out.condition, data_rows):
                        try:
                            await asyncio.to_thread(
                                notifier.send,
                                channel=out.channel or 'dingtalk',
                                title=f"{m.name} - {task.name}",
                                records=len(data_rows),
                                adapter_name=m.name,
                                task_name=task.name,
                                sample_rows=data_rows[:3] if data_rows else None,
                            )
                            log(f"Notification sent via {out.channel}")
                        except Exception as e:
                            log(f"Notification failed: {e}")
                    else:
                        log(f"Notification skipped (condition not met: {out.condition})")
            return files

        async def finalize_output_files(data_rows, runtime_files, exported_files):
            if adapter_id == 'amazon-ops-assistant' and task_id == 'amazon_label_batch_process':
                try:
                    generated_refs = list(run_params.get("__amazon_generated_refs") or [])
                    packaged_refs = await asyncio.to_thread(
                        copy_amazon_label_outputs_to_export_folder,
                        generated_refs=generated_refs,
                        exported_files=exported_files,
                        run_params=run_params,
                        log=log,
                    )
                    return merge_output_file_refs(packaged_refs)
                except Exception as package_error:
                    log(f"[warn] 亚马逊标签后处理输出失败，回退到原始输出: {package_error}")
                    return merge_output_file_refs(runtime_files, exported_files)

            if adapter_id == 'semir-cloud-drive':
                try:
                    packaged_refs = await asyncio.to_thread(
                        _finalize_semir_cloud_drive_outputs,
                        task_id=task_id,
                        data_rows=data_rows,
                        runtime_files=runtime_files,
                        exported_files=exported_files,
                        run_params=run_params,
                        runtime_artifact_dir=runtime_artifact_dir,
                        log=log,
                    )
                    return merge_output_file_refs(packaged_refs)
                except Exception as package_error:
                    log(f"[warn] semir 后处理失败，回退到原始输出: {package_error}")
                    return merge_output_file_refs(runtime_files, exported_files)

            if adapter_id == 'mop-ops-assistant' and task_id == 'cloud_folder_download':
                try:
                    packaged_refs = await asyncio.to_thread(
                        _finalize_mop_cloud_folder_download_outputs,
                        data_rows=data_rows,
                        runtime_files=runtime_files,
                        exported_files=exported_files,
                        run_params=run_params,
                        runtime_artifact_dir=runtime_artifact_dir,
                        log=log,
                    )
                    return merge_output_file_refs(packaged_refs)
                except Exception as package_error:
                    log(f"[warn] MOP 云盘下载后处理失败，回退到原始输出: {package_error}")
                    return merge_output_file_refs(runtime_files, exported_files)

            if adapter_id == 'bala-ai-video-assistant':
                try:
                    packaged_refs = await asyncio.to_thread(
                        _finalize_bala_ai_video_assistant_outputs,
                        task_id=task_id,
                        data_rows=data_rows,
                        runtime_files=runtime_files,
                        exported_files=exported_files,
                        run_params=run_params,
                        runtime_artifact_dir=runtime_artifact_dir,
                        log=log,
                    )
                    return merge_output_file_refs(packaged_refs)
                except Exception as package_error:
                    log(f"[warn] 巴拉AI视频素材后处理失败，回退到原始输出: {package_error}")
                    return merge_output_file_refs(runtime_files, exported_files)

            if adapter_id == 'plm-ops-assistant' and task_id == 'size_chart_downloader':
                try:
                    packaged_refs = await asyncio.to_thread(
                        _finalize_plm_size_chart_downloader_outputs,
                        runtime_files=runtime_files,
                        exported_files=exported_files,
                        run_params=run_params,
                        log=log,
                    )
                    return merge_output_file_refs(packaged_refs)
                except Exception as package_error:
                    log(f"[warn] PLM 尺码表导出目录复制失败，回退到原始输出: {package_error}")
                    return merge_output_file_refs(runtime_files, exported_files)

            if adapter_id == 'tmall-ops-assistant' and task_id in {
                'tmall_packaging_upload',
                'tmall_ai_image_test_chain',
                'tmall_material_test_data_export',
                'tmall_compete_paid_monitor',
            }:
                try:
                    packaged_refs = await asyncio.to_thread(
                        _finalize_tmall_ops_assistant_outputs,
                        task_id=task_id,
                        data_rows=data_rows,
                        runtime_files=runtime_files,
                        exported_files=exported_files,
                        run_params=run_params,
                        runtime_artifact_dir=runtime_artifact_dir,
                        log=log,
                    )
                    return merge_output_file_refs(packaged_refs)
                except Exception as package_error:
                    log(f"[warn] 天猫任务后处理失败，回退到原始输出: {package_error}")
                    return merge_output_file_refs(runtime_files, exported_files)

            if adapter_id == 'shenhui-new-arrival':
                try:
                    packaged_refs = await asyncio.to_thread(
                        _finalize_shenhui_new_arrival_outputs,
                        task_id=task_id,
                        data_rows=data_rows,
                        runtime_files=runtime_files,
                        exported_files=exported_files,
                        run_params=run_params,
                        runtime_artifact_dir=runtime_artifact_dir,
                        log=log,
                    )
                    return merge_output_file_refs(packaged_refs)
                except Exception as package_error:
                    log(f"[warn] 深绘图包后处理失败，回退到原始输出: {package_error}")
                    return merge_output_file_refs(runtime_files, exported_files)

            if adapter_id == 'tiktok-ops-assistant' and task_id == 'creator_video_download':
                try:
                    packaged_refs = await asyncio.to_thread(
                        _finalize_tiktok_creator_video_outputs,
                        data_rows=data_rows,
                        runtime_files=runtime_files,
                        exported_files=exported_files,
                        run_params=run_params,
                        runtime_artifact_dir=runtime_artifact_dir,
                        log=log,
                    )
                    return merge_output_file_refs(packaged_refs)
                except Exception as package_error:
                    log(f"[warn] TikTok 视频打包后处理失败，回退到原始输出: {package_error}")
                    return merge_output_file_refs(runtime_files, exported_files)

            return merge_output_file_refs(runtime_files, exported_files)

        # 可选登录检测：若 manifest 配置了 auth.check_script，则最多等 5 分钟
        if not task.skip_auth and m.auth and m.auth.check_script:
            try:
                auth_path = adapter_loader.resolve_adapter_file(adapter_id, m.auth.check_script)
            except FileNotFoundError:
                auth_path = None
            if auth_path:
                await recover_temu_no_auth("登录检测前")
                await wait_for_control()
                log(f"运行登录检测：{m.auth.check_script}")
                auth_result = await runner.evaluate(auth_path.read_text(encoding='utf-8'))
                logged_in = False
                if auth_result.success:
                    if isinstance(auth_result.meta, dict):
                        logged_in = bool(auth_result.meta.get('logged_in'))
                    if not logged_in and auth_result.data and isinstance(auth_result.data, list):
                        logged_in = bool((auth_result.data[0] or {}).get('logged_in'))

                if not logged_in:
                    login_url = (
                        target_entry_url
                        if adapter_id == 'mop-ops-assistant' and target_entry_url.startswith('https://fmp.semirapp.com/')
                        else (m.auth.login_url or target_entry_url)
                    )
                    if mode == 'new':
                        await wait_for_control()
                        log(f"检测到未登录，跳转登录页：{login_url}")
                        login_nav = await navigate_runner_to(login_url, wait_seconds=2)
                        if not login_nav.success:
                            raise RuntimeError(login_nav.error or f"无法跳转登录页：{login_url}")
                        await asyncio.sleep(2)
                    else:
                        log(f"检测到未登录，请在当前页面完成 {platform_name} 登录…")
                    wait_seconds = 300
                    for i in range(wait_seconds):
                        await wait_for_control({"records": 0})
                        auth_result = await runner.evaluate(auth_path.read_text(encoding='utf-8'))
                        ok = False
                        if auth_result.success:
                            if isinstance(auth_result.meta, dict):
                                ok = bool(auth_result.meta.get('logged_in'))
                            if not ok and auth_result.data and isinstance(auth_result.data, list):
                                ok = bool((auth_result.data[0] or {}).get('logged_in'))
                        if ok:
                            log("用户已完成登录，继续执行任务")
                            break
                        if i in (0, 4) or (i + 1) % 15 == 0:
                            log(f"等待用户登录 {platform_name}… {i + 1}/{wait_seconds}s")
                        await asyncio.sleep(1)
                    else:
                        raise RuntimeError(f"等待登录超时（5分钟）。请完成 {platform_name} 登录后重试。")

                should_return_to_entry = mode == 'new'
                if mode == 'current':
                    current_url = await get_runner_current_url()
                    should_return_to_entry = bool(target_entry_url and not is_task_entry_url(current_url))
                    if should_return_to_entry:
                        log(f"登录后当前页不在业务入口，导航回：{target_entry_url}（当前：{current_url[:120]}）")

                if should_return_to_entry:
                    # 登录完成后，重新导航回业务入口页
                    await wait_for_control()
                    await recover_temu_no_auth("导航业务入口前")
                    if mode == 'new':
                        log(f"导航回业务入口：{target_entry_url}")
                    entry_nav = await navigate_runner_to(
                        target_entry_url,
                        wait_seconds=3,
                        via_page=adapter_id == 'temu' and _is_temu_backend_url(target_entry_url),
                    )
                    if not entry_nav.success:
                        raise RuntimeError(entry_nav.error or f"无法导航回业务入口：{target_entry_url}")
                    await asyncio.sleep(3)
                    await recover_temu_no_auth("导航业务入口后")
        elif adapter_id == 'temu' and mode == 'new' and _is_temu_backend_url(target_entry_url):
            await recover_temu_no_auth("跳过登录检测后")
            log(f"导航回业务入口：{target_entry_url}")
            entry_nav = await navigate_runner_to(
                target_entry_url,
                wait_seconds=3,
                via_page=_is_temu_backend_url(target_entry_url),
            )
            if not entry_nav.success:
                raise RuntimeError(entry_nav.error or f"无法导航回业务入口：{target_entry_url}")
            await asyncio.sleep(3)
            await recover_temu_no_auth("导航业务入口后")

        script_path = adapter_loader.resolve_adapter_file(adapter_id, task.script)

        await wait_for_control({"records": 0})
        log(f"Injecting script: {task.script}")
        if (adapter_id, task_id) == ("tmall-ops-assistant", "buyer_reviews") and per_item_review_urls:
            aggregated_data = []
            total_item_count = len(per_item_review_urls)
            for item_index, item_url in enumerate(per_item_review_urls, start=1):
                item_params = dict(run_params)
                item_params["item_links"] = item_url
                item_params["__item_index__"] = item_index
                item_params["__item_total__"] = total_item_count
                log(f"抓取商品 {item_index}/{total_item_count}: {item_url}")
                item_nav = await runner.navigate(item_url, wait_seconds=2 if item_index == 1 else 1.5)
                if not item_nav.success:
                    raise RuntimeError(item_nav.error or f"无法导航到商品页：{item_url}")
                item_rows = await runner.run_script_file(script_path, params=item_params, control_hook=wait_for_control)
                aggregated_data.extend(item_rows)
            data = aggregated_data
        else:
            data = await runner.run_script_file(script_path, params=run_params, control_hook=wait_for_control)
        raw_count = len(data)
        data = _apply_final_export_guards(adapter_id, task_id, data)
        if adapter_id == 'tiktok-ops-assistant' and task_id == 'creator_video_download':
            data = _verify_tiktok_creator_video_download_rows(data, log=log)
        if adapter_id == 'amazon-ops-assistant' and task_id == 'amazon_label_batch_process':
            data, amazon_generated_refs = split_amazon_label_rows(
                data_rows=data,
                run_params=run_params,
                runtime_artifact_dir=runtime_artifact_dir,
                log=log,
            )
            run_params["__amazon_generated_refs"] = amazon_generated_refs
            raw_count = len(data)
        if (adapter_id, task_id) == ('tmall-ops-assistant', 'tmall_ai_image_generation'):
            data = await _apply_tmall_ai_image_generation(data, run_params, wait_for_control, log)
            raw_count = len(data)
        if (adapter_id, task_id) == ('tmall-ops-assistant', 'tmall_ai_image_test_chain'):
            data = await _apply_tmall_ai_image_test_chain(run_params, runtime_artifact_dir, wait_for_control, log)
            raw_count = len(data)
        if (adapter_id, task_id) == (BALA_AI_VIDEO_ADAPTER_ID, BALA_AI_FACE_BACKGROUND_TASK_ID):
            data = await _apply_bala_ai_face_background_generate(run_params, wait_for_control, log)
            raw_count = len(data)
        deduped_count = len(data)
        log(f"Script complete. Records: {raw_count}")
        if deduped_count != raw_count:
            log(f"Final export guard removed {raw_count - deduped_count} duplicate rows before export")

        runtime_files = list(getattr(runner, 'runtime_output_files', []) or [])
        skip_table_export = _is_tmall_ai_generation_confirmation_result(adapter_id, task_id, data)
        if skip_table_export:
            log("确认提交生图阶段不导出表格文件，等待最终创建结果后再输出")
            exported_files = []
            output_files = merge_output_file_refs([
                ref for ref in runtime_files
                if not _is_table_output_file_ref(ref)
            ])
        else:
            exported_files = await export_outputs(data)
            output_files = await finalize_output_files(data, runtime_files, exported_files)
        approval_board_urls = _extract_tmall_approval_board_urls(data) if (adapter_id, task_id) == ('tmall-ops-assistant', 'tmall_ai_image_test_chain') else {}
        approval_board_url = approval_board_urls.get("approval_board_url", "") if approval_board_urls else ""
        bala_review_board_url = _extract_bala_review_board_url(data) if (adapter_id, task_id) == (BALA_AI_VIDEO_ADAPTER_ID, BALA_AI_FACE_BACKGROUND_TASK_ID) else ""
        if approval_board_url:
            output_files = merge_output_file_refs([approval_board_url], output_files)
        if bala_review_board_url:
            output_files = merge_output_file_refs([bala_review_board_url], output_files)
        data_sink.finish_run(run_id, len(data), output_files)
        if instance_uid:
            _sync_task_instance_artifacts(instance_uid, output_files, run_id)
        done_status = {
            'status': 'done',
            'run_id': run_id,
            'adapter_id': adapter_id,
            'task_id': task_id,
            'instance_uid': instance_uid,
            'records': len(data),
            'last_seen_at': datetime.now().isoformat(),
            'current_row': len(data),
            'phase': 'done',
        }
        if approval_board_url:
            done_status['approval_board_url'] = approval_board_url
        if bala_review_board_url:
            done_status['bala_review_board_url'] = bala_review_board_url
        _set_live_status(live_jids, done_status)
        if instance_uid:
            approval_summary = _parse_tmall_approval_board_url(
                approval_board_urls.get("local_board_url", "") or approval_board_url
            )
            generation_confirmation_pending = any(
                isinstance(row, dict) and str(row.get("阶段") or "").strip() == "确认提交生图"
                for row in data or []
            )
            instance_summary = {
                "records": len(data),
                "output_files": output_files,
                "approval_board_url": approval_board_url,
                "cloud_board_url": approval_board_urls.get("cloud_board_url", ""),
                "local_board_url": approval_board_urls.get("local_board_url", ""),
                "approval_stage": "confirm_generation" if generation_confirmation_pending else "approval",
                "run_id": run_id,
                **approval_summary,
            }
            if approval_board_url:
                data_sink.update_task_instance(
                    instance_uid,
                    status="waiting_generation_confirmation" if generation_confirmation_pending else "waiting_approval",
                    current_step="confirm" if generation_confirmation_pending else "approval",
                    summary=instance_summary,
                )
            else:
                data_sink.update_task_instance(
                    instance_uid,
                    status="completed",
                    current_step="create",
                    summary=instance_summary,
                )
        if schedule_uid:
            _notify_task_schedule_result(
                schedule_uid=schedule_uid,
                instance_uid=instance_uid,
                run_id=run_id,
                status="completed",
                records=len(data),
                output_files=output_files,
                error="",
                started_at=run_started_at,
                finished_at=datetime.now().isoformat(),
                log=log,
            )
        log(f"[{adapter_id}/{task_id}] Done. {len(data)} records.")

    except RunAbortedError as e:
        err = e.reason or str(e)
        data = list(e.partial_data or data or [])
        raw_count = len(data)
        data = _apply_final_export_guards(adapter_id, task_id, data)
        if adapter_id == 'tiktok-ops-assistant' and task_id == 'creator_video_download':
            data = _verify_tiktok_creator_video_download_rows(data, log=log)
        deduped_count = len(data)
        if deduped_count != raw_count:
            log(f"Final export guard removed {raw_count - deduped_count} duplicate rows before partial export")
        if run_control:
            run_control['pause_requested'] = False
            run_control['stop_requested'] = False
            run_control['resume_event'].set()
            run_control['pause_logged'] = False

        try:
            runtime_files = list(getattr(runner, 'runtime_output_files', []) or []) if runner else []
            exported = await export_outputs(data) if runner else []
            output_files = await finalize_output_files(data, runtime_files, exported) if runner else []
        except Exception as export_error:
            output_files = []
            log(f"[warn] 停止任务时导出部分结果失败: {export_error}")

        data_sink.stop_run(run_id, len(data), output_files, err)
        stopped_status = {
            'status': 'stopped',
            'run_id': run_id,
            'adapter_id': adapter_id,
            'task_id': task_id,
            'instance_uid': instance_uid,
            'records': len(data),
            'error': err,
            'last_seen_at': datetime.now().isoformat(),
            'current_row': len(data),
            'phase': 'stopped',
        }
        _set_live_status(live_jids, stopped_status)
        if instance_uid:
            _sync_task_instance_artifacts(instance_uid, output_files, run_id)
            data_sink.update_task_instance(
                instance_uid,
                status="stopped",
                summary={"records": len(data), "output_files": output_files, "error": err, "run_id": run_id},
            )
        if schedule_uid:
            _notify_task_schedule_result(
                schedule_uid=schedule_uid,
                instance_uid=instance_uid,
                run_id=run_id,
                status="stopped",
                records=len(data),
                output_files=output_files,
                error=err,
                started_at=run_started_at,
                finished_at=datetime.now().isoformat(),
                log=log,
            )
        log(f"[{adapter_id}/{task_id}] Stopped. {len(data)} records.")

    except asyncio.CancelledError:
        err = "任务已停止"
        if run_control:
            run_control['pause_requested'] = False
            run_control['stop_requested'] = False
            run_control['resume_event'].set()
            run_control['pause_logged'] = False
        raw_count = len(data)
        data = _apply_final_export_guards(adapter_id, task_id, data)
        if adapter_id == 'tiktok-ops-assistant' and task_id == 'creator_video_download':
            data = _verify_tiktok_creator_video_download_rows(data, log=log)
        deduped_count = len(data)
        if deduped_count != raw_count:
            log(f"Final export guard removed {raw_count - deduped_count} duplicate rows before partial export")
        try:
            runtime_files = list(getattr(runner, 'runtime_output_files', []) or []) if runner else []
            exported = await _await_cleanup_after_cancel(export_outputs(data)) if runner else []
            output_files = await _await_cleanup_after_cancel(finalize_output_files(data, runtime_files, exported)) if runner else []
        except Exception as export_error:
            output_files = []
            log(f"[warn] 停止任务时导出部分结果失败: {export_error}")
        data_sink.stop_run(run_id, len(data), output_files, err)
        stopped_status = {
            'status': 'stopped',
            'run_id': run_id,
            'adapter_id': adapter_id,
            'task_id': task_id,
            'instance_uid': instance_uid,
            'records': len(data),
            'error': err,
            'last_seen_at': datetime.now().isoformat(),
            'current_row': len(data),
            'phase': 'stopped',
        }
        _set_live_status(live_jids, stopped_status)
        if instance_uid:
            _sync_task_instance_artifacts(instance_uid, output_files, run_id)
            data_sink.update_task_instance(
                instance_uid,
                status="stopped",
                summary={"records": len(data), "output_files": output_files, "error": err, "run_id": run_id},
            )
        if schedule_uid:
            _notify_task_schedule_result(
                schedule_uid=schedule_uid,
                instance_uid=instance_uid,
                run_id=run_id,
                status="stopped",
                records=len(data),
                output_files=output_files,
                error=err,
                started_at=run_started_at,
                finished_at=datetime.now().isoformat(),
                log=log,
            )
        log(f"[{adapter_id}/{task_id}] Stopped. {len(data)} records.")
        raise

    except Exception as e:
        err = str(e)
        log(f"[{adapter_id}/{task_id}] ERROR: {err}")
        data_sink.fail_run(run_id, err)
        error_status = {
            'status': 'error',
            'run_id': run_id,
            'adapter_id': adapter_id,
            'task_id': task_id,
            'instance_uid': instance_uid,
            'records': 0,
            'error': err,
            'last_seen_at': datetime.now().isoformat(),
            'current_row': 0,
            'phase': 'error',
        }
        _set_live_status(live_jids, error_status)
        if instance_uid:
            data_sink.update_task_instance(
                instance_uid,
                status="failed",
                summary={"records": 0, "error": err, "run_id": run_id},
            )
        if schedule_uid:
            _notify_task_schedule_result(
                schedule_uid=schedule_uid,
                instance_uid=instance_uid,
                run_id=run_id,
                status="failed",
                records=0,
                output_files=[],
                error=err,
                started_at=run_started_at,
                finished_at=datetime.now().isoformat(),
                log=log,
            )
        raise
    finally:
        if run_control is not None:
            run_control['task'] = None
            run_control['pause_requested'] = False
            run_control['stop_requested'] = False
            run_control['resume_event'].set()
            run_control['pause_logged'] = False


@asynccontextmanager
async def lifespan(app: FastAPI):
    instance_lock = _acquire_backend_instance_lock()
    if _api_auth_required():
        _get_api_token()
    # Init DB
    data_sink.init_db()
    owns_backend_instance = bool(instance_lock.acquired)
    app.state.owns_backend_instance = owns_backend_instance
    if not owns_backend_instance:
        logger.warning("Another crawshrimp backend instance is already active; skip startup side effects in this process")

    if owns_backend_instance:
        orphaned_active_runs = data_sink.list_active_runs(ACTIVE_LIVE_STATUSES)
        orphaned_runs = data_sink.stop_orphaned_active_runs()
        if orphaned_active_runs:
            _cleanup_orphaned_runtime_artifacts(orphaned_active_runs)
        if orphaned_runs:
            logger.warning("Marked %s orphaned active task run(s) as stopped after backend startup", orphaned_runs)

    # Install built-in adapters
    if owns_backend_instance:
        built_in = Path(__file__).parent.parent / "adapters"
        if built_in.exists():
            for d in adapter_loader.iter_manifest_dirs(built_in):
                try:
                    adapter_loader.install_from_dir(str(d), install_mode="copy", preserve_existing_link=True)
                except Exception as e:
                    logger.warning(f"Built-in adapter failed {d.name}: {e}")

        adapter_loader.scan_all()
        try:
            ensure_knowledge_index()
        except Exception:
            logger.exception("knowledge index initialization failed; continuing without prebuilt index")

    # Register scheduled tasks
    if owns_backend_instance:
        try:
            for item in adapter_loader.list_all():
                if item['enabled']:
                    m = adapter_loader.get_adapter(item['id'])
                    if m:
                        sched_module.register_adapter(m, _run_scheduled_task)
            try:
                persisted_schedules = data_sink.list_task_schedules(enabled=True)
            except Exception:
                logger.exception("failed to load persisted task schedules; continuing without user schedules")
                persisted_schedules = []
            sched_module.register_task_schedules(persisted_schedules, _run_task_schedule)

            sched_module.start()
        except Exception:
            logger.exception("scheduler startup failed; continuing without scheduled jobs")
        try:
            _start_cloud_machine_if_enabled()
        except Exception:
            logger.exception("cloud approval machine auto-start failed; continuing without cloud machine loop")
        try:
            ai_video_generation_service.ensure_worker_started()
        except Exception:
            logger.exception("ai video generation recovery failed; continuing without active video recovery")
    logger.info("crawshrimp core started")
    try:
        yield
    finally:
        try:
            if owns_backend_instance:
                sched_module.shutdown()
        except Exception:
            logger.exception("scheduler shutdown failed")
        try:
            ai_video_generation_service.stop_worker()
        except Exception:
            logger.exception("ai video worker shutdown failed")
        app.state.owns_backend_instance = False
        instance_lock.close()


app = FastAPI(title="crawshrimp", version=API_VERSION, lifespan=lifespan)
app.add_middleware(
    CORSMiddleware,
    allow_origins=_allowed_cors_origins(),
    allow_methods=["*"],
    allow_headers=["Content-Type", "X-Crawshrimp-Token"],
)


@app.middleware("http")
async def require_local_api_token(request: Request, call_next):
    if request.method == "OPTIONS" or _is_public_api_path(request.url.path):
        return await call_next(request)

    expected = _get_api_token()
    if not expected:
        if _api_auth_required():
            return _add_local_cors_headers(
                request,
                JSONResponse(
                    {"detail": "CRAWSHRIMP_API_TOKEN is required before using crawshrimp API"},
                    status_code=503,
                ),
            )
        return await call_next(request)

    supplied = str(request.headers.get(API_TOKEN_HEADER) or "").strip()
    if not supplied or not hmac.compare_digest(supplied, expected):
        return _add_local_cors_headers(
            request,
            JSONResponse({"detail": "Unauthorized"}, status_code=401),
        )

    return await call_next(request)


def _runtime_operation_exempt_path(path: str) -> bool:
    normalized = str(path or "")
    return (
        normalized == "/runtime/install-readiness"
        or normalized == "/runtime/update-drain"
        or normalized == "/health"
        or normalized.startswith("/health/")
    )


def _request_requires_runtime_operation(request: Request) -> bool:
    method = str(request.method or "").upper()
    path = str(request.url.path or "")
    if method == "OPTIONS" or _runtime_operation_exempt_path(path):
        return False
    if method in {"POST", "PUT", "PATCH", "DELETE"}:
        return True
    return method == "GET" and re.match(r"^/data/[^/]+/[^/]+/export$", path) is not None


def _update_pending_response() -> JSONResponse:
    return JSONResponse(
        {
            "detail": {
                "code": "update_pending",
                "message": "抓虾正在准备安装更新，请稍后再启动新任务。",
            }
        },
        status_code=409,
    )


@app.middleware("http")
async def track_runtime_operation(request: Request, call_next):
    token = ""
    if not _request_requires_runtime_operation(request):
        return await call_next(request)
    operation_id = f"{request.method.upper()} {request.url.path}"
    try:
        token = runtime_install_guard.begin_operation("http_request", operation_id, operation_id)
    except UpdateDrainActive:
        return _add_local_cors_headers(request, _update_pending_response())
    try:
        return await call_next(request)
    finally:
        runtime_install_guard.end_operation(token)


def _is_public_api_path(path: str) -> bool:
    normalized = str(path or "")
    if normalized in PRIVATE_API_PATHS:
        return False
    if re.fullmatch(r"/bala-ai-video-review/api/[^/]+/export-video-input", normalized):
        return False
    return normalized in PUBLIC_API_PATHS or any(normalized.startswith(prefix) for prefix in PUBLIC_API_PREFIXES)


def _adapter_asset_headers() -> dict[str, str]:
    return {
        "Access-Control-Allow-Origin": "*",
        "Cross-Origin-Resource-Policy": "cross-origin",
        "Cache-Control": "public, max-age=31536000, immutable",
    }


@app.get("/adapter-assets/{adapter_id}/{asset_path:path}")
def get_adapter_asset(adapter_id: str, asset_path: str):
    normalized = str(asset_path or "").replace("\\", "/").lstrip("/")
    if not normalized.startswith(("vendor/", "assets/")):
        raise HTTPException(400, "adapter asset path must be under vendor/ or assets/")

    adapter_dir = adapter_loader.get_adapter_dir(adapter_id)
    if not adapter_dir:
        raise HTTPException(404, f"Adapter not found: {adapter_id}")

    try:
        path = adapter_loader.resolve_adapter_relative_file(
            Path(adapter_dir),
            normalized,
            allowed_suffixes=None,
            require_exists=True,
        )
    except FileNotFoundError as exc:
        raise HTTPException(404, str(exc))
    except ValueError as exc:
        raise HTTPException(400, str(exc))

    return FileResponse(path, headers=_adapter_asset_headers())


class BalaMaterialBatchFromRowsRequest(BaseModel):
    rows: list = []
    source_task: dict = {}


class BalaMaterialSelectionRequest(BaseModel):
    selected_asset_ids: List[str] = []


class BalaMaterialExportAiInputRequest(BaseModel):
    operation_type: str = "face_swap"
    selected_asset_ids: List[str] = []
    model_ref_ids: List[str] = []
    background_prompt: str = ""
    garment_images: dict = {}
    outfit_reference_images: dict = {}
    variant_reference_images: dict = {}
    pose_prompt: str = ""
    prompt_cards: list = []
    prompt_extra: str = ""


class BalaReviewDecisionsRequest(BaseModel):
    decisions: dict = {}


class BalaReviewRegenerateRequest(BaseModel):
    asset_id: str = ""
    prompt: str = ""
    reference_paths: Optional[List[str]] = None
    submit_async: bool = False


class BalaReviewExportVideoInputRequest(BaseModel):
    provider: str = "qn_img2video"
    output_dir: str = ""
    template_id: str = ""
    template_match: str = ""
    prompt: str = ""
    download_videos: bool = True


class BalaSeedanceVideoRequest(BaseModel):
    style_code: str = ""
    prompt: str = ""
    image_paths: List[str] = []
    output_dir: str = ""
    output_path: str = ""
    model: str = BALA_SEEDANCE_DEFAULT_MODEL
    ratio: str = "9:16"
    resolution: str = "720p"
    duration: int = 5
    generate_audio: bool = True
    watermark: bool = False
    wait: bool = True
    interval_seconds: int = 5
    timeout_seconds: int = 1800


class BalaHappyHorseVideoRequest(BaseModel):
    provider: str = "happyhorse"
    style_code: str = ""
    mode: str = "i2v"
    prompt: str = ""
    image_paths: List[str] = []
    video_paths: List[str] = []
    image_urls: List[str] = []
    output_dir: str = ""
    output_path: str = ""
    resolution: str = "720P"
    ratio: str = "3:4"
    duration: int = 5
    generate_audio: bool = False
    kling_mode: str = "std"
    pixverse_image_path: str = ""
    pixverse_video_path: str = ""
    pixverse_image_url: str = ""
    pixverse_video_url: str = ""
    watermark: bool = False
    wait: bool = True
    interval_seconds: int = 5
    timeout_seconds: int = 1800


class BalaVideoProviderTaskRequest(BaseModel):
    provider: str = ""
    task_id: str = ""
    style_code: str = ""
    output_dir: str = ""
    output_path: str = ""
    download: bool = False
    interval_seconds: int = 5
    timeout_seconds: int = 1800


class BalaVideoProviderPreflightRequest(BaseModel):
    provider: str = ""
    style_code: str = ""
    mode: str = "i2v"
    prompt: str = ""
    image_paths: List[str] = []
    video_paths: List[str] = []
    output_dir: str = ""
    resolution: str = "720P"
    ratio: str = "9:16"
    duration: int = 5
    generate_audio: bool = True
    kling_mode: str = "std"
    pixverse_image_path: str = ""
    pixverse_video_path: str = ""
    pixverse_image_url: str = ""
    pixverse_video_url: str = ""
    watermark: bool = False


def _bala_ai_video_base_url() -> str:
    return str(os.environ.get("CRAWSHRIMP_PUBLIC_BASE_URL") or f"http://127.0.0.1:{os.environ.get('CRAWSHRIMP_PORT', '18765')}").rstrip("/")


def _safe_bala_template_text(value: object) -> str:
    return str(value or "").replace("\ufeff", "").strip()


def _normalize_bala_template(item: dict, index: int) -> dict:
    template_id = _safe_bala_template_text(item.get("templateId") or item.get("template_id") or item.get("模板ID") or item.get("id"))
    slot_description = _safe_bala_template_text(item.get("slotDescription") or item.get("slot_description") or item.get("槽位说明"))
    return {
        "index": int(item.get("index") or item.get("序号") or index),
        "templateId": template_id,
        "id": template_id,
        "title": _safe_bala_template_text(item.get("title") or item.get("模板标题") or template_id),
        "description": _safe_bala_template_text(item.get("description") or item.get("描述") or slot_description),
        "slotDescription": slot_description,
        "type": _safe_bala_template_text(item.get("type") or item.get("类型") or "action"),
        "ratio": _safe_bala_template_text(item.get("ratio") or item.get("比例") or "3:4"),
        "duration": int(float(str(item.get("duration") or item.get("时长秒") or 0) or 0)),
        "provider": _safe_bala_template_text(item.get("provider") or "content"),
        "videoUrl": _safe_bala_template_text(item.get("videoUrl") or item.get("远程预览视频")),
        "coverUrl": _safe_bala_template_text(item.get("coverUrl") or item.get("远程封面")),
        "localPreviewVideo": _safe_bala_template_text(item.get("localPreviewVideo") or item.get("本地预览视频")),
        "localCoverImage": _safe_bala_template_text(item.get("localCoverImage") or item.get("本地封面")),
        "slots": item.get("slots") if isinstance(item.get("slots"), list) else [],
    }


def _load_bala_video_template_catalog() -> dict:
    if BALA_VIDEO_TEMPLATE_CATALOG_JSON.is_file():
        payload = json.loads(BALA_VIDEO_TEMPLATE_CATALOG_JSON.read_text(encoding="utf-8-sig"))
        source_templates = payload.get("templates") if isinstance(payload, dict) else []
        templates = [
            _normalize_bala_template(item, index)
            for index, item in enumerate(source_templates or [], start=1)
            if isinstance(item, dict)
        ]
        return {
            "ok": True,
            "source": str(BALA_VIDEO_TEMPLATE_CATALOG_JSON),
            "mainCategory": _safe_bala_template_text(payload.get("mainCategory") if isinstance(payload, dict) else ""),
            "templates": [item for item in templates if item.get("templateId")],
            "count": len([item for item in templates if item.get("templateId")]),
        }
    if BALA_VIDEO_TEMPLATE_CATALOG_CSV.is_file():
        with BALA_VIDEO_TEMPLATE_CATALOG_CSV.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.DictReader(handle))
        templates = [
            _normalize_bala_template(row, index)
            for index, row in enumerate(rows, start=1)
        ]
        return {
            "ok": True,
            "source": str(BALA_VIDEO_TEMPLATE_CATALOG_CSV),
            "mainCategory": "童装/婴儿装/亲子装",
            "templates": [item for item in templates if item.get("templateId")],
            "count": len([item for item in templates if item.get("templateId")]),
        }
    return {
        "ok": False,
        "source": str(BALA_VIDEO_TEMPLATE_DIR),
        "mainCategory": "",
        "templates": [],
        "count": 0,
        "error": "未找到本地生意管家模板目录",
    }


def _parse_seedance_cli_json_objects(text: str) -> list[dict]:
    decoder = json.JSONDecoder()
    index = 0
    objects: list[dict] = []
    source = str(text or "")
    while index < len(source):
        brace = source.find("{", index)
        if brace < 0:
            break
        try:
            value, end = decoder.raw_decode(source[brace:])
        except json.JSONDecodeError:
            index = brace + 1
            continue
        if isinstance(value, dict):
            objects.append(value)
        index = brace + end
    return objects


def _bala_video_provider_secrets() -> dict[str, str]:
    cfg = load_config()
    return {
        "seedance": str(_nested_config_value(cfg, "ai.video.seedance_api_key") or "").strip(),
        "happyhorse": str(_nested_config_value(cfg, "ai.video.bailian_api_key") or "").strip(),
    }


def _bala_video_provider_status() -> dict:
    cfg = load_config()
    secrets = _bala_video_provider_secrets()

    def source_for(config_key: str) -> str:
        if str(_nested_config_value(cfg, config_key) or "").strip():
            return "AI 能力"
        return "未配置"

    return {
        "seedance": {
            "configured": bool(secrets["seedance"]),
            "source": source_for("ai.video.seedance_api_key"),
        },
        "happyhorse": {
            "configured": bool(secrets["happyhorse"]),
            "source": source_for("ai.video.bailian_api_key"),
        },
    }


def _bala_video_provider_env(provider: str) -> tuple[dict, list[str]]:
    cfg = load_config()
    secrets = _bala_video_provider_secrets()
    status_key = _bala_provider_status_key(provider)
    env = dict(os.environ)
    env.pop("ELECTRON_RUN_AS_NODE", None)
    for environment_name in (
        "ARK_API_KEY",
        "VOLCENGINE_ARK_API_KEY",
        "DASHSCOPE_API_KEY",
        "BAILIAN_API_KEY",
    ):
        env.pop(environment_name, None)
    secret_values = [value for value in secrets.values() if value]
    if status_key == "seedance":
        if secrets["seedance"]:
            env["ARK_API_KEY"] = secrets["seedance"]
        base_url = str(_nested_config_value(cfg, "ai.video.seedance_base_url") or "").strip()
        if base_url:
            env["ARK_BASE_URL"] = base_url
    elif status_key == "happyhorse":
        if secrets["happyhorse"]:
            env["DASHSCOPE_API_KEY"] = secrets["happyhorse"]
        mappings = {
            "BAILIAN_WORKSPACE_ID": "ai.video.bailian_workspace_id",
            "BAILIAN_REGION": "ai.video.bailian_region",
            "BAILIAN_BASE_URL": "ai.video.bailian_base_url",
        }
        for environment_name, config_key in mappings.items():
            value = str(_nested_config_value(cfg, config_key) or "").strip()
            if value:
                env[environment_name] = value
    return env, secret_values


def _bala_video_node_runtime() -> tuple[str, dict[str, str]]:
    bundled = str(os.environ.get("CRAWSHRIMP_NODE_EXECUTABLE") or "").strip()
    if bundled:
        return bundled, {"ELECTRON_RUN_AS_NODE": "1"}
    return shutil.which("node") or "node", {}


async def _communicate_video_provider_process(
    process: object,
    *,
    timeout_seconds: float,
    provider_label: str,
) -> tuple[bytes, bytes]:
    try:
        return await asyncio.wait_for(
            process.communicate(),
            timeout=max(float(timeout_seconds), 0.001),
        )
    except asyncio.TimeoutError as exc:
        try:
            process.kill()
        except (ProcessLookupError, OSError):
            pass
        try:
            await process.wait()
        except Exception:
            pass
        raise HTTPException(504, f"{provider_label} CLI 执行超时，进程已终止，可从当前视频任务重试") from exc


def _validate_bala_video_provider_image_path(raw_path: object, provider_label: str) -> Path:
    value = str(raw_path or "").strip()
    path = Path(value).expanduser()
    if not path.is_file():
        raise HTTPException(400, f"{provider_label} 参考图片不存在：{value}")
    if path.suffix.lower() not in BALA_AI_IMAGE_EXTS:
        raise HTTPException(400, f"{provider_label} 参考文件不是支持的图片：{value}")
    if path.stat().st_size > 20 * 1024 * 1024:
        raise HTTPException(400, f"{provider_label} 参考图片超过 20MB：{value}")
    try:
        from PIL import Image

        with Image.open(path) as image:
            image.verify()
    except Exception as exc:
        raise HTTPException(400, f"{provider_label} 参考文件不是有效图片：{value}") from exc
    return path


def _validate_bala_video_provider_video_path(raw_path: object, provider_label: str) -> Path:
    value = str(raw_path or "").strip()
    path = Path(value).expanduser()
    if not path.is_file():
        raise HTTPException(400, f"{provider_label} 参考视频不存在：{value}")
    if path.suffix.lower() not in {".mp4", ".mov", ".m4v", ".webm"}:
        raise HTTPException(400, f"{provider_label} 参考文件不是支持的视频：{value}")
    if path.stat().st_size > 200 * 1024 * 1024:
        raise HTTPException(400, f"{provider_label} 参考视频超过 200MB：{value}")
    mime = mimetypes.guess_type(str(path))[0] or ""
    if path.suffix.lower() in {".mp4", ".m4v"}:
        mime = "video/mp4"
    elif path.suffix.lower() == ".mov":
        mime = "video/quicktime"
    elif path.suffix.lower() == ".webm":
        mime = "video/webm"
    if mime not in {"video/mp4", "video/quicktime", "video/webm"}:
        raise HTTPException(400, f"{provider_label} 参考文件不是支持的视频：{value}")
    return path


def _sanitize_video_provider_output(text: object, secret_values: Optional[List[str]] = None) -> str:
    sanitized = str(text or "")
    for secret in secret_values or []:
        if secret:
            sanitized = sanitized.replace(secret, "[redacted]")
    sanitized = re.sub(r"(?i)(bearer\s+)[A-Za-z0-9._~+/=-]+", r"\1[redacted]", sanitized)
    sanitized = re.sub(
        r"(?i)((?:api[_ -]?key|authorization)\s*[:=]\s*)[^\s,;]+",
        r"\1[redacted]",
        sanitized,
    )
    return sanitized


def _normalize_bala_bailian_provider(value: object) -> str:
    provider = str(value or "happyhorse").strip().lower() or "happyhorse"
    if provider not in BALA_BAILIAN_VIDEO_PROVIDERS:
        raise HTTPException(400, "百炼视频模型必须是 HappyHorse、Kling 或 PixVerse")
    return provider


def _bala_provider_status_key(provider: str) -> str:
    provider = str(provider or "").strip().lower()
    return "happyhorse" if provider in BALA_BAILIAN_VIDEO_PROVIDERS else provider


def _bala_video_provider_label(provider: str) -> str:
    provider = str(provider or "").strip().lower()
    if provider == "seedance":
        return "Seedance"
    if provider == "happyhorse":
        return "百炼 HappyHorse"
    if provider == "kling-v3":
        return "百炼 Kling v3"
    if provider == "kling-omni":
        return "百炼 Kling Omni"
    if provider == BALA_PIXVERSE_PROVIDER:
        return "百炼 PixVerse"
    return "视频供应商"


def _bala_is_http_url(value: object) -> bool:
    raw = str(value or "").strip()
    if not raw:
        return False
    try:
        parsed = urlparse(raw)
    except Exception:
        return False
    return parsed.scheme in {"http", "https"} and bool(parsed.netloc)


def _bala_is_provider_media_url(value: object) -> bool:
    raw = str(value or "").strip()
    if not raw:
        return False
    try:
        parsed = urlparse(raw)
    except Exception:
        return False
    if parsed.scheme in {"http", "https"}:
        return bool(parsed.netloc)
    return parsed.scheme == "oss" and bool(parsed.netloc) and bool(parsed.path.strip("/"))


def _bala_bailian_upload_api_key() -> str:
    secrets = ai_video_generation_service.provider_secrets()
    return str(secrets.get("bailian_upload") or secrets.get("happyhorse") or "").strip()


def _upload_bala_bailian_asset(*, model: str, path: Path, provider_label: str) -> str:
    api_key = _bala_bailian_upload_api_key()
    if not api_key:
        raise HTTPException(400, f"{provider_label} 本地素材上传需要先在 AI 能力配置 OSS 上传百炼 key")
    try:
        return ai_video_generation_service._upload_file_to_bailian_temp_storage(
            api_key=api_key,
            model=model,
            path=path,
        )
    except ai_video_generation_service.AiVideoError as exc:
        raise HTTPException(exc.http_status, str(exc)) from exc


def _bala_preflight_media_url(path: Path) -> str:
    safe_name = re.sub(r"[^A-Za-z0-9._~-]+", "-", path.name).strip("-") or "local"
    return f"oss://bala-preflight/{safe_name}"


def _seedance_output_path(req: BalaSeedanceVideoRequest) -> Path:
    explicit = str(req.output_path or "").strip()
    if explicit:
        return Path(explicit).expanduser()
    output_dir = Path(str(req.output_dir or Path.home() / "Downloads" / "巴拉AI视频成片")).expanduser()
    style = re.sub(r"[^A-Za-z0-9._~-]+", "-", str(req.style_code or "seedance").strip()).strip("-") or "seedance"
    return output_dir / f"{style}_seedance_{datetime.now().strftime('%Y%m%d-%H%M%S')}.mp4"


def _build_seedance_payload(req: BalaSeedanceVideoRequest) -> dict:
    prompt = str(req.prompt or "").strip()
    if not prompt:
        raise HTTPException(400, "Seedance 任务需要填写 Prompt")
    content: list[dict] = [{"type": "text", "text": prompt}]
    for raw_path in req.image_paths or []:
        value = str(raw_path or "").strip()
        if not value:
            continue
        path = _validate_bala_video_provider_image_path(value, "Seedance")
        content.append({
            "type": "image_url",
            "image_url": {"url": file_to_data_url(str(path))},
            "role": "reference_image",
        })
    resolution = str(req.resolution or "720p").strip().lower() or "720p"
    if resolution not in {"480p", "720p", "1080p"}:
        raise HTTPException(400, "Seedance 清晰度仅支持 480p / 720p / 1080p")
    return {
        "model": str(req.model or BALA_SEEDANCE_DEFAULT_MODEL).strip() or BALA_SEEDANCE_DEFAULT_MODEL,
        "content": content,
        "generate_audio": bool(req.generate_audio),
        "ratio": str(req.ratio or "9:16").strip() or "9:16",
        "resolution": resolution,
        "duration": max(4, min(int(req.duration or 5), 15)),
        "watermark": bool(req.watermark),
    }


async def _run_seedance_cli(req: BalaSeedanceVideoRequest) -> dict:
    if not BALA_SEEDANCE_CLI_DIR.is_dir():
        raise HTTPException(500, "Seedance CLI 未安装到项目 integrations 目录")
    payload = _build_seedance_payload(req)
    output_path = _seedance_output_path(req)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    run_dir = runtime_paths.child_dir("bala-ai-video-seedance")
    run_dir.mkdir(parents=True, exist_ok=True)
    payload_path = run_dir / f"seedance-payload-{uuid4().hex}.json"
    payload_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    node_executable, node_env = _bala_video_node_runtime()
    command = [
        node_executable,
        "bin/seedance.js",
        "submit",
        str(payload_path),
    ]
    if req.wait:
        command.append("--wait")
    command.extend([
        "--download",
        str(output_path),
        "--interval",
        str(max(1, min(int(req.interval_seconds or 5), 60))),
        "--timeout",
        str(max(30, min(int(req.timeout_seconds or 1800), 7200))),
    ])
    env, secret_values = _bala_video_provider_env("seedance")
    env.update(node_env)
    if not str(env.get("ARK_API_KEY") or "").strip():
        return {
            "ok": False,
            "provider": "seedance",
            "status": "needs_config",
            "error": "请先在 AI 能力配置 Seedance 凭据",
        }
    process = await asyncio.create_subprocess_exec(
        *command,
        cwd=str(BALA_SEEDANCE_CLI_DIR),
        env=env,
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.PIPE,
    )
    stdout_bytes, stderr_bytes = await _communicate_video_provider_process(
        process,
        timeout_seconds=max(60, min(int(req.timeout_seconds or 1800), 7200) + 120),
        provider_label="Seedance",
    )
    stdout = stdout_bytes.decode("utf-8", errors="replace")
    stderr = stderr_bytes.decode("utf-8", errors="replace")
    objects = _parse_seedance_cli_json_objects(stdout)
    created = objects[0] if objects else {}
    final = objects[-1] if objects else {}
    if process.returncode != 0:
        sanitized = _sanitize_video_provider_output(stderr or stdout, secret_values)
        raise HTTPException(500, sanitized.strip() or "Seedance CLI 执行失败")
    return {
        "ok": True,
        "provider": "seedance",
        "task_id": str(final.get("id") or created.get("id") or ""),
        "status": str(final.get("status") or created.get("status") or ""),
        "video_url": str(((final.get("content") or {}).get("video_url")) or ""),
        "local_video_path": str(output_path) if output_path.is_file() else "",
        "payload_path": str(payload_path),
        "stderr_tail": "\n".join(_sanitize_video_provider_output(stderr, secret_values).splitlines()[-20:]),
        "raw": final or created,
    }


def _happyhorse_media_urls(req: object) -> list[str]:
    urls = [str(value or "").strip() for value in getattr(req, "image_urls", []) or []]
    urls = [value for value in urls if value]
    for raw_path in getattr(req, "image_paths", []) or []:
        value = str(raw_path or "").strip()
        if not value:
            continue
        path = _validate_bala_video_provider_image_path(value, "HappyHorse")
        urls.append(file_to_data_url(str(path)))
    return urls


def _build_happyhorse_payload(req: object) -> dict:
    mode = str(getattr(req, "mode", "i2v") or "i2v").strip().lower()
    if mode not in BALA_HAPPYHORSE_MODELS:
        raise HTTPException(400, "HappyHorse 生成方式必须是文生、图生或参考生视频")
    prompt = str(getattr(req, "prompt", "") or "").strip()
    if not prompt:
        raise HTTPException(400, "HappyHorse 任务需要填写 Prompt")

    urls = _happyhorse_media_urls(req)
    input_payload: dict = {"prompt": prompt}
    if mode == "i2v":
        if len(urls) != 1:
            raise HTTPException(400, "HappyHorse 图生视频只允许选择 1 张首帧图")
        input_payload["media"] = [{"type": "first_frame", "url": urls[0]}]
    elif mode == "r2v":
        if not 1 <= len(urls) <= 9:
            raise HTTPException(400, "HappyHorse 参考生视频需要选择 1-9 张参考图")
        input_payload["media"] = [
            {"type": "reference_image", "url": url}
            for url in urls
        ]

    resolution = str(getattr(req, "resolution", "720P") or "720P").strip().upper()
    if resolution not in {"720P", "1080P"}:
        raise HTTPException(400, "HappyHorse 清晰度只支持 720P 或 1080P")
    parameters: dict = {
        "resolution": resolution,
        "duration": max(3, min(int(getattr(req, "duration", 5) or 5), 15)),
        "watermark": bool(getattr(req, "watermark", False)),
    }
    if mode != "i2v":
        parameters["ratio"] = str(getattr(req, "ratio", "3:4") or "3:4").strip() or "3:4"
    return {
        "model": BALA_HAPPYHORSE_MODELS[mode],
        "input": input_payload,
        "parameters": parameters,
    }


def _bala_bailian_local_media_url(
    *,
    model: str,
    path: Path,
    provider_label: str,
    upload_local_assets: bool,
) -> str:
    if upload_local_assets:
        return _upload_bala_bailian_asset(
            model=model,
            path=path,
            provider_label=provider_label,
        )
    return _bala_preflight_media_url(path)


def _build_kling_payload(req: object, provider: str, *, upload_local_assets: bool = False) -> dict:
    prompt = str(getattr(req, "prompt", "") or "").strip()
    if not prompt:
        raise HTTPException(400, f"{_bala_video_provider_label(provider)} 任务需要填写 Prompt")
    model = BALA_KLING_VIDEO_MODELS[provider]
    provider_label = _bala_video_provider_label(provider)
    local_image_paths = [
        _validate_bala_video_provider_image_path(value, provider_label)
        for value in (getattr(req, "image_paths", []) or [])
        if str(value or "").strip()
    ]
    if len(local_image_paths) > 2:
        raise HTTPException(400, "Kling 本地图片最多选择 2 张（首帧 / 尾帧）")
    media: list[dict] = []
    for index, path in enumerate(local_image_paths):
        media.append({
            "type": "first_frame" if index == 0 else "last_frame",
            "url": _bala_bailian_local_media_url(
                model=model,
                path=path,
                provider_label=provider_label,
                upload_local_assets=upload_local_assets,
            ),
        })
    image_urls = [
        str(value or "").strip()
        for value in (getattr(req, "image_urls", []) or [])
        if str(value or "").strip()
    ]
    if image_urls:
        if len(local_image_paths) + len(image_urls) > 2:
            raise HTTPException(400, "Kling 图片素材最多 2 张（首帧 / 尾帧）")
        for index, url in enumerate(image_urls, start=len(local_image_paths)):
            if not _bala_is_provider_media_url(url):
                raise HTTPException(400, "Kling 图片 URL 必须是 HTTP/HTTPS 或百炼临时 oss:// URL")
            media.append({"type": "first_frame" if index == 0 else "last_frame", "url": url})
    mode = str(getattr(req, "kling_mode", "") or getattr(req, "mode", "") or "std").strip().lower()
    if mode not in BALA_KLING_MODES:
        raise HTTPException(400, "Kling 模式仅支持 std 或 pro")
    ratio = str(getattr(req, "ratio", "") or "16:9").strip() or "16:9"
    if ratio not in BALA_KLING_RATIOS:
        raise HTTPException(400, "Kling 比例仅支持 16:9 / 9:16 / 1:1")
    input_payload = {"prompt": prompt}
    if media:
        input_payload["media"] = media
    return {
        "model": model,
        "input": input_payload,
        "parameters": {
            "mode": mode,
            "aspect_ratio": ratio,
            "duration": max(3, min(int(getattr(req, "duration", 5) or 5), 15)),
            "audio": bool(getattr(req, "generate_audio", False)),
            "watermark": bool(getattr(req, "watermark", False)),
        },
    }


def _build_pixverse_payload(req: object, *, upload_local_assets: bool = False) -> dict:
    provider_label = _bala_video_provider_label(BALA_PIXVERSE_PROVIDER)
    local_image_values = [
        str(value or "").strip()
        for value in [
            getattr(req, "pixverse_image_path", ""),
            *(getattr(req, "image_paths", []) or []),
        ]
        if str(value or "").strip()
    ]
    if len(local_image_values) > 1:
        raise HTTPException(400, "PixVerse 只能提供 1 张角色图")
    local_video_values = [
        str(value or "").strip()
        for value in [
            getattr(req, "pixverse_video_path", ""),
            *(getattr(req, "video_paths", []) or []),
        ]
        if str(value or "").strip()
    ]
    if len(local_video_values) > 1:
        raise HTTPException(400, "PixVerse 只能提供 1 条动作视频")
    image_url = str(getattr(req, "pixverse_image_url", "") or "").strip()
    video_url = str(getattr(req, "pixverse_video_url", "") or "").strip()
    if image_url and not _bala_is_provider_media_url(image_url):
        raise HTTPException(400, "PixVerse 角色图 URL 必须是 HTTP/HTTPS 或百炼临时 oss:// URL")
    if video_url and not _bala_is_provider_media_url(video_url):
        raise HTTPException(400, "PixVerse 动作视频 URL 必须是 HTTP/HTTPS 或百炼临时 oss:// URL")
    if local_image_values:
        if image_url:
            raise HTTPException(400, "PixVerse 角色图请在本地上传和 URL 中二选一")
        image_path = _validate_bala_video_provider_image_path(local_image_values[0], provider_label)
        image_url = _bala_bailian_local_media_url(
            model=BALA_PIXVERSE_MODEL,
            path=image_path,
            provider_label=provider_label,
            upload_local_assets=upload_local_assets,
        )
    if local_video_values:
        if video_url:
            raise HTTPException(400, "PixVerse 动作视频请在本地上传和 URL 中二选一")
        video_path = _validate_bala_video_provider_video_path(local_video_values[0], provider_label)
        video_url = _bala_bailian_local_media_url(
            model=BALA_PIXVERSE_MODEL,
            path=video_path,
            provider_label=provider_label,
            upload_local_assets=upload_local_assets,
        )
    if not image_url or not video_url:
        raise HTTPException(400, "PixVerse 需要 1 张角色图和 1 条动作视频，可本地上传或填写 URL")
    resolution = str(getattr(req, "resolution", "720P") or "720P").strip().upper()
    if resolution not in BALA_PIXVERSE_RESOLUTIONS:
        raise HTTPException(400, "PixVerse 清晰度仅支持 360P / 540P / 720P")
    return {
        "model": BALA_PIXVERSE_MODEL,
        "input": {
            "media": [
                {"type": "image_url", "url": image_url},
                {"type": "video_url", "url": video_url},
            ],
        },
        "parameters": {
            "resolution": resolution,
            "watermark": bool(getattr(req, "watermark", False)),
        },
    }


def _build_bailian_video_payload(req: object, *, upload_local_assets: bool = False) -> dict:
    provider = _normalize_bala_bailian_provider(getattr(req, "provider", "happyhorse"))
    if provider == "happyhorse":
        return _build_happyhorse_payload(req)
    if provider in BALA_KLING_VIDEO_MODELS:
        return _build_kling_payload(req, provider, upload_local_assets=upload_local_assets)
    return _build_pixverse_payload(req, upload_local_assets=upload_local_assets)


def _preflight_bala_video_provider(req: BalaVideoProviderPreflightRequest) -> dict:
    provider = str(req.provider or "").strip().lower()
    if provider not in {"seedance", *BALA_BAILIAN_VIDEO_PROVIDERS}:
        raise HTTPException(400, "视频供应商必须是 Seedance、HappyHorse、Kling 或 PixVerse")
    status_key = _bala_provider_status_key(provider)
    status = _bala_video_provider_status()[status_key]
    if not status["configured"]:
        return {
            "ok": False,
            "provider": provider,
            "status": "needs_config",
            "configured": False,
            "source": status["source"],
            "error": f"请先在 AI 能力配置 {_bala_video_provider_label(provider)} 凭据",
        }

    cli_dir = BALA_SEEDANCE_CLI_DIR if provider == "seedance" else BALA_HAPPYHORSE_CLI_DIR
    cli_file = cli_dir / "bin" / ("seedance.js" if provider == "seedance" else "bailian.js")
    if not cli_file.is_file():
        raise HTTPException(500, "视频供应商共享能力未完整安装")

    output_dir = Path(str(req.output_dir or "").strip()).expanduser()
    if not str(req.output_dir or "").strip():
        raise HTTPException(400, "请先选择视频输出目录")
    if output_dir.exists() and not output_dir.is_dir():
        raise HTTPException(400, "视频输出路径不是文件夹")
    writable_parent = output_dir
    while not writable_parent.exists() and writable_parent != writable_parent.parent:
        writable_parent = writable_parent.parent
    if not writable_parent.is_dir() or not os.access(writable_parent, os.W_OK):
        raise HTTPException(400, "视频输出目录不可写")

    if provider == "seedance":
        payload = _build_seedance_payload(BalaSeedanceVideoRequest(
            style_code=req.style_code,
            prompt=req.prompt,
            image_paths=list(req.image_paths or []),
            output_dir=str(output_dir),
            ratio=req.ratio,
            resolution=getattr(req, "resolution", None) or "720p",
            duration=req.duration,
            generate_audio=req.generate_audio,
            watermark=bool(getattr(req, "watermark", False)),
            wait=False,
        ))
        mode = "reference"
    else:
        payload = _build_bailian_video_payload(BalaHappyHorseVideoRequest(
            provider=provider,
            style_code=req.style_code,
            mode=req.mode,
            prompt=req.prompt,
            image_paths=list(req.image_paths or []),
            video_paths=list(req.video_paths or []),
            output_dir=str(output_dir),
            resolution=req.resolution,
            ratio=req.ratio,
            duration=req.duration,
            generate_audio=req.generate_audio,
            kling_mode=req.kling_mode,
            pixverse_image_path=req.pixverse_image_path,
            pixverse_video_path=req.pixverse_video_path,
            pixverse_image_url=req.pixverse_image_url,
            pixverse_video_url=req.pixverse_video_url,
            watermark=req.watermark,
            wait=False,
        ), upload_local_assets=False)
        mode = (
            str(req.mode or "i2v").strip().lower()
            if provider == "happyhorse"
            else str(req.kling_mode or "std").strip().lower()
            if provider in BALA_KLING_VIDEO_MODELS
            else "motioncontrol"
        )
    return {
        "ok": True,
        "provider": provider,
        "status": "ready",
        "configured": True,
        "source": status["source"],
        "mode": mode,
        "model": str(payload.get("model") or ""),
        "image_count": len(req.image_paths or []),
        "output_dir": str(output_dir),
    }


def _happyhorse_output_path(req: BalaHappyHorseVideoRequest) -> Path:
    return _bailian_output_path(req)


def _bailian_output_path(req: BalaHappyHorseVideoRequest) -> Path:
    explicit = str(req.output_path or "").strip()
    if explicit:
        return Path(explicit).expanduser()
    output_dir = Path(str(req.output_dir or Path.home() / "Downloads" / "巴拉AI视频成片")).expanduser()
    provider = _normalize_bala_bailian_provider(getattr(req, "provider", "happyhorse"))
    style = re.sub(r"[^A-Za-z0-9._~-]+", "-", str(req.style_code or provider).strip()).strip("-") or provider
    if provider == "happyhorse":
        mode = str(req.mode or "i2v").strip().lower() or "i2v"
    elif provider in BALA_KLING_VIDEO_MODELS:
        mode = str(req.kling_mode or "std").strip().lower() or "std"
    else:
        mode = "motioncontrol"
    provider_slug = re.sub(r"[^a-z0-9]+", "-", provider).strip("-") or "bailian"
    mode_slug = re.sub(r"[^a-z0-9]+", "-", mode).strip("-") or "video"
    return output_dir / f"{style}_{provider_slug}_{mode_slug}_{datetime.now().strftime('%Y%m%d-%H%M%S')}.mp4"


async def _run_happyhorse_cli(req: BalaHappyHorseVideoRequest) -> dict:
    if not BALA_HAPPYHORSE_CLI_DIR.is_dir():
        raise HTTPException(500, "百炼视频能力未安装到项目共享目录")
    provider = _normalize_bala_bailian_provider(getattr(req, "provider", "happyhorse"))
    provider_label = _bala_video_provider_label(provider)
    node_executable, node_env = _bala_video_node_runtime()
    env, secret_values = _bala_video_provider_env(provider)
    env.update(node_env)
    if not str(env.get("DASHSCOPE_API_KEY") or "").strip():
        return {
            "ok": False,
            "provider": provider,
            "status": "needs_config",
            "error": f"请先在 AI 能力配置 {provider_label} 凭据",
        }
    payload = _build_bailian_video_payload(req, upload_local_assets=True)
    output_path = _bailian_output_path(req)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    run_dir = runtime_paths.child_dir("bala-ai-video-bailian")
    run_dir.mkdir(parents=True, exist_ok=True)
    provider_slug = re.sub(r"[^a-z0-9]+", "-", provider).strip("-") or "bailian"
    payload_path = run_dir / f"{provider_slug}-payload-{uuid4().hex}.json"
    payload_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    command = [node_executable, "bin/bailian.js", "submit", str(payload_path)]
    if req.wait:
        command.append("--wait")
    command.extend([
        "--download",
        str(output_path),
        "--interval",
        str(max(1, min(int(req.interval_seconds or 5), 60))),
        "--timeout",
        str(max(30, min(int(req.timeout_seconds or 1800), 7200))),
    ])
    process = await asyncio.create_subprocess_exec(
        *command,
        cwd=str(BALA_HAPPYHORSE_CLI_DIR),
        env=env,
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.PIPE,
    )
    stdout_bytes, stderr_bytes = await _communicate_video_provider_process(
        process,
        timeout_seconds=max(60, min(int(req.timeout_seconds or 1800), 7200) + 120),
        provider_label=provider_label,
    )
    stdout = stdout_bytes.decode("utf-8", errors="replace")
    stderr = stderr_bytes.decode("utf-8", errors="replace")
    objects = _parse_seedance_cli_json_objects(stdout)
    created = objects[0] if objects else {}
    final = objects[-1] if objects else {}
    if process.returncode != 0:
        sanitized = _sanitize_video_provider_output(stderr or stdout, secret_values)
        raise HTTPException(500, sanitized.strip() or f"{provider_label} CLI 执行失败")
    created_output = created.get("output") if isinstance(created.get("output"), dict) else {}
    final_output = final.get("output") if isinstance(final.get("output"), dict) else {}
    return {
        "ok": True,
        "provider": provider,
        "mode": (
            str(req.mode or "i2v").strip().lower()
            if provider == "happyhorse"
            else str(req.kling_mode or "std").strip().lower()
            if provider in BALA_KLING_VIDEO_MODELS
            else "motioncontrol"
        ),
        "task_id": str(final_output.get("task_id") or created_output.get("task_id") or ""),
        "status": str(final_output.get("task_status") or created_output.get("task_status") or ""),
        "video_url": str(final_output.get("video_url") or ""),
        "local_video_path": str(output_path) if output_path.is_file() else "",
        "payload_path": str(payload_path),
        "stderr_tail": "\n".join(_sanitize_video_provider_output(stderr, secret_values).splitlines()[-20:]),
        "raw": final or created,
    }


def _video_provider_task_output_path(req: BalaVideoProviderTaskRequest) -> Path:
    explicit = str(req.output_path or "").strip()
    if explicit:
        return Path(explicit).expanduser()
    root = Path(str(req.output_dir or Path.home() / "Downloads" / "巴拉AI视频成片")).expanduser()
    style = re.sub(r"[^A-Za-z0-9._~-]+", "-", str(req.style_code or "video").strip()).strip("-") or "video"
    provider = re.sub(r"[^a-z0-9]+", "-", str(req.provider or "provider").strip().lower()).strip("-") or "provider"
    task_id = re.sub(r"[^A-Za-z0-9._~-]+", "-", str(req.task_id or "task").strip()).strip("-") or "task"
    return root / f"{style}_{provider}_{task_id}.mp4"


async def _run_video_provider_task_cli(req: BalaVideoProviderTaskRequest) -> dict:
    provider = str(req.provider or "").strip().lower()
    task_id = str(req.task_id or "").strip()
    if provider not in {"seedance", *BALA_BAILIAN_VIDEO_PROVIDERS}:
        raise HTTPException(400, "视频供应商必须是 Seedance、HappyHorse、Kling 或 PixVerse")
    if not task_id:
        raise HTTPException(400, "刷新视频状态需要 provider task ID")

    cli_dir = BALA_SEEDANCE_CLI_DIR if provider == "seedance" else BALA_HAPPYHORSE_CLI_DIR
    cli_file = "bin/seedance.js" if provider == "seedance" else "bin/bailian.js"
    if not cli_dir.is_dir():
        raise HTTPException(500, "视频供应商共享 CLI 未安装")

    node_executable, node_env = _bala_video_node_runtime()
    output_path = _video_provider_task_output_path(req)
    command = [node_executable, cli_file, "wait" if req.download else "get", task_id]
    if req.download:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        command.extend([
            "--download",
            str(output_path),
            "--interval",
            str(max(1, min(int(req.interval_seconds or 5), 60))),
            "--timeout",
            str(max(30, min(int(req.timeout_seconds or 1800), 7200))),
        ])

    status_key = _bala_provider_status_key(provider)
    provider_label = _bala_video_provider_label(provider)
    env, secret_values = _bala_video_provider_env(provider)
    env.update(node_env)
    required_key = "ARK_API_KEY" if status_key == "seedance" else "DASHSCOPE_API_KEY"
    if not str(env.get(required_key) or "").strip():
        return {
            "ok": False,
            "provider": provider,
            "task_id": task_id,
            "status": "needs_config",
            "error": f"请先在 AI 能力配置 {provider_label} 凭据",
        }

    process = await asyncio.create_subprocess_exec(
        *command,
        cwd=str(cli_dir),
        env=env,
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.PIPE,
    )
    timeout = max(60, min(int(req.timeout_seconds or 1800), 7200) + 120) if req.download else 120
    stdout_bytes, stderr_bytes = await _communicate_video_provider_process(
        process,
        timeout_seconds=timeout,
        provider_label=provider_label,
    )
    stdout = stdout_bytes.decode("utf-8", errors="replace")
    stderr = stderr_bytes.decode("utf-8", errors="replace")
    objects = _parse_seedance_cli_json_objects(stdout)
    final = objects[-1] if objects else {}
    if process.returncode != 0:
        sanitized = _sanitize_video_provider_output(stderr or stdout, secret_values)
        raise HTTPException(500, sanitized.strip() or "视频供应商任务刷新失败")

    if status_key == "seedance":
        final_task_id = str(final.get("id") or task_id)
        status = str(final.get("status") or "")
        content = final.get("content") if isinstance(final.get("content"), dict) else {}
        video_url = str(content.get("video_url") or "")
    else:
        output = final.get("output") if isinstance(final.get("output"), dict) else {}
        final_task_id = str(output.get("task_id") or task_id)
        status = str(output.get("task_status") or "")
        video_url = str(output.get("video_url") or "")

    return {
        "ok": True,
        "provider": provider,
        "task_id": final_task_id,
        "status": status,
        "video_url": video_url,
        "local_video_path": str(output_path) if req.download and output_path.is_file() else "",
        "stderr_tail": "\n".join(_sanitize_video_provider_output(stderr, secret_values).splitlines()[-20:]),
        "raw": final,
    }


def _bala_material_asset_path(batch: dict, asset_id: str) -> Path:
    target = str(asset_id or "").strip()
    for item in batch.get("items") or []:
        if not isinstance(item, dict):
            continue
        for asset in item.get("assets") or []:
            if not isinstance(asset, dict):
                continue
            if str(asset.get("id") or "").strip() == target:
                path = Path(str(asset.get("path") or "")).expanduser()
                if not path.is_file():
                    raise HTTPException(404, "Bala material image not found")
                return path
    raise HTTPException(404, "Bala material asset not found")


def _bala_review_asset_path(batch: dict, asset_id: str) -> Path:
    target = str(asset_id or "").strip()
    for item in batch.get("items") or []:
        if not isinstance(item, dict):
            continue
        for asset in item.get("assets") or []:
            if not isinstance(asset, dict):
                continue
            if str(asset.get("id") or "").strip() == target:
                path = Path(str(asset.get("path") or asset.get("source_path") or "")).expanduser()
                if not path.is_file():
                    raise HTTPException(404, "Bala review image not found")
                return path
    raise HTTPException(404, "Bala review asset not found")


def _load_bala_review_batch_checked(batch_id: str, token: str) -> dict:
    try:
        batch = bala_ai_video_review.load_review_batch(batch_id)
        bala_ai_video_review.validate_token(batch, token)
        return batch
    except PermissionError as exc:
        raise HTTPException(403, str(exc)) from exc
    except FileNotFoundError as exc:
        raise HTTPException(404, str(exc)) from exc


def _mutate_bala_review_batch_checked(batch_id: str, token: str, mutation: Callable[[dict], dict | None]) -> dict:
    try:
        return bala_ai_video_review.mutate_review_batch(batch_id, token, mutation)
    except PermissionError as exc:
        raise HTTPException(403, str(exc)) from exc
    except FileNotFoundError as exc:
        raise HTTPException(404, str(exc)) from exc


def _bala_review_find_asset(batch: dict, asset_id: str) -> tuple[dict, dict]:
    target = str(asset_id or "").strip()
    for item in batch.get("items") or []:
        if not isinstance(item, dict):
            continue
        for asset in item.get("assets") or []:
            if isinstance(asset, dict) and str(asset.get("id") or "").strip() == target:
                return item, asset
    raise HTTPException(404, "Bala review asset not found")


def _bala_review_reference_paths(asset: dict, requested: Optional[List[str]]) -> list[str]:
    if requested is not None:
        source = requested
    else:
        source = [
            asset.get("model_image_path"),
            *(asset.get("garment_paths") or []),
            *(asset.get("outfit_reference_paths") or []),
            *(asset.get("variant_reference_paths") or []),
        ]
    paths: list[str] = []
    for raw in source or []:
        value = str(raw or "").strip()
        if not value:
            continue
        path = Path(value).expanduser()
        if path.is_file() and path.suffix.lower() in BALA_AI_IMAGE_EXTS:
            paths.append(str(path))
    return paths


def _append_bala_review_regenerate_asset(batch: dict, req: BalaReviewRegenerateRequest) -> dict:
    item, source_asset = _bala_review_find_asset(batch, req.asset_id)
    source_path = str(source_asset.get("source_path") or source_asset.get("path") or "").strip()
    if not source_path or not Path(source_path).expanduser().is_file():
        raise HTTPException(400, "重跑缺少可用源图")
    operation_type = bala_ai_video_review.normalize_operation_type(source_asset.get("operation_type"))
    prompt = str(req.prompt or source_asset.get("prompt") or source_asset.get("background_prompt") or source_asset.get("pose_prompt") or "").strip()
    if not prompt:
        prompt = _bala_ai_operation_prompt(Path(source_path).expanduser(), {}, {
            "background_prompt": source_asset.get("background_prompt") or "",
            "pose_prompt": source_asset.get("pose_prompt") or "",
            "prompt_extra": source_asset.get("review_note") or "",
        }, operation_type)
    reference_paths = _bala_review_reference_paths(source_asset, req.reference_paths)
    title = _safe_local_name(
        f"巴拉审核重跑 {item.get('style_code') or ''} {operation_type}",
        "巴拉审核重跑",
    )
    job = data_sink.create_ai_image_job({
        "title": title,
        "prompt": prompt,
        "model_key": "gpt-image-2",
        "status": "draft",
        "output_dir": str(Path(str(batch.get("artifact_dir") or runtime_paths.child_dir("bala-ai-video-review"))).expanduser() / "regenerated"),
        "params": {
            "prompt": prompt,
            "main_image_path": source_path,
            "reference_image_paths": reference_paths,
            "workflow": BALA_AI_FACE_BACKGROUND_TASK_ID,
            "surface": "ai-video-workflow",
            "workspace_dir": str(batch.get("workspace_dir") or ""),
            "operation_type": operation_type,
            "style_code": item.get("style_code") or "",
        },
        "summary": {
            "workflow": BALA_AI_FACE_BACKGROUND_TASK_ID,
            "surface": "ai-video-workflow",
            "workspace_dir": str(batch.get("workspace_dir") or ""),
            "operation_type": operation_type,
            "source_path": source_path,
            "regenerate_from_asset_id": source_asset.get("id"),
        },
    })
    job_uid = str(job.get("job_uid") or "")
    data_sink.create_ai_image_asset({
        "job_uid": job_uid,
        "kind": "main",
        "source_type": "local",
        "path": source_path,
        "sort_order": 0,
        "meta": {"role": "source_product_model", "style_code": item.get("style_code") or ""},
    })
    for index, path in enumerate(reference_paths, start=1):
        data_sink.create_ai_image_asset({
            "job_uid": job_uid,
            "kind": "reference",
            "source_type": "local",
            "path": path,
            "sort_order": index,
            "meta": {"role": "reference"},
        })

    submit_note = "已创建重跑任务，未提交"
    retry_status = "failed"
    run_uid = ""
    if req.submit_async:
        try:
            run_result = ai_image_service.submit_workbench_batch(
                job_uid,
                [{"title": title, "prompt": prompt, "count": 1}],
                request_uid=f"bala-review-regenerate-{uuid4().hex}",
            )
            first_run = next(iter(run_result.get("runs") or []), {})
            run_uid = str(first_run.get("run_uid") or "")
            if run_result.get("accepted") and run_uid:
                retry_status = "generating"
                submit_note = "已异步提交重跑任务"
            else:
                submit_note = "重跑提交未接受" if not run_result.get("accepted") else "重跑提交未返回运行记录"
        except ai_image_service.MissingModelKeyError as exc:
            submit_note = str(exc)
        except Exception as exc:
            submit_note = str(exc)

    new_asset_id = f"{source_asset.get('id') or 'asset'}-retry-{secrets.token_hex(3)}"
    token = str(batch.get("token") or "")
    item.setdefault("assets", []).append({
        "id": new_asset_id,
        "kind": "ai",
        "style_code": item.get("style_code") or "",
        "source_asset_id": source_asset.get("source_asset_id") or source_asset.get("id") or "",
        "source_path": source_path,
        "path": "",
        "filename": "",
        "status": retry_status,
        "operation_type": operation_type,
        "job_uid": job_uid,
        "run_uid": run_uid,
        "prompt": prompt,
        "image_url": f"{_bala_ai_video_base_url()}/bala-ai-video-review/api/{batch.get('batch_id')}/image/{new_asset_id}?token={token}",
        "review_note": submit_note,
        "regenerated_from_asset_id": source_asset.get("id") or "",
    })
    return bala_ai_video_review.refresh_generated_assets(batch)


@app.post("/bala-ai-video-materials/api/from-rows")
def create_bala_ai_video_material_batch(req: BalaMaterialBatchFromRowsRequest):
    batch = bala_ai_video_materials.build_material_batch(
        list(req.rows or []),
        str(runtime_paths.child_dir("bala-ai-video-materials")),
        _bala_ai_video_base_url(),
    )
    batch["source_task"] = dict(req.source_task or {})
    bala_ai_video_materials.save_material_batch(batch)
    return batch


@app.get("/bala-ai-video-materials/api/{batch_id}")
def get_bala_ai_video_material_batch(batch_id: str, token: str = ""):
    try:
        batch = bala_ai_video_materials.load_material_batch(batch_id)
        bala_ai_video_materials.validate_token(batch, token)
        return bala_ai_video_materials.attach_material_preview_urls(batch, _bala_ai_video_base_url())
    except PermissionError as exc:
        raise HTTPException(403, str(exc))
    except FileNotFoundError as exc:
        raise HTTPException(404, str(exc))


@app.get("/bala-ai-video-materials/api/{batch_id}/image/{asset_id}")
def get_bala_ai_video_material_image(batch_id: str, asset_id: str, token: str = ""):
    try:
        batch = bala_ai_video_materials.load_material_batch(batch_id)
        bala_ai_video_materials.validate_token(batch, token)
        path = _bala_material_asset_path(batch, asset_id)
    except PermissionError as exc:
        raise HTTPException(403, str(exc))
    except FileNotFoundError as exc:
        raise HTTPException(404, str(exc))
    media_type = mimetypes.guess_type(str(path))[0] or "image/jpeg"
    return FileResponse(path, media_type=media_type, filename=path.name)


@app.get("/bala-ai-video-materials/api/{batch_id}/thumbnail/{asset_id}")
def get_bala_ai_video_material_thumbnail(batch_id: str, asset_id: str, token: str = ""):
    try:
        batch = bala_ai_video_materials.load_material_batch(batch_id)
        bala_ai_video_materials.validate_token(batch, token)
        source_path = _bala_material_asset_path(batch, asset_id)
        thumbnail = bala_ai_video_materials.ensure_material_thumbnail(batch, asset_id, source_path)
    except PermissionError as exc:
        raise HTTPException(403, str(exc))
    except FileNotFoundError as exc:
        raise HTTPException(404, str(exc))
    except OSError as exc:
        raise HTTPException(422, "素材缩略图生成失败") from exc
    return FileResponse(
        thumbnail,
        media_type="image/webp",
        filename=thumbnail.name,
        headers={"Cache-Control": "private, max-age=31536000, immutable"},
    )


@app.post("/bala-ai-video-materials/api/{batch_id}/selection")
def save_bala_ai_video_material_selection(batch_id: str, req: BalaMaterialSelectionRequest, token: str = ""):
    try:
        batch = bala_ai_video_materials.load_material_batch(batch_id)
        bala_ai_video_materials.validate_token(batch, token)
        updated = bala_ai_video_materials.update_material_selection(batch, list(req.selected_asset_ids or []))
        bala_ai_video_materials.save_material_batch(updated)
        return updated
    except PermissionError as exc:
        raise HTTPException(403, str(exc))
    except FileNotFoundError as exc:
        raise HTTPException(404, str(exc))


@app.post("/bala-ai-video-materials/api/{batch_id}/export-ai-input")
def export_bala_ai_video_material_ai_input(batch_id: str, req: BalaMaterialExportAiInputRequest, token: str = ""):
    try:
        batch = bala_ai_video_materials.load_material_batch(batch_id)
        bala_ai_video_materials.validate_token(batch, token)
        return bala_ai_video_materials.export_ai_input(
            batch,
            req.operation_type,
            _model_payload(req),
        )
    except PermissionError as exc:
        raise HTTPException(403, str(exc))
    except FileNotFoundError as exc:
        raise HTTPException(404, str(exc))


@app.get("/bala-ai-video-model-library/api")
def list_bala_ai_model_library(age_label: str = "", gender: str = "", group: str = "", search: str = ""):
    data = bala_ai_model_library.load_model_library()
    return bala_ai_model_library.filter_model_library(
        data,
        age_label=age_label,
        gender=gender,
        group=group,
        search=search,
    )


@app.get("/bala-ai-video-templates/api")
def list_bala_ai_video_templates(search: str = "", template_type: str = ""):
    payload = _load_bala_video_template_catalog()
    keyword = str(search or "").strip().lower()
    desired_type = str(template_type or "").strip()
    templates = []
    for item in payload.get("templates") or []:
        haystack = " ".join([
            str(item.get("templateId") or ""),
            str(item.get("title") or ""),
            str(item.get("description") or ""),
            str(item.get("slotDescription") or ""),
            str(item.get("type") or ""),
        ]).lower()
        if desired_type and str(item.get("type") or "") != desired_type:
            continue
        if keyword and keyword not in haystack:
            continue
        templates.append(item)
    return {
        **payload,
        "templates": templates,
        "count": len(templates),
    }


@app.post("/bala-ai-video-seedance/api/run")
async def run_bala_ai_video_seedance(req: BalaSeedanceVideoRequest):
    return await _run_seedance_cli(req)


@app.get("/bala-ai-video-providers/api/status")
def get_bala_ai_video_provider_status():
    return _bala_video_provider_status()


@app.post("/bala-ai-video-providers/api/preflight")
def preflight_bala_ai_video_provider(req: BalaVideoProviderPreflightRequest):
    return _preflight_bala_video_provider(req)


@app.post("/bala-ai-video-providers/api/task")
async def refresh_bala_ai_video_provider_task(req: BalaVideoProviderTaskRequest):
    return await _run_video_provider_task_cli(req)


@app.post("/bala-ai-video-happyhorse/api/run")
async def run_bala_ai_video_happyhorse(req: BalaHappyHorseVideoRequest):
    return await _run_happyhorse_cli(req)


@app.get("/bala-ai-video-model-library/api/image/{model_id:path}")
def get_bala_ai_model_library_image(model_id: str):
    try:
        path = bala_ai_model_library.resolve_model_image_path(unquote(str(model_id or "")))
    except ValueError as exc:
        raise HTTPException(400, str(exc))
    except FileNotFoundError as exc:
        raise HTTPException(404, str(exc))
    media_type = mimetypes.guess_type(str(path))[0] or "image/jpeg"
    return FileResponse(path, media_type=media_type, filename=path.name)


@app.get("/bala-ai-video-review/api/{batch_id}")
def get_bala_ai_video_review_batch(batch_id: str, token: str = ""):
    return _load_bala_review_batch_checked(batch_id, token)


@app.get("/bala-ai-video-review-workspace/api/batches")
def list_bala_ai_video_review_workspace_batches(style_codes: str = "", workspace_dir: str = "", limit: int = 100):
    requested_styles = [
        value.strip()
        for value in re.split(r"[,，\s]+", str(style_codes or ""))
        if value.strip()
    ]
    return {
        "items": bala_ai_video_review.list_review_batches(
            style_codes=requested_styles,
            workspace_dir=workspace_dir,
            limit=limit,
        ),
    }


@app.get("/bala-ai-video-review/api/{batch_id}/image/{asset_id}")
def get_bala_ai_video_review_image(batch_id: str, asset_id: str, token: str = ""):
    batch = _load_bala_review_batch_checked(batch_id, token)
    path = _bala_review_asset_path(batch, asset_id)
    media_type = mimetypes.guess_type(str(path))[0] or "image/jpeg"
    return FileResponse(path, media_type=media_type, filename=path.name)


@app.post("/bala-ai-video-review/api/{batch_id}/decisions")
def save_bala_ai_video_review_decisions(batch_id: str, req: BalaReviewDecisionsRequest, token: str = ""):
    decisions = dict(req.decisions or {})
    return _mutate_bala_review_batch_checked(
        batch_id,
        token,
        lambda batch: bala_ai_video_review.update_review_decisions(batch, decisions),
    )


@app.delete("/bala-ai-video-review/api/{batch_id}/asset/{asset_id}")
def delete_bala_ai_video_review_asset(batch_id: str, asset_id: str, token: str = ""):
    try:
        return _mutate_bala_review_batch_checked(
            batch_id,
            token,
            lambda batch: bala_ai_video_review.remove_review_asset(batch, asset_id),
        )
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    except KeyError as exc:
        raise HTTPException(404, "Bala review asset not found") from exc


@app.post("/bala-ai-video-review/api/{batch_id}/refresh")
def refresh_bala_ai_video_review_batch(batch_id: str, token: str = ""):
    return _mutate_bala_review_batch_checked(
        batch_id,
        token,
        bala_ai_video_review.refresh_generated_assets,
    )


@app.post("/bala-ai-video-review/api/{batch_id}/regenerate")
async def regenerate_bala_ai_video_review_asset(batch_id: str, req: BalaReviewRegenerateRequest, token: str = ""):
    updated = await asyncio.to_thread(
        _mutate_bala_review_batch_checked,
        batch_id,
        token,
        lambda batch: _append_bala_review_regenerate_asset(batch, req),
    )
    matching_retry_ids = [
        str(candidate.get("id") or "")
        for item in updated.get("items") or []
        for candidate in item.get("assets") or []
        if str(candidate.get("regenerated_from_asset_id") or "") == str(req.asset_id or "").strip()
    ]
    if not matching_retry_ids:
        raise HTTPException(400, "重跑任务已创建但未找到回写资产")
    _, asset = _bala_review_find_asset(updated, matching_retry_ids[-1])
    return {"ok": True, "batch": updated, "asset": asset}


@app.post("/bala-ai-video-review/api/{batch_id}/export-video-input")
def export_bala_ai_video_review_video_input(batch_id: str, req: BalaReviewExportVideoInputRequest, token: str = ""):
    batch = _load_bala_review_batch_checked(batch_id, token)
    manifest = bala_ai_video_review.export_video_input_manifest(
        batch,
        req.output_dir,
        provider=req.provider or batch.get("video_provider") or "qn_img2video",
    )
    params = bala_ai_video_review.build_qn_img2video_initial_params(manifest)
    if req.template_id:
        params["template_id"] = req.template_id
    if req.template_match:
        params["template_match"] = req.template_match
    if req.prompt:
        params["prompt"] = req.prompt
        params["custom_prompt"] = req.prompt
    params["download_template_previews"] = True
    params["download_videos"] = bool(req.download_videos)
    return {
        "ok": True,
        **manifest,
        "images": manifest.get("image_paths") or [],
        "next_task": {
            "adapter_id": BALA_AI_VIDEO_ADAPTER_ID,
            "task_id": "qn_img2video_batch",
            "params": params,
        },
    }


@app.get("/bala-ai-video-review/{batch_id}")
def get_bala_ai_video_review_board(batch_id: str, token: str = ""):
    batch = _load_bala_review_batch_checked(batch_id, token)
    title = html.escape(f"巴拉 AI 图片审核 - {batch.get('batch_id') or batch_id}", quote=True)
    rows = []
    for item in batch.get("items") or []:
        style_code = html.escape(str(item.get("style_code") or ""), quote=True)
        for asset in item.get("assets") or []:
            if not isinstance(asset, dict) or asset.get("kind") != "ai":
                continue
            rows.append(
                "<tr>"
                f"<td>{style_code}</td>"
                f"<td>{html.escape(str(asset.get('operation_type') or ''), quote=True)}</td>"
                f"<td>{html.escape(str(asset.get('status') or ''), quote=True)}</td>"
                f"<td>{html.escape(str(asset.get('filename') or Path(str(asset.get('path') or '')).name), quote=True)}</td>"
                "</tr>"
            )
    body = "\n".join(rows) or "<tr><td colspan=\"4\">暂无生成图片</td></tr>"
    return HTMLResponse(f"""<!doctype html>
<html lang=\"zh-CN\">
<head><meta charset=\"utf-8\"><title>{title}</title></head>
<body>
<h1>{title}</h1>
<p>请在抓虾客户端中打开审核抽屉完成选择、重试和进入视频生成。</p>
<table border=\"1\" cellspacing=\"0\" cellpadding=\"6\">
<thead><tr><th>款号</th><th>操作</th><th>状态</th><th>文件</th></tr></thead>
<tbody>{body}</tbody>
</table>
</body>
</html>""")


class AiImageJobRequest(BaseModel):
    title: str = ""
    prompt: str = ""
    model_key: str = "gpt-image-2"
    status: str = "draft"
    output_dir: str = ""
    params: dict = {}
    summary: dict = {}


class AiImageJobPatchRequest(BaseModel):
    title: Optional[str] = None
    prompt: Optional[str] = None
    model_key: Optional[str] = None
    status: Optional[str] = None
    output_dir: Optional[str] = None
    params: Optional[dict] = None
    summary: Optional[dict] = None


class AiImagePinRequest(BaseModel):
    pinned: bool


class AiImageBatchPromptRequest(BaseModel):
    title: str = ""
    prompt: str = ""
    count: int = 1


class AiImageBatchRunRequest(BaseModel):
    request_uid: str = ""
    prompts: list[AiImageBatchPromptRequest] = []


class AiImageAssetRequest(BaseModel):
    job_uid: str
    kind: str = "reference"
    source_type: str = "local"
    path: str = ""
    url: str = ""
    mime_type: str = ""
    sort_order: int = 0
    meta: dict = {}


class AiImageCanvasRequest(BaseModel):
    job_uid: str
    title: str = ""
    canvas: dict = {}


class AiImageSaveAsRequest(BaseModel):
    directory: str
    files: list[str] = []


class AiImageMaterializeRequest(BaseModel):
    file: str = ""
    url: str = ""
    data_url: str = ""
    filename: str = ""


class LocalImagePreviewRequest(BaseModel):
    path: str


def _model_payload(model: BaseModel, **kwargs) -> dict:
    if hasattr(model, "model_dump"):
        return model.model_dump(**kwargs)
    return model.dict(**kwargs)


def _resolve_local_preview_path(raw_path: str) -> Path:
    value = str(raw_path or "").strip()
    if not value:
        raise HTTPException(400, "图片路径不能为空")
    if value.lower().startswith("file://"):
        parsed = urlparse(value)
        value = unquote(parsed.path or "")
    path = Path(value).expanduser()
    return path if path.is_absolute() else path.resolve(strict=False)


def _local_image_mime(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in {".jpg", ".jpeg"}:
        return "image/jpeg"
    if suffix == ".png":
        return "image/png"
    if suffix == ".webp":
        return "image/webp"
    if suffix == ".gif":
        return "image/gif"
    guessed = mimetypes.guess_type(str(path))[0] or ""
    return guessed if guessed in {"image/jpeg", "image/png", "image/webp", "image/gif"} else ""


@app.post("/files/local-image-preview")
def read_local_image_preview(req: LocalImagePreviewRequest):
    path = _resolve_local_preview_path(req.path)
    mime = _local_image_mime(path)
    if not mime:
        raise HTTPException(400, "请选择 PNG、JPG、WEBP 或 GIF 图片")
    try:
        stat = path.stat()
    except FileNotFoundError as exc:
        raise HTTPException(404, "图片文件不存在") from exc
    if not path.is_file():
        raise HTTPException(400, "图片文件不存在")
    if stat.st_size > 25 * 1024 * 1024:
        raise HTTPException(400, "图片超过 25MB，无法预览")
    data_url = f"data:{mime};base64,{base64.b64encode(path.read_bytes()).decode('ascii')}"
    return {"ok": True, "path": str(path), "data_url": data_url}


@app.get("/ai-image/jobs")
def list_ai_image_jobs():
    return data_sink.list_ai_image_jobs()


@app.post("/ai-image/jobs")
def create_ai_image_job(req: AiImageJobRequest):
    payload = _model_payload(req, exclude_none=True)
    payload.pop("summary", None)
    return data_sink.create_ai_image_job(payload)


@app.get("/ai-image/jobs/{job_uid}")
def get_ai_image_job(job_uid: str):
    job = data_sink.get_ai_image_job(job_uid)
    if not job:
        raise HTTPException(404, f"AI image job not found: {job_uid}")
    job["assets"] = data_sink.list_ai_image_assets(job_uid)
    job["canvases"] = data_sink.list_ai_image_canvases(job_uid)
    return job


@app.patch("/ai-image/jobs/{job_uid}")
def update_ai_image_job(job_uid: str, req: AiImageJobPatchRequest):
    payload = _model_payload(req, exclude_unset=True, exclude_none=True)
    payload.pop("summary", None)
    job = data_sink.update_ai_image_job(job_uid, payload)
    if not job:
        raise HTTPException(404, f"AI image job not found: {job_uid}")
    return job


@app.patch("/ai-image/jobs/{job_uid}/pin")
def pin_ai_image_job(job_uid: str, req: AiImagePinRequest):
    job = data_sink.set_ai_image_job_pinned(job_uid, req.pinned)
    if not job:
        raise HTTPException(404, f"AI image job not found: {job_uid}")
    return job


@app.delete("/ai-image/jobs/{job_uid}")
def delete_ai_image_job(job_uid: str):
    try:
        return ai_image_service.delete_workbench_job(job_uid)
    except ai_image_service.ActiveAiImageJobError as exc:
        raise HTTPException(409, str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(404, str(exc)) from exc
    except RuntimeError as exc:
        raise HTTPException(500, str(exc)) from exc


@app.post("/ai-image/jobs/{job_uid}/run")
def run_ai_image_job(job_uid: str):
    if not data_sink.get_ai_image_job(job_uid):
        raise HTTPException(404, f"AI image job not found: {job_uid}")
    try:
        return ai_image_service.run_job_with_one_xm(job_uid)
    except ai_image_service.MissingModelKeyError as exc:
        raise HTTPException(400, {
            "message": str(exc),
            "config_id": exc.config_id,
        }) from exc
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc


@app.post("/ai-image/jobs/{job_uid}/batch-run")
def batch_run_ai_image_job(job_uid: str, req: AiImageBatchRunRequest):
    if not data_sink.get_ai_image_job(job_uid):
        raise HTTPException(404, f"AI image job not found: {job_uid}")
    prompts = [_model_payload(item) for item in req.prompts]
    if not 1 <= len(prompts) <= 100:
        raise HTTPException(400, "批量 Prompt 数量必须为 1–100 条")
    try:
        return ai_image_service.submit_workbench_batch(
            job_uid,
            prompts,
            request_uid=req.request_uid,
        )
    except ai_image_service.MissingModelKeyError as exc:
        raise HTTPException(400, {
            "message": str(exc),
            "config_id": exc.config_id,
        }) from exc
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc


@app.post("/ai-image/jobs/{job_uid}/runs/{run_uid}/retry")
def retry_ai_image_run(job_uid: str, run_uid: str):
    if not data_sink.get_ai_image_job(job_uid):
        raise HTTPException(404, f"AI image job not found: {job_uid}")
    try:
        return ai_image_service.retry_workbench_run(job_uid, run_uid)
    except ai_image_service.MissingModelKeyError as exc:
        raise HTTPException(400, {
            "message": str(exc),
            "config_id": exc.config_id,
        }) from exc
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc


@app.post("/ai-image/jobs/{job_uid}/save-as")
def save_as_ai_image_job(job_uid: str, req: AiImageSaveAsRequest):
    job = data_sink.get_ai_image_job(job_uid)
    if not job:
        raise HTTPException(404, f"AI image job not found: {job_uid}")
    summary = job.get("summary") if isinstance(job.get("summary"), dict) else {}
    output_assets = [{"path": str(path)} for path in summary.get("output_files") or []]
    output_assets.extend({"url": str(url)} for url in summary.get("image_urls") or [])
    output_assets.extend(data_sink.list_ai_image_assets(job_uid))
    requested_files = {str(path or "").strip() for path in (req.files or []) if str(path or "").strip()}
    if requested_files:
        output_assets = [
            asset for asset in output_assets
            if str(asset.get("path") or asset.get("url") or "").strip() in requested_files
        ]
    files = ai_image_service.copy_assets_to_directory(output_assets, req.directory)
    return {"ok": True, "job_uid": job_uid, "files": files}


@app.post("/ai-image/jobs/{job_uid}/materialize")
def materialize_ai_image_job(job_uid: str, req: AiImageMaterializeRequest):
    data_url = str(req.data_url or "").strip()
    if data_url:
        job = data_sink.get_ai_image_job(job_uid)
        try:
            if not job:
                return ai_image_service.materialize_data_url_image(
                    job_uid,
                    data_url,
                    filename=req.filename,
                    allow_unlisted=True,
                )
            return ai_image_service.materialize_data_url_image(job_uid, data_url, filename=req.filename)
        except ValueError as exc:
            raise HTTPException(400, str(exc)) from exc
    source = str(req.url or req.file or "").strip()
    if not source:
        raise HTTPException(400, "图片地址不能为空")
    job = data_sink.get_ai_image_job(job_uid)
    try:
        if not job:
            return ai_image_service.materialize_remote_image(job_uid, source, allow_unlisted=True)
        return ai_image_service.materialize_remote_image(job_uid, source)
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc


@app.post("/ai-image/assets")
def create_ai_image_asset(req: AiImageAssetRequest):
    if not data_sink.get_ai_image_job(req.job_uid):
        raise HTTPException(404, f"AI image job not found: {req.job_uid}")
    return data_sink.create_ai_image_asset(_model_payload(req, exclude_none=True))


@app.post("/ai-image/canvases")
def create_ai_image_canvas(req: AiImageCanvasRequest):
    if not data_sink.get_ai_image_job(req.job_uid):
        raise HTTPException(404, f"AI image job not found: {req.job_uid}")
    return data_sink.create_ai_image_canvas(_model_payload(req, exclude_none=True))


class AiVideoAssetInputRequest(BaseModel):
    model_config = ConfigDict(extra="forbid")

    role: str = "reference_image"
    sourceType: str = "local_file"
    fileToken: Optional[str] = None
    sortOrder: int = 0


class AiVideoMediaInputRequest(BaseModel):
    model_config = ConfigDict(extra="forbid")

    type: str
    url: str
    keep_original_sound: Optional[str] = None
    keepOriginalSound: Optional[str] = None


class AiVideoPublicParametersRequest(BaseModel):
    model_config = ConfigDict(extra="forbid")

    ratio: Optional[str] = None
    aspect_ratio: Optional[str] = None
    aspectRatio: Optional[str] = None
    resolution: Optional[str] = None
    duration: Optional[int] = None
    mode: Optional[str] = None
    audio: Optional[bool] = None
    watermark: Optional[bool] = None
    generateAudio: Optional[bool] = None
    imageUrl: Optional[str] = None
    image_url: Optional[str] = None
    videoUrl: Optional[str] = None
    video_url: Optional[str] = None
    media: List[AiVideoMediaInputRequest] = Field(default_factory=list)


class AiVideoCreateJobRequest(BaseModel):
    model_config = ConfigDict(extra="forbid")

    requestUid: str
    provider: str
    model: str
    prompt: str
    assets: List[AiVideoAssetInputRequest] = Field(default_factory=list)
    parameters: AiVideoPublicParametersRequest = Field(default_factory=AiVideoPublicParametersRequest)
    outputDirToken: str = ""


class AiVideoValidateRequest(BaseModel):
    model_config = ConfigDict(extra="forbid")

    provider: str
    model: str
    prompt: str = ""
    assets: List[AiVideoAssetInputRequest] = Field(default_factory=list)
    parameters: AiVideoPublicParametersRequest = Field(default_factory=AiVideoPublicParametersRequest)
    outputDirToken: str = ""


class AiVideoUpdateJobRequest(BaseModel):
    model_config = ConfigDict(extra="forbid")

    provider: Optional[str] = None
    model: Optional[str] = None
    prompt: Optional[str] = None
    assets: Optional[List[AiVideoAssetInputRequest]] = None
    parameters: Optional[AiVideoPublicParametersRequest] = None
    outputDirToken: Optional[str] = None


class AiVideoRetryRequest(BaseModel):
    model_config = ConfigDict(extra="forbid")

    requestUid: str


def _require_ai_video_worker_owner() -> None:
    """Only the process holding the shared backend lock may mutate video work."""
    if getattr(app.state, "owns_backend_instance", True) is False:
        raise HTTPException(
            status_code=503,
            detail="当前核心服务不是任务执行主实例，请连接已持有数据目录锁的桌面端",
        )


def _ai_video_http_error(exc: Exception) -> HTTPException:
    if isinstance(exc, ai_video_generation_service.AiVideoError):
        return HTTPException(
            status_code=int(exc.http_status or 400),
            detail={
                "ok": False,
                "error": {
                    "code": exc.code,
                    "message": str(exc),
                    "detail": exc.provider_message or None,
                },
            },
        )
    return HTTPException(status_code=500, detail=ai_video_generation_service.envelope_error(exc))


@app.get("/ai-video/config")
def ai_video_get_config():
    return ai_video_generation_service.get_config()


@app.post("/ai-video/validate")
def ai_video_validate(req: AiVideoValidateRequest):
    try:
        payload = _model_payload(req, exclude_none=True)
        result = ai_video_generation_service.validate_public_input(payload)
        return {"ok": True, "data": result}
    except Exception as exc:
        raise _ai_video_http_error(exc) from exc


@app.post("/ai-video/jobs")
def ai_video_create_job(req: AiVideoCreateJobRequest):
    _require_ai_video_worker_owner()
    try:
        payload = _model_payload(req, exclude_none=True)
        return ai_video_generation_service.create_job(payload)
    except Exception as exc:
        raise _ai_video_http_error(exc) from exc


@app.get("/ai-video/jobs")
def ai_video_list_jobs(status: str = "", provider: str = "", limit: int = 50):
    try:
        return ai_video_generation_service.list_jobs(status=status, provider=provider, limit=limit)
    except Exception as exc:
        raise _ai_video_http_error(exc) from exc


@app.get("/ai-video/jobs/{job_id}")
def ai_video_get_job(job_id: str):
    try:
        return ai_video_generation_service.get_job(job_id)
    except Exception as exc:
        raise _ai_video_http_error(exc) from exc


@app.patch("/ai-video/jobs/{job_id}")
def ai_video_update_job(job_id: str, req: AiVideoUpdateJobRequest):
    _require_ai_video_worker_owner()
    try:
        payload = _model_payload(req, exclude_none=True, exclude_unset=True)
        return ai_video_generation_service.update_job(job_id, payload)
    except Exception as exc:
        raise _ai_video_http_error(exc) from exc


@app.post("/ai-video/jobs/{job_id}/duplicate")
def ai_video_duplicate_job(job_id: str):
    _require_ai_video_worker_owner()
    try:
        return ai_video_generation_service.duplicate_job(job_id)
    except Exception as exc:
        raise _ai_video_http_error(exc) from exc


@app.post("/ai-video/jobs/{job_id}/retry")
def ai_video_retry_job(job_id: str, req: AiVideoRetryRequest):
    _require_ai_video_worker_owner()
    try:
        return ai_video_generation_service.retry_job(job_id, req.requestUid)
    except Exception as exc:
        raise _ai_video_http_error(exc) from exc


@app.delete("/ai-video/jobs/{job_id}")
def ai_video_delete_job(job_id: str):
    _require_ai_video_worker_owner()
    try:
        return ai_video_generation_service.delete_job(job_id)
    except Exception as exc:
        raise _ai_video_http_error(exc) from exc


@app.get("/ai-video/runs/{run_id}")
def ai_video_get_run(run_id: str):
    try:
        return ai_video_generation_service.get_run(run_id)
    except Exception as exc:
        raise _ai_video_http_error(exc) from exc


@app.post("/ai-video/runs/{run_id}/archive")
def ai_video_retry_archive(run_id: str):
    _require_ai_video_worker_owner()
    try:
        return ai_video_generation_service.retry_archive(run_id)
    except Exception as exc:
        raise _ai_video_http_error(exc) from exc


def _normalize_tmall_approval_batch_id(batch_id: str) -> str:
    normalized = re.sub(r"[^A-Za-z0-9._~-]+", "", str(batch_id or "").strip())
    if not normalized:
        raise HTTPException(400, "Invalid approval batch id")
    return normalized


def _find_tmall_approval_batch_path(batch_id: str) -> Path:
    normalized = _normalize_tmall_approval_batch_id(batch_id)
    pattern = f"tmall-ai-image-approval-batch-{normalized}.json"
    roots = [runtime_paths.data_root(), Path.cwd()]
    matches: list[Path] = []
    for root in roots:
        if not root.exists():
            continue
        try:
            matches.extend(path for path in root.rglob(pattern) if path.is_file())
        except OSError:
            logger.debug("approval batch search skipped unavailable root %s", root, exc_info=True)
    if not matches:
        raise HTTPException(404, f"Approval batch not found: {batch_id}")
    matches.sort(key=lambda path: path.stat().st_mtime, reverse=True)
    return matches[0]


def _load_tmall_approval_batch(batch_id: str) -> dict:
    path = _find_tmall_approval_batch_path(batch_id)
    try:
        batch = json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:
        raise HTTPException(500, f"Unable to read approval batch: {path}") from exc
    if not batch.get("json_path"):
        batch["json_path"] = str(path)
    return batch


def _validate_tmall_approval_token(batch: dict, token: str) -> None:
    expected = str((batch or {}).get("token") or "").strip()
    supplied = str(token or "").strip()
    if not expected or not supplied or not hmac.compare_digest(expected, supplied):
        raise HTTPException(403, "Invalid approval token")


def _tmall_approval_asset_path(batch: dict, asset_id: str) -> Path:
    module = _load_tmall_ai_image_chain_module()
    try:
        _item, asset = module.find_approval_asset(batch, asset_id)
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc
    path = Path(str(asset.get("path") or "")).expanduser()
    if not path.is_file():
        raise HTTPException(404, f"Approval image not found: {asset_id}")
    return path


def _tmall_approval_allowed_reference_paths(batch: dict) -> set[str]:
    allowed: set[str] = set()
    module = _load_tmall_ai_image_chain_module()
    parse_list = getattr(module, "parse_list", None)

    def add_path(value) -> None:
        text = str(value or "").strip()
        if not text:
            return
        try:
            allowed.add(str(Path(text).expanduser().resolve()))
        except Exception:
            allowed.add(str(Path(text).expanduser()))

    for item in batch.get("items") or []:
        if not isinstance(item, dict):
            continue
        add_path(item.get("origin_path"))
        add_path(item.get("detail_reference_path"))
        for asset in item.get("assets") or []:
            if not isinstance(asset, dict):
                continue
            add_path(asset.get("path"))
            refs = asset.get("reference_paths")
            if callable(parse_list):
                refs = parse_list(refs)
            elif isinstance(refs, str):
                refs = [part.strip() for part in re.split(r"[\n\r,，、；;]+", refs) if part.strip()]
            for ref in refs or []:
                add_path(ref)
            generation_row = asset.get("generation_row") if isinstance(asset.get("generation_row"), dict) else {}
            generation_refs = generation_row.get("__1xm_reference_paths") or generation_row.get("参考图文件")
            if callable(parse_list):
                generation_refs = parse_list(generation_refs)
            elif isinstance(generation_refs, str):
                generation_refs = [part.strip() for part in re.split(r"[\n\r,，、；;]+", generation_refs) if part.strip()]
            for ref in generation_refs or []:
                add_path(ref)
    for ref in batch.get("imported_reference_paths") or []:
        add_path(ref)
    return allowed


def _resolve_tmall_approval_path(value: str) -> str:
    path = Path(str(value or "").strip()).expanduser()
    try:
        return str(path.resolve())
    except Exception:
        return str(path)


def _safe_tmall_approval_paths(batch: dict, paths: list[str], *, field_label: str = "图片") -> list[str]:
    allowed_refs = _tmall_approval_allowed_reference_paths(batch)
    safe_paths: list[str] = []
    rejected: list[str] = []
    for raw in paths or []:
        text = str(raw or "").strip()
        if not text:
            continue
        if _resolve_tmall_approval_path(text) in allowed_refs:
            safe_paths.append(text)
        else:
            rejected.append(Path(text).name or text)
    if rejected:
        raise HTTPException(403, f"{field_label}未登记在当前审批批次：{', '.join(rejected[:3])}")
    return safe_paths


class TmallApprovalDecisionsRequest(BaseModel):
    decisions: dict = {}


class TmallApprovalReferenceImportRequest(BaseModel):
    paths: List[str] = []


class TmallApprovalRegenerateRequest(BaseModel):
    asset_id: str
    prompt: Optional[str] = ""
    reference_paths: Optional[List[str]] = None


class TmallApprovalGenerateRequest(BaseModel):
    item_id: str = ""
    style_code: str = ""
    prompt: str = ""
    main_image_path: str = ""
    reference_paths: Optional[List[str]] = None


class TmallApprovalGenerationConfirmRequest(BaseModel):
    items: List[dict] = []


def _update_tmall_generation_task_instance(batch_id: str, result: dict) -> None:
    instance = data_sink.find_task_instance_by_approval_batch_id(batch_id)
    if not instance:
        return
    summary = _parse_json_object(instance.get("summary_json"))
    summary.update({
        "approval_batch_id": batch_id,
        "generation_status": result.get("status"),
        "generated": int(result.get("generated") or 0),
        "generation_failed": int(result.get("failed") or 0),
    })
    status = "waiting_approval" if result.get("status") == "pending_approval" else (
        "running" if result.get("status") == "generating" else "generation_failed"
    )
    current_step = "approval" if result.get("status") in {"pending_approval", "generating"} else "confirm"
    data_sink.update_task_instance(
        instance["instance_uid"],
        status=status,
        current_step=current_step,
        summary=summary,
    )


async def _run_tmall_generation_confirmation_job(batch_id: str, batch: dict) -> None:
    module = _load_tmall_ai_image_chain_module()
    try:
        result = await asyncio.to_thread(module.submit_generation_confirmation_batch, batch, log=logger.info)
        _update_tmall_generation_task_instance(batch_id, result)
    except Exception as exc:
        logger.exception("Tmall AI image generation confirmation job failed: %s", batch_id)
        try:
            failed_batch = _load_tmall_approval_batch(batch_id)
            failed_batch["status"] = "generation_failed"
            failed_batch["submit_progress"] = {
                **dict(failed_batch.get("submit_progress") or {}),
                "status": "failed",
                "message": str(exc),
            }
            failed_batch["generation_summary"] = {
                **dict(failed_batch.get("generation_summary") or {}),
                "failed": int((failed_batch.get("generation_summary") or {}).get("requested_prompts") or 1),
            }
            failed_batch["updated_at"] = datetime.now().isoformat()
            module.save_approval_batch(failed_batch)
        except Exception:
            logger.debug("Failed to persist Tmall AI generation failure state", exc_info=True)
        _update_tmall_generation_task_instance(batch_id, {
            "status": "generation_failed",
            "generated": 0,
            "failed": 1,
        })


def _safe_tmall_generation_confirmation_items(batch: dict, items: list[dict]) -> list[dict]:
    safe_items: list[dict] = []
    for item in items or []:
        if not isinstance(item, dict):
            continue
        safe_item = dict(item)
        for key, label in [("origin_path", "主图"), ("detail_reference_path", "参考图")]:
            if key not in safe_item:
                continue
            value = str(safe_item.get(key) or "").strip()
            safe_item[key] = (_safe_tmall_approval_paths(batch, [value], field_label=label) or [""])[0] if value else ""
        safe_reference_assets = []
        for asset in safe_item.get("reference_assets") or []:
            if not isinstance(asset, dict):
                continue
            safe_asset = dict(asset)
            value = str(safe_asset.get("path") or "").strip()
            safe_path = (_safe_tmall_approval_paths(batch, [value], field_label="参考图") or [""])[0] if value else ""
            if not safe_path:
                continue
            safe_asset["path"] = safe_path
            safe_reference_assets.append(safe_asset)
        if "reference_assets" in safe_item:
            safe_item["reference_assets"] = safe_reference_assets
        safe_prompts = []
        for prompt in safe_item.get("generation_prompts") or []:
            if not isinstance(prompt, dict):
                continue
            safe_prompt = dict(prompt)
            try:
                image_count = int(safe_prompt.get("image_count") or safe_prompt.get("generation_image_count") or safe_prompt.get("count") or 1)
            except Exception:
                image_count = 1
            safe_prompt["image_count"] = max(1, min(8, image_count))
            safe_prompt["reference_paths"] = _safe_tmall_approval_paths(
                batch,
                [str(path or "").strip() for path in (safe_prompt.get("reference_paths") or []) if str(path or "").strip()],
                field_label="参考图",
            )
            safe_prompts.append(safe_prompt)
        safe_item["generation_prompts"] = safe_prompts
        safe_items.append(safe_item)
    return safe_items


@app.get("/tmall-ai-image-approval/api/{batch_id}")
def get_tmall_ai_image_approval_batch(batch_id: str, token: str = ""):
    batch = _load_tmall_approval_batch(batch_id)
    _validate_tmall_approval_token(batch, token)
    return batch


@app.get("/tmall-ai-image-approval/api/{batch_id}/image/{asset_id}")
def get_tmall_ai_image_approval_image(batch_id: str, asset_id: str, token: str = ""):
    batch = _load_tmall_approval_batch(batch_id)
    _validate_tmall_approval_token(batch, token)
    path = _tmall_approval_asset_path(batch, asset_id)
    media_type = "image/jpeg"
    guessed, _encoding = mimetypes.guess_type(str(path))
    if guessed:
        media_type = guessed
    return FileResponse(path, media_type=media_type, filename=path.name)


@app.get("/tmall-ai-image-approval/api/{batch_id}/reference-image")
def get_tmall_ai_image_approval_reference_image(batch_id: str, path: str, token: str = ""):
    batch = _load_tmall_approval_batch(batch_id)
    _validate_tmall_approval_token(batch, token)
    image_path = Path(str(path or "")).expanduser()
    try:
        resolved = str(image_path.resolve())
    except Exception:
        resolved = str(image_path)
    if resolved not in _tmall_approval_allowed_reference_paths(batch):
        raise HTTPException(403, "Reference image is not registered in this approval batch")
    if not image_path.is_file():
        raise HTTPException(404, "Reference image not found")
    media_type = "image/jpeg"
    guessed, _encoding = mimetypes.guess_type(str(image_path))
    if guessed:
        media_type = guessed
    return FileResponse(image_path, media_type=media_type, filename=image_path.name)


@app.post("/tmall-ai-image-approval/api/{batch_id}/decisions")
def save_tmall_ai_image_approval_decisions(batch_id: str, req: TmallApprovalDecisionsRequest, token: str = ""):
    batch = _load_tmall_approval_batch(batch_id)
    _validate_tmall_approval_token(batch, token)
    module = _load_tmall_ai_image_chain_module()
    allowed_refs = _tmall_approval_allowed_reference_paths(batch)
    decisions = dict(req.decisions or {})
    for patch in decisions.values():
        if not isinstance(patch, dict) or "reference_paths" not in patch:
            continue
        safe_refs = []
        for ref in getattr(module, "parse_list")(patch.get("reference_paths")):
            try:
                resolved = str(Path(str(ref or "")).expanduser().resolve())
            except Exception:
                resolved = str(Path(str(ref or "")).expanduser())
            if resolved in allowed_refs:
                safe_refs.append(ref)
        patch["reference_paths"] = safe_refs
    batch = module.update_approval_decisions(batch, decisions)
    return {"ok": True, "status": batch.get("status"), "updated_at": batch.get("updated_at")}


@app.post("/tmall-ai-image-approval-local/api/{batch_id}/reference-files")
def import_tmall_ai_image_approval_reference_files(batch_id: str, req: TmallApprovalReferenceImportRequest, token: str = ""):
    batch = _load_tmall_approval_batch(batch_id)
    _validate_tmall_approval_token(batch, token)
    module = _load_tmall_ai_image_chain_module()
    artifact_dir_raw = str(batch.get("artifact_dir") or "").strip()
    if artifact_dir_raw:
        artifact_dir = Path(artifact_dir_raw).expanduser()
    else:
        artifact_dir = Path(str(batch.get("json_path") or ".")).expanduser().parent
    target_dir = artifact_dir / "reference-imports"
    target_dir.mkdir(parents=True, exist_ok=True)

    imported: list[str] = []
    allowed_suffixes = {".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp", ".tif", ".tiff"}
    for raw in req.paths or []:
        src = Path(str(raw or "")).expanduser()
        if not src.is_file() or src.suffix.lower() not in allowed_suffixes:
            continue
        stem = re.sub(r"[^A-Za-z0-9._~-]+", "-", src.stem).strip("-") or "reference"
        target = target_dir / f"{stem}-{secrets.token_hex(3)}{src.suffix.lower()}"
        shutil.copy2(src, target)
        imported.append(str(target))

    existing = [str(item or "").strip() for item in (batch.get("imported_reference_paths") or []) if str(item or "").strip()]
    batch["imported_reference_paths"] = list(dict.fromkeys([*existing, *imported]))
    batch["updated_at"] = datetime.now().isoformat()
    module.save_approval_batch(batch)
    return {"ok": True, "paths": imported}


@app.post("/tmall-ai-image-approval/api/{batch_id}/regenerate")
async def regenerate_tmall_ai_image_approval_asset(batch_id: str, req: TmallApprovalRegenerateRequest, token: str = ""):
    batch = _load_tmall_approval_batch(batch_id)
    _validate_tmall_approval_token(batch, token)
    module = _load_tmall_ai_image_chain_module()
    try:
        safe_reference_paths = _safe_tmall_approval_paths(
            batch,
            list(req.reference_paths or []),
            field_label="参考图",
        ) if req.reference_paths is not None else []
        asset = await asyncio.to_thread(
            module.regenerate_approval_asset,
            batch,
            req.asset_id,
            prompt=req.prompt or "",
            reference_paths=safe_reference_paths,
        )
        return {"ok": True, "asset": asset}
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(400, str(exc)) from exc


@app.post("/tmall-ai-image-approval/api/{batch_id}/generate")
async def generate_tmall_ai_image_approval_asset(batch_id: str, req: TmallApprovalGenerateRequest, token: str = ""):
    batch = _load_tmall_approval_batch(batch_id)
    _validate_tmall_approval_token(batch, token)
    module = _load_tmall_ai_image_chain_module()
    main_image_path = str(req.main_image_path or "").strip()
    if main_image_path:
        main_image_path = (_safe_tmall_approval_paths(batch, [main_image_path], field_label="主图") or [""])[0]
    safe_reference_paths = _safe_tmall_approval_paths(
        batch,
        list(req.reference_paths or []),
        field_label="参考图",
    )
    try:
        asset = await asyncio.to_thread(
            module.generate_approval_asset_for_item,
            batch,
            item_id=req.item_id or "",
            style_code=req.style_code or "",
            prompt=req.prompt or "",
            main_image_path=main_image_path,
            reference_paths=safe_reference_paths,
        )
        return {"ok": True, "asset": asset}
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(400, str(exc)) from exc


@app.post("/tmall-ai-image-approval/api/{batch_id}/submit-generation")
async def submit_tmall_ai_image_generation_confirmation(batch_id: str, req: TmallApprovalGenerationConfirmRequest, token: str = ""):
    batch = _load_tmall_approval_batch(batch_id)
    _validate_tmall_approval_token(batch, token)
    module = _load_tmall_ai_image_chain_module()
    try:
        safe_items = _safe_tmall_generation_confirmation_items(batch, list(req.items or []))
        if safe_items:
            batch = module.update_generation_confirmation(batch, safe_items)
        batch = module.prepare_generation_confirmation_placeholders(batch)
        _update_tmall_generation_task_instance(batch_id, {
            "status": "generating",
            "generated": 0,
            "failed": 0,
        })
        asyncio.create_task(_run_tmall_generation_confirmation_job(batch_id, batch))
        return {
            "ok": True,
            "accepted": True,
            "status": "generating",
            "generated": 0,
            "failed": 0,
            "batch": batch,
        }
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(400, str(exc)) from exc


@app.post("/tmall-ai-image-approval/api/{batch_id}/submit")
async def submit_tmall_ai_image_approval_batch(batch_id: str, token: str = ""):
    batch = _load_tmall_approval_batch(batch_id)
    _validate_tmall_approval_token(batch, token)
    module = _load_tmall_ai_image_chain_module()
    try:
        result = await module.upload_approved_tmall_batch(batch, log=logger.info)
    except Exception as exc:
        raise HTTPException(400, str(exc)) from exc
    instance = data_sink.find_task_instance_by_approval_batch_id(batch_id)
    if instance:
        failed = int(result.get("failed") or 0)
        attempted = int(result.get("attempted") or 0)
        status = "completed" if attempted > 0 and failed == 0 else "partial_failed"
        result_path = str(result.get("excel_path") or result.get("result_path") or "").strip()
        audit_path = str(result.get("audit_path") or "").strip()
        summary = _parse_json_object(instance.get("summary_json"))
        summary_files = list(summary.get("output_files") or [])
        if result_path and result_path not in summary_files:
            summary_files.append(result_path)
        summary.update({
            "approval_batch_id": batch_id,
            "submit_status": result.get("status"),
            "attempted": attempted,
            "succeeded": int(result.get("succeeded") or 0),
            "failed": failed,
            "submitted": int(result.get("submitted") or 0),
            "result_path": result_path,
            "result_excel_path": result_path,
            "audit_result_path": audit_path,
            "output_files": summary_files,
        })
        data_sink.update_task_instance(
            instance["instance_uid"],
            status=status,
            current_step="create",
            summary=summary,
        )
        if result_path:
            data_sink.add_task_instance_artifact(
                instance["instance_uid"],
                kind=_task_instance_artifact_kind(result_path),
                label=Path(result_path).name,
                path=result_path,
                meta={"approval_batch_id": batch_id},
            )
    return result


@app.get("/tmall-ai-image-approval/{batch_id}")
def get_tmall_ai_image_approval_board(batch_id: str, token: str = ""):
    batch = _load_tmall_approval_batch(batch_id)
    _validate_tmall_approval_token(batch, token)
    board_path = Path(str(batch.get("board_path") or "")).expanduser()
    if not board_path.is_file():
        module = _load_tmall_ai_image_chain_module()
        board_path = Path(str(batch.get("json_path") or "")).with_name(f"tmall-ai-image-approval-board-{batch.get('batch_id', batch_id)}.html")
        board_path.write_text(module.render_approval_board_html(batch), encoding="utf-8")
    return FileResponse(board_path, media_type="text/html; charset=utf-8")


# ─── Health ───

@app.get("/health")
def health(probe: bool = False):
    if probe:
        return {
            "status": "ok",
            "runtime": _backend_runtime_info(),
        }

    warnings = []
    try:
        chrome_ok = CDPBridge().is_available(timeout=0.2)
    except Exception as e:
        logger.debug("health chrome probe failed", exc_info=True)
        warnings.append(str(e))
        chrome_ok = False

    try:
        adapters_count = len(adapter_loader.list_all())
    except Exception as e:
        logger.debug("health adapter status failed", exc_info=True)
        warnings.append(str(e))
        adapters_count = 0

    try:
        scheduled_jobs_count = len(sched_module.list_jobs())
    except Exception as e:
        logger.debug("health scheduler status failed", exc_info=True)
        warnings.append(str(e))
        scheduled_jobs_count = 0

    return {
        "status": "ok",
        "chrome": chrome_ok,
        "adapters": adapters_count,
        "scheduled_jobs": scheduled_jobs_count,
        "warnings": warnings,
        "runtime": _backend_runtime_info(),
    }


# ─── Adapters ───

@app.get("/script-favorites")
def get_script_favorites():
    return {"favorites": script_favorites.list_favorites()}


@app.put("/script-favorites/{adapter_id}")
def favorite_script(adapter_id: str):
    if not adapter_loader.get_adapter(adapter_id):
        raise HTTPException(404, f"Adapter not found: {adapter_id}")
    return {"favorites": script_favorites.favorite(adapter_id)}


@app.delete("/script-favorites/{adapter_id}")
def unfavorite_script(adapter_id: str):
    if not adapter_loader.get_adapter(adapter_id):
        raise HTTPException(404, f"Adapter not found: {adapter_id}")
    return {"favorites": script_favorites.unfavorite(adapter_id)}

@app.get("/adapters")
def list_adapters():
    return adapter_loader.list_all()


class InstallRequest(BaseModel):
    path: Optional[str] = None
    zip_base64: Optional[str] = None
    install_mode: Optional[str] = None


@app.post("/adapters/install")
def install_adapter(req: InstallRequest):
    try:
        if req.path:
            m = adapter_loader.install_from_dir(req.path, install_mode=str(req.install_mode or "copy"))
        elif req.zip_base64:
            raw = base64.b64decode(req.zip_base64)
            with tempfile.NamedTemporaryFile(suffix=".zip", delete=False) as f:
                f.write(raw)
                tmp = f.name
            try:
                m = adapter_loader.install_from_zip(tmp, install_mode=str(req.install_mode or "copy"))
            finally:
                os.unlink(tmp)
        else:
            raise HTTPException(400, "Provide 'path' or 'zip_base64'")

        # Register scheduler jobs for newly installed adapter
        sched_module.register_adapter(m, _run_scheduled_task)
        adapter_meta = adapter_loader.get_install_metadata(m.id)
        return {
            "ok": True,
            "adapter": {
                "id": m.id,
                "name": m.name,
                "version": m.version,
                "install_mode": adapter_meta.get("install_mode", "copy"),
                "source_path": adapter_meta.get("source_path", ""),
                "runtime_path": adapter_meta.get("runtime_path", ""),
            }
        }
    except Exception as e:
        raise HTTPException(400, str(e))


@app.delete("/adapters/{adapter_id}")
def uninstall_adapter(adapter_id: str):
    if not adapter_loader.get_adapter(adapter_id):
        raise HTTPException(404, f"Adapter not found: {adapter_id}")
    sched_module.unregister_adapter(adapter_id)
    adapter_loader.uninstall(adapter_id)
    script_favorites.unfavorite(adapter_id)
    return {"ok": True}


class EnableRequest(BaseModel):
    enabled: bool


@app.patch("/adapters/{adapter_id}/enable")
def enable_adapter(adapter_id: str, req: EnableRequest):
    m = adapter_loader.get_adapter(adapter_id)
    if not m:
        raise HTTPException(404, f"Adapter not found: {adapter_id}")
    adapter_loader.set_enabled(adapter_id, req.enabled)
    if req.enabled:
        sched_module.register_adapter(m, _run_scheduled_task)
    else:
        sched_module.unregister_adapter(adapter_id)
    return {"ok": True, "enabled": req.enabled}


# ─── Tasks ───

@app.get("/tasks")
def list_tasks():
    adapter_loader.scan_all()
    result = []
    scheduled = {j['job_id']: j for j in sched_module.list_jobs()}
    for item in adapter_loader.list_all():
        m = adapter_loader.get_adapter(item['id'])
        if not m:
            continue
        for task in m.tasks:
            jid = f"{m.id}::{task.id}"
            last_run = data_sink.get_latest_run(m.id, task.id)
            result.append({
                "adapter_id": m.id,
                "adapter_name": m.name,
                "adapter_version": m.version,
                "task_id": task.id,
                "task_name": task.name,
                "description": task.description,
                "param_probe_script": task.param_probe_script,
                "execution_ui_mode": task.execution_ui_mode,
                "validation_only_label": task.validation_only_label,
                "auto_precheck_note": task.auto_precheck_note,
                "params": [_serialize_task_param(m.id, p) for p in task.params],
                "trigger": task.trigger.model_dump(),
                "enabled": item['enabled'],
                "next_run": scheduled.get(jid, {}).get('next_run'),
                "last_run": last_run,
                "live": _run_status.get(jid),
            })
    return result


class RuntimeDrainReleaseRequest(BaseModel):
    drain_token: str


def _active_task_item_from_live(jid: str, live: dict) -> dict:
    live_snapshot = dict(live or {})
    adapter_id = str(live_snapshot.get("adapter_id") or "")
    task_id = str(live_snapshot.get("task_id") or "")
    if not adapter_id and "::" in jid:
        adapter_id, task_id = jid.split("::", 1)
    elif not adapter_id:
        adapter_id = jid
    return {
        "jid": jid,
        "adapter_id": adapter_id,
        "task_id": task_id,
        "instance_uid": live_snapshot.get("instance_uid") or "",
        "status": live_snapshot.get("status") or "running",
        "run_id": live_snapshot.get("run_id"),
        "records": int(live_snapshot.get("records") or 0),
        "phase": live_snapshot.get("phase"),
        "last_seen_at": live_snapshot.get("last_seen_at"),
        "current_row": int(live_snapshot.get("current_row") or live_snapshot.get("row_no") or 0),
    }


def _active_task_items() -> list[dict]:
    tasks = []
    seen_controls = set()
    seen_jids = set()
    for jid, live in sorted(
        _run_status.items(),
        key=lambda item: (str(item[0]).startswith("instance::"), str(item[0])),
    ):
        live_snapshot = dict(live or {})
        if live_snapshot.get("status") not in ACTIVE_LIVE_STATUSES:
            continue
        control = _run_controls.get(jid)
        task = control.get("task") if control else None
        has_live_task = bool(task and not task.done())
        if not has_live_task and not _task_lock(jid).locked():
            continue
        if control:
            control_key = id(control)
            if control_key in seen_controls:
                continue
            seen_controls.add(control_key)
        tasks.append(_active_task_item_from_live(jid, live_snapshot))
        seen_jids.add(jid)

    for jid, control in sorted(_run_controls.items()):
        if jid in seen_jids:
            continue
        task = control.get("task") if isinstance(control, dict) else None
        if not task or task.done():
            continue
        control_key = id(control)
        if control_key in seen_controls:
            continue
        seen_controls.add(control_key)
        tasks.append(_active_task_item_from_live(jid, {
            "status": "running",
            "adapter_id": str(control.get("adapter_id") or ""),
            "task_id": str(control.get("task_id") or ""),
            "instance_uid": str(control.get("instance_uid") or ""),
        }))
    return tasks


@app.get("/tasks/active")
def active_tasks():
    tasks = _active_task_items()
    return {"active": bool(tasks), "tasks": tasks}


def _install_blocker(kind: str, blocker_id: str, label: str, status: str = "running", **extra) -> dict:
    item = {
        "kind": str(kind or "operation"),
        "id": str(blocker_id or label or "operation"),
        "label": str(label or blocker_id or "后台操作"),
        "status": str(status or "running"),
    }
    for key, value in extra.items():
        if value not in (None, ""):
            item[key] = value
    return item


def _ai_image_job_is_active(job: dict) -> bool:
    active_statuses = {"queued", "running"}
    if str((job or {}).get("status") or "").strip().lower() in active_statuses:
        return True
    summary = job.get("summary") if isinstance(job.get("summary"), dict) else {}
    if str(summary.get("status") or "").strip().lower() in active_statuses:
        return True
    return any(
        str((run or {}).get("status") or "").strip().lower() in active_statuses
        for run in summary.get("runs") or []
        if isinstance(run, dict)
    )


def _collect_external_install_state() -> dict:
    blockers = []
    failed_sources = []
    for task in _active_task_items():
        jid = str(task.get("jid") or "")
        blockers.append(_install_blocker(
            "task",
            jid,
            jid,
            str(task.get("status") or "running"),
            instance_uid=task.get("instance_uid") or "",
        ))

    try:
        running_instances = data_sink.list_task_instances(status_group="running", limit=500)
    except Exception:
        logger.debug("failed to collect running task instances for install readiness", exc_info=True)
        failed_sources.append("task_instances")
        running_instances = []
    for instance in running_instances:
        uid = str(instance.get("instance_uid") or "")
        if uid:
            blockers.append(_install_blocker(
                "task",
                f"instance::{uid}",
                str(instance.get("title") or uid),
                "running",
                instance_uid=uid,
            ))

    try:
        ai_jobs = data_sink.list_active_ai_image_jobs()
    except Exception:
        logger.debug("failed to collect ai image jobs for install readiness", exc_info=True)
        failed_sources.append("ai_image_jobs")
        ai_jobs = []
    for job in ai_jobs:
        if not _ai_image_job_is_active(job):
            continue
        uid = str(job.get("job_uid") or "")
        if uid:
            blockers.append(_install_blocker(
                "ai_image",
                uid,
                str(job.get("title") or uid),
                str(job.get("status") or "running"),
            ))

    if str(getattr(cloud_machine_controller, "last_health", "") or "") == "online_busy":
        blockers.append(_install_blocker("cloud_job", "cloud_machine", "云端任务机", "running"))

    seen = set()
    deduped = []
    for blocker in blockers:
        key = (blocker.get("kind"), blocker.get("id"))
        if key in seen:
            continue
        seen.add(key)
        deduped.append(blocker)
    return {
        "blockers": deduped,
        "failed_sources": failed_sources,
    }


def _collect_external_install_blockers() -> list[dict]:
    return _collect_external_install_state()["blockers"]


def _install_readiness() -> dict:
    external_state = _collect_external_install_state()
    external_blockers = external_state["blockers"]
    failed_sources = external_state["failed_sources"]
    guard_readiness = runtime_install_guard.readiness()
    blockers = []
    seen = set()
    for blocker in [*guard_readiness["blockers"], *external_blockers]:
        key = (blocker.get("kind"), blocker.get("id"))
        if key in seen:
            continue
        seen.add(key)
        blockers.append(blocker)
    return {
        **guard_readiness,
        "ready": len(blockers) == 0 and not guard_readiness["draining"] and not failed_sources,
        "blockers": blockers,
        **({
            "error": "install_readiness_unavailable",
            "failed_sources": failed_sources,
        } if failed_sources else {}),
    }


@app.get("/runtime/install-readiness")
def runtime_install_readiness():
    return _install_readiness()


@app.post("/runtime/update-drain")
def acquire_runtime_update_drain():
    external_state = _collect_external_install_state()
    if external_state["failed_sources"]:
        raise HTTPException(409, {
            "code": "install_readiness_unavailable",
            "message": "无法确认是否仍有任务正在运行，已阻止安装更新。",
            "blockers": external_state["blockers"],
            "failed_sources": external_state["failed_sources"],
        })
    blockers = external_state["blockers"]
    if blockers:
        raise HTTPException(409, {
            "code": "runtime_busy",
            "message": "仍有任务正在运行。",
            "blockers": blockers,
        })
    try:
        drain_token = runtime_install_guard.acquire_drain()
    except InstallRuntimeBusy as exc:
        raise HTTPException(409, {
            "code": "runtime_busy",
            "message": "仍有任务正在运行。",
            "blockers": exc.blockers,
        })
    except UpdateDrainActive:
        raise HTTPException(409, {
            "code": "update_pending",
            "message": "抓虾正在准备安装更新，请稍后再启动新任务。",
        })
    return {
        "ok": True,
        "drain_token": drain_token,
        "readiness": {
            **runtime_install_guard.readiness(),
            "ready": True,
            "install_ready": True,
            "blockers": [],
        },
    }


@app.delete("/runtime/update-drain")
def release_runtime_update_drain(req: RuntimeDrainReleaseRequest):
    if not runtime_install_guard.release_drain(req.drain_token):
        raise HTTPException(409, {
            "code": "runtime_drain_token_mismatch",
            "message": "更新安装令牌不匹配。",
        })
    return {
        "ok": True,
        "readiness": _install_readiness(),
    }


class RunTaskRequest(BaseModel):
    params: Optional[dict] = None
    current_tab_id: Optional[str] = None


class TaskInstanceCreateRequest(BaseModel):
    adapter_id: str
    task_id: str
    title: str
    params: dict = {}


class TaskInstancePatchRequest(BaseModel):
    title: Optional[str] = None
    status: Optional[str] = None
    current_step: Optional[str] = None
    params: Optional[dict] = None
    summary: Optional[dict] = None
    archived: Optional[bool] = None


class TaskScheduleCreateRequest(BaseModel):
    adapter_id: str = TMALL_OPS_ADAPTER_ID
    task_id: str = TMALL_MATERIAL_EXPORT_TASK_ID
    title: str
    frequency: str
    time_of_day: str
    weekday: Optional[int] = None
    params: Optional[dict] = None
    sync_to_cloud: bool = True
    notify_channel: str = "dingtalk"
    notify_template: str = DEFAULT_TASK_SCHEDULE_NOTIFY_TEMPLATE
    enabled: bool = True


class TaskSchedulePatchRequest(BaseModel):
    title: Optional[str] = None
    enabled: Optional[bool] = None
    frequency: Optional[str] = None
    time_of_day: Optional[str] = None
    weekday: Optional[int] = None
    params: Optional[dict] = None
    notify_channel: Optional[str] = None
    notify_template: Optional[str] = None


class ProbeTaskParamsRequest(BaseModel):
    params: Optional[dict] = None
    current_tab_id: Optional[str] = None


def _parse_json_object(value) -> dict:
    try:
        parsed = json.loads(value) if isinstance(value, str) else value
    except Exception:
        return {}
    return parsed if isinstance(parsed, dict) else {}


def _serialize_task_instance(row: dict) -> dict:
    data = dict(row or {})
    data["params"] = _parse_json_object(data.get("params_json"))
    data["summary"] = _parse_json_object(data.get("summary_json"))
    return data


def _serialize_task_instance_detail(detail: dict) -> dict:
    data = _serialize_task_instance(detail)
    data["runs"] = list((detail or {}).get("runs") or [])
    data["artifacts"] = list((detail or {}).get("artifacts") or [])
    data["events"] = list((detail or {}).get("events") or [])
    return data


def _serialize_task_schedule(row: dict, next_runs: Optional[dict[str, str]] = None) -> dict:
    data = dict(row or {})
    data["params"] = _parse_json_object(data.get("params_json"))
    data["enabled"] = int(data.get("enabled") or 0)
    data["archived"] = int(data.get("archived") or 0)
    next_runs = next_runs if next_runs is not None else _schedule_next_run_map()
    data["next_run"] = next_runs.get(str(data.get("schedule_uid") or ""), "")
    return data


def _model_patch(req: BaseModel) -> dict:
    if hasattr(req, "model_dump"):
        return req.model_dump(exclude_unset=True)
    return req.dict(exclude_unset=True)


def _normalize_schedule_params(adapter_id: str, task_id: str, params: Optional[dict]) -> dict:
    if adapter_id == TMALL_OPS_ADAPTER_ID and task_id == TMALL_MATERIAL_EXPORT_TASK_ID:
        return _default_tmall_material_export_params(params or {})
    return dict(params or {})


def _normalize_task_schedule_create(req: TaskScheduleCreateRequest) -> dict:
    adapter_id = str(req.adapter_id or TMALL_OPS_ADAPTER_ID).strip()
    task_id = str(req.task_id or TMALL_MATERIAL_EXPORT_TASK_ID).strip()
    frequency, weekday = _validate_schedule_frequency(req.frequency, req.weekday)
    time_of_day = _validate_schedule_time(req.time_of_day)
    params = dict(req.params or {})
    if adapter_id == TMALL_OPS_ADAPTER_ID and task_id == TMALL_MATERIAL_EXPORT_TASK_ID:
        if req.sync_to_cloud:
            params["sync_to_cloud"] = ["enabled"]
        else:
            params.pop("sync_to_cloud", None)
    return {
        "adapter_id": adapter_id,
        "task_id": task_id,
        "title": str(req.title or "").strip() or "巴拉-AI测图数据抓取导出定时任务",
        "frequency": frequency,
        "time_of_day": time_of_day,
        "weekday": weekday,
        "params": _normalize_schedule_params(adapter_id, task_id, params),
        "notify_channel": str(req.notify_channel or "dingtalk").strip() or "dingtalk",
        "notify_template": str(req.notify_template or DEFAULT_TASK_SCHEDULE_NOTIFY_TEMPLATE),
        "enabled": bool(req.enabled),
    }


def _normalize_task_schedule_patch(current: dict, patch: dict) -> dict:
    merged_frequency = patch.get("frequency", current.get("frequency"))
    merged_weekday = patch.get("weekday", current.get("weekday"))
    if "time_of_day" in patch:
        patch["time_of_day"] = _validate_schedule_time(patch.get("time_of_day"))
    if "frequency" in patch or "weekday" in patch:
        frequency, weekday = _validate_schedule_frequency(merged_frequency, merged_weekday)
        patch["frequency"] = frequency
        patch["weekday"] = weekday
    if "params" in patch:
        patch["params"] = _normalize_schedule_params(
            str(current.get("adapter_id") or ""),
            str(current.get("task_id") or ""),
            patch.get("params"),
        )
    return patch


def _refresh_task_schedule_registration(schedule: dict) -> None:
    sched_module.register_task_schedule(schedule, _run_task_schedule)


def _create_instance_for_task_schedule(schedule_uid: str) -> dict:
    schedule = data_sink.get_task_schedule_detail(schedule_uid)
    if not schedule or int(schedule.get("archived") or 0) == 1:
        raise HTTPException(404, "Task schedule not found")
    if int(schedule.get("enabled") or 0) != 1:
        raise HTTPException(409, "定时任务已停用，无法运行")

    params = _normalize_schedule_params(
        str(schedule.get("adapter_id") or ""),
        str(schedule.get("task_id") or ""),
        schedule.get("params") or {},
    )
    params["__task_schedule_uid"] = schedule["schedule_uid"]
    title = f"{schedule.get('title') or '定时任务'} {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    instance = data_sink.create_task_instance(
        str(schedule.get("adapter_id") or ""),
        str(schedule.get("task_id") or ""),
        title,
        params,
    )
    instance_uid = str(instance.get("instance_uid") or "")
    params["__task_instance_uid"] = instance_uid
    data_sink.update_task_instance(instance_uid, status="running", current_step="config", params=params)
    data_sink.record_task_schedule_run(
        schedule["schedule_uid"],
        instance_uid=instance_uid,
        status="queued",
        error="",
    )
    return {**schedule, "instance_uid": instance_uid, "run_params": params}


def _notify_task_schedule_result(
    schedule_uid: str,
    instance_uid: str,
    run_id: Optional[int],
    status: str,
    records: int,
    output_files: list,
    error: str = "",
    started_at: str = "",
    finished_at: str = "",
    log=None,
) -> None:
    schedule = data_sink.get_task_schedule_detail(schedule_uid)
    if not schedule:
        return
    data_sink.record_task_schedule_run(
        schedule_uid,
        run_id=run_id,
        instance_uid=instance_uid,
        status=status,
        error=error,
    )
    channel = str(schedule.get("notify_channel") or "dingtalk").strip()
    if not channel or channel == "none":
        return

    adapter_name, task_name = _task_display_name(str(schedule.get("adapter_id") or ""), str(schedule.get("task_id") or ""))
    output_text = "\n".join(str(item) for item in (output_files or []) if str(item or "").strip()) or "无"
    export_dir = str((schedule.get("params") or {}).get("output_dir") or "")
    message = _render_task_schedule_template(schedule.get("notify_template") or DEFAULT_TASK_SCHEDULE_NOTIFY_TEMPLATE, {
        "schedule_title": schedule.get("title") or "",
        "task_name": task_name,
        "status": status,
        "records": records,
        "run_id": run_id or "",
        "instance_uid": instance_uid,
        "output_files": output_text,
        "export_dir": export_dir,
        "started_at": started_at,
        "finished_at": finished_at or datetime.now().isoformat(),
        "error": error or "",
    })
    try:
        notifier.send(
            channel=channel,
            title=f"{schedule.get('title') or task_name} - {status}",
            records=records,
            adapter_name=adapter_name,
            task_name=task_name,
            sample_rows=None,
            error=error or None,
            message=message,
        )
    except Exception as exc:
        warning = f"定时任务通知发送失败：{exc}"
        if log:
            log(f"[warn] {warning}")
        if instance_uid:
            data_sink.add_task_instance_event(
                instance_uid,
                "notify_failed",
                warning,
                {"schedule_uid": schedule_uid, "channel": channel},
            )


@app.get("/task-instances")
def list_task_instances_endpoint(
    status_group: str = "",
    adapter_id: str = "",
    task_id: str = "",
    keyword: str = "",
    limit: int = 100,
):
    items = data_sink.list_task_instances(
        status_group=status_group,
        adapter_id=adapter_id,
        task_id=task_id,
        keyword=keyword,
        limit=limit,
    )
    return {"items": [_serialize_task_instance(row) for row in items]}


@app.post("/task-instances")
def create_task_instance_endpoint(req: TaskInstanceCreateRequest):
    row = data_sink.create_task_instance(req.adapter_id, req.task_id, req.title, req.params)
    return _serialize_task_instance(row)


@app.get("/task-instances/{instance_uid}")
def get_task_instance_endpoint(instance_uid: str):
    detail = data_sink.get_task_instance_detail(instance_uid)
    if not detail:
        raise HTTPException(404, "Task instance not found")
    return _serialize_task_instance_detail(detail)


@app.patch("/task-instances/{instance_uid}")
def patch_task_instance_endpoint(instance_uid: str, req: TaskInstancePatchRequest):
    patch = _model_patch(req)
    if patch.get("archived") is True:
        patch["status"] = "archived"
    row = data_sink.update_task_instance(instance_uid, **patch)
    if not row:
        raise HTTPException(404, "Task instance not found")
    return _serialize_task_instance(row)


@app.get("/task-schedules")
def list_task_schedules_endpoint(
    adapter_id: str = "",
    task_id: str = "",
    enabled: Optional[bool] = None,
    keyword: str = "",
    include_archived: bool = False,
    limit: int = 100,
):
    items = data_sink.list_task_schedules(
        adapter_id=adapter_id,
        task_id=task_id,
        enabled=enabled,
        keyword=keyword,
        include_archived=include_archived,
        limit=limit,
    )
    next_runs = _schedule_next_run_map()
    return {"items": [_serialize_task_schedule(row, next_runs=next_runs) for row in items]}


@app.post("/task-schedules")
def create_task_schedule_endpoint(req: TaskScheduleCreateRequest):
    values = _normalize_task_schedule_create(req)
    row = data_sink.create_task_schedule(**values)
    _refresh_task_schedule_registration(row)
    return _serialize_task_schedule(row)


@app.get("/task-schedules/{schedule_uid}")
def get_task_schedule_endpoint(schedule_uid: str):
    detail = data_sink.get_task_schedule_detail(schedule_uid)
    if not detail:
        raise HTTPException(404, "Task schedule not found")
    return _serialize_task_schedule(detail)


@app.patch("/task-schedules/{schedule_uid}")
def patch_task_schedule_endpoint(schedule_uid: str, req: TaskSchedulePatchRequest):
    current = data_sink.get_task_schedule_detail(schedule_uid)
    if not current:
        raise HTTPException(404, "Task schedule not found")
    patch = _normalize_task_schedule_patch(current, _model_patch(req))
    row = data_sink.update_task_schedule(schedule_uid, **patch)
    _refresh_task_schedule_registration(row)
    return _serialize_task_schedule(row)


@app.delete("/task-schedules/{schedule_uid}")
def delete_task_schedule_endpoint(schedule_uid: str):
    row = data_sink.archive_task_schedule(schedule_uid)
    if not row:
        raise HTTPException(404, "Task schedule not found")
    sched_module.unregister_task_schedule(schedule_uid)
    return {"ok": True, "schedule": _serialize_task_schedule(row)}


@app.post("/task-schedules/{schedule_uid}/run-now")
async def run_task_schedule_now_endpoint(schedule_uid: str):
    schedule = data_sink.get_task_schedule_detail(schedule_uid)
    if not schedule or int(schedule.get("archived") or 0) == 1:
        raise HTTPException(404, "Task schedule not found")
    jid = f"{schedule.get('adapter_id')}::{schedule.get('task_id')}"
    if _task_is_active(jid):
        raise HTTPException(409, "任务正在运行中，请稍后再运行该定时任务")

    schedule_run = _create_instance_for_task_schedule(schedule_uid)
    try:
        start_result = await _start_task_run(
            str(schedule_run.get("adapter_id") or ""),
            str(schedule_run.get("task_id") or ""),
            dict(schedule_run.get("run_params") or {}),
            {},
        )
    except Exception as exc:
        message = getattr(exc, "detail", None) or str(exc)
        data_sink.update_task_instance(
            schedule_run["instance_uid"],
            status="failed",
            summary={"error": message, "schedule_uid": schedule_uid},
        )
        data_sink.record_task_schedule_run(
            schedule_uid,
            instance_uid=schedule_run["instance_uid"],
            status="failed",
            error=message,
        )
        raise
    return {
        "ok": True,
        "instance_uid": schedule_run["instance_uid"],
        "start": start_result,
    }


@app.post("/probe/run")
async def run_probe(req: ProbeRequest):
    try:
        result = await run_probe_request(req)
        try:
            rebuild_knowledge_index()
        except Exception as knowledge_error:
            logger.warning("probe 完成后重建知识索引失败: %s", knowledge_error)
        return result.model_dump()
    except ValueError as exc:
        raise HTTPException(404, str(exc))
    except RuntimeError as exc:
        raise HTTPException(400, str(exc))


@app.get("/probe/{probe_id}")
def get_probe(probe_id: str):
    try:
        result = read_probe_bundle(probe_id)
        return result.model_dump()
    except FileNotFoundError as exc:
        raise HTTPException(404, str(exc))


@app.get("/probe/{probe_id}/bundle")
def get_probe_bundle(probe_id: str):
    try:
        result = read_probe_bundle_full(probe_id)
        return result.model_dump()
    except FileNotFoundError as exc:
        raise HTTPException(404, str(exc))


@app.post("/knowledge/rebuild")
def rebuild_knowledge():
    try:
        return rebuild_knowledge_index()
    except RuntimeError as exc:
        raise HTTPException(400, str(exc))


@app.get("/knowledge/search")
def get_knowledge(
    q: str = "",
    adapter_id: str = "",
    task_id: str = "",
    url: str = "",
    limit: int = 8,
):
    try:
        return search_knowledge(q, adapter_id=adapter_id, task_id=task_id, url=url, limit=limit)
    except RuntimeError as exc:
        raise HTTPException(400, str(exc))


@app.post("/dev-harness/snapshot")
async def dev_harness_snapshot(req: DevHarnessSnapshotRequest):
    try:
        return await run_harness_snapshot(req)
    except ValueError as exc:
        raise HTTPException(404, str(exc))
    except RuntimeError as exc:
        raise HTTPException(400, str(exc))


@app.post("/dev-harness/eval")
async def dev_harness_eval(req: DevHarnessEvalRequest):
    try:
        return await run_harness_eval(req)
    except ValueError as exc:
        raise HTTPException(404, str(exc))
    except RuntimeError as exc:
        raise HTTPException(400, str(exc))


@app.post("/dev-harness/capture")
async def dev_harness_capture(req: DevHarnessCaptureRequest):
    try:
        return await run_harness_capture(req)
    except ValueError as exc:
        raise HTTPException(404, str(exc))
    except RuntimeError as exc:
        raise HTTPException(400, str(exc))


@app.post("/tasks/{adapter_id}/{task_id}/params/probe")
async def probe_task_params(adapter_id: str, task_id: str, req: ProbeTaskParamsRequest):
    adapter_loader.scan_all()
    adapter = adapter_loader.get_adapter(adapter_id)
    if not adapter:
        raise HTTPException(404, f"Adapter not found: {adapter_id}")

    task = next((item for item in adapter.tasks if item.id == task_id), None)
    if not task:
        raise HTTPException(404, f"Task not found: {task_id}")

    probe_script = str(task.param_probe_script or "").strip()
    if not probe_script:
        raise HTTPException(404, f"Task does not declare param_probe_script: {adapter_id}/{task_id}")

    try:
        script_path = adapter_loader.resolve_adapter_file(adapter_id, probe_script)
    except FileNotFoundError:
        raise HTTPException(404, f"Param probe script not found: {probe_script}")
    except ValueError as exc:
        raise HTTPException(400, str(exc))

    run_params = dict(req.params or {})
    current_tab_id = str(req.current_tab_id or "").strip()

    try:
        session = await open_browser_session(
            adapter_id,
            task_id,
            mode=str(run_params.get("mode") or "current").strip().lower() or "current",
            current_tab_id=current_tab_id,
        )
        prelude = (
            f"window.__CRAWSHRIMP_PARAMS__ = {json.dumps(run_params, ensure_ascii=False)};\n"
            f"window.__CRAWSHRIMP_PARAM_PROBE__ = true;\n"
        )
        result = await session.runner.evaluate_with_reconnect(
            prelude + script_path.read_text(encoding="utf-8"),
            allow_navigation_retry=True,
        )
    except ValueError as exc:
        raise HTTPException(404, str(exc))
    except RuntimeError as exc:
        raise HTTPException(400, str(exc))
    except Exception as exc:
        raise HTTPException(500, str(exc))

    if not result.success:
        raise HTTPException(400, result.error or "task param probe failed")

    return {
        "ok": True,
        "patches": result.data if isinstance(result.data, list) else [],
        "meta": result.meta or {},
    }


async def _run_scheduled_task(adapter_id: str, task_id: str):
    jid = f"{adapter_id}::{task_id}"
    if _task_is_active(jid):
        logger.warning("Skip scheduled task %s because a run is already active", jid)
        return

    lock = _task_lock(jid)
    async with lock:
        if _task_has_live_control(jid):
            logger.warning("Skip scheduled task %s because a manual run is already active", jid)
            return
        await _execute_task(adapter_id, task_id, {}, {}, run_control=None)


async def _run_task_schedule(schedule_uid: str):
    try:
        schedule_run = _create_instance_for_task_schedule(schedule_uid)
    except HTTPException as exc:
        logger.warning("Skip task schedule %s: %s", schedule_uid, exc.detail)
        return

    adapter_id = str(schedule_run.get("adapter_id") or "")
    task_id = str(schedule_run.get("task_id") or "")
    jid = f"{adapter_id}::{task_id}"
    if _task_is_active(jid):
        logger.warning("Skip task schedule %s because %s is already active", schedule_uid, jid)
        data_sink.update_task_instance(
            schedule_run["instance_uid"],
            status="failed",
            summary={"error": "任务正在运行中，本次定时触发已跳过", "schedule_uid": schedule_uid},
        )
        data_sink.record_task_schedule_run(
            schedule_uid,
            instance_uid=schedule_run["instance_uid"],
            status="skipped",
            error="任务正在运行中，本次定时触发已跳过",
        )
        return

    lock = _task_lock(jid)
    async with lock:
        if _task_has_live_control(jid):
            logger.warning("Skip task schedule %s because a manual run is active", schedule_uid)
            data_sink.update_task_instance(
                schedule_run["instance_uid"],
                status="failed",
                summary={"error": "任务正在运行中，本次定时触发已跳过", "schedule_uid": schedule_uid},
            )
            data_sink.record_task_schedule_run(
                schedule_uid,
                instance_uid=schedule_run["instance_uid"],
                status="skipped",
                error="任务正在运行中，本次定时触发已跳过",
            )
            return
        try:
            await _execute_task(
                adapter_id,
                task_id,
                dict(schedule_run.get("run_params") or {}),
                {},
                run_control=None,
            )
        except Exception as exc:
            logger.exception("Task schedule %s failed: %s", schedule_uid, exc)


async def _run_task_background(
    adapter_id: str,
    task_id: str,
    params: dict,
    runtime_options: dict,
    run_control: Optional[dict],
    runtime_operation_token: str = "",
):
    instance_uid = str((params or {}).get("__task_instance_uid") or "").strip()
    jid = _run_jid(adapter_id, task_id, instance_uid)
    control_jids = _run_jids(adapter_id, task_id, instance_uid)
    lock = _task_lock(_task_jid(adapter_id, task_id))
    try:
        async with lock:
            await _execute_task(adapter_id, task_id, params, runtime_options, run_control=run_control)
    except Exception as exc:
        live = dict(_run_status.get(jid) or {})
        if live.get('status') in ACTIVE_LIVE_STATUSES or not live:
            _run_status[jid] = {
                'status': 'error',
                'run_id': live.get('run_id'),
                'adapter_id': adapter_id,
                'task_id': task_id,
                'instance_uid': instance_uid,
                'records': int(live.get('records') or 0),
                'error': str(exc),
                'last_seen_at': datetime.now().isoformat(),
                'current_row': int(live.get('current_row') or live.get('row_no') or 0),
                'phase': live.get('phase') or 'error',
            }
        _run_logs.setdefault(jid, []).append(f"[{adapter_id}/{task_id}] FATAL: {exc}")
        logger.exception("Background task crashed before cleanup: %s", jid)
    finally:
        for control_jid in control_jids:
            if _run_controls.get(control_jid) is run_control:
                _run_controls.pop(control_jid, None)
        if runtime_operation_token:
            runtime_install_guard.end_operation(runtime_operation_token)


async def _start_task_run(adapter_id: str, task_id: str, params: Optional[dict] = None, runtime_options: Optional[dict] = None):
    adapter_loader.scan_all()
    m = adapter_loader.get_adapter(adapter_id)
    if not m:
        raise HTTPException(404, f"Adapter not found: {adapter_id}")
    if not any(t.id == task_id for t in m.tasks):
        raise HTTPException(404, f"Task not found: {task_id}")

    run_params = dict(params or {})
    instance_uid = str(run_params.get("__task_instance_uid") or "").strip()
    control_jids = _run_jids(adapter_id, task_id, instance_uid)
    if any(_task_is_active(control_jid) for control_jid in control_jids):
        raise HTTPException(409, "任务正在运行中，请先暂停/继续/停止当前任务")

    run_control = _build_run_control()
    run_control["adapter_id"] = adapter_id
    run_control["task_id"] = task_id
    run_control["instance_uid"] = instance_uid
    primary_jid = _run_jid(adapter_id, task_id, instance_uid)
    runtime_operation_token = runtime_install_guard.begin_operation(
        "task",
        primary_jid,
        primary_jid,
        "running",
    )
    background_coro = _run_task_background(
        adapter_id,
        task_id,
        run_params,
        runtime_options or {},
        run_control,
        runtime_operation_token,
    )
    try:
        task_handle = asyncio.create_task(background_coro)
    except BaseException:
        background_coro.close()
        runtime_install_guard.end_operation(runtime_operation_token)
        raise
    run_control['task'] = task_handle
    for control_jid in control_jids:
        _run_controls[control_jid] = run_control
    return {"ok": True, "message": "Task started in background"}


@app.post("/tasks/{adapter_id}/{task_id}/run")
async def run_task(adapter_id: str, task_id: str, req: RunTaskRequest = RunTaskRequest()):
    return await _start_task_run(
        adapter_id,
        task_id,
        req.params or {},
        {"current_tab_id": req.current_tab_id},
    )


@app.post("/task-instances/{instance_uid}/run")
async def run_task_instance(instance_uid: str, req: RunTaskRequest = RunTaskRequest()):
    instance = data_sink.get_task_instance(instance_uid)
    if not instance:
        raise HTTPException(404, "Task instance not found")
    params = _parse_json_object(instance.get("params_json"))
    if req.params:
        params.update(dict(req.params or {}))
    params["__task_instance_uid"] = instance_uid
    runtime_options = {"current_tab_id": req.current_tab_id}
    previous_status = str(instance.get("status") or "draft").strip() or "draft"
    data_sink.update_task_instance(instance_uid, status="running", current_step="config", params=params)
    try:
        result = await _start_task_run(instance["adapter_id"], instance["task_id"], params, runtime_options)
    except HTTPException as exc:
        message = getattr(exc, "detail", None) or str(exc)
        summary = _parse_json_object(instance.get("summary_json"))
        summary.update({"error": message})
        if exc.status_code == 409:
            data_sink.update_task_instance(instance_uid, status=previous_status, summary=summary)
        else:
            data_sink.update_task_instance(instance_uid, status="failed", current_step="config", summary=summary)
        raise
    except Exception as exc:
        message = getattr(exc, "detail", None) or str(exc)
        summary = _parse_json_object(instance.get("summary_json"))
        summary.update({"error": message})
        data_sink.update_task_instance(instance_uid, status="failed", current_step="config", summary=summary)
        raise
    return result


def _pause_run_jid(jid: str):
    control = _run_controls.get(jid)
    live = _run_status.get(jid) or {}
    task = control.get('task') if control else None
    if not control or not task or task.done() or live.get('status') not in ACTIVE_LIVE_STATUSES:
        raise HTTPException(409, "任务当前未在运行，无法暂停")
    if live.get('status') in {'paused', 'pausing'}:
        return {"ok": True, "status": live.get('status')}

    control['pause_requested'] = True
    control['resume_event'].clear()
    _merge_live_status([jid], {'status': 'pausing', 'last_seen_at': datetime.now().isoformat()})
    _append_run_log([jid], '[control] 收到暂停指令')
    return {"ok": True, "status": "pausing"}


def _resume_run_jid(jid: str):
    control = _run_controls.get(jid)
    live = _run_status.get(jid) or {}
    task = control.get('task') if control else None
    if not control or not task or task.done() or live.get('status') not in {'paused', 'pausing'}:
        raise HTTPException(409, "任务当前未暂停，无法继续")

    control['pause_requested'] = False
    control['resume_event'].set()
    control['pause_logged'] = False
    _merge_live_status([jid], {'status': 'running', 'last_seen_at': datetime.now().isoformat()})
    _append_run_log([jid], '[control] 收到继续指令')
    return {"ok": True, "status": "running"}


def _stop_run_jid(jid: str):
    control = _run_controls.get(jid)
    live = _run_status.get(jid) or {}
    task = control.get('task') if control else None
    if not control or not task or task.done() or live.get('status') not in ACTIVE_LIVE_STATUSES:
        raise HTTPException(409, "任务当前未在运行，无法停止")

    control['stop_requested'] = True
    control['pause_requested'] = False
    control['resume_event'].set()
    _merge_live_status([jid], {'status': 'stopping', 'last_seen_at': datetime.now().isoformat()})
    _append_run_log([jid], '[control] 收到停止指令')
    task.cancel()
    return {"ok": True, "status": "stopping"}


def _get_task_instance_or_404(instance_uid: str) -> dict:
    instance = data_sink.get_task_instance(instance_uid)
    if not instance:
        raise HTTPException(404, "Task instance not found")
    return instance


@app.post("/tasks/{adapter_id}/{task_id}/pause")
async def pause_task(adapter_id: str, task_id: str):
    return _pause_run_jid(_task_jid(adapter_id, task_id))


@app.post("/task-instances/{instance_uid}/pause")
async def pause_task_instance(instance_uid: str):
    _get_task_instance_or_404(instance_uid)
    return _pause_run_jid(_instance_jid(instance_uid))


@app.post("/tasks/{adapter_id}/{task_id}/resume")
async def resume_task(adapter_id: str, task_id: str):
    return _resume_run_jid(_task_jid(adapter_id, task_id))


@app.post("/task-instances/{instance_uid}/resume")
async def resume_task_instance(instance_uid: str):
    _get_task_instance_or_404(instance_uid)
    return _resume_run_jid(_instance_jid(instance_uid))


@app.post("/tasks/{adapter_id}/{task_id}/stop")
async def stop_task(adapter_id: str, task_id: str):
    return _stop_run_jid(_task_jid(adapter_id, task_id))


@app.post("/task-instances/{instance_uid}/stop")
async def stop_task_instance(instance_uid: str):
    _get_task_instance_or_404(instance_uid)
    return _stop_run_jid(_instance_jid(instance_uid))


@app.get("/tasks/{adapter_id}/{task_id}/status")
def task_status(adapter_id: str, task_id: str):
    jid = _task_jid(adapter_id, task_id)
    live = _run_status.get(jid)
    last = data_sink.get_latest_run(adapter_id, task_id)
    return {"live": live, "last_run": last}


@app.get("/task-instances/{instance_uid}/run-status")
def task_instance_run_status(instance_uid: str):
    _get_task_instance_or_404(instance_uid)
    jid = _instance_jid(instance_uid)
    live = _run_status.get(jid)
    last = _latest_instance_run(instance_uid)
    return {"live": live, "last_run": last}


@app.get("/tasks/{adapter_id}/{task_id}/logs")
def task_logs(adapter_id: str, task_id: str):
    jid = _task_jid(adapter_id, task_id)
    return {"logs": _run_logs.get(jid, [])}


@app.get("/task-instances/{instance_uid}/logs")
def task_instance_logs(instance_uid: str):
    _get_task_instance_or_404(instance_uid)
    return {"logs": _run_logs.get(_instance_jid(instance_uid), [])}


@app.delete("/tasks/{adapter_id}/{task_id}/logs")
def clear_task_logs(adapter_id: str, task_id: str):
    jid = _task_jid(adapter_id, task_id)
    _run_logs[jid] = []
    return {"ok": True}


@app.delete("/task-instances/{instance_uid}/logs")
def clear_task_instance_logs(instance_uid: str):
    _get_task_instance_or_404(instance_uid)
    _run_logs[_instance_jid(instance_uid)] = []
    return {"ok": True}


# ─── Data ───

@app.get("/data/{adapter_id}/{task_id}")
def get_data(adapter_id: str, task_id: str, limit: int = 20):
    runs = data_sink.list_runs(adapter_id, task_id, limit=limit)
    return {"runs": runs}


@app.get("/data/{adapter_id}/{task_id}/export")
def export_data(adapter_id: str, task_id: str, format: str = "excel"):
    last = data_sink.get_latest_run(adapter_id, task_id)
    if not last or not last.get('output_files'):
        raise HTTPException(404, "No exported data found. Run the task first.")
    import json as _json
    files = _json.loads(last['output_files']) if isinstance(last['output_files'], str) else last['output_files']
    for f in files:
        if format == 'excel' and f.endswith('.xlsx'):
            return FileResponse(f, filename=Path(f).name)
        if format == 'json' and f.endswith('.json'):
            return FileResponse(f, filename=Path(f).name)
    raise HTTPException(404, f"No {format} export found")



# ─── File utilities ───

class ReadExcelRequest(BaseModel):
    path: str
    sheet: Optional[str] = None   # 不传则读第一个 sheet
    header_row: int = 1           # 第几行作为表头（1-indexed）


class DeleteFilesRequest(BaseModel):
    paths: List[str]


class SyncDataFilesRequest(BaseModel):
    adapter_id: str
    task_id: str
    paths: List[str]
    endpoint: Optional[str] = None
    app_code: Optional[str] = None




@app.post("/files/read-excel")
def read_excel(req: ReadExcelRequest):
    """读取本地 Excel / CSV 文件，返回 headers + rows 数组，供 JS 脚本批量处理"""
    res = _read_local_excel(req.path, req.sheet, req.header_row)
    if "error" in res:
        raise HTTPException(400, res["error"])
    return res


@app.post("/files/delete")
def delete_files(req: DeleteFilesRequest):
    raw_paths = [str(path).strip() for path in (req.paths or []) if str(path).strip()]
    if not raw_paths:
        return {"ok": False, "error": "No file paths provided"}

    seen = set()
    normalized_paths = []
    for raw in raw_paths:
        path = _resolve_allowed_file_delete_path(raw)
        _ensure_registered_output_file_path(path)
        normalized = str(path)
        if normalized in seen:
            continue
        seen.add(normalized)
        normalized_paths.append(normalized)

    deleted_paths = []
    missing_paths = []
    failed_paths = []

    for item in normalized_paths:
        path = Path(item)
        try:
            if not path.exists():
                missing_paths.append(item)
                continue
            if path.is_dir():
                shutil.rmtree(path)
            else:
                path.unlink()
            deleted_paths.append(item)
        except Exception as e:
            failed_paths.append({"path": item, "error": str(e)})

    prune_targets = deleted_paths + missing_paths
    prune_result = data_sink.remove_output_files(prune_targets) if prune_targets else {"updated_runs": 0, "removed_refs": 0}

    return {
        "ok": bool(deleted_paths or missing_paths),
        "deleted_count": len(deleted_paths),
        "missing_count": len(missing_paths),
        "failed_count": len(failed_paths),
        "deleted_paths": deleted_paths,
        "missing_paths": missing_paths,
        "failed": failed_paths,
        "updated_runs": prune_result.get("updated_runs", 0),
        "removed_refs": prune_result.get("removed_refs", 0),
    }


@app.post("/data-sync/odps")
def sync_data_files_to_odps(req: SyncDataFilesRequest):
    cfg = load_config()
    odps_cfg = cfg.get("odps") if isinstance(cfg.get("odps"), dict) else {}
    endpoint = str(
        req.endpoint
        or os.environ.get("CRAWSHRIMP_ODPS_ENDPOINT")
        or odps_sync.DEFAULT_DATAWORKS_ENDPOINT
    ).strip()
    app_code = str(
        req.app_code
        or odps_cfg.get("app_code")
        or cfg.get("odps.app_code")
        or os.environ.get("CRAWSHRIMP_ODPS_APP_CODE")
        or ""
    ).strip()
    try:
        return odps_sync.sync_files(
            adapter_id=req.adapter_id,
            task_id=req.task_id,
            file_paths=req.paths,
            endpoint=endpoint,
            app_code=app_code,
        )
    except odps_sync.OdpsSyncError as e:
        raise HTTPException(400, str(e))


# ─── Settings ───

DEFAULT_CLOUD_APPROVAL_CAPABILITIES = ["generate_ai_image", "regenerate_ai_image", "submit_tmall_material_test", "crawl_tmall_material_test_data", "read_prompt_library"]


class CloudApprovalConfigRequest(BaseModel):
    base_url: str = ""
    registration_token: str = ""
    machine_name: str = ""
    machine_enabled: bool = False
    capabilities: list[str] = []


class CloudApprovalEnrollRequest(BaseModel):
    registration_token: str = ""
    machine_name: str = ""
    capabilities: list[str] = []


class CloudApprovalSyncBatchRequest(BaseModel):
    batch_id: str
    token: str


class CloudMachineLoopController:
    def __init__(self):
        self._lock = threading.Lock()
        self._thread: threading.Thread | None = None
        self._stop_event: threading.Event | None = None
        self.last_health = "stopped"
        self.last_error = ""

    def is_running(self) -> bool:
        with self._lock:
            if self._thread and self._thread.is_alive():
                return True
            self._thread = None
            self._stop_event = None
            if self.last_health == "running":
                self.last_health = "stopped"
            return False

    def start(self, agent: CloudMachineAgent) -> bool:
        with self._lock:
            if self._thread and self._thread.is_alive():
                return False
            stop_event = threading.Event()
            self._stop_event = stop_event
            self.last_health = "running"
            self.last_error = ""

            def run_loop():
                try:
                    agent.run_forever(stop_event)
                except Exception as exc:
                    logger.exception("cloud approval machine loop stopped with error")
                    self.last_health = "error"
                    self.last_error = str(exc)
                finally:
                    if stop_event.is_set() and self.last_health != "error":
                        self.last_health = "stopped"

            self._thread = threading.Thread(target=run_loop, name="cloud-approval-machine", daemon=True)
            self._thread.start()
            return True

    def stop(self) -> bool:
        with self._lock:
            thread = self._thread
            stop_event = self._stop_event
            if not thread or not thread.is_alive() or not stop_event:
                self._thread = None
                self._stop_event = None
                self.last_health = "stopped"
                return False
            stop_event.set()
        thread.join(timeout=2.0)
        with self._lock:
            if not thread.is_alive():
                self._thread = None
                self._stop_event = None
                self.last_health = "stopped"
        return True


cloud_machine_controller = CloudMachineLoopController()


def _cloud_approval_config() -> dict:
    cfg = load_config()
    cloud = cfg.get("cloud_approval") if isinstance(cfg.get("cloud_approval"), dict) else {}
    return cloud if isinstance(cloud, dict) else {}


def _cloud_approval_resolution(*, refresh: bool = False):
    cloud = _cloud_approval_config()
    return resolve_cloud_approval_url(
        str(cloud.get("base_url") or ""),
        refresh=refresh,
    )


def _cloud_capabilities(*sources) -> list[str]:
    for source in sources:
        raw = source.get("capabilities") if isinstance(source, dict) else source
        if not isinstance(raw, list):
            continue
        capabilities: list[str] = []
        for item in raw:
            text = str(item or "").strip()
            if text and text not in capabilities:
                capabilities.append(text)
        if capabilities:
            return capabilities
    return list(DEFAULT_CLOUD_APPROVAL_CAPABILITIES)


def _build_cloud_client() -> CloudApprovalClient:
    resolution = _cloud_approval_resolution()
    saved = data_sink.get_cloud_machine_credentials() or {}
    return CloudApprovalClient(
        base_url=resolution.base_url,
        machine_token=str(saved.get("machine_token") or ""),
        on_transport_error=invalidate_cloud_approval_url_cache,
    )


def _cloud_prompt_libraries(client: CloudApprovalClient | None = None) -> dict:
    active_client = client or _build_cloud_client()
    return active_client.request_json("GET", "/api/prompt-libraries", token_type="machine")


def _cloud_prompt_templates(
    library_id: str | int,
    *,
    category: str = "",
    gender: str = "",
    limit: int = 200,
    client: CloudApprovalClient | None = None,
) -> dict:
    active_client = client or _build_cloud_client()
    query = {
        "limit": max(1, min(int(limit or 200), 500)),
    }
    if str(category or "").strip():
        query["category"] = str(category or "").strip()
    if str(gender or "").strip():
        query["gender"] = str(gender or "").strip()
    path = f"/api/prompt-libraries/{library_id}/resolved?{urlencode(query)}"
    return active_client.request_json("GET", path, token_type="machine")


def _cloud_prompt_library_export(library_id: str | int, *, client: CloudApprovalClient | None = None) -> dict:
    active_client = client or _build_cloud_client()
    return active_client.request_json("GET", f"/api/prompt-libraries/{library_id}/export", token_type="machine")


def _begin_runtime_operation_or_skip(kind: str, operation_id: str, label: str, status: str = "running") -> str:
    try:
        return runtime_install_guard.begin_operation(kind, operation_id, label, status)
    except UpdateDrainActive:
        return ""


def _end_runtime_operation(token: str) -> None:
    runtime_install_guard.end_operation(token)


def _begin_cloud_job_operation() -> str:
    return _begin_runtime_operation_or_skip("cloud_job", "cloud_machine_claim", "云端任务机")


sched_module.set_runtime_operation_hooks(
    begin_operation=_begin_runtime_operation_or_skip,
    end_operation=_end_runtime_operation,
)


def _cloud_machine_agent(client: CloudApprovalClient) -> CloudMachineAgent:
    return CloudMachineAgent(
        client,
        app_version=API_VERSION,
        heartbeat_callback=lambda health: setattr(cloud_machine_controller, "last_health", str(health or "")),
        begin_job_operation=_begin_cloud_job_operation,
        end_job_operation=_end_runtime_operation,
    )


def _safe_cloud_enrollment_response(response: dict) -> dict:
    safe = dict(response or {})
    safe.pop("machine_token", None)
    return safe


def _cloud_approval_status(*, refresh: bool = False) -> dict:
    cloud = _cloud_approval_config()
    resolution = _cloud_approval_resolution(refresh=refresh)
    saved = data_sink.get_cloud_machine_credentials() or {}
    machine_name = str(cloud.get("machine_name") or saved.get("machine_name") or "").strip()
    machine_id = str(saved.get("machine_id") or "").strip()
    token_present = bool(str(saved.get("machine_token") or "").strip())
    running = cloud_machine_controller.is_running()
    health = "running" if running and cloud_machine_controller.last_health == "stopped" else cloud_machine_controller.last_health
    return {
        **resolution.status_fields(),
        "running": running,
        "auth": "enrolled" if token_present else "missing_token",
        "health": health or ("running" if running else "stopped"),
        "machine_name": machine_name,
        "machine_id": machine_id,
        "token_present": token_present,
        "machine_enabled": bool(cloud.get("machine_enabled")),
        "capabilities": _cloud_capabilities(saved, cloud),
        "last_error": cloud_machine_controller.last_error,
    }


@app.get("/cloud-approval/status")
def get_cloud_approval_status(refresh: bool = False):
    return _cloud_approval_status(refresh=refresh)


@app.post("/cloud-approval/config")
def configure_cloud_approval(req: CloudApprovalConfigRequest):
    patch_config({
        "cloud_approval": {
            **_cloud_approval_config(),
            "registration_token": str(req.registration_token or "").strip(),
            "machine_name": str(req.machine_name or "").strip(),
            "machine_enabled": bool(req.machine_enabled),
            "capabilities": _cloud_capabilities(req.capabilities),
        },
    })
    return {"ok": True, "status": _cloud_approval_status()}


@app.post("/cloud-approval/enroll-machine")
def enroll_cloud_machine(req: CloudApprovalEnrollRequest):
    cloud = _cloud_approval_config()
    resolution = _cloud_approval_resolution(refresh=True)
    if not resolution.configured:
        raise HTTPException(400, f"cloud approval address is invalid: {resolution.service_error}")
    if not resolution.service_reachable:
        raise HTTPException(502, f"cloud approval service is unavailable: {resolution.service_error}")
    registration_token = str(req.registration_token or cloud.get("registration_token") or "").strip()
    if not registration_token:
        raise HTTPException(400, "registration token is required")
    machine_name = str(req.machine_name or cloud.get("machine_name") or "").strip()
    if not machine_name:
        raise HTTPException(400, "machine name is required")
    capabilities = _cloud_capabilities(req.capabilities, cloud)
    client = CloudApprovalClient(
        base_url=resolution.base_url,
        on_transport_error=invalidate_cloud_approval_url_cache,
    )
    try:
        response = _cloud_machine_agent(client).enroll(registration_token, machine_name, capabilities)
    except CloudApprovalError as exc:
        raise HTTPException(502, str(exc)) from exc
    patch_config({
        "cloud_approval": {
            **cloud,
            "registration_token": "",
            "machine_name": machine_name,
            "capabilities": capabilities,
        },
    })
    return {**_safe_cloud_enrollment_response(response), "status": _cloud_approval_status()}


@app.post("/cloud-approval/sync-batch")
def sync_cloud_approval_batch(req: CloudApprovalSyncBatchRequest):
    batch = _load_tmall_approval_batch(req.batch_id)
    _validate_tmall_approval_token(batch, req.token)
    try:
        result = sync_local_approval_batch(batch, _build_cloud_client())
    except (CloudApprovalError, ValueError) as exc:
        raise HTTPException(502, str(exc)) from exc
    return {"ok": True, "result": result}


@app.get("/cloud-approval/prompt-libraries")
def get_cloud_prompt_libraries():
    try:
        return _cloud_prompt_libraries()
    except CloudApprovalError as exc:
        raise HTTPException(502, str(exc)) from exc


@app.get("/cloud-approval/prompt-libraries/{library_id}/resolved")
def get_cloud_prompt_templates(library_id: str, category: str = "", gender: str = "", limit: int = 200):
    try:
        return _cloud_prompt_templates(library_id, category=category, gender=gender, limit=limit)
    except CloudApprovalError as exc:
        raise HTTPException(502, str(exc)) from exc


@app.get("/cloud-approval/prompt-libraries/{library_id}/export")
def get_cloud_prompt_library_export(library_id: str):
    try:
        return _cloud_prompt_library_export(library_id)
    except CloudApprovalError as exc:
        raise HTTPException(502, str(exc)) from exc


@app.post("/cloud-approval/machine/start")
def start_cloud_machine():
    status = _cloud_approval_status()
    if not status["configured"]:
        raise HTTPException(400, "cloud approval base_url is required")
    if not status["service_reachable"]:
        raise HTTPException(502, f"cloud approval service is unavailable: {status['service_error']}")
    if not status["token_present"]:
        raise HTTPException(400, "machine enrollment is required before start")
    client = _build_cloud_client()
    cloud_machine_controller.start(_cloud_machine_agent(client))
    patch_config({"cloud_approval": {**_cloud_approval_config(), "machine_enabled": True}})
    return {"ok": True, "status": _cloud_approval_status()}


def _start_cloud_machine_if_enabled() -> bool:
    cloud = _cloud_approval_config()
    if not bool(cloud.get("machine_enabled")):
        return False
    status = _cloud_approval_status()
    if not status["configured"] or not status["service_reachable"] or not status["token_present"]:
        logger.warning("Cloud approval machine is enabled but not configured or enrolled; skip auto-start")
        return False
    return cloud_machine_controller.start(_cloud_machine_agent(_build_cloud_client()))


@app.post("/cloud-approval/machine/stop")
def stop_cloud_machine():
    cloud_machine_controller.stop()
    patch_config({"cloud_approval": {**_cloud_approval_config(), "machine_enabled": False}})
    return {"ok": True, "status": _cloud_approval_status()}


_AI_VIDEO_PRIVATE_SETTING_FIELDS = frozenset({
    "seedance_api_key",
    "seedance_base_url",
    "bailian_api_key",
    "bailian_workspace_id",
    "bailian_region",
    "bailian_base_url",
    "bailian_upload_api_key",
    "bailian_uploads_url",
})


def _public_settings() -> dict:
    """Return renderer-safe settings without provider credentials or routing data."""
    saved = load_config()
    public = json.loads(json.dumps(saved, ensure_ascii=False))
    ai = public.get("ai") if isinstance(public.get("ai"), dict) else {}
    source_ai = saved.get("ai") if isinstance(saved.get("ai"), dict) else {}
    source_video = source_ai.get("video") if isinstance(source_ai.get("video"), dict) else {}
    ai["video"] = {
        "seedance_configured": bool(str(source_video.get("seedance_api_key") or "").strip()),
        "happyhorse_configured": bool(str(source_video.get("bailian_api_key") or "").strip()),
        "bailian_upload_configured": bool(str(source_video.get("bailian_upload_api_key") or "").strip()),
    }
    public["ai"] = ai
    return public


def _safe_settings_write_patch(cfg: dict) -> dict:
    """Treat blank write-only AI-video fields as unchanged, never as secret erasure."""
    patch = json.loads(json.dumps(cfg if isinstance(cfg, dict) else {}, ensure_ascii=False))
    for field in _AI_VIDEO_PRIVATE_SETTING_FIELDS:
        dotted = f"ai.video.{field}"
        if dotted in patch and not str(patch.get(dotted) or "").strip():
            patch.pop(dotted, None)
    ai = patch.get("ai") if isinstance(patch.get("ai"), dict) else None
    video = ai.get("video") if isinstance(ai, dict) and isinstance(ai.get("video"), dict) else None
    if isinstance(video, dict):
        for field in _AI_VIDEO_PRIVATE_SETTING_FIELDS:
            if field in video and not str(video.get(field) or "").strip():
                video.pop(field, None)
        # These are response-only status flags and must never be persisted.
        video.pop("seedance_configured", None)
        video.pop("happyhorse_configured", None)
        video.pop("bailian_upload_configured", None)
    return patch


@app.get("/settings")
def get_settings():
    return _public_settings()


@app.put("/settings")
def put_settings(cfg: dict):
    # Renderer reads a redacted document, so PUT must merge rather than erase
    # provider secrets that were intentionally omitted from that document.
    patch_config(_safe_settings_write_patch(cfg))
    reset_bridge()
    return {"ok": True}


@app.patch("/settings")
def patch_settings(cfg: dict):
    patch_config(_safe_settings_write_patch(cfg))
    reset_bridge()
    return {"ok": True}


class TestNotifyRequest(BaseModel):
    channel: str  # dingtalk | feishu | webhook


@app.post("/settings/test-notify")
def test_notify(req: TestNotifyRequest):
    try:
        notifier.send(
            channel=req.channel,
            title="🦐 抓虾通知测试",
            records=0,
            adapter_name="crawshrimp",
            task_name="测试通知",
            sample_rows=[{"状态": "✅ 通知配置正常", "渠道": req.channel}],
        )
        return {"ok": True, "msg": f"{req.channel} 通知发送成功"}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get("/settings/chrome-tabs")
def chrome_tabs():
    try:
        tabs = get_bridge().get_tabs()
        return [{'id': t.get('id'), 'url': t.get('url'), 'title': t.get('title')}
                for t in tabs if t.get('type') == 'page']
    except ConnectionError as e:
        raise HTTPException(503, str(e))


if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("CRAWSHRIMP_PORT", 18765))
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s"
    )
    uvicorn.run(app, host="127.0.0.1", port=port, access_log=False)
