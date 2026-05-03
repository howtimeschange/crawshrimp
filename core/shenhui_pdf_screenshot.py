"""Shenhui PDF screenshot packaging helpers.

This module is intentionally independent from ``api_server.py``.  The API
server should only route task completion into this service; PDF rendering,
cropping, 800px canvas composition, and naming rules live here.
"""

from __future__ import annotations

import json
import logging
import re
import shutil
import subprocess
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)

CANVAS_SIZE = 800
CANVAS_PADDING = 24
PDF_RENDER_TARGET_LONG_EDGE = 3600
PDF_RENDER_MIN_SCALE = 3.0
PDF_RENDER_MAX_SCALE = 8.0
DEFAULT_WASH_CROP_BOXES = [(0.0892, 0.2084, 0.4189, 0.7546)]
DEFAULT_TAG_CROP_BOXES = [(0.0113, 0.2352, 0.1535, 0.5058)]


def safe_local_name(value: str, fallback: str = "item") -> str:
    text = str(value or "").strip()
    text = re.sub(r"[\x00-\x1f]+", "", text)
    text = re.sub(r'[\\/:*?"<>|]+', "_", text)
    text = re.sub(r"\s+", " ", text).strip(" ._")
    return text or fallback


def ensure_unique_local_path(path: Path) -> Path:
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    index = 2
    while True:
        candidate = path.with_name(f"{stem}_{index}{suffix}")
        if not candidate.exists():
            return candidate
        index += 1


def ensure_unique_local_dir(path: Path) -> Path:
    if not path.exists():
        return path
    index = 2
    while True:
        candidate = path.with_name(f"{path.name}_{index}")
        if not candidate.exists():
            return candidate
        index += 1


def copy_file_to_unique_target(source: Path, target: Path) -> Path:
    target.parent.mkdir(parents=True, exist_ok=True)
    unique_target = ensure_unique_local_path(target)
    shutil.copy2(source, unique_target)
    return unique_target


def _normalize_code_text(value: str) -> str:
    return str(value or "").replace("－", "-").replace("–", "-").replace("—", "-")


def extract_style_color_code(source_name: str = "", text: str = "") -> tuple[str, str]:
    """Extract ``12-digit style`` and ``5-digit color`` from PDF text/name.

    Hang-tag PDFs usually expose a visible ``201226109105-00322`` code in text.
    Wash-label PDFs may only carry the color in the filename, for example
    ``201226109105-2（88201牛仔中蓝）洗唛.pdf``.
    """

    normalized_text = _normalize_code_text(text)
    normalized_name = _normalize_code_text(source_name)

    direct_code = re.search(r"(?<!\d)(\d{12})\s*-\s*(\d{5})(?!\d)", normalized_text)
    if direct_code:
        return direct_code.group(1), direct_code.group(2)

    filename_code = re.search(r"(?<!\d)(\d{12})(?!\d).{0,80}?(?<!\d)(\d{5})(?!\d)", normalized_name)
    if filename_code:
        return filename_code.group(1), filename_code.group(2)

    fallback_style = re.search(r"(?<!\d)(\d{12})(?!\d)", f"{normalized_text}\n{normalized_name}")
    if fallback_style:
        return fallback_style.group(1), ""

    return "", ""


def build_style_color_base(style_code: str, color_code: str, fallback: str) -> str:
    style = str(style_code or "").strip()
    color = str(color_code or "").strip()
    if style and color:
        return safe_local_name(f"{style}-{color}", safe_local_name(fallback, "pdf"))
    return safe_local_name(fallback, "pdf")


def split_style_color_code(value: str) -> tuple[str, str]:
    match = re.fullmatch(r"\s*(\d{12})(?:-(\d{5}))?\s*", _normalize_code_text(value))
    if not match:
        return "", ""
    return match.group(1), match.group(2) or ""


def _pdf_role(pdf_type: str) -> str:
    return "wash_label" if str(pdf_type or "").strip().lower() == "wash_label" else "hang_tag"


def build_pdf_yq_filename(
    style_code: str,
    color_code: str,
    pdf_type: str,
    sequence: int,
    has_multiple_colors: bool,
    extension: str = ".png",
) -> str:
    index = max(1, int(sequence or 1))
    ext = extension if str(extension or "").startswith(".") else f".{extension or 'png'}"
    token = "yq(2)" if _pdf_role(pdf_type) == "wash_label" else "yq(1)"
    style = str(style_code or "").strip()
    color = str(color_code or "").strip()

    if has_multiple_colors:
        prefix = f"{style}-{color}" if style and color else style
        fallback = f"{token}-{index}{ext}"
        return safe_local_name(f"{prefix}{token}-{index}{ext}", fallback)

    suffix = "" if index == 1 else f"-{index}"
    return safe_local_name(f"{token}{suffix}{ext}", f"yq-{index}{ext}")


def build_yq_filename(base_code: str, sequence: int, extension: str = ".png") -> str:
    index = max(1, int(sequence or 1))
    ext = extension if str(extension or "").startswith(".") else f".{extension or 'png'}"
    return safe_local_name(f"{base_code}yq-{index}{ext}", f"yq-{index}.png")


def _is_style_color_code(value: str) -> bool:
    return bool(re.fullmatch(r"\d{12}-\d{5}", str(value or "").strip()))


def parse_style_color_overrides(raw_value) -> dict[str, str]:
    """Parse manual filename-to-style-color mappings.

    Supported lines:
    - ``hang-tag.pdf=201226109105-00322``
    - ``hang-tag.pdf,201226109105-00322``
    - ``201226109105-00322`` as a default for rows without a mapping
    """

    text = str(raw_value or "").strip()
    if not text:
        return {}

    overrides: dict[str, str] = {}
    for raw_line in re.split(r"[\r\n]+", text):
        line = raw_line.strip()
        if not line:
            continue
        if "=" in line:
            key, code = line.split("=", 1)
        elif "," in line:
            key, code = line.split(",", 1)
        elif "，" in line:
            key, code = line.split("，", 1)
        elif _is_style_color_code(line):
            key, code = "__default__", line
        else:
            continue

        clean_code = _normalize_code_text(code).strip()
        if not _is_style_color_code(clean_code):
            continue
        clean_key = str(key or "__default__").strip()
        overrides[clean_key] = clean_code
    return overrides


def _style_color_override_for_path(pdf_path: Path, run_params: dict) -> str:
    overrides = parse_style_color_overrides((run_params or {}).get("style_color_overrides"))
    if not overrides:
        return ""
    candidates = [
        str(pdf_path),
        pdf_path.name,
        pdf_path.stem,
    ]
    normalized = {str(key).strip().lower(): value for key, value in overrides.items()}
    for candidate in candidates:
        value = normalized.get(str(candidate).strip().lower())
        if value:
            return value
    return overrides.get("__default__", "")


def _coerce_number(value) -> Optional[float]:
    try:
        number = float(value)
    except Exception:
        return None
    return number if number == number else None


def parse_crop_boxes(raw_value) -> list[tuple[float, float, float, float]]:
    """Parse crop boxes as ``x,y,width,height`` values.

    Values can be normalized ratios (0-1) or pixels.  Accepted input forms:
    ``"0.1,0.2,0.4,0.5"``, a JSON array, a dict with x/y/width/height, or a
    list of those dicts/arrays.  This is the backend contract for the future
    visual box-selection UI.
    """

    if raw_value is None:
        return []
    parsed = raw_value
    if isinstance(raw_value, str):
        text = raw_value.strip()
        if not text:
            return []
        try:
            parsed = json.loads(text)
        except Exception:
            parsed = [part.strip() for part in re.split(r"[,，\s]+", text) if part.strip()]

    if isinstance(parsed, dict):
        parsed = [parsed]

    if not isinstance(parsed, list):
        return []

    if parsed and all(_coerce_number(item) is not None for item in parsed) and len(parsed) >= 4:
        parsed = [parsed[:4]]

    boxes = []
    for item in parsed:
        if isinstance(item, dict):
            x = _coerce_number(item.get("x", item.get("left")))
            y = _coerce_number(item.get("y", item.get("top")))
            width = _coerce_number(item.get("width", item.get("w")))
            height = _coerce_number(item.get("height", item.get("h")))
            right = _coerce_number(item.get("right"))
            bottom = _coerce_number(item.get("bottom"))
            if width is None and x is not None and right is not None:
                width = right - x
            if height is None and y is not None and bottom is not None:
                height = bottom - y
        elif isinstance(item, (list, tuple)) and len(item) >= 4:
            x, y, width, height = (_coerce_number(value) for value in item[:4])
        else:
            continue

        if None in (x, y, width, height) or width <= 0 or height <= 0:
            continue
        boxes.append((float(x), float(y), float(width), float(height)))

    return boxes


def _crop_box_to_pixels(image, box: tuple[float, float, float, float]) -> tuple[int, int, int, int]:
    x, y, width, height = box
    normalized = all(0 <= value <= 1.000001 for value in (x, y, width, height)) and x + width <= 1.000001 and y + height <= 1.000001
    if normalized:
        left = int(round(x * image.width))
        top = int(round(y * image.height))
        right = int(round((x + width) * image.width))
        bottom = int(round((y + height) * image.height))
    else:
        left = int(round(x))
        top = int(round(y))
        right = int(round(x + width))
        bottom = int(round(y + height))

    left = max(0, min(image.width - 1, left))
    top = max(0, min(image.height - 1, top))
    right = max(left + 1, min(image.width, right))
    bottom = max(top + 1, min(image.height, bottom))
    return left, top, right, bottom


def _crop_whitespace(image, pad: int = 12):
    try:
        from PIL import Image, ImageChops
    except Exception:
        return image

    background = Image.new(image.mode, image.size, image.getpixel((0, 0)))
    diff = ImageChops.difference(image, background)
    bbox = diff.getbbox()
    if not bbox:
        return image
    left = max(0, bbox[0] - pad)
    top = max(0, bbox[1] - pad)
    right = min(image.width, bbox[2] + pad)
    bottom = min(image.height, bbox[3] + pad)
    return image.crop((left, top, right, bottom))


def _split_wide_label(image):
    if image.width < image.height * 1.25:
        return [image]

    gray = image.convert("L")
    columns = []
    for x in range(gray.width):
        dark = 0
        for y in range(gray.height):
            if gray.getpixel((x, y)) < 245:
                dark += 1
        columns.append(dark)

    low_threshold = max(2, int(gray.height * 0.01))
    min_gap_width = max(8, int(gray.width * 0.012))
    gaps = []
    run_start = None
    for x, dark in enumerate(columns):
        if dark <= low_threshold:
            if run_start is None:
                run_start = x
        elif run_start is not None:
            if x - run_start >= min_gap_width:
                gaps.append((run_start, x))
            run_start = None
    if run_start is not None and gray.width - run_start >= min_gap_width:
        gaps.append((run_start, gray.width))

    split_points = [
        (start + end) // 2
        for start, end in gaps
        if gray.width * 0.08 <= ((start + end) // 2) <= gray.width * 0.92
    ]
    if not split_points:
        return [image]

    edges = [0, *split_points, image.width]
    parts = []
    for left, right in zip(edges, edges[1:]):
        if right - left < max(20, int(image.width * 0.05)):
            continue
        part = _crop_whitespace(image.crop((left, 0, right, image.height)))
        if part.width > 20 and part.height > 20:
            parts.append(part)
    return parts or [image]


def _fit_on_white_canvas(image, canvas_size: int = CANVAS_SIZE):
    try:
        from PIL import Image
    except Exception:
        return image

    source = image.convert("RGB")
    canvas = Image.new("RGB", (canvas_size, canvas_size), "white")
    max_size = max(1, canvas_size - CANVAS_PADDING * 2)
    scale = min(max_size / source.width, max_size / source.height)
    target_width = max(1, int(round(source.width * scale)))
    target_height = max(1, int(round(source.height * scale)))

    resampling = getattr(getattr(Image, "Resampling", Image), "LANCZOS", Image.BICUBIC)
    resized = source.resize((target_width, target_height), resampling)
    left = (canvas_size - target_width) // 2
    top = (canvas_size - target_height) // 2
    canvas.paste(resized, (left, top))
    return canvas


def crop_image_to_yq_pages(
    image_path: Path,
    output_dir: Path,
    stem: str = "yq",
    *,
    crop_boxes: Optional[list[tuple[float, float, float, float]]] = None,
    canvas_size: int = CANVAS_SIZE,
) -> list[Path]:
    try:
        from PIL import Image
    except Exception:
        return [image_path]

    try:
        output_dir.mkdir(parents=True, exist_ok=True)
        base = Image.open(image_path).convert("RGB")
        selected_parts = []

        boxes = crop_boxes or []
        if boxes:
            for box in boxes:
                selected_parts.append(_crop_whitespace(base.crop(_crop_box_to_pixels(base, box))))
        else:
            cropped = _crop_whitespace(base)
            selected_parts = _split_wide_label(cropped)

        outputs = []
        for index, part in enumerate(selected_parts, start=1):
            suffix = "" if index == 1 else f"_{index}"
            target = ensure_unique_local_path(output_dir / f"{stem}{suffix}.png")
            _fit_on_white_canvas(part, canvas_size=canvas_size).save(target, "PNG")
            outputs.append(target)
        return outputs or [image_path]
    except Exception:
        logger.debug("Failed to crop shenhui PDF preview image %s", image_path, exc_info=True)
        return [image_path]


def _pdf_page_render_scale(page) -> float:
    rect = page.rect
    long_edge = max(float(rect.width or 0), float(rect.height or 0), 1.0)
    target_scale = PDF_RENDER_TARGET_LONG_EDGE / long_edge
    return max(PDF_RENDER_MIN_SCALE, min(PDF_RENDER_MAX_SCALE, target_scale))


def render_pdf_pages_with_pymupdf(pdf_path: Path, output_dir: Path) -> list[Path]:
    try:
        import fitz
    except Exception:
        return []

    try:
        output_dir.mkdir(parents=True, exist_ok=True)
        doc = fitz.open(str(pdf_path))
        if doc.page_count < 1:
            doc.close()
            return []
        outputs = []
        for page_index in range(doc.page_count):
            page = doc.load_page(page_index)
            scale = _pdf_page_render_scale(page)
            pixmap = page.get_pixmap(matrix=fitz.Matrix(scale, scale), alpha=False)
            page_suffix = "" if doc.page_count == 1 else f"_p{page_index + 1}"
            target = ensure_unique_local_path(output_dir / f"{pdf_path.stem}{page_suffix}.png")
            pixmap.save(str(target))
            if target.is_file():
                outputs.append(target)
        doc.close()
        return outputs
    except Exception:
        logger.debug("Failed to render PDF with PyMuPDF: %s", pdf_path, exc_info=True)
        return []


def render_pdf_with_pymupdf(pdf_path: Path, output_dir: Path) -> Optional[Path]:
    pages = render_pdf_pages_with_pymupdf(pdf_path, output_dir)
    return pages[0] if pages else None


def render_pdf_with_quicklook(pdf_path: Path, output_dir: Path) -> Optional[Path]:
    if shutil.which("qlmanage") is None:
        return None
    try:
        output_dir.mkdir(parents=True, exist_ok=True)
        subprocess.run(
            ["qlmanage", "-t", "-s", "1800", "-o", str(output_dir), str(pdf_path)],
            check=False,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            timeout=45,
        )
        candidates = [
            output_dir / f"{pdf_path.name}.png",
            output_dir / f"{pdf_path.stem}.png",
        ]
        candidates.extend(sorted(output_dir.glob(f"{pdf_path.name}*.png")))
        return next((path for path in candidates if path.is_file()), None)
    except Exception:
        logger.debug("Failed to render PDF with Quick Look: %s", pdf_path, exc_info=True)
        return None


def render_pdf_pages_with_quicklook(pdf_path: Path, output_dir: Path) -> list[Path]:
    rendered = render_pdf_with_quicklook(pdf_path, output_dir)
    return [rendered] if rendered else []


def extract_pdf_text(pdf_path: Path) -> str:
    try:
        import fitz
    except Exception:
        return ""

    try:
        doc = fitz.open(str(pdf_path))
        chunks = [page.get_text("text") for page in doc]
        doc.close()
        return "\n".join(chunks)
    except Exception:
        logger.debug("Failed to extract PDF text: %s", pdf_path, exc_info=True)
        return ""


def infer_pdf_type(pdf_path: Path, row: dict, run_params: dict) -> str:
    requested = str(run_params.get("pdf_type") or row.get("__pdf_type") or "").strip().lower()
    if requested in {"wash_label", "hang_tag"}:
        return requested
    name = f"{pdf_path.name} {row.get('PDF类型') or ''}"
    if any(token in name for token in ("洗唛", "洗标", "水洗")):
        return "wash_label"
    if any(token in name for token in ("吊牌", "合格证")):
        return "hang_tag"
    return "auto"


def crop_boxes_for_pdf_type(pdf_type: str, run_params: dict) -> list[tuple[float, float, float, float]]:
    if pdf_type == "wash_label":
        boxes = parse_crop_boxes(run_params.get("wash_crop_boxes") or run_params.get("wash_crop_box"))
        if boxes:
            return boxes
        return list(DEFAULT_WASH_CROP_BOXES)
    if pdf_type == "hang_tag":
        boxes = parse_crop_boxes(run_params.get("tag_crop_boxes") or run_params.get("tag_crop_box"))
        if boxes:
            return boxes
        return list(DEFAULT_TAG_CROP_BOXES)
    return parse_crop_boxes(run_params.get("crop_boxes") or run_params.get("crop_box"))


def convert_pdf_to_yq_images(
    pdf_path: Path,
    work_dir: Path,
    log,
    *,
    crop_boxes: Optional[list[tuple[float, float, float, float]]] = None,
) -> list[Path]:
    try:
        with pdf_path.open("rb") as handle:
            if handle.read(5) != b"%PDF-":
                log(f"[warn] PDF 文件头无效，保留原文件待人工处理: {pdf_path.name}")
                return []
    except Exception:
        log(f"[warn] PDF 文件无法读取，保留原文件待人工处理: {pdf_path.name}")
        return []

    render_dir = work_dir / "rendered"
    crop_dir = work_dir / "cropped"
    rendered_pages = render_pdf_pages_with_pymupdf(pdf_path, render_dir) or render_pdf_pages_with_quicklook(pdf_path, render_dir)
    if not rendered_pages:
        log(f"[warn] PDF 裁图不可用，保留原 PDF 待人工处理: {pdf_path.name}")
        return []

    outputs = []
    for page_index, rendered in enumerate(rendered_pages, start=1):
        outputs.extend(
            crop_image_to_yq_pages(
                rendered,
                crop_dir,
                f"yq_p{page_index}",
                crop_boxes=crop_boxes,
            )
        )
    log(f"Shenhui PDF cropped: {pdf_path.name} -> {len(outputs)} image(s) from {len(rendered_pages)} page(s)")
    return [path for path in outputs if path.is_file()]


def _style_color_for_row(row: dict, pdf_path: Path, run_params: dict) -> tuple[str, str]:
    explicit_code = str(row.get("__style_color_code") or row.get("款色编码") or "").strip()
    style_code, color_code = split_style_color_code(explicit_code)
    if style_code:
        return style_code, color_code

    explicit_style = str(row.get("__style_code") or row.get("款号") or "").strip()
    style_code, color_code = split_style_color_code(explicit_style)
    if style_code:
        return style_code, color_code

    override = _style_color_override_for_path(pdf_path, run_params)
    style_code, color_code = split_style_color_code(override)
    if style_code:
        return style_code, color_code

    text = extract_pdf_text(pdf_path)
    return extract_style_color_code(source_name=pdf_path.name, text=text)


def _pdf_output_mode(run_params: dict) -> str:
    raw = str((run_params or {}).get("output_mode") or "").strip().lower()
    if raw in {"merge_existing_zips", "merge_zip", "append_to_zip", "append_zip"}:
        return "merge_existing_zips"
    return "create_package"


def _pdf_path_for_row(row: dict) -> Path:
    raw = str(row.get("__pdf_path") or row.get("原始路径") or "").strip()
    if not raw and str(row.get("__shenhui_asset_role") or "").strip().lower() == "pdf_yq":
        raw = str(row.get("本地文件") or "").strip()
    return Path(raw).expanduser() if raw else Path()


def convert_pdf_rows_to_yq_output_root(
    data_rows: list,
    output_root: Path,
    pdf_work_dir: Path,
    run_params: dict,
    log,
) -> int:
    output_root.mkdir(parents=True, exist_ok=True)
    pdf_work_dir.mkdir(parents=True, exist_ok=True)

    work_items = []
    colors_by_style_role: dict[tuple[str, str], set[str]] = {}
    for row in data_rows or []:
        if not isinstance(row, dict):
            continue
        pdf_path = _pdf_path_for_row(row)
        if not str(pdf_path):
            continue
        if pdf_path.suffix.lower() != ".pdf":
            continue
        row["__pdf_path"] = str(pdf_path)
        if not pdf_path.is_file():
            log(f"[warn] PDF 文件不存在，跳过截图: {pdf_path}")
            row["处理动作"] = "截图失败"
            row["备注"] = f"PDF 文件不存在：{pdf_path}"
            continue

        pdf_type = infer_pdf_type(pdf_path, row, run_params or {})
        style_code, color_code = _style_color_for_row(row, pdf_path, run_params or {})
        style_folder = safe_local_name(style_code or row.get("__shenhui_group_code") or pdf_path.stem, "pdf")
        if style_code and color_code:
            colors_by_style_role.setdefault((style_code, _pdf_role(pdf_type)), set()).add(color_code)
        work_items.append({
            "row": row,
            "pdf_path": pdf_path,
            "pdf_type": pdf_type,
            "style_code": style_code,
            "color_code": color_code,
            "style_folder": style_folder,
        })

    converted_count = 0
    sequence_by_name_scope: dict[tuple[str, str, str], int] = {}
    for item in work_items:
        pdf_path = item["pdf_path"]
        pdf_type = item["pdf_type"]
        style_code = item["style_code"]
        color_code = item["color_code"]
        style_folder = item["style_folder"]
        has_multiple_colors = bool(
            style_code
            and any(
                len(colors) > 1
                for (role_style_code, _role), colors in colors_by_style_role.items()
                if role_style_code == style_code
            )
        )
        crop_boxes = crop_boxes_for_pdf_type(pdf_type, run_params or {})
        target_dir = output_root / style_folder
        converted = convert_pdf_to_yq_images(
            pdf_path,
            pdf_work_dir / style_folder / safe_local_name(pdf_path.stem, "pdf"),
            log,
            crop_boxes=crop_boxes,
        )
        if not converted:
            log(f"[warn] PDF 未生成截图: {pdf_path.name}")
            item["row"]["处理动作"] = "截图失败"
            item["row"]["备注"] = "PDF 未生成截图，请检查文件内容或截图框模板"
            continue

        sequence_key = (
            style_folder,
            _pdf_role(pdf_type),
            color_code if has_multiple_colors else "",
        )
        next_sequence = sequence_by_name_scope.get(sequence_key, 1)
        copied_count = 0
        for converted_path in converted:
            target_name = build_pdf_yq_filename(
                style_code or style_folder,
                color_code,
                pdf_type,
                next_sequence,
                has_multiple_colors,
            )
            copied = copy_file_to_unique_target(converted_path, target_dir / target_name)
            converted_count += 1
            copied_count += 1
            next_sequence += 1
            log(f"Shenhui PDF screenshot copied: {copied}")
        sequence_by_name_scope[sequence_key] = next_sequence
        item["row"]["处理动作"] = "截图完成"
        item["row"]["备注"] = f"生成 {copied_count} 张截图；款号目录：{style_folder}"

    return converted_count


def _find_style_zip(target_root: Path, style_code: str) -> Optional[Path]:
    style = safe_local_name(style_code, "")
    if not style:
        return None

    direct = target_root / f"{style}.zip"
    if direct.is_file():
        return direct

    matches = sorted(
        path for path in target_root.rglob("*.zip")
        if path.is_file() and path.stem == style
    )
    return matches[0] if matches else None


def _replace_files_in_zip(zip_path: Path, source_dir: Path) -> int:
    files = sorted(path for path in source_dir.rglob("*") if path.is_file())
    if not files:
        return 0

    replacements = {
        str(path.relative_to(source_dir)).replace("\\", "/"): path
        for path in files
    }
    temp_zip = ensure_unique_local_path(zip_path.with_suffix(".tmp.zip"))
    try:
        with zipfile.ZipFile(zip_path, "r") as source, zipfile.ZipFile(temp_zip, "w", compression=zipfile.ZIP_DEFLATED) as target:
            replaced_names = set(replacements)
            copied_names = set()
            for info in source.infolist():
                normalized_name = str(info.filename).replace("\\", "/")
                if normalized_name in replaced_names:
                    continue
                if normalized_name in copied_names:
                    continue
                target.writestr(info, source.read(info.filename))
                copied_names.add(normalized_name)

            for arcname, source_file in replacements.items():
                target.write(source_file, arcname=arcname)
        temp_zip.replace(zip_path)
        return len(replacements)
    finally:
        if temp_zip.exists():
            temp_zip.unlink(missing_ok=True)


def _merge_pdf_outputs_into_existing_style_zips(output_root: Path, target_root: Path, log) -> list[str]:
    merged_refs = []
    for style_dir in sorted(path for path in output_root.iterdir() if path.is_dir()):
        target_zip = _find_style_zip(target_root, style_dir.name)
        if not target_zip:
            log(f"[warn] 未找到同款号 ZIP，跳过合并: {style_dir.name}.zip")
            continue
        replaced_count = _replace_files_in_zip(target_zip, style_dir)
        if replaced_count:
            merged_refs.append(str(target_zip))
            log(f"Shenhui PDF screenshots merged into ZIP: {target_zip} ({replaced_count} file(s))")
    return merged_refs


def _rewrite_pdf_summary_excels(exported_files: list, data_rows: list, log) -> None:
    public_rows = [row for row in data_rows or [] if isinstance(row, dict)]
    if not public_rows:
        return

    for file_path in exported_files or []:
        path = Path(str(file_path or "")).expanduser()
        if path.suffix.lower() != ".xlsx" or not path.is_file():
            continue
        try:
            from openpyxl import load_workbook

            workbook = load_workbook(path)
            for worksheet in workbook.worksheets:
                header_cells = list(worksheet[1])
                header_map = {
                    str(cell.value or "").strip(): cell.column
                    for cell in header_cells
                    if str(cell.value or "").strip()
                }
                target_columns = [
                    column_name for column_name in ("处理动作", "备注")
                    if column_name in header_map
                ]
                if not target_columns:
                    continue
                for row_index, source_row in enumerate(public_rows, start=2):
                    if row_index > worksheet.max_row:
                        break
                    for column_name in target_columns:
                        value = source_row.get(column_name, "")
                        worksheet.cell(row=row_index, column=header_map[column_name]).value = "" if value is None else str(value)
            workbook.save(path)
            log(f"Shenhui PDF summary Excel updated: {path}")
        except Exception:
            logger.debug("Failed to rewrite Shenhui PDF summary Excel %s", path, exc_info=True)
            log(f"[warn] PDF 截图结果 Excel 状态回写失败: {path.name}")


def finalize_pdf_batch_screenshot_outputs(
    data_rows: list,
    exported_files: list,
    run_params: dict,
    runtime_artifact_dir: str,
    log,
) -> list[str]:
    runtime_dir = Path(runtime_artifact_dir)
    runtime_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    package_base = safe_local_name(
        run_params.get("package_name") or f"深绘PDF截图_{timestamp}",
        f"深绘PDF截图_{timestamp}",
    )
    output_root = ensure_unique_local_dir(runtime_dir / package_base)
    output_root.mkdir(parents=True, exist_ok=True)
    pdf_work_dir = ensure_unique_local_dir(runtime_dir / f"{output_root.name}_pdf_work")

    converted_count = convert_pdf_rows_to_yq_output_root(
        data_rows=data_rows,
        output_root=output_root,
        pdf_work_dir=pdf_work_dir,
        run_params=run_params or {},
        log=log,
    )

    zip_path = None
    if converted_count:
        zip_path = ensure_unique_local_path(runtime_dir / f"{output_root.name}.zip")
        with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            for file_path in sorted(output_root.rglob("*")):
                if not file_path.is_file():
                    continue
                archive.write(file_path, arcname=str(file_path.relative_to(output_root.parent)))
        log(f"Shenhui PDF screenshot package created: {zip_path}")

    final_refs = []
    if zip_path:
        final_refs.append(str(zip_path))
    final_refs.extend(str(path) for path in exported_files or [] if str(path or "").strip())

    _rewrite_pdf_summary_excels(exported_files, data_rows, log)

    export_folder = str((run_params or {}).get("export_folder") or "").strip()
    if export_folder:
        target_root = Path(export_folder).expanduser()
        target_root.mkdir(parents=True, exist_ok=True)
        external_refs = []
        if converted_count and output_root.exists() and _pdf_output_mode(run_params or {}) == "merge_existing_zips":
            external_refs.extend(_merge_pdf_outputs_into_existing_style_zips(output_root, target_root, log))
        elif converted_count and output_root.exists():
            external_dir = ensure_unique_local_dir(target_root / output_root.name)
            shutil.copytree(output_root, external_dir)
            if zip_path and zip_path.exists():
                copied_zip = copy_file_to_unique_target(zip_path, target_root / zip_path.name)
                external_refs.append(str(copied_zip))
            log(f"Shenhui PDF screenshots copied to export folder: {external_dir}")
        for file_path in exported_files or []:
            source = Path(str(file_path or "")).expanduser()
            if not source.is_file():
                continue
            copied = copy_file_to_unique_target(source, target_root / source.name)
            external_refs.append(str(copied))
        if external_refs:
            final_refs = external_refs
    elif _pdf_output_mode(run_params or {}) == "merge_existing_zips":
        log("[warn] 合并同款号 ZIP 模式需要选择导出目录；本次仅生成独立截图包。")

    if pdf_work_dir.exists():
        shutil.rmtree(pdf_work_dir, ignore_errors=True)

    return final_refs
