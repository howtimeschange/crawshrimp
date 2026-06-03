"""ODPS/MaxCompute data sync helpers."""
from __future__ import annotations

import json
import re
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Callable, Optional
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

from openpyxl import load_workbook


TASK_TABLE_MAP = {
    ("temu", "mall_flux"): "imp_ods_temu_mall_flux",
    ("tiktok-ops-assistant", "product_analytics"): "imp_ods_tiktok_product_analytics",
    ("aliexpress-ops-assistant", "deal_analysis"): "imp_ods_aliexpress_deal_analysis",
    ("aliexpress-ops-assistant", "product_ranking"): "imp_ods_aliexpress_product_ranking",
    ("lazada-plus-v1", "business_advisor"): "imp_ods_lazada_business_advisor",
    ("shopee-plus-v2", "business_analysis"): "imp_ods_shopee_business_analysis",
}

LAZADA_BUSINESS_ADVISOR_FIELDS = [
    ("platform_name", "string", "平台名称"),
    ("sheet_name", "string", "数据表"),
    ("country_code", "string", "国家编码"),
    ("stat_date_range", "string", "统计日期范围"),
    ("stat_date", "string", "统计日期"),
    ("metric_name", "string", "指标名称"),
    ("metric_value", "double", "指标值"),
    ("unit", "string", "单位"),
    ("definition", "string", "指标定义"),
    ("raw_metric_header", "string", "原始指标表头"),
    ("captured_at", "datetime", "抓取时间"),
]

SHOPEE_BUSINESS_ANALYSIS_FIELDS = [
    ("platform_name", "string", "平台名称"),
    ("sheet_name", "string", "数据表"),
    ("dimension_type", "string", "维度类型"),
    ("stat_time", "string", "统计时间"),
    ("stat_date", "string", "统计日期"),
    ("shop_name", "string", "店铺名称"),
    ("market", "string", "市场"),
    ("metric_name", "string", "指标名称"),
    ("order_status", "string", "订单状态"),
    ("currency", "string", "币种"),
    ("metric_value", "double", "指标值"),
    ("raw_metric_header", "string", "原始指标表头"),
    ("captured_at", "datetime", "抓取时间"),
]

ALIEXPRESS_COMMON_FIELD_MAP = {
    "平台名称": "platform_name",
    "店铺名称": "shop_name",
    "channelId": "channel_id",
    "统计日期": "stat_date",
    "统计日期范围": "stat_date_range",
    "时间筛选": "date_type_label",
    "国家": "country_label",
    "国家编码": "country_code",
    "业务模式": "biz_type_label",
    "业务模式编码": "biz_type_code",
    "抓取时间": "captured_at",
}

ALIEXPRESS_PRODUCT_RANKING_EXTRA_FIELD_MAP = {
    "榜单类型": "rank_type_label",
    "平台筛选": "platform_label",
    "平台编码": "platform_code",
    "类目": "category_name",
    "类目路径": "category_path",
    "类目ID": "category_id",
    "商品筛选": "product_filter",
    "页码": "page_no",
    "页内序号": "row_no",
    "排行": "rank_no",
    "商品ID": "item_id",
    "商品名称": "item_title",
    "商品图片": "image_url",
    "商品链接": "detail_url",
    "最低价": "min_price",
    "最高价": "max_price",
    "商品状态": "item_status",
    "Choice商品": "is_choice",
    "首次上架时间": "first_online_date",
    "支付金额": "pay_amt",
    "支付金额环比": "pay_amt_chain_ratio",
    "税费": "tax_amt",
    "税费环比": "tax_amt_chain_ratio",
    "商品访客数": "uv",
    "商品访客数环比": "uv_chain_ratio",
    "新访客数": "new_visitor_cnt",
    "新访客数环比": "new_visitor_cnt_chain_ratio",
    "老访客数": "old_visitor_cnt",
    "老访客数环比": "old_visitor_cnt_chain_ratio",
    "支付买家数": "pay_buyer_cnt",
    "支付买家数环比": "pay_buyer_cnt_chain_ratio",
    "新支付买家数": "pay_new_buyer_cnt",
    "新支付买家数环比": "pay_new_buyer_cnt_chain_ratio",
    "老支付买家数": "pay_old_buyer_cnt",
    "老支付买家数环比": "pay_old_buyer_cnt_chain_ratio",
    "支付转化率": "pay_conversion_rate",
    "支付转化率环比": "pay_conversion_rate_chain_ratio",
    "客单价": "pay_per_buyer_amt",
    "客单价环比": "pay_per_buyer_amt_chain_ratio",
    "支付件数": "pay_item_qty",
    "支付件数环比": "pay_item_qty_chain_ratio",
    "支付订单数": "pay_order_cnt",
    "支付订单数环比": "pay_order_cnt_chain_ratio",
    "下单金额": "order_amt",
    "下单金额环比": "order_amt_chain_ratio",
    "下单买家数": "order_buyer_cnt",
    "下单买家数环比": "order_buyer_cnt_chain_ratio",
    "下单件数": "order_item_qty",
    "下单件数环比": "order_item_qty_chain_ratio",
    "下单转化率": "order_conversion_rate",
    "下单转化率环比": "order_conversion_rate_chain_ratio",
    "搜索曝光次数": "search_exp_pv",
    "搜索曝光次数环比": "search_exp_pv_chain_ratio",
    "加购买家数": "add_cart_buyer_cnt",
    "加购买家数环比": "add_cart_buyer_cnt_chain_ratio",
    "收藏买家数": "wishlist_buyer_cnt",
    "收藏买家数环比": "wishlist_buyer_cnt_chain_ratio",
    "详情到下单转化率": "l2d_uv_rate",
    "详情到下单转化率环比": "l2d_uv_rate_chain_ratio",
    "退款金额": "refund_amt",
    "退款金额环比": "refund_amt_chain_ratio",
    "停留时长": "avg_stay_time",
    "停留时长环比": "avg_stay_time_chain_ratio",
    "环比标签": "chain_ratio_label",
}

TASK_FIELD_MAP = {
    ("temu", "mall_flux"): {
        "平台名称": "platform_name",
        "店铺名称": "shop_name",
        "店铺ID": "shop_id",
        "外层站点": "outer_site",
        "统计日期范围": "stat_date_range",
        "统计粒度": "stat_grain",
        "列表页码": "page_no",
        "抓取时间": "captured_at",
        "列表行号": "row_no",
        "日期": "stat_date",
        "总数据/总浏览量": "total_views",
        "总数据/总访客数": "total_visitors",
        "总数据/总支付买家数": "total_paid_buyers",
        "总数据/总支付转化率": "total_payment_conversion_rate",
        "总数据/总支付件数": "total_paid_items",
        "商品数据/商品浏览量": "product_views",
        "商品数据/商品访客数": "product_visitors",
        "商品数据/商详支付买家数": "product_detail_paid_buyers",
        "商品数据/商详支付转化率": "product_detail_payment_conversion_rate",
        "店铺数据/店铺页浏览量": "shop_page_views",
        "店铺数据/店铺页面访客数": "shop_page_visitors",
        "店铺数据/店铺页支付买家数": "shop_page_paid_buyers",
        "店铺数据/店铺页支付转化率": "shop_page_payment_conversion_rate",
    },
    ("tiktok-ops-assistant", "product_analytics"): {
        "平台名称": "platform_name",
        "区域": "region",
        "店铺ID": "shop_id",
        "店铺名称": "shop_name",
        "统计日期范围": "stat_date_range",
        "对比日期范围": "compare_date_range",
        "抓取时间": "captured_at",
        "订单数": "orders",
        "商品曝光次数": "product_impressions",
        "商品点击量": "product_clicks",
    },
    ("aliexpress-ops-assistant", "deal_analysis"): {
        **ALIEXPRESS_COMMON_FIELD_MAP,
        "数据类型": "data_type",
        "指标分组": "metric_group",
        "指标编码": "metric_code",
        "指标名称": "metric_name",
        "指标值": "metric_value",
        "环比标签": "cycle_label",
        "环比值": "cycle_value",
        "同比标签": "line_label",
        "同比值": "line_value",
    },
    ("aliexpress-ops-assistant", "product_ranking"): {
        **ALIEXPRESS_COMMON_FIELD_MAP,
        **ALIEXPRESS_PRODUCT_RANKING_EXTRA_FIELD_MAP,
    },
}

ALIEXPRESS_COMMON_FIELD_TYPE_MAP = {
    "平台名称": "string",
    "店铺名称": "string",
    "channelId": "string",
    "统计日期": "string",
    "统计日期范围": "string",
    "时间筛选": "string",
    "国家": "string",
    "国家编码": "string",
    "业务模式": "string",
    "业务模式编码": "string",
    "抓取时间": "datetime",
}

ALIEXPRESS_PRODUCT_RANKING_EXTRA_FIELD_TYPE_MAP = {
    "榜单类型": "string",
    "平台筛选": "string",
    "平台编码": "string",
    "类目": "string",
    "类目路径": "string",
    "类目ID": "string",
    "商品筛选": "string",
    "页码": "bigint",
    "页内序号": "bigint",
    "排行": "bigint",
    "商品ID": "string",
    "商品名称": "string",
    "商品图片": "string",
    "商品链接": "string",
    "最低价": "string",
    "最高价": "string",
    "商品状态": "string",
    "Choice商品": "string",
    "首次上架时间": "string",
    "支付金额": "double",
    "支付金额环比": "string",
    "税费": "double",
    "税费环比": "string",
    "商品访客数": "bigint",
    "商品访客数环比": "string",
    "新访客数": "bigint",
    "新访客数环比": "string",
    "老访客数": "bigint",
    "老访客数环比": "string",
    "支付买家数": "bigint",
    "支付买家数环比": "string",
    "新支付买家数": "bigint",
    "新支付买家数环比": "string",
    "老支付买家数": "bigint",
    "老支付买家数环比": "string",
    "支付转化率": "string",
    "支付转化率环比": "string",
    "客单价": "double",
    "客单价环比": "string",
    "支付件数": "bigint",
    "支付件数环比": "string",
    "支付订单数": "bigint",
    "支付订单数环比": "string",
    "下单金额": "double",
    "下单金额环比": "string",
    "下单买家数": "bigint",
    "下单买家数环比": "string",
    "下单件数": "bigint",
    "下单件数环比": "string",
    "下单转化率": "string",
    "下单转化率环比": "string",
    "搜索曝光次数": "bigint",
    "搜索曝光次数环比": "string",
    "加购买家数": "bigint",
    "加购买家数环比": "string",
    "收藏买家数": "bigint",
    "收藏买家数环比": "string",
    "详情到下单转化率": "string",
    "详情到下单转化率环比": "string",
    "退款金额": "double",
    "退款金额环比": "string",
    "停留时长": "double",
    "停留时长环比": "string",
    "环比标签": "string",
}

TASK_FIELD_TYPE_MAP = {
    ("temu", "mall_flux"): {
        "平台名称": "string",
        "店铺名称": "string",
        "店铺ID": "string",
        "外层站点": "string",
        "统计日期范围": "string",
        "统计粒度": "string",
        "列表页码": "bigint",
        "抓取时间": "datetime",
        "列表行号": "bigint",
        "日期": "string",
        "总数据/总浏览量": "bigint",
        "总数据/总访客数": "bigint",
        "总数据/总支付买家数": "bigint",
        "总数据/总支付转化率": "string",
        "总数据/总支付件数": "bigint",
        "商品数据/商品浏览量": "bigint",
        "商品数据/商品访客数": "bigint",
        "商品数据/商详支付买家数": "bigint",
        "商品数据/商详支付转化率": "string",
        "店铺数据/店铺页浏览量": "bigint",
        "店铺数据/店铺页面访客数": "bigint",
        "店铺数据/店铺页支付买家数": "bigint",
        "店铺数据/店铺页支付转化率": "string",
    },
    ("tiktok-ops-assistant", "product_analytics"): {
        "平台名称": "string",
        "区域": "string",
        "店铺ID": "string",
        "店铺名称": "string",
        "统计日期范围": "string",
        "对比日期范围": "string",
        "抓取时间": "datetime",
        "订单数": "bigint",
        "商品曝光次数": "bigint",
        "商品点击量": "bigint",
    },
    ("aliexpress-ops-assistant", "deal_analysis"): {
        **ALIEXPRESS_COMMON_FIELD_TYPE_MAP,
        "数据类型": "string",
        "指标分组": "string",
        "指标编码": "string",
        "指标名称": "string",
        "指标值": "double",
        "环比标签": "string",
        "环比值": "double",
        "同比标签": "string",
        "同比值": "double",
    },
    ("aliexpress-ops-assistant", "product_ranking"): {
        **ALIEXPRESS_COMMON_FIELD_TYPE_MAP,
        **ALIEXPRESS_PRODUCT_RANKING_EXTRA_FIELD_TYPE_MAP,
    },
}

SHOPEE_DIMENSION_COLUMNS = {
    "统计时间": "stat_time",
    "店铺名称": "shop_name",
    "市场": "market",
    "指标": "metric_definition_name",
    "定义": "metric_definition",
}

DEFAULT_WRITE_MODE = "append"
DEFAULT_DATAWORKS_PATH = "/api/v1/dataworks/write_odps"
DEFAULT_DATAWORKS_ENDPOINT = "http://dataworksapi.semirapp.com/api/v1/dataworks/write_odps"


class OdpsSyncError(Exception):
    """Raised when a data file cannot be converted or synced."""


def get_table_name(adapter_id: str, task_id: str) -> str:
    table_name = TASK_TABLE_MAP.get((str(adapter_id or "").strip(), str(task_id or "").strip()))
    if not table_name:
        raise OdpsSyncError(f"暂不支持同步该任务：{adapter_id}/{task_id}")
    return table_name


def get_field_name(adapter_id: str, task_id: str, header: str, index: int) -> str:
    mapping = TASK_FIELD_MAP.get((str(adapter_id or "").strip(), str(task_id or "").strip()), {})
    mapped = str(mapping.get(str(header or "").strip()) or "").strip()
    if mapped:
        return mapped
    slug = re.sub(r"[^0-9A-Za-z_]+", "_", str(header or "").strip()).strip("_").lower()
    if slug:
        return slug
    return f"col_{index + 1}"


def get_field_type(adapter_id: str, task_id: str, header: str, values: list[Any]) -> str:
    mapping = TASK_FIELD_TYPE_MAP.get((str(adapter_id or "").strip(), str(task_id or "").strip()), {})
    mapped = str(mapping.get(str(header or "").strip()) or "").strip()
    if mapped:
        return mapped
    return infer_odps_type(header, values)


def normalize_endpoint(endpoint: str) -> str:
    url = str(endpoint or "").strip()
    if not url:
        return ""
    return url.rstrip("/")


def resolve_write_odps_url(endpoint: str) -> str:
    url = normalize_endpoint(endpoint)
    if not url:
        return ""
    if url.endswith("/write_odps"):
        return url
    if url.endswith("/api/v1/dataworks"):
        return f"{url}/write_odps"
    if "dataworksapi.semirapp.com" in url or "alicloudapi.com" in url:
        return f"{url}{DEFAULT_DATAWORKS_PATH}"
    return f"{url}/write_odps"


def build_request_headers(app_code: str = "") -> dict[str, str]:
    headers = {"Content-Type": "application/json"}
    code = str(app_code or "").strip()
    if code:
        headers["Authorization"] = f"APPCODE {code}"
    return headers


def _clean_header(value: Any, index: int) -> str:
    text = str(value or "").strip()
    return text or f"列{index + 1}"


def _cell_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return str(value)


def read_excel_rows(file_path: str) -> tuple[list[str], list[dict[str, Any]]]:
    path = Path(file_path).expanduser()
    if not path.exists() or not path.is_file():
        raise OdpsSyncError(f"文件不存在：{file_path}")
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise OdpsSyncError(f"暂只支持同步 xlsx/xlsm Excel 文件：{path.name}")

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        raw_rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if not raw_rows:
        raise OdpsSyncError(f"Excel 文件为空：{path.name}")

    header_row = next((row for row in raw_rows if any(str(cell or "").strip() for cell in row)), None)
    if not header_row:
        raise OdpsSyncError(f"Excel 文件缺少表头：{path.name}")

    header_index = raw_rows.index(header_row)
    headers = [_clean_header(value, index) for index, value in enumerate(header_row)]
    rows: list[dict[str, Any]] = []
    for raw in raw_rows[header_index + 1:]:
        if all(cell is None or str(cell).strip() == "" for cell in raw):
            continue
        row = {}
        for index, header in enumerate(headers):
            row[header] = _cell_value(raw[index] if index < len(raw) else None)
        rows.append(row)

    if not rows:
        raise OdpsSyncError(f"Excel 文件没有可同步的数据行：{path.name}")
    return headers, rows


def _normalize_shopee_sheet_name(value: str) -> str:
    return str(value or "").strip()


def _split_shopee_metric_header(header: str) -> tuple[str, str, str]:
    text = str(header or "").strip()
    currency = ""
    order_status = ""
    currency_match = re.search(r"\(([A-Z]{3})\)", text)
    if currency_match:
        currency = currency_match.group(1)
        text = text.replace(currency_match.group(0), "").strip()
    status_match = re.search(r"（([^）]+订单)）", text)
    if status_match:
        order_status = status_match.group(1)
        text = text.replace(status_match.group(0), "").strip()
    metric_name = re.sub(r"\s+", " ", text).strip()
    return metric_name, order_status, currency


def _shopee_dimension_type(sheet_name: str) -> str:
    name = _normalize_shopee_sheet_name(sheet_name)
    if name == "关键指标":
        return "overview"
    if name == "趋势指标":
        return "trend"
    if name == "店铺维度指标分析":
        return "shop_share"
    if name == "市场维度指标分析":
        return "market_share"
    if name == "详情":
        return "detail"
    if name == "指标定义":
        return "definition"
    return "unknown"


def _normalize_shopee_stat_date(value: Any) -> str:
    text = str(value or "").strip()
    match = re.search(r"\d{4}-\d{2}-\d{2}", text)
    return match.group(0) if match else ""


def _shopee_business_analysis_rows(file_path: str) -> tuple[list[dict[str, str]], list[dict[str, Any]]]:
    path = Path(file_path).expanduser()
    if not path.exists() or not path.is_file():
        raise OdpsSyncError(f"文件不存在：{file_path}")
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise OdpsSyncError(f"暂只支持同步 xlsx/xlsm Excel 文件：{path.name}")

    wb = load_workbook(path, read_only=True, data_only=True)
    rows: list[dict[str, Any]] = []
    captured_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    try:
        for ws in wb.worksheets:
            sheet_name = _normalize_shopee_sheet_name(ws.title)
            raw_rows = list(ws.iter_rows(values_only=True))
            header_row = next((row for row in raw_rows if any(str(cell or "").strip() for cell in row)), None)
            if not header_row:
                continue
            header_index = raw_rows.index(header_row)
            headers = [_clean_header(value, index) for index, value in enumerate(header_row)]
            dimension_headers = {
                header for header in headers
                if header in SHOPEE_DIMENSION_COLUMNS
            }
            metric_headers = [
                header for header in headers
                if header not in dimension_headers and str(header or "").strip()
            ]
            if not metric_headers:
                continue

            for raw in raw_rows[header_index + 1:]:
                if all(cell is None or str(cell).strip() == "" for cell in raw):
                    continue
                raw_map = {
                    header: _cell_value(raw[index] if index < len(raw) else None)
                    for index, header in enumerate(headers)
                }
                for header in metric_headers:
                    raw_value = raw_map.get(header, "")
                    if str(raw_value or "").strip() == "":
                        continue
                    metric_name, order_status, currency = _split_shopee_metric_header(header)
                    rows.append({
                        "platform_name": "Shopee",
                        "sheet_name": sheet_name,
                        "dimension_type": _shopee_dimension_type(sheet_name),
                        "stat_time": raw_map.get("统计时间", ""),
                        "stat_date": _normalize_shopee_stat_date(raw_map.get("统计时间", "")),
                        "shop_name": raw_map.get("店铺名称", ""),
                        "market": raw_map.get("市场", ""),
                        "metric_name": metric_name,
                        "order_status": order_status,
                        "currency": currency,
                        "metric_value": raw_value,
                        "raw_metric_header": header,
                        "captured_at": captured_at,
                    })
    finally:
        wb.close()

    if not rows:
        raise OdpsSyncError(f"Excel 文件没有可同步的数据行：{path.name}")

    fields = [
        {"name": name, "type": field_type, "comment": comment}
        for name, field_type, comment in SHOPEE_BUSINESS_ANALYSIS_FIELDS
    ]
    return fields, rows


def _read_workbook_sheets(file_path: str) -> list[tuple[str, list[tuple[Any, ...]]]]:
    path = Path(file_path).expanduser()
    if not path.exists() or not path.is_file():
        raise OdpsSyncError(f"文件不存在：{file_path}")

    def read_with_openpyxl(source: Any) -> list[tuple[str, list[tuple[Any, ...]]]]:
        wb = load_workbook(source, read_only=True, data_only=True)
        try:
            return [
                (ws.title, list(ws.iter_rows(values_only=True)))
                for ws in wb.worksheets
            ]
        finally:
            wb.close()

    try:
        return read_with_openpyxl(path)
    except Exception as openpyxl_error:
        try:
            with path.open("rb") as handle:
                signature = handle.read(4)
            if signature.startswith(b"PK"):
                return read_with_openpyxl(BytesIO(path.read_bytes()))
        except Exception:
            pass
        if path.suffix.lower() not in {".xls", ".xlt"}:
            raise openpyxl_error
        try:
            import xlrd  # type: ignore
        except ImportError as import_error:
            raise OdpsSyncError("同步 Lazada 生意参谋 .xls 文件需要安装 xlrd，请更新后端依赖") from import_error
        try:
            book = xlrd.open_workbook(str(path))
            sheets: list[tuple[str, list[tuple[Any, ...]]]] = []
            for sheet in book.sheets():
                rows = [
                    tuple(sheet.cell_value(row_index, col_index) for col_index in range(sheet.ncols))
                    for row_index in range(sheet.nrows)
                ]
                sheets.append((sheet.name, rows))
            return sheets
        except Exception as xlrd_error:
            raise OdpsSyncError(f"无法读取 Lazada 生意参谋 Excel 文件：{xlrd_error}") from xlrd_error


def _normalize_lazada_sheet_name(value: str) -> str:
    return str(value or "").strip()


def _normalize_lazada_stat_date_range(value: Any) -> str:
    text = str(value or "").strip()
    dates = []
    for match in re.finditer(r"(\d{1,2})[-/](\d{1,2})[-/](\d{4})|(\d{4})[-/](\d{1,2})[-/](\d{1,2})", text):
        if match.group(4):
            year, month, day = match.group(4), match.group(5), match.group(6)
        else:
            day, month, year = match.group(1), match.group(2), match.group(3)
        dates.append(f"{int(year):04d}-{int(month):02d}-{int(day):02d}")
    if len(dates) >= 2:
        return f"{dates[0]} ~ {dates[1]}"
    if len(dates) == 1:
        return dates[0]
    return ""


def _lazada_country_from_filename(file_path: str) -> str:
    name = Path(file_path).stem.upper()
    for code in ("MY", "SG", "ID", "PH", "TH", "VN"):
        if re.search(rf"(?:^|[_\-\s]){code}(?:$|[_\-\s])", name):
            return code
    return ""


def _looks_like_lazada_header(row: tuple[Any, ...]) -> bool:
    headers = {str(cell or "").strip().lower() for cell in row if str(cell or "").strip()}
    if "date" in headers and len(headers - {"date"}) >= 1:
        return True
    if "metric name" in headers and "definition" in headers:
        return True
    return bool(headers & {"metric", "metrics", "indicator", "name"}) and bool(headers & {"value", "definition", "unit"})


def _lazada_metric_name(raw_map: dict[str, Any]) -> str:
    for key in ("Metric", "Metrics", "Metric Name", "Indicator", "Name", "指标", "指标名称"):
        value = str(raw_map.get(key, "") or "").strip()
        if value:
            return value
    for key, value in raw_map.items():
        if str(key or "").strip().lower() in {"value", "unit", "definition"}:
            continue
        text = str(value or "").strip()
        if text:
            return text
    return ""


def _lazada_metric_value(raw_map: dict[str, Any]) -> Any:
    for key in ("Value", "值", "指标值"):
        if key in raw_map:
            return raw_map.get(key, "")
    return ""


def _lazada_unit(raw_map: dict[str, Any]) -> str:
    for key in ("Unit", "Currency", "单位", "币种"):
        value = str(raw_map.get(key, "") or "").strip()
        if value:
            return value
    return ""


def _lazada_definition(raw_map: dict[str, Any]) -> str:
    for key in ("Definition", "Description", "定义", "指标定义"):
        value = str(raw_map.get(key, "") or "").strip()
        if value:
            return value
    return ""


def _lazada_business_advisor_rows(file_path: str) -> tuple[list[dict[str, str]], list[dict[str, Any]]]:
    path = Path(file_path).expanduser()
    if path.suffix.lower() not in {".xls", ".xlsx", ".xlsm"}:
        raise OdpsSyncError(f"暂只支持同步 Excel 文件：{path.name}")

    captured_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    country_code = _lazada_country_from_filename(file_path)
    rows: list[dict[str, Any]] = []
    workbook_sheets = _read_workbook_sheets(file_path)

    for sheet_name_raw, raw_rows in workbook_sheets:
        sheet_name = _normalize_lazada_sheet_name(sheet_name_raw)
        stat_date_range = ""
        for raw in raw_rows[:10]:
            line = " ".join(str(cell or "").strip() for cell in raw if str(cell or "").strip())
            if "date range" in line.lower() or "日期" in line:
                stat_date_range = _normalize_lazada_stat_date_range(line)
                if stat_date_range:
                    break

        header_index = None
        for index, raw in enumerate(raw_rows):
            if _looks_like_lazada_header(raw):
                header_index = index
                break
        if header_index is None:
            continue

        headers = [_clean_header(value, index) for index, value in enumerate(raw_rows[header_index])]
        lower_headers = {header.lower(): header for header in headers}
        date_header = lower_headers.get("date")
        definition_header = lower_headers.get("definition")
        metric_name_header = lower_headers.get("metric name") or lower_headers.get("metric")
        if date_header:
            metric_headers = [header for header in headers if header != date_header and str(header or "").strip()]
            for raw in raw_rows[header_index + 1:]:
                if all(cell is None or str(cell).strip() == "" for cell in raw):
                    continue
                raw_map = {
                    header: _cell_value(raw[index] if index < len(raw) else None)
                    for index, header in enumerate(headers)
                }
                stat_value = str(raw_map.get(date_header, "") or "").strip()
                if not stat_value:
                    continue
                row_stat_range = _normalize_lazada_stat_date_range(stat_value)
                row_stat_date = normalize_partition_value(row_stat_range or stat_value)
                for header in metric_headers:
                    raw_value = raw_map.get(header, "")
                    if str(raw_value or "").strip() == "":
                        continue
                    rows.append({
                        "platform_name": "Lazada",
                        "sheet_name": sheet_name,
                        "country_code": country_code,
                        "stat_date_range": row_stat_range or stat_date_range,
                        "stat_date": row_stat_date,
                        "metric_name": header,
                        "metric_value": raw_value,
                        "unit": "",
                        "definition": "",
                        "raw_metric_header": header,
                        "captured_at": captured_at,
                    })
            continue

        if definition_header and metric_name_header:
            for raw in raw_rows[header_index + 1:]:
                if all(cell is None or str(cell).strip() == "" for cell in raw):
                    continue
                raw_map = {
                    header: _cell_value(raw[index] if index < len(raw) else None)
                    for index, header in enumerate(headers)
                }
                metric_name = str(raw_map.get(metric_name_header, "") or "").strip()
                definition = str(raw_map.get(definition_header, "") or "").strip()
                if not metric_name and not definition:
                    continue
                rows.append({
                    "platform_name": "Lazada",
                    "sheet_name": sheet_name,
                    "country_code": country_code,
                    "stat_date_range": stat_date_range,
                    "stat_date": normalize_partition_value(stat_date_range),
                    "metric_name": metric_name,
                    "metric_value": "",
                    "unit": "",
                    "definition": definition,
                    "raw_metric_header": definition_header,
                    "captured_at": captured_at,
                })
            continue

        for raw in raw_rows[header_index + 1:]:
            if all(cell is None or str(cell).strip() == "" for cell in raw):
                continue
            raw_map = {
                header: _cell_value(raw[index] if index < len(raw) else None)
                for index, header in enumerate(headers)
            }
            metric_name = _lazada_metric_name(raw_map)
            definition = _lazada_definition(raw_map)
            metric_value = _lazada_metric_value(raw_map)
            if not metric_name and not definition:
                continue
            rows.append({
                "platform_name": "Lazada",
                "sheet_name": sheet_name,
                "country_code": country_code,
                "stat_date_range": stat_date_range,
                "stat_date": normalize_partition_value(stat_date_range),
                "metric_name": metric_name,
                "metric_value": metric_value,
                "unit": _lazada_unit(raw_map),
                "definition": definition,
                "raw_metric_header": "Value" if str(metric_value or "").strip() else "Definition",
                "captured_at": captured_at,
            })

    if not rows:
        raise OdpsSyncError(f"Excel 文件没有可同步的 Lazada 生意参谋数据行：{path.name}")

    fields = [
        {"name": name, "type": field_type, "comment": comment}
        for name, field_type, comment in LAZADA_BUSINESS_ADVISOR_FIELDS
    ]
    return fields, rows


def infer_odps_type(field_name: str, values: list[Any]) -> str:
    name = str(field_name or "")
    clean_values = [str(value or "").strip() for value in values if str(value or "").strip()]
    if not clean_values:
        return "string"
    if any(token in name for token in ["日期范围", "范围"]):
        return "string"
    if "时间" in name:
        return "datetime"
    if any(token in name for token in ["转化率", "率", "占比"]):
        return "string"
    numeric_values = [value.replace(",", "") for value in clean_values]
    if all(re.fullmatch(r"-?\d+", value) for value in numeric_values):
        return "bigint"
    if all(re.fullmatch(r"-?\d+(\.\d+)?", value) for value in numeric_values):
        return "double"
    return "string"


def cast_odps_value(value: Any, field_type: str) -> Any:
    if value is None:
        return None
    text = str(value).strip()
    if text == "" or text in {"-", "--", "—", "N/A", "n/a", "NA", "na", "null", "NULL"}:
        return None
    normalized_type = str(field_type or "").strip().lower()
    if normalized_type in {"int", "integer", "bigint"}:
        numeric_text = text.replace(",", "")
        if re.fullmatch(r"-?\d+", numeric_text):
            return int(numeric_text)
        return text
    if normalized_type in {"float", "double", "decimal", "numeric"}:
        numeric_text = text.replace(",", "")
        if numeric_text.endswith("%"):
            numeric_text = numeric_text[:-1].strip()
        if re.fullmatch(r"-?\d+(\.\d+)?", numeric_text):
            return float(numeric_text)
        return text
    if normalized_type in {"datetime", "date", "timestamp"}:
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}", text):
            return f"{text} 00:00:00"
        return text
    return text


def normalize_partition_value(value: str) -> str:
    text = str(value or "").strip()
    match = re.search(r"\d{4}-\d{2}-\d{2}", text)
    if match:
        return match.group(0)
    return datetime.now().strftime("%Y-%m-%d")


def build_sync_payload(
    adapter_id: str,
    task_id: str,
    file_path: str,
    *,
    write_mode: str = DEFAULT_WRITE_MODE,
) -> dict[str, Any]:
    if (str(adapter_id or "").strip(), str(task_id or "").strip()) == ("lazada-plus-v1", "business_advisor"):
        fields, rows = _lazada_business_advisor_rows(file_path)
        field_types = {field["name"]: field["type"] for field in fields}
        data_rows = [
            {
                field["name"]: cast_odps_value(row.get(field["name"], ""), field_types[field["name"]])
                for field in fields
            }
            for row in rows
        ]
        date_value = rows[0].get("stat_date") or rows[0].get("stat_date_range") or ""
        return {
            "table_name": get_table_name(adapter_id, task_id),
            "fields": fields,
            "data": data_rows,
            "write_mode": write_mode,
            "partition_spec": {"dt": normalize_partition_value(str(date_value))},
        }

    if (str(adapter_id or "").strip(), str(task_id or "").strip()) == ("shopee-plus-v2", "business_analysis"):
        fields, rows = _shopee_business_analysis_rows(file_path)
        field_types = {field["name"]: field["type"] for field in fields}
        data_rows = [
            {
                field["name"]: cast_odps_value(row.get(field["name"], ""), field_types[field["name"]])
                for field in fields
            }
            for row in rows
        ]
        date_value = rows[0].get("stat_date") or rows[0].get("stat_time") or ""
        return {
            "table_name": get_table_name(adapter_id, task_id),
            "fields": fields,
            "data": data_rows,
            "write_mode": write_mode,
            "partition_spec": {"dt": normalize_partition_value(str(date_value))},
        }

    headers, rows = read_excel_rows(file_path)
    field_names = [get_field_name(adapter_id, task_id, header, index) for index, header in enumerate(headers)]
    field_types = [
        get_field_type(adapter_id, task_id, header, [row.get(header, "") for row in rows])
        for header in headers
    ]
    data_rows = [
        {
            field_names[index]: cast_odps_value(row.get(header, ""), field_types[index])
            for index, header in enumerate(headers)
        }
        for row in rows
    ]
    fields = [
        {
            "name": field_names[index],
            "type": field_types[index],
            "comment": header,
        }
        for index, header in enumerate(headers)
    ]
    date_value = rows[0].get("日期") or rows[0].get("统计日期") or rows[0].get("统计日期范围") or ""
    return {
        "table_name": get_table_name(adapter_id, task_id),
        "fields": fields,
        "data": data_rows,
        "write_mode": write_mode,
        "partition_spec": {"dt": normalize_partition_value(str(date_value))},
    }


def _post_json(
    url: str,
    payload: dict[str, Any],
    timeout: int = 30,
    headers: Optional[dict[str, str]] = None,
) -> dict[str, Any]:
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    req = Request(url, data=body, headers=headers or build_request_headers(), method="POST")
    try:
        with urlopen(req, timeout=timeout) as resp:
            text = resp.read().decode("utf-8")
    except HTTPError as error:
        detail = error.read().decode("utf-8", errors="replace")
        raise OdpsSyncError(f"接口返回 HTTP {error.code}: {detail or error.reason}") from error
    except URLError as error:
        raise OdpsSyncError(f"接口请求失败：{error.reason}") from error

    if not text:
        return {}
    try:
        parsed = json.loads(text)
    except json.JSONDecodeError as error:
        raise OdpsSyncError(f"接口返回不是合法 JSON：{text[:200]}") from error
    return parsed if isinstance(parsed, dict) else {"result": parsed}


def sync_file(
    adapter_id: str,
    task_id: str,
    file_path: str,
    *,
    endpoint: str,
    app_code: str = "",
    post_json: Optional[Callable[..., dict[str, Any]]] = None,
) -> dict[str, Any]:
    url = resolve_write_odps_url(endpoint)
    if not url:
        raise OdpsSyncError("未配置 ODPS 同步接口地址")
    if not str(app_code or "").strip():
        raise OdpsSyncError("未配置 ODPS AppCode，请在设置里填写 ODPS AppCode")

    payload = build_sync_payload(adapter_id, task_id, file_path)
    sender = post_json or _post_json
    response = sender(url, payload, 30, headers=build_request_headers(app_code))
    if response.get("success") is False:
        raise OdpsSyncError(response.get("error") or response.get("message") or "接口同步失败")
    return {
        "ok": True,
        "file": str(file_path),
        "table_name": payload["table_name"],
        "count": len(payload["data"]),
        "partition_spec": payload["partition_spec"],
        "response": response,
    }


def sync_files(
    adapter_id: str,
    task_id: str,
    file_paths: list[str],
    *,
    endpoint: str,
    app_code: str = "",
) -> dict[str, Any]:
    paths = [str(path or "").strip() for path in file_paths or [] if str(path or "").strip()]
    if not paths:
        raise OdpsSyncError("请选择需要同步的 Excel 文件")

    results = []
    failed = []
    for path in paths:
        try:
            results.append(sync_file(adapter_id, task_id, path, endpoint=endpoint, app_code=app_code))
        except Exception as error:
            failed.append({"file": path, "error": str(error)})

    return {
        "ok": not failed,
        "synced_count": len(results),
        "failed_count": len(failed),
        "results": results,
        "failed": failed,
    }
