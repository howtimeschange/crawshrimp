"""Local executors for cloud approval machine jobs."""
from __future__ import annotations

import asyncio
import importlib.util
import mimetypes
import sys
from pathlib import Path
from typing import Any, Mapping

from core import runtime_paths
from core.cloud_approval_client import CloudApprovalClient, CloudApprovalError

MATERIAL_IMPORT_DETAIL_CHUNK_SIZE = 1000


class CloudJobBlocked(RuntimeError):
    def __init__(self, status: str, message: str = ""):
        super().__init__(message or status)
        self.status = status
        self.message = message or status


class CloudJobTerminalFailure(RuntimeError):
    pass


class CloudJobCancelled(RuntimeError):
    pass


class CloudJobExecutor:
    def __init__(
        self,
        client: CloudApprovalClient,
        work_dir: Path | None = None,
        *,
        tmall_module=None,
        material_test_runner=None,
    ):
        self.client = client
        self.work_dir = Path(work_dir) if work_dir is not None else runtime_paths.data_root() / "cloud-jobs"
        self.tmall_module = tmall_module
        self.material_test_runner = material_test_runner

    def execute(self, job: Mapping[str, Any]) -> dict:
        job_type = str(job.get("job_type") or "")
        if job_type == "generate_ai_image":
            return self.execute_generate_ai_image(job)
        if job_type == "regenerate_ai_image":
            return self.execute_regenerate_ai_image(job)
        if job_type == "submit_tmall_material_test":
            return self.execute_submit_tmall_material_test(job)
        if job_type == "crawl_tmall_material_test_data":
            return self.execute_crawl_tmall_material_test_data(job)
        raise ValueError(f"Unsupported cloud job type: {job_type}")

    def execute_regenerate_ai_image(self, job: Mapping[str, Any]) -> dict:
        payload = _payload(job)
        asset_uid = _required_text(payload, "asset_uid")
        rejected_asset_uid = str(payload.get("rejected_asset_uid") or payload.get("parent_asset_uid") or "")
        batch_uid = str(payload.get("batch_uid") or job.get("batch_uid") or "")
        style_id = _positive_int(payload.get("style_id"))
        task_dir = self._task_dir(job)
        self._progress(job, "downloading_assets", "下载参考素材")
        reference_paths = [
            str(self._download_asset(job, uid, task_dir / "references"))
            for uid in _text_list(payload.get("reference_asset_uids"))
        ]
        self._renew(job)

        module = self._tmall_module()
        batch = self._regeneration_batch(job, payload, task_dir, reference_paths)
        self._progress(job, "regenerating", "本地重新生图")
        asset = module.regenerate_approval_asset(
            batch,
            rejected_asset_uid or asset_uid,
            prompt=str(payload.get("prompt_text") or ""),
            reference_paths=reference_paths,
        )
        self._renew(job)

        output_path = _local_file(asset.get("path"), task_dir)
        self._progress(job, "uploading_results", "上传重新生图结果")
        upload_result = self._upload_result_asset(
            job=job,
            batch_uid=batch_uid,
            style_id=style_id,
            asset_uid=asset_uid,
            path=output_path,
            prompt_text=str(asset.get("prompt") or payload.get("prompt_text") or ""),
            parent_asset_uid=str(payload.get("parent_asset_uid") or rejected_asset_uid),
            generation_job_id=str((asset.get("generation_row") or {}).get("任务ID") or (asset.get("generation_row") or {}).get("__1xm_task_id") or ""),
        )
        self._renew(job)
        return {
            "status": "succeeded",
            "result": {
                "asset_uid": asset_uid,
                "filename": output_path.name,
                "object_key": upload_result.get("object_key", ""),
                "local_result_path": str(output_path),
            },
        }

    def execute_generate_ai_image(self, job: Mapping[str, Any]) -> dict:
        payload = _payload(job)
        batch_uid = str(payload.get("batch_uid") or job.get("batch_uid") or "")
        style_id = _positive_int(payload.get("style_id"))
        source_asset_uid = _required_text(payload, "source_asset_uid")
        prompt_text = _required_text(payload, "prompt_text")
        task_dir = self._task_dir(job)
        self._progress(job, "downloading_assets", "下载在线生图素材")
        source_path = self._download_asset(job, source_asset_uid, task_dir / "source")
        reference_paths = [
            str(self._download_asset(job, uid, task_dir / "references"))
            for uid in _text_list(payload.get("reference_asset_uids"))
        ]
        self._renew(job)

        module = self._tmall_module()
        batch = self._generation_batch(job, payload, task_dir, str(source_path), reference_paths)
        self._progress(job, "generating", "本地执行在线生图")
        asset = module.generate_approval_asset_for_item(
            batch,
            item_id=str(payload.get("item_id") or ""),
            style_code=str(payload.get("style_code") or ""),
            prompt=prompt_text,
            main_image_path=str(source_path),
            reference_paths=reference_paths,
        )
        self._renew(job)

        output_path = _local_file(asset.get("path"), task_dir)
        result_asset_uid = _safe_generated_asset_uid(str(payload.get("request_uid") or ""), _job_uid(job))
        generation_row = asset.get("generation_row") if isinstance(asset.get("generation_row"), Mapping) else {}
        self._progress(job, "uploading_results", "上传在线生图结果")
        upload_result = self._upload_result_asset(
            job=job,
            batch_uid=batch_uid,
            style_id=style_id,
            asset_uid=result_asset_uid,
            path=output_path,
            prompt_text=str(asset.get("prompt") or prompt_text),
            parent_asset_uid=source_asset_uid,
            generation_job_id=str(generation_row.get("任务ID") or generation_row.get("1XM任务ID") or generation_row.get("__1xm_task_id") or ""),
            prompt_template_version_id=_positive_int(payload.get("prompt_template_version_id")) or None,
        )
        self._renew(job)
        return {
            "status": "succeeded",
            "result": {
                "asset_uid": result_asset_uid,
                "filename": output_path.name,
                "object_key": upload_result.get("object_key", ""),
                "local_result_path": str(output_path),
                "generation": generation_row,
            },
        }

    def execute_submit_tmall_material_test(self, job: Mapping[str, Any]) -> dict:
        payload = _payload(job)
        submit_plan = payload.get("submit_plan") if isinstance(payload.get("submit_plan"), Mapping) else None
        if not submit_plan:
            raise ValueError("submit_tmall_material_test payload is missing")
        task_dir = self._task_dir(job)
        self._progress(job, "downloading_assets", "下载已确认图片")
        batch = self._submit_batch(job, submit_plan, task_dir)
        self._renew(job)

        module = self._tmall_module()
        self._progress(job, "submitting", "本地提交天猫测图任务")
        try:
            result = _run_maybe_async(module.upload_approved_tmall_batch(batch))
        except RuntimeError as exc:
            if _looks_like_login_block(str(exc)):
                raise CloudJobBlocked("needs_login", str(exc)) from exc
            raise CloudJobTerminalFailure(str(exc)) from exc
        if _submit_result_failed(result):
            raise CloudJobTerminalFailure(_submit_failure_message(result))
        self._renew(job)
        return {
            "status": "succeeded",
            "result": {
                "status": str(result.get("status") or ""),
                "result_path": _safe_result_path(result.get("result_path"), task_dir),
                "submitted": result.get("submitted", result.get("succeeded", 0)),
                "attempted": result.get("attempted", 0),
                "succeeded": result.get("succeeded", 0),
                "failed": result.get("failed", 0),
            },
        }

    def execute_crawl_tmall_material_test_data(self, job: Mapping[str, Any]) -> dict:
        payload = _payload(job)
        task_dir = self._task_dir(job)
        export_dir = task_dir / "exports"
        export_dir.mkdir(parents=True, exist_ok=True)
        run_params = {
            "mode": "new",
            "output_dir": str(export_dir),
            "test_status": "1",
            "statistic_type": "ACCUMULATE_30_DAYS",
            "page_size": 20,
        }
        provided_params = payload.get("run_params")
        if isinstance(provided_params, Mapping):
            run_params.update({str(key): value for key, value in provided_params.items()})
        run_params["output_dir"] = str(export_dir)

        self._progress(job, "crawling", "本地抓取天猫测图数据")
        try:
            raw_result = _run_maybe_async(self._material_test_runner()(run_params, task_dir, job))
        except RuntimeError as exc:
            if _looks_like_login_block(str(exc)):
                raise CloudJobBlocked("needs_login", str(exc)) from exc
            raise CloudJobTerminalFailure(str(exc)) from exc
        self._renew(job)

        workbook_path = _material_workbook_path(raw_result, export_dir, task_dir)
        parsed = parse_tmall_material_test_workbook(workbook_path)
        asset_uid = _safe_generated_asset_uid(f"material-test-{_job_uid(job)}", _job_uid(job))
        self._progress(job, "uploading_results", "上传测图数据工作簿")
        upload_result = self._upload_material_result_workbook(job, workbook_path, asset_uid)
        self._renew(job)

        import_payload_base = {
            "job_uid": _job_uid(job),
            "lease_id": _lease_id(job),
            "source": {
                "source_uid": asset_uid,
                "filename": workbook_path.name,
                "object_key": upload_result.get("object_key", ""),
            },
        }
        self._progress(job, "importing", "导入测图数据到云端")
        import_result = self._import_material_rows(import_payload_base, parsed["overview_rows"], parsed["detail_rows"])
        self._renew(job)
        return {
            "status": "succeeded",
            "result": {
                "workbook_asset_uid": asset_uid,
                "workbook_filename": workbook_path.name,
                "workbook_object_key": upload_result.get("object_key", ""),
                "overview_rows": import_result.get("overview_rows", len(parsed["overview_rows"])),
                "detail_rows": import_result.get("detail_rows", len(parsed["detail_rows"])),
                "inserted_or_updated": import_result.get("inserted_or_updated", 0),
            },
        }

    def _import_material_rows(self, payload_base: Mapping[str, Any], overview_rows: list[dict], detail_rows: list[dict]) -> dict:
        totals = {"overview_rows": 0, "detail_rows": 0, "inserted_or_updated": 0}
        chunks = list(_chunks(detail_rows, MATERIAL_IMPORT_DETAIL_CHUNK_SIZE)) or [[]]
        for index, detail_chunk in enumerate(chunks):
            payload = {
                **payload_base,
                "overview_rows": overview_rows if index == 0 else [],
                "detail_rows": detail_chunk,
            }
            result = self.client.request_json("POST", "/api/material-test/import", payload)
            totals["overview_rows"] += int(result.get("overview_rows") or 0)
            totals["detail_rows"] += int(result.get("detail_rows") or 0)
            totals["inserted_or_updated"] += int(result.get("inserted_or_updated") or 0)
        return totals

    def _task_dir(self, job: Mapping[str, Any]) -> Path:
        job_uid = _safe_segment(str(job.get("job_uid") or "job"))
        path = self.work_dir / job_uid
        path.mkdir(parents=True, exist_ok=True)
        return path

    def _download_asset(self, job: Mapping[str, Any], asset_uid: str, target_dir: Path) -> Path:
        target_dir.mkdir(parents=True, exist_ok=True)
        safe_uid = _safe_segment(asset_uid)
        target = target_dir / f"{safe_uid}.jpg"
        if hasattr(self.client, "download_asset"):
            return Path(self.client.download_asset(asset_uid, target, job_uid=_job_uid(job), lease_id=_lease_id(job)))
        raise CloudApprovalError("cloud client does not support asset download")

    def _upload_result_asset(
        self,
        *,
        job: Mapping[str, Any],
        batch_uid: str,
        style_id: int,
        asset_uid: str,
        path: Path,
        prompt_text: str,
        parent_asset_uid: str,
        generation_job_id: str,
        prompt_template_version_id: int | None = None,
    ) -> dict:
        content_type = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
        presign = self.client.request_json("POST", "/api/assets/presign", {
            "batch_uid": batch_uid,
            "style_id": style_id,
            "asset_uid": asset_uid,
            "job_uid": _job_uid(job),
            "lease_id": _lease_id(job),
            "kind": "ai",
            "filename": path.name,
            "content_type": content_type,
            "prompt_text": prompt_text,
            "prompt_template_version_id": prompt_template_version_id,
            "parent_asset_uid": parent_asset_uid,
            "generation_job_id": generation_job_id,
        })
        upload_url = str(presign.get("upload_url") or "")
        if not upload_url:
            raise CloudApprovalError(f"cloud presign response missing upload_url for asset {asset_uid}")
        self.client.upload_asset(upload_url, path, content_type)
        return presign

    def _upload_material_result_workbook(self, job: Mapping[str, Any], path: Path, asset_uid: str) -> dict:
        content_type = mimetypes.guess_type(path.name)[0] or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        presign = self.client.request_json("POST", "/api/assets/presign", {
            "batch_uid": "material-test",
            "style_id": 1,
            "asset_uid": asset_uid,
            "job_uid": _job_uid(job),
            "lease_id": _lease_id(job),
            "kind": "result",
            "filename": path.name,
            "content_type": content_type,
            "meta": {"job_type": "crawl_tmall_material_test_data"},
        })
        upload_url = str(presign.get("upload_url") or "")
        if not upload_url:
            raise CloudApprovalError("cloud presign response missing upload_url for material test workbook")
        self.client.upload_asset(upload_url, path, content_type)
        return presign

    def _progress(self, job: Mapping[str, Any], phase: str, message: str) -> dict:
        response = self.client.request_json("POST", f"/api/jobs/{_job_uid(job)}/progress", {
            "job_uid": _job_uid(job),
            "lease_id": _lease_id(job),
            "message": message,
            "result": {"phase": phase},
        })
        _raise_if_cancel_requested(response)
        return response

    def _renew(self, job: Mapping[str, Any]) -> dict:
        response = self.client.request_json("POST", f"/api/jobs/{_job_uid(job)}/renew", {
            "job_uid": _job_uid(job),
            "lease_id": _lease_id(job),
        })
        _raise_if_cancel_requested(response)
        return response

    def _regeneration_batch(self, job: Mapping[str, Any], payload: Mapping[str, Any], task_dir: Path, reference_paths: list[str]) -> dict:
        batch_uid = str(payload.get("batch_uid") or job.get("batch_uid") or "")
        style_id = payload.get("style_id")
        asset_uid = str(payload.get("asset_uid") or "")
        return {
            "batch_id": batch_uid,
            "task_run_uid": _job_uid(job),
            "artifact_dir": str(task_dir),
            "run_params": {},
            "items": [{
                "id": str(style_id or ""),
                "style_id": style_id,
                "assets": [{
                    "id": asset_uid,
                    "kind": "ai",
                    "status": "rejected",
                    "prompt": str(payload.get("prompt_text") or ""),
                    "reference_paths": reference_paths,
                    "generation_row": dict(payload.get("generation_row") or {"完整Prompt": str(payload.get("prompt_text") or "")}),
                }],
            }],
        }

    def _generation_batch(self, job: Mapping[str, Any], payload: Mapping[str, Any], task_dir: Path, source_path: str, reference_paths: list[str]) -> dict:
        batch_uid = str(payload.get("batch_uid") or job.get("batch_uid") or "")
        style_id = payload.get("style_id")
        style_code = str(payload.get("style_code") or "")
        item_id = str(payload.get("item_id") or style_id or "")
        return {
            "batch_id": batch_uid,
            "task_run_uid": _job_uid(job),
            "artifact_dir": str(task_dir),
            "run_params": {},
            "items": [{
                "id": item_id,
                "style_id": style_id,
                "style_code": style_code,
                "item_id": item_id,
                "skc_code": str(payload.get("skc_code") or ""),
                "category": str(payload.get("category") or ""),
                "gender": str(payload.get("gender") or ""),
                "origin_path": source_path,
                "workflow": {
                    "style_code": style_code,
                    "item_id": item_id,
                    "skc_code": str(payload.get("skc_code") or ""),
                    "category": str(payload.get("category") or ""),
                    "gender": str(payload.get("gender") or ""),
                },
                "assets": [
                    {
                        "id": str(payload.get("source_asset_uid") or ""),
                        "kind": "source",
                        "status": "uploaded",
                        "path": source_path,
                        "reference_paths": reference_paths,
                    }
                ],
            }],
        }

    def _submit_batch(self, job: Mapping[str, Any], submit_plan: Mapping[str, Any], task_dir: Path) -> dict:
        batch_uid = str(submit_plan.get("batch_uid") or job.get("batch_uid") or "")
        styles_by_id = {
            int(style.get("id")): style
            for style in submit_plan.get("styles") or []
            if isinstance(style, Mapping) and _positive_int(style.get("id")) > 0
        }
        source_by_style: dict[int, Mapping[str, Any]] = {}
        for asset in submit_plan.get("assets") or []:
            if not isinstance(asset, Mapping):
                continue
            kind = str(asset.get("kind") or "")
            if kind not in ("source", "origin", "main"):
                continue
            style_id = _positive_int(asset.get("style_id"))
            if style_id <= 0 or style_id in source_by_style:
                continue
            source_by_style[style_id] = asset
        items_by_style: dict[int, dict] = {}
        for asset in submit_plan.get("assets") or []:
            if not isinstance(asset, Mapping) or str(asset.get("kind") or "") != "ai":
                continue
            if str(asset.get("status") or "") != "approved":
                continue
            style_id = _positive_int(asset.get("style_id"))
            style = styles_by_id.get(style_id, {})
            local_path = self._download_asset(job, str(asset.get("asset_uid") or ""), task_dir / "approved")
            source_asset = source_by_style.get(style_id)
            if not source_asset:
                raise ValueError(f"approved style {style_id or 'unknown'} is missing a source asset for submit origin")
            origin_path = str(self._download_asset(job, str(source_asset.get("asset_uid") or ""), task_dir / "origins"))
            item = items_by_style.setdefault(style_id, {
                "id": str(style.get("style_uid") or style_id),
                "style_id": style_id,
                "style_code": str(style.get("style_code") or ""),
                "item_id": str(style.get("item_id") or ""),
                "origin_path": origin_path,
                "workflow": {
                    "row_no": style.get("row_no", ""),
                    "style_code": str(style.get("style_code") or ""),
                    "item_id": str(style.get("item_id") or ""),
                },
                "semir_data": {},
                "assets": [],
            })
            item["assets"].append({
                "id": str(asset.get("asset_uid") or ""),
                "kind": "ai",
                "status": "approved",
                "path": str(local_path),
                "filename": str(asset.get("filename") or local_path.name),
                "prompt": str(asset.get("prompt_text") or ""),
                "generation_row": dict(asset.get("meta") or {}),
            })
        return {
            "batch_id": batch_uid,
            "artifact_dir": str(task_dir),
            "run_params": {},
            "items": list(items_by_style.values()),
        }

    def _tmall_module(self):
        if self.tmall_module is None:
            self.tmall_module = _load_tmall_module()
        return self.tmall_module

    def _material_test_runner(self):
        if self.material_test_runner is None:
            self.material_test_runner = _default_material_test_runner
        return self.material_test_runner


def _load_tmall_module():
    script = Path(__file__).resolve().parents[1] / "adapters" / "tmall-ops-assistant" / "tools" / "run_tmall_ai_image_test_chain.py"
    spec = importlib.util.spec_from_file_location("cloud_tmall_ai_image_test_chain", script)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Unable to load Tmall cloud executor module: {script}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def _run_maybe_async(value):
    if asyncio.iscoroutine(value):
        return asyncio.run(value)
    return value


def _default_material_test_runner(run_params: Mapping[str, Any], _task_dir: Path, _job: Mapping[str, Any]) -> Any:
    from core import api_server

    return api_server._execute_task("tmall-ops-assistant", "tmall_material_test_data_export", dict(run_params), {}, run_control=None)


def _material_workbook_path(result: Any, export_dir: Path, task_dir: Path) -> Path:
    candidates: list[Path] = []
    if isinstance(result, (str, Path)):
        candidates.append(Path(result))
    if isinstance(result, Mapping):
        for key in ("workbook_path", "result_path", "path"):
            if result.get(key):
                candidates.append(Path(str(result.get(key))))
        for key in ("output_files", "exported_files", "files"):
            value = result.get(key)
            if isinstance(value, list):
                candidates.extend(Path(str(item)) for item in value if str(item or ""))
    candidates.extend(sorted(export_dir.glob("*.xlsx"), key=lambda path: path.stat().st_mtime, reverse=True))
    for candidate in candidates:
        path = candidate.expanduser().resolve()
        if path.suffix.lower() in {".xlsx", ".xlsm", ".xls"} and path.is_file() and (
            path.is_relative_to(task_dir.resolve()) or path.is_relative_to(export_dir.resolve())
        ):
            return path
    raise ValueError("material test crawl did not produce an XLSX workbook in the task directory")


def parse_tmall_material_test_workbook(path: Path) -> dict:
    try:
        import openpyxl
    except Exception as exc:  # pragma: no cover - dependency is declared in core/requirements.txt
        raise RuntimeError("openpyxl is required to parse material test workbooks") from exc

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    overview_rows = _sheet_records(workbook, "概览", _normalize_material_overview_row)
    detail_rows = _sheet_records(workbook, "明细", _normalize_material_detail_row)
    workbook.close()
    return {"overview_rows": overview_rows, "detail_rows": detail_rows}


def _sheet_records(workbook: Any, sheet_name: str, normalizer) -> list[dict]:
    if sheet_name not in workbook.sheetnames:
        return []
    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(cell or "").strip() for cell in next(rows, [])]
    records = []
    for values in rows:
        source = {headers[index]: values[index] if index < len(values) else "" for index in range(len(headers)) if headers[index]}
        record = normalizer(source)
        if record:
            records.append(record)
    return records


def _normalize_material_overview_row(row: Mapping[str, Any]) -> dict:
    record = {
        "record_type": _text(row.get("记录类型")),
        "row_no": _nullable_int(row.get("表格行号")),
        "style_code": _text(row.get("款号")),
        "item_id": _text(row.get("商品ID")),
        "item_title": _text(row.get("商品标题")),
        "task_id": _text(row.get("任务ID")),
        "test_status": _text(row.get("测试状态")),
        "test_channel": _text(row.get("测试渠道")),
        "material_count": _int(row.get("测试素材数")),
        "statistic_type": _text(row.get("统计口径") or row.get("测试渠道")),
        "best_material": _text(row.get("最优素材")),
        "execution_result": _text(row.get("执行结果")),
        "remark": _text(row.get("备注")),
    }
    return record if record["item_id"] and record["task_id"] else {}


def _normalize_material_detail_row(row: Mapping[str, Any]) -> dict:
    record = {
        **_normalize_material_overview_row(row),
        "statistic_type": _text(row.get("统计口径")),
        "statistic_date": _text(row.get("统计日期")),
        "image_type": _text(row.get("图片类型")),
        "material_id": _text(row.get("素材ID")),
        "material_ratio": _text(row.get("素材比例")),
        "material_share": _decimal(row.get("素材占比")),
        "material_url": _text(row.get("素材URL")),
        "search_impressions": _int(row.get("搜索曝光")),
        "search_clicks": _int(row.get("搜索点击")),
        "search_ctr": _decimal(row.get("搜索点击率")),
        "detail_impressions": _int(row.get("详情曝光")),
        "detail_clicks": _int(row.get("详情点击")),
        "detail_ctr": _decimal(row.get("详情点击率")),
        "detail_add_to_cart": _int(row.get("详情加购")),
        "detail_pay_conversion": _int(row.get("详情支付转化")),
        "detail_pay_conversion_rate": _decimal(row.get("详情支付转化率")),
        "data_download_url": _text(row.get("数据下载链接")),
    }
    return record if record.get("item_id") and record.get("task_id") and record.get("statistic_type") and record.get("material_url") else {}


def _payload(job: Mapping[str, Any]) -> Mapping[str, Any]:
    payload = job.get("payload")
    if not isinstance(payload, Mapping) or not payload:
        raise ValueError(f"{job.get('job_type') or 'cloud job'} payload is missing")
    return payload


def _required_text(payload: Mapping[str, Any], key: str) -> str:
    value = str(payload.get(key) or "")
    if not value:
        raise ValueError(f"cloud job payload is missing {key}")
    return value


def _job_uid(job: Mapping[str, Any]) -> str:
    return _required_text(job, "job_uid")


def _lease_id(job: Mapping[str, Any]) -> str:
    return _required_text(job, "lease_id")


def _chunks(rows: list[dict], size: int):
    for index in range(0, len(rows), size):
        yield rows[index:index + size]


def _text_list(value: Any) -> list[str]:
    if not isinstance(value, list):
        return []
    return [str(item) for item in value if str(item or "")]


def _positive_int(value: Any) -> int:
    try:
        number = int(value)
    except (TypeError, ValueError):
        return 0
    return number if number > 0 else 0


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _int(value: Any) -> int:
    try:
        text = str(value or "").replace(",", "").strip()
        return int(float(text)) if text else 0
    except (TypeError, ValueError):
        return 0


def _nullable_int(value: Any) -> int | None:
    number = _int(value)
    return number if number > 0 else None


def _decimal(value: Any) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value or "").replace(",", "").strip()
    if not text or text == "-":
        return 0.0
    if text.endswith("%"):
        try:
            return float(text[:-1]) / 100
        except ValueError:
            return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def _safe_segment(value: str) -> str:
    text = str(value or "").strip().replace("\\", "/").split("/")[-1]
    safe = "".join(ch if ch.isalnum() or ch in "._-" else "-" for ch in text)
    return safe.strip(".-") or "item"


def _safe_generated_asset_uid(request_uid: str, job_uid: str) -> str:
    source = request_uid or f"gen-{job_uid}"
    safe = _safe_segment(source)
    return safe if safe.startswith("gen-") else f"gen-{safe}"


def _raise_if_cancel_requested(response: Mapping[str, Any]) -> None:
    if response.get("cancel_requested") is True:
        raise CloudJobCancelled("cloud job cancellation requested")


def _local_file(value: Any, root: Path) -> Path:
    path = Path(str(value or ""))
    if not path.is_file() or not path.resolve().is_relative_to(root.resolve()):
        raise ValueError("cloud job result must be a local task file")
    return path


def _safe_result_path(value: Any, root: Path) -> str:
    if not value:
        return ""
    path = Path(str(value))
    if not path.is_file() or not path.resolve().is_relative_to(root.resolve()):
        return ""
    return str(path)


def _looks_like_login_block(message: str) -> bool:
    lowered = str(message or "").lower()
    return any(marker in lowered for marker in ("login", "登录", "cdp", "chrome", "cookie"))


def _submit_result_failed(result: Any) -> bool:
    if not isinstance(result, Mapping):
        return False
    if result.get("ok") is False:
        return True
    if _positive_int(result.get("failed")) > 0:
        return True
    return str(result.get("status") or "") in {"create_failed", "partial_failed"}


def _submit_failure_message(result: Any) -> str:
    if not isinstance(result, Mapping):
        return "tmall_submit_failed"
    status = str(result.get("status") or "tmall_submit_failed")
    failed = _positive_int(result.get("failed"))
    attempted = _positive_int(result.get("attempted"))
    parts = [status]
    if failed:
        parts.append(f"failed={failed}")
    if attempted:
        parts.append(f"attempted={attempted}")
    message = str(result.get("message") or result.get("error") or "")
    if message:
        parts.append(message)
    return "; ".join(parts)
