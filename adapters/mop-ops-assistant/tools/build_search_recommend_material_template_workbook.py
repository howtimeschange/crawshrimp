#!/usr/bin/env python3
"""Build the MOP search recommendation material Excel template workbook."""

from __future__ import annotations

import argparse
from copy import copy
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
NOTE_FILL = PatternFill("solid", fgColor="EAF4FF")
HEAD_FONT = Font(color="FFFFFF", bold=True)
GRID_SIDE = Side(style="thin", color="808080")
GRID_BORDER = Border(left=GRID_SIDE, right=GRID_SIDE, top=GRID_SIDE, bottom=GRID_SIDE)
TASK_LAST_ROW = 30


def set_header(row) -> None:
    for cell in row:
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = GRID_BORDER


def freeze_and_filter(ws, ref: str) -> None:
    ws.auto_filter.ref = ref
    ws.freeze_panes = "A2"


def apply_grid(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = GRID_BORDER
            alignment = copy(cell.alignment)
            alignment.vertical = alignment.vertical or "top"
            alignment.wrap_text = True
            cell.alignment = alignment


def build_workbook(output: Path) -> None:
    wb = Workbook()
    ws_task = wb.active
    ws_task.title = "填写任务"
    ws_guide = wb.create_sheet("字段说明")

    task_headers = ["商品ID", "商家编码", "达人", "素材图片", "素材张数", "添加标题", "内容描述", "裁剪比例", "备注"]
    ws_task.append(task_headers)
    set_header(ws_task[1])
    ws_task.append([
        "1051467606993",
        "455133A2114Z",
        "",
        "",
        3,
        "亚麻连衣裙穿搭",
        "清爽亚麻质感搭配简洁版型，日常通勤和周末出游都很适合，突出自然垂坠和轻盈气质。",
        "3:4",
        "示例：素材图片可留空；同款单行会按主图下每个达人文件夹各发布一条；同款多行会按行顺序分配达人包，保留每行标题/描述",
    ])
    ws_task.append([
        "776047586897",
        "564231A5355Z",
        "达人A",
        "/Users/you/Desktop/搜推素材/776047586897/01.jpg;/Users/you/Desktop/搜推素材/776047586897/02.jpg;/Users/you/Desktop/搜推素材/776047586897/03.jpg",
        "",
        "凉感Polo上新",
        "轻薄凉感面料适合夏季穿着，版型利落不闷热，适合通勤、休闲和日常搭配。",
        "1:1",
        "示例：直接填写三张图片路径",
    ])
    for idx, width in enumerate([18, 18, 16, 72, 12, 24, 80, 12, 56, 92], 1):
        ws_task.column_dimensions[get_column_letter(idx)].width = width
    for row in ws_task.iter_rows(min_row=2, max_row=3):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    crop_col = get_column_letter(task_headers.index("裁剪比例") + 1)
    for cell in ws_task[f"{crop_col}2:{crop_col}{TASK_LAST_ROW}"]:
        cell[0].number_format = "@"
    crop_validation = DataValidation(type="list", formula1='"3:4,1:1"', allow_blank=True)
    crop_validation.error = "裁剪比例仅支持 3:4 或 1:1"
    crop_validation.errorTitle = "裁剪比例填写错误"
    crop_validation.prompt = "请选择 3:4 或 1:1；留空则使用运行界面默认值。"
    crop_validation.promptTitle = "裁剪比例"
    ws_task.add_data_validation(crop_validation)
    crop_validation.add(f"{crop_col}2:{crop_col}{TASK_LAST_ROW}")

    notes = [
        "搜推图文素材要求至少 3 张图片，最多 9 张图片。",
        "添加标题必填，上限 20 个中文字符；内容描述必填，上限 1000 个中文字符。",
        "商品ID和商家编码二选一必填；商品ID为空时，脚本会自动按商家编码去千牛“我的商品”解析商品ID。",
        "素材图片可写本地绝对路径或 http(s) URL；多张可用换行、分号、竖线或空格分隔。",
        "业务图包约定：素材根目录/含商家编码的商品文件夹/图片/主图/达人/图片；顶层目录可用 + 关联多个商家编码。",
        "如果同款只填一行，脚本会按主图下的每个达人文件夹拆成多条内容，每个达人文件夹各取前 N 张素材。",
        "如果同款在表格里填写多行，脚本会按行顺序分配达人图片包；每行自己的标题、内容描述、裁剪比例、备注会保留。",
        "可选填写“达人”列精确指定达人文件夹名。",
        "未匹配业务图包时，兼容旧规则：素材根目录/商品ID或商家编码/01.jpg、02.jpg、03.jpg。",
        "上传前会自动居中裁剪到 3:4 或 1:1；单行“裁剪比例”列可覆盖运行界面默认值。",
        "手动多选命名约定：商品ID_01.jpg 或 商家编码_01.jpg，或父文件夹名为商品ID/商家编码。",
    ]
    note_col = len(task_headers) + 1
    ws_task.cell(1, note_col).value = "填写说明"
    ws_task.cell(1, note_col).fill = HEAD_FILL
    ws_task.cell(1, note_col).font = HEAD_FONT
    ws_task.cell(1, note_col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_task.cell(2, note_col).value = "\n".join(notes[:2])
    ws_task.cell(2, note_col).fill = WARN_FILL
    ws_task.cell(3, note_col).value = "\n".join(notes[2:])
    ws_task.cell(3, note_col).fill = NOTE_FILL
    for row_idx in range(1, TASK_LAST_ROW + 1):
        ws_task.row_dimensions[row_idx].height = 24
    ws_task.row_dimensions[2].height = 96
    ws_task.row_dimensions[3].height = 102
    apply_grid(ws_task, 1, TASK_LAST_ROW, 1, note_col)
    freeze_and_filter(ws_task, f"A1:I{TASK_LAST_ROW}")

    guide_rows = [
        ("商品ID", "二选一", "千牛/淘宝商品 ID；有值时优先使用。为空但填写商家编码时，脚本会自动解析商品ID", "1051467606993"),
        ("商家编码", "二选一", "店铺内部商家编码；商品ID为空时，脚本会去千牛“我的商品”按商家编码搜索并取回商品ID", "455133A2114Z"),
        ("达人", "否", "填写业务图包主图下的达人文件夹名时，会精确使用该达人图片包；不填时，单行同款会展开全部达人，多行重复同款会按行顺序分配达人包", "AHOYECHO"),
        ("素材图片", "否", "本地绝对路径或 http(s) 图片 URL；多张用换行、分号、竖线或空格分隔。也可以留空，改在 UI 选择素材根目录或手动多选素材图片", "/Users/you/Desktop/搜推素材/1051467606993/01.jpg;/Users/you/Desktop/搜推素材/1051467606993/02.jpg;/Users/you/Desktop/搜推素材/1051467606993/03.jpg"),
        ("素材张数", "否", "素材图片为空且填写了“素材根目录”时生效；单行同款展开全部达人，多行重复同款按行顺序分配达人包；每条取对应达人前 N 张；搜推图文至少 3 张、最多 9 张", "3"),
        ("添加标题", "是", "发布弹窗里的“添加标题”；必填，上限 20 个中文字符；页面建议 8-10 个中文字符", "亚麻连衣裙穿搭"),
        ("内容描述", "是", "发布弹窗里的“内容描述”；必填，上限 1000 个中文字符；页面建议 10-1000 个中文字符", "清爽亚麻质感搭配简洁版型，日常通勤和周末出游都很适合。"),
        ("裁剪比例", "否", "上传前自动居中裁剪；支持 3:4 或 1:1；留空使用运行界面的默认裁剪比例", "3:4"),
        ("备注", "否", "原样带到结果里，便于人工追踪", "第一批搜推素材"),
        ("素材包摆放", "建议", "推荐业务图包：素材根目录/MOP655137Z4106Z+656939D1213Y/图片/主图/达人/图片；顶层目录可用 + 关联多个商家编码；脚本只取主图图片，忽略视频和买家秀；多个达人文件夹可自动展开或按重复行分配", "/Users/you/Desktop/搜推素材/MOP655137Z4106Z+656939D1213Y/图片/主图/AHOYECHO/xxx.jpg"),
        ("手动多选命名", "建议", "使用“手动选择素材图片”时，文件名以商品ID/商家编码开头或父文件夹为商品ID/商家编码；脚本会自动归组", "455133A2114Z_01.jpg 或 455133A2114Z/01.jpg"),
    ]
    ws_guide.append(["字段", "是否必填", "说明", "示例"])
    set_header(ws_guide[1])
    for row in guide_rows:
        ws_guide.append(row)
    for idx, width in enumerate([18, 12, 86, 64], 1):
        ws_guide.column_dimensions[get_column_letter(idx)].width = width
    for row in ws_guide.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    apply_grid(ws_guide, 1, len(guide_rows) + 1, 1, 4)
    freeze_and_filter(ws_guide, f"A1:D{len(guide_rows) + 1}")

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = True

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()
    build_workbook(args.output)
    print(args.output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
