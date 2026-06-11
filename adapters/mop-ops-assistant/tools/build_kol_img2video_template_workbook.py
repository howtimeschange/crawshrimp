#!/usr/bin/env python3
"""Build the MOP KOL display-video Excel template workbook."""

from __future__ import annotations

import argparse
from copy import copy
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
NOTE_FILL = PatternFill("solid", fgColor="EAF4FF")
HEAD_FONT = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
GRID_SIDE = Side(style="thin", color="808080")
GRID_BORDER = Border(left=GRID_SIDE, right=GRID_SIDE, top=GRID_SIDE, bottom=GRID_SIDE)
TASK_LAST_ROW = 30


def set_header(row) -> None:
    for cell in row:
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = GRID_BORDER


def freeze_and_filter(ws, ref: str) -> None:
    ws.auto_filter.ref = ref
    ws.freeze_panes = "A2"


def apply_grid(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = GRID_BORDER
            alignment = copy(cell.alignment)
            alignment.vertical = alignment.vertical or "top"
            alignment.wrap_text = True
            cell.alignment = alignment


def build_workbook(output: Path) -> None:
    wb = Workbook()
    ws_task = wb.active
    ws_task.title = "填写任务"
    ws_guide = wb.create_sheet("字段说明")

    task_headers = ["商品ID", "商家编码", "达人", "素材图片", "素材张数", "比例", "提示词", "备注"]
    ws_task.append(task_headers)
    set_header(ws_task[1])
    ws_task.append([
        "728857154429",
        "46X096070266",
        "",
        "",
        3,
        "3:4",
        "",
        "素材图片可留空：同款单行会按主图下每个达人文件夹各生成一条；同款多行会按行顺序分配达人包，也可填写达人精确指定",
    ])
    ws_task.append([
        "741042967594",
        "564231A5355Z",
        "达人A",
        "/Users/you/Desktop/KOL素材/741042967594/01.jpg;/Users/you/Desktop/KOL素材/741042967594/02.jpg",
        "",
        "3:4",
        "突出版型和面料质感",
        "示例：直接填写多张素材图片路径",
    ])
    for idx, width in enumerate([18, 18, 16, 68, 12, 10, 36, 54, 92], 1):
        ws_task.column_dimensions[get_column_letter(idx)].width = width
    for row in ws_task.iter_rows(min_row=2, max_row=3):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    notes = [
        "现在脚本按千牛页面的“选商品 + 图片生成展示视频”逻辑提交，不再选择或指定模板。",
        "素材图片不是必填；如果留空，请在运行界面选择“素材根目录”或“手动选择素材图片”。",
        "商品ID和商家编码二选一必填；商品ID为空时，脚本会自动按商家编码去千牛“我的商品”解析商品ID。",
        "业务图包约定：素材根目录/含商家编码的商品文件夹/图片/主图/达人/图片；顶层目录可用 + 关联多个商家编码。",
        "如果同款只填一行，脚本会按主图下的每个达人文件夹拆成多条内容，每个达人文件夹各取前 N 张素材。",
        "如果同款在表格里填写多行，脚本会按行顺序分配达人图片包；每行自己的比例、提示词、备注会保留。",
        "可选填写“达人”列精确指定达人文件夹名。",
        "未匹配业务图包时，兼容旧规则：素材根目录/商品ID或商家编码/01.jpg、02.jpg、03.jpg。",
        "手动多选命名约定：商品ID_01.jpg 或 商家编码_01.jpg，或父文件夹名为商品ID/商家编码。",
        "素材图片列支持本地绝对路径或 http(s) 图片 URL；多张可用换行、分号、竖线或空格分隔。",
    ]
    note_col = len(task_headers) + 1
    ws_task.cell(1, note_col).value = "填写说明"
    ws_task.cell(1, note_col).fill = HEAD_FILL
    ws_task.cell(1, note_col).font = HEAD_FONT
    ws_task.cell(1, note_col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_task.cell(2, note_col).value = "\n".join(notes[:2])
    ws_task.cell(2, note_col).fill = WARN_FILL
    ws_task.cell(3, note_col).value = "\n".join(notes[2:])
    ws_task.cell(3, note_col).fill = NOTE_FILL
    for row_idx in range(1, TASK_LAST_ROW + 1):
        ws_task.row_dimensions[row_idx].height = 24
    ws_task.row_dimensions[2].height = 92
    ws_task.row_dimensions[3].height = 88
    apply_grid(ws_task, 1, TASK_LAST_ROW, 1, note_col)
    freeze_and_filter(ws_task, f"A1:H{TASK_LAST_ROW}")

    dv = DataValidation(type="list", formula1='"3:4,1:1,9:16,16:9"', allow_blank=True)
    ws_task.add_data_validation(dv)
    dv.add("F2:F500")

    guide_rows = [
        ("商品ID", "二选一", "千牛/淘宝商品 ID；有值时优先使用。为空但填写商家编码时，脚本会自动解析商品ID", "728857154429"),
        ("商家编码", "二选一", "店铺内部商家编码；商品ID为空时，脚本会去千牛“我的商品”按商家编码搜索并取回商品ID", "46X096070266"),
        ("达人", "否", "填写业务图包主图下的达人文件夹名时，会精确使用该达人图片包；不填时，单行同款会展开全部达人，多行重复同款会按行顺序分配达人包", "AHOYECHO"),
        ("素材图片", "否", "本地绝对路径或 http(s) 图片 URL；多张用换行、分号、竖线或空格分隔。也可以留空，改在 UI 选择素材根目录或手动多选素材图片", "/Users/you/Desktop/KOL素材/728857154429/01.jpg;/Users/you/Desktop/KOL素材/728857154429/02.jpg"),
        ("素材张数", "否", "素材图片为空且填写了“素材根目录”时生效；单行同款展开全部达人，多行重复同款按行顺序分配达人包；每条取对应达人前 N 张；旧规则读取 01.jpg 到 NN.jpg", "3"),
        ("比例", "否", "展示视频生成比例；为空使用运行界面的默认比例", "3:4"),
        ("提示词", "否", "传给每张图片的生成提示；为空则不传", "突出商品版型和面料质感"),
        ("备注", "否", "原样带到结果里，便于人工追踪", "第一批 KOL 素材"),
        ("素材包摆放", "建议", "推荐业务图包：素材根目录/MOP655137Z4106Z+656939D1213Y/图片/主图/达人/图片；顶层目录可用 + 关联多个商家编码；脚本只取主图图片，忽略视频和买家秀；多个达人文件夹可自动展开或按重复行分配", "/Users/you/Desktop/KOL素材/MOP655137Z4106Z+656939D1213Y/图片/主图/AHOYECHO/xxx.jpg"),
        ("手动多选命名", "建议", "使用“手动选择素材图片”时，文件名以商品ID/商家编码开头或父文件夹为商品ID/商家编码；脚本会自动归组", "46X096070266_01.jpg 或 46X096070266/01.jpg"),
    ]
    ws_guide.append(["字段", "是否必填", "说明", "示例"])
    set_header(ws_guide[1])
    for row in guide_rows:
        ws_guide.append(row)
    for idx, width in enumerate([18, 12, 84, 58], 1):
        ws_guide.column_dimensions[get_column_letter(idx)].width = width
    for row in ws_guide.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    apply_grid(ws_guide, 1, len(guide_rows) + 1, 1, 4)
    freeze_and_filter(ws_guide, f"A1:D{len(guide_rows) + 1}")

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = True

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--catalog", required=False, type=Path, help="Ignored; kept for old build commands")
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()
    build_workbook(args.output)
    print(args.output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
