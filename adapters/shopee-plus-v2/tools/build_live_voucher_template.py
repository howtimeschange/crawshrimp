from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "Shopee直播优惠券批量设置模板（26.4.8）.xlsx"


HEADERS = [
    "站点",
    "店铺",
    "优惠券品类",
    "优惠券名称（可选）",
    "优惠券领取期限（开始）精确到分",
    "优惠券领取期限（结束）精确到分",
    "是否提前显示优惠券",
    "主播用户名列表",
    "奖励类型",
    "折扣类型",
    "优惠限额",
    "最高优惠金额（无限制填nocap）",
    "最低消费金额",
    "可使用总数",
    "每个买家可用的优惠券数量上限",
    "优惠商品范围",
    "是否覆盖已有券",
]


EXAMPLE_ROWS = [
    [
        "新加坡",
        "your-shop.sg",
        "直播优惠券",
        "",
        "2026-04-09 20:00",
        "2026-04-09 23:59",
        "0",
        "",
        "折扣",
        "折扣金额",
        "10",
        "",
        "59",
        "200",
        "1",
        "全部商品",
        "0",
    ],
    [
        "新加坡",
        "your-shop.sg",
        "限定的主播优惠券",
        "",
        "2026-04-10 20:00",
        "2026-04-10 23:59",
        "1",
        "streamer_a,streamer_b",
        "Shopee币回扣",
        "扣除百分比",
        "10",
        "nocap",
        "99",
        "300",
        "1",
        "全部商品",
        "0",
    ],
]


GUIDE_ROWS = [
    ["字段", "说明", "示例 / 允许值"],
    ["站点", "Shopee 站点名称", "新加坡 / 马来西亚 / 泰国 ..."],
    ["店铺", "后台当前店铺名称，建议与店铺切换器展示一致", "your-shop.sg"],
    ["优惠券品类", "仅支持 usecase=6 的两种券型", "直播优惠券 / 限定的主播优惠券"],
    ["优惠券名称（可选）", "可留空；留空时脚本按“折扣类型 + 优惠限额”自动生成", "折扣金额：10"],
    ["优惠券领取期限（开始/结束）精确到分", "格式必须是 YYYY-MM-DD HH:MM", "2026-04-09 20:00"],
    ["是否提前显示优惠券", "1=是，0=否", "1"],
    ["主播用户名列表", "仅“限定的主播优惠券”必填，推荐使用英文逗号分隔；同时兼容换行、中文逗号、分号和 |", "streamer_a,streamer_b"],
    ["奖励类型", "直播券支持折扣 / Shopee币回扣", "折扣 / Shopee币回扣"],
    ["折扣类型", "Shopee币回扣仅支持“扣除百分比”", "折扣金额 / 扣除百分比"],
    ["优惠限额", "折扣金额填金额；扣除百分比填百分数", "10"],
    ["最高优惠金额（无限制填nocap）", "百分比或 Shopee币回扣场景建议填写；无限制填 nocap", "nocap / 30"],
    ["最低消费金额", "订单最低消费金额", "59"],
    ["可使用总数", "优惠券总量", "200"],
    ["每个买家可用的优惠券数量上限", "每位买家可用上限", "1"],
    ["优惠商品范围", "当前版本仅支持“全部商品”", "全部商品"],
    ["是否覆盖已有券", "1=发现时间重叠券时先结束/删除旧券，0=不处理", "0 / 1"],
]


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    note_fill = PatternFill("solid", fgColor="FFF2CC")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 24

    widths = {
        "A": 12,
        "B": 22,
        "C": 18,
        "D": 22,
        "E": 22,
        "F": 22,
        "G": 14,
        "H": 30,
        "I": 16,
        "J": 14,
        "K": 12,
        "L": 22,
        "M": 12,
        "N": 12,
        "O": 18,
        "P": 16,
        "Q": 14,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="center")

    optional_columns = {4, 8, 12}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for idx in optional_columns:
            if len(row) >= idx:
                row[idx - 1].fill = note_fill


def build_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "LiveVouchers"
    ws.append(HEADERS)
    for row in EXAMPLE_ROWS:
        ws.append(row)
    style_sheet(ws)

    guide = wb.create_sheet("填写说明")
    for row in GUIDE_ROWS:
        guide.append(row)
    style_sheet(guide)
    guide.column_dimensions["A"].width = 28
    guide.column_dimensions["B"].width = 52
    guide.column_dimensions["C"].width = 36

    for idx, row in enumerate(guide.iter_rows(min_row=2, max_row=guide.max_row), start=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        guide.row_dimensions[idx].height = 34

    return wb


if __name__ == "__main__":
    wb = build_workbook()
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(OUT)
