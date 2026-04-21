from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "templates" / "lazada-voucher-template.xlsx"
ENTRY_MAX_ROW = 500
ENUM_SHEET_TITLE = "_枚举值"


VOUCHER_HEADERS = [
    "唯一键",
    "站点",
    "优惠工具",
    "促销名称",
    "领取生效类型",
    "券生效开始时间",
    "券生效结束时间",
    "最早可领取时间",
    "领取开始时间",
    "领取结束时间",
    "领取后可用天数",
    "适用范围",
    "折扣类型",
    "折扣值",
    "最低消费金额",
    "最高优惠金额",
    "每人限用次数",
    "预算模式",
    "预算金额",
    "总发行量",
    "Flexi主玩法",
    "Flexi子玩法",
    "Flexi条件类型",
    "Flexi是否叠加折扣",
    "Flexi总单量",
    "Flexi赠品是否免邮",
    "Flexi固定件数",
    "Flexi固定总价",
    "备注",
]

VOUCHER_EXAMPLES = [
    [
        "REGULAR_MY_001",
        "MY",
        "Regular Voucher",
        "MY Regular Voucher Demo",
        "FIXED_TIME",
        "2026-04-05 10:00",
        "2026-04-30 23:59",
        "2026-04-05 08:00",
        "",
        "",
        "",
        "ENTIRE_SHOP",
        "PERCENT_OFF",
        "10",
        "99",
        "20",
        "1",
        "",
        "",
        "300",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "示例：Regular Voucher 百分比折扣",
    ],
    [
        "FLEXI_SG_001",
        "SG",
        "Flexi Combo",
        "SG Flexi Combo Demo",
        "",
        "2026-04-05 10:00",
        "2026-04-30 23:59",
        "",
        "",
        "",
        "",
        "ENTIRE_SHOP",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "MONEY_DISCOUNT_OFF",
        "PercentageDiscountOff",
        "ITEM_QUANTITY",
        "0",
        "100",
        "",
        "",
        "",
        "示例：Flexi Combo 百分比阶梯折扣，具体门槛在 FlexiTiers",
    ],
    [
        "NEWBUYER_MY_001",
        "MY",
        "Store New Buyer Voucher",
        "MY New Buyer Demo",
        "FIXED_TIME",
        "2026-04-05 10:00",
        "2026-05-05 23:59",
        "2026-04-05 09:00",
        "",
        "",
        "",
        "ENTIRE_SHOP",
        "PERCENT_OFF",
        "15",
        "49",
        "20",
        "1",
        "",
        "",
        "500",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "示例：新买家券百分比折扣",
    ],
    [
        "FOLLOWER_MY_001",
        "MY",
        "Store Follower Voucher",
        "MY Follower Demo",
        "FIXED_TIME",
        "2026-04-05 10:00",
        "2026-05-05 23:59",
        "2026-04-05 09:00",
        "",
        "",
        "",
        "ENTIRE_SHOP",
        "PERCENT_OFF",
        "8",
        "59",
        "15",
        "1",
        "",
        "",
        "800",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "示例：粉丝券百分比折扣",
    ],
]

FLEXI_TIER_HEADERS = [
    "唯一键",
    "阶梯序号",
    "门槛值",
    "优惠值",
    "赠品数量",
    "买家可选赠品数",
    "赠品SKU列表",
    "备注",
]

FLEXI_TIER_EXAMPLES = [
    ["FLEXI_SG_001", "1", "2", "5", "", "", "", "买 2 件打 5%"],
    ["FLEXI_SG_001", "2", "3", "8", "", "", "", "买 3 件打 8%"],
]

INSTRUCTION_ROWS = [
    ["Sheet", "字段", "说明"],
    ["Vouchers", "枚举型字段", "模板里的站点、优惠工具、领取生效类型、适用范围、折扣类型、预算模式、Flexi 主玩法 / 子玩法 / 条件类型等字段已预置下拉框。"],
    ["Vouchers", "站点", "页面右侧国家栏可切换 Indonesia / Malaysia / Philippine / Singapore / Thailand / Vietnam；当前版本支持六站点切换，建议先按小批量首跑。"],
    ["Vouchers", "优惠工具", "支持填写 Regular Voucher、Flexi Combo、Store New Buyer Voucher、Store Follower Voucher。"],
    ["Vouchers", "启用", "已移除。导入模板里的每一行默认都会参与执行。"],
    ["Vouchers", "领取生效类型", "仅 Voucher 类工具使用；FIXED_TIME 对应固定时间生效，USE_AFTER_COLLECTION 对应领取后生效。"],
    ["Vouchers", "适用范围", "Specific Products 已完成字段探查，但 live phase 仍需补商品选择。"],
    ["Vouchers", "预算模式/预算金额", "仅 Regular Voucher + MONEY_OFF 使用；Regular 的百分比折扣页面改为总发行量。"],
    ["Vouchers", "总发行量", "Store New Buyer / Store Follower 必填；Regular Voucher 的百分比折扣场景也需要填写。"],
    ["Vouchers", "Flexi主玩法", "Flexi Combo 页面有四张主卡：Money/Discount Off、Free gift/sample、Combo Buy、Fixed Price。"],
    ["Vouchers", "Flexi子玩法", "不同主玩法下可选值不同，详情见 lazada-voucher-field-guide.csv。"],
    ["FlexiTiers", "唯一键", "必须和 Vouchers sheet 的“唯一键”一致。"],
    ["FlexiTiers", "门槛值", "Item Quantity 模式填件数；Order Value 模式填订单金额。"],
    ["FlexiTiers", "赠品SKU列表", "当前只做字段固化；live 商品选择弹层后续接入。"],
]

FILL_GUIDE_HEADERS = ["Sheet", "字段", "中文解释", "何时必填", "可选值/枚举", "示例值", "备注"]

FILL_GUIDE_ROWS = [
    ["Vouchers", "唯一键", "主表唯一标识；FlexiTiers 用它关联回主表。", "每行必填", "自定义唯一值", "REGULAR_MY_001", "不要重复。建议用 券类型_站点_序号。"],
    ["Vouchers", "站点", "要在哪个 Lazada 站点创建活动。", "每行必填", "ID, MY, PH, SG, TH, VN", "MY", "页面右侧国家栏实测可见 6 个站点；当前版本支持六站点切换，建议先按小批量首跑。"],
    ["Vouchers", "优惠工具", "活动类型。", "每行必填", "Regular Voucher, Flexi Combo, Store New Buyer Voucher, Store Follower Voucher", "Regular Voucher", "四种工具共用同一张主表。"],
    ["Vouchers", "促销名称", "页面里的 Promotion Name。", "每行必填", "文本", "MY Regular Voucher Demo", "建议同站点内保持可识别且唯一。"],
    ["Vouchers", "领取生效类型", "Voucher Use Time Type。", "Regular / New Buyer / Follower 必填", "FIXED_TIME, USE_AFTER_COLLECTION", "FIXED_TIME", "Flexi Combo 留空。当前 live 最稳的是 FIXED_TIME。"],
    ["Vouchers", "券生效开始时间", "活动生效开始时间。", "Voucher 的 FIXED_TIME；Flexi Combo", "YYYY-MM-DD HH:MM", "2026-04-05 10:00", "对应 Voucher Redeem Period / Effective Period 的开始时间。"],
    ["Vouchers", "券生效结束时间", "活动生效结束时间。", "Voucher 的 FIXED_TIME；Flexi Combo", "YYYY-MM-DD HH:MM", "2026-04-30 23:59", "对应 Voucher Redeem Period / Effective Period 的结束时间。"],
    ["Vouchers", "最早可领取时间", "最早可开始领取的时间。", "Voucher 的 FIXED_TIME 可选", "YYYY-MM-DD HH:MM", "2026-04-05 08:00", "对应 Collect Start Time；不填则沿用页面默认。"],
    ["Vouchers", "领取开始时间", "领取窗口开始时间。", "USE_AFTER_COLLECTION 必填", "YYYY-MM-DD HH:MM", "2026-04-05 08:00", "Flexi Combo 留空。"],
    ["Vouchers", "领取结束时间", "领取窗口结束时间。", "USE_AFTER_COLLECTION 必填", "YYYY-MM-DD HH:MM", "2026-04-30 23:59", "Flexi Combo 留空。"],
    ["Vouchers", "领取后可用天数", "买家领取后可在几天内使用。", "USE_AFTER_COLLECTION 必填", "正整数", "7", "Flexi Combo 留空。"],
    ["Vouchers", "适用范围", "活动作用范围。", "每行必填", "ENTIRE_SHOP, SPECIFIC_PRODUCTS", "ENTIRE_SHOP", "当前 live 只稳定支持 ENTIRE_SHOP。"],
    ["Vouchers", "折扣类型", "Voucher 折扣类型。", "Regular / New Buyer / Follower 必填", "MONEY_OFF, PERCENT_OFF", "PERCENT_OFF", "Flexi Combo 留空。当前 live 最稳的是 PERCENT_OFF。"],
    ["Vouchers", "折扣值", "折扣具体数值。", "Regular / New Buyer / Follower 必填", "数值", "10", "金额减填金额；百分比填数字，不要写 %。"],
    ["Vouchers", "最低消费金额", "订单最低消费门槛。", "Regular / New Buyer / Follower 必填", "数值", "99", "对应 If Order Min.Spend。"],
    ["Vouchers", "最高优惠金额", "百分比折扣时的每单封顶金额。", "Regular / New Buyer / Follower 且 PERCENT_OFF", "数值", "20", "MONEY_OFF 留空。"],
    ["Vouchers", "每人限用次数", "单个买家最多可用几次。", "Regular / New Buyer / Follower 必填", "正整数", "1", "对应 Voucher Limit per Customer。"],
    ["Vouchers", "预算模式", "Regular 券金额减时的预算控制方式。", "Regular Voucher 且 MONEY_OFF", "LIMITED, UNLIMITED", "LIMITED", "PERCENT_OFF 留空。"],
    ["Vouchers", "预算金额", "Regular 券金额减且有限预算时的总预算。", "Regular Voucher 且 MONEY_OFF 且 LIMITED", "数值", "5000", "其余场景留空。"],
    ["Vouchers", "总发行量", "最多发放多少张券。", "Regular Voucher 且 PERCENT_OFF；或 Store New Buyer；或 Store Follower", "正整数", "300", "对应 Total Voucher to be Issued。"],
    ["Vouchers", "Flexi主玩法", "Flexi Combo 顶部四张主玩法卡。", "Flexi Combo 必填", "MONEY_DISCOUNT_OFF, FREE_GIFT_SAMPLE, COMBO_BUY, FIXED_PRICE", "MONEY_DISCOUNT_OFF", "当前 live 已跑通的是 MONEY_DISCOUNT_OFF。"],
    ["Vouchers", "Flexi子玩法", "Flexi Combo 在主玩法下的具体子选项。", "Flexi Combo 大多数主玩法必填", "按主玩法决定", "PercentageDiscountOff", "当前 live 已跑通的是 MONEY_DISCOUNT_OFF + PercentageDiscountOff。"],
    ["Vouchers", "Flexi条件类型", "Deal Criteria 条件类型。", "Flexi Combo 且主玩法 != FIXED_PRICE", "ITEM_QUANTITY, ORDER_VALUE", "ITEM_QUANTITY", "对应按件数或按订单金额达门槛。"],
    ["Vouchers", "Flexi是否叠加折扣", "是否允许叠加折扣。", "Flexi Combo 部分玩法可填", "1, 0", "0", "1=是，0=否。当前稳定值建议填 0。"],
    ["Vouchers", "Flexi总单量", "Flexi Combo 活动总订单量上限。", "Flexi Combo 必填", "正整数", "100", "对应 Total number of Flexi Combo Orders。"],
    ["Vouchers", "Flexi赠品是否免邮", "赠品 / 样品是否免邮。", "Flexi Combo 赠品类玩法", "1, 0", "1", "当前赠品选择 live 尚未接入。"],
    ["Vouchers", "Flexi固定件数", "Fixed Price 玩法的固定购买件数。", "Flexi Combo 且 FIXED_PRICE", "正整数", "3", "当前 FIXED_PRICE 未纳入稳定 live。"],
    ["Vouchers", "Flexi固定总价", "Fixed Price 玩法的总价。", "Flexi Combo 且 FIXED_PRICE", "数值", "99", "当前 FIXED_PRICE 未纳入稳定 live。"],
    ["Vouchers", "备注", "业务备注。", "选填", "文本", "示例：MY 站常规券", "不参与页面填写。"],
    ["FlexiTiers", "唯一键", "关联到主表中的某一条 Flexi Combo。", "每行必填", "引用 Vouchers.唯一键", "FLEXI_SG_001", "同一张 Flexi 可有多条阶梯。"],
    ["FlexiTiers", "阶梯序号", "第几阶梯。", "每行必填", "正整数", "1", "从 1 开始递增。"],
    ["FlexiTiers", "门槛值", "达到多少件 / 金额后触发当前阶梯。", "主玩法 != FIXED_PRICE", "数值", "2", "ITEM_QUANTITY 填件数；ORDER_VALUE 填订单金额。"],
    ["FlexiTiers", "优惠值", "当前阶梯的优惠值。", "Money/Discount Off 或 Combo Buy", "数值", "5", "百分比折扣填数字，不要写 %。"],
    ["FlexiTiers", "赠品数量", "当前阶梯送几件赠品 / 样品。", "赠品类玩法", "正整数", "1", "当前 live 尚未接赠品商品选择。"],
    ["FlexiTiers", "买家可选赠品数", "允许买家从赠品池里选几件。", "赠品类玩法可选", "正整数", "1", "对应 Let buyer choose out of free gifts。"],
    ["FlexiTiers", "赠品SKU列表", "赠品 SKU，多个用逗号分隔。", "赠品类玩法可选", "逗号分隔文本", "SKU123,SKU456", "当前先固化字段，后续再接 live 选品弹层。"],
    ["FlexiTiers", "备注", "阶梯备注。", "选填", "文本", "买 2 件打 5%", "不参与页面填写。"],
]

ENUM_LISTS = {
    "Enum_SiteCodes": ["ID", "MY", "PH", "SG", "TH", "VN"],
    "Enum_ToolLabels": ["Regular Voucher", "Flexi Combo", "Store New Buyer Voucher", "Store Follower Voucher"],
    "Enum_UseTimeTypes": ["FIXED_TIME", "USE_AFTER_COLLECTION"],
    "Enum_ApplyScopes": ["ENTIRE_SHOP", "SPECIFIC_PRODUCTS"],
    "Enum_DiscountTypes": ["MONEY_OFF", "PERCENT_OFF"],
    "Enum_BudgetModes": ["LIMITED", "UNLIMITED"],
    "Enum_FlexiMainTypes": ["MONEY_DISCOUNT_OFF", "FREE_GIFT_SAMPLE", "COMBO_BUY", "FIXED_PRICE"],
    "Enum_FlexiCriteriaTypes": ["ITEM_QUANTITY", "ORDER_VALUE"],
    "Enum_BooleanFlags": ["1", "0"],
    "FlexiSubtype_MONEY_DISCOUNT_OFF": ["MoneyValueOff", "PercentageDiscountOff"],
    "FlexiSubtype_FREE_GIFT_SAMPLE": [
        "FreeGiftsOnly",
        "FreeSamplesOnly",
        "PercentOffAndFreeGift",
        "MoneyOffAndFreeGift",
        "PercentOffAndFreeSample",
        "MoneyOffAndFreeSample",
    ],
    "FlexiSubtype_COMBO_BUY": [
        "MoneyValueOff",
        "PercentageDiscountOff",
        "FreeGiftsOnly",
        "FreeSamplesOnly",
        "PercentOffAndFreeGift",
        "MoneyOffAndFreeGift",
        "PercentOffAndFreeSample",
        "MoneyOffAndFreeSample",
    ],
    "FlexiSubtype_FIXED_PRICE": [""],
}

VOUCHER_DROPDOWNS = {
    "站点": "=Enum_SiteCodes",
    "优惠工具": "=Enum_ToolLabels",
    "领取生效类型": "=Enum_UseTimeTypes",
    "适用范围": "=Enum_ApplyScopes",
    "折扣类型": "=Enum_DiscountTypes",
    "预算模式": "=Enum_BudgetModes",
    "Flexi主玩法": "=Enum_FlexiMainTypes",
    "Flexi子玩法": '=INDIRECT("FlexiSubtype_"&$U2)',
    "Flexi条件类型": "=Enum_FlexiCriteriaTypes",
    "Flexi是否叠加折扣": "=Enum_BooleanFlags",
    "Flexi赠品是否免邮": "=Enum_BooleanFlags",
}


def style_header(ws):
    fill = PatternFill("solid", fgColor="1A71FF")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="center", horizontal="center")


def autosize(ws):
    for column_cells in ws.columns:
        max_len = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[column_letter].width = min(max(max_len + 2, 12), 36)


def build_enum_sheet(wb):
    enum_ws = wb.create_sheet(ENUM_SHEET_TITLE)
    for col_idx, (range_name, values) in enumerate(ENUM_LISTS.items(), start=1):
        col_letter = get_column_letter(col_idx)
        enum_ws.cell(row=1, column=col_idx, value=range_name)
        for row_idx, value in enumerate(values, start=2):
            enum_ws.cell(row=row_idx, column=col_idx, value=value)
        end_row = max(len(values) + 1, 2)
        wb.defined_names.add(
            DefinedName(
                range_name,
                attr_text=f"'{ENUM_SHEET_TITLE}'!${col_letter}$2:${col_letter}${end_row}",
            )
        )
    enum_ws.sheet_state = "hidden"


def add_dropdowns(ws, headers, dropdowns, max_row=ENTRY_MAX_ROW):
    header_to_col = {header: idx + 1 for idx, header in enumerate(headers)}
    for header, formula in dropdowns.items():
        col_idx = header_to_col.get(header)
        if not col_idx:
            continue
        col_letter = get_column_letter(col_idx)
        dv = DataValidation(
            type="list",
            formula1=formula,
            allow_blank=True,
            showDropDown=False,
        )
        dv.promptTitle = f"{header} 可选值"
        dv.prompt = "请优先从下拉框中选择枚举值。"
        dv.errorTitle = "无效选项"
        dv.error = f"{header} 请输入下拉列表中的有效枚举值。"
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}{max_row}")


def build():
    wb = Workbook()

    ws = wb.active
    ws.title = "Vouchers"
    ws.append(VOUCHER_HEADERS)
    for row in VOUCHER_EXAMPLES:
        ws.append(row)
    ws.freeze_panes = "A2"
    style_header(ws)
    autosize(ws)

    build_enum_sheet(wb)
    add_dropdowns(ws, VOUCHER_HEADERS, VOUCHER_DROPDOWNS)

    flexi = wb.create_sheet("FlexiTiers")
    flexi.append(FLEXI_TIER_HEADERS)
    for row in FLEXI_TIER_EXAMPLES:
        flexi.append(row)
    flexi.freeze_panes = "A2"
    style_header(flexi)
    autosize(flexi)

    ins = wb.create_sheet("Instructions")
    for row in INSTRUCTION_ROWS:
        ins.append(row)
    style_header(ins)
    ins.freeze_panes = "A2"
    autosize(ins)

    guide = wb.create_sheet("填写说明")
    guide.append(FILL_GUIDE_HEADERS)
    for row in FILL_GUIDE_ROWS:
        guide.append(row)
    guide.freeze_panes = "A2"
    style_header(guide)
    autosize(guide)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(OUT)


if __name__ == "__main__":
    build()
